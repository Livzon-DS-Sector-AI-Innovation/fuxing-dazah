"""Spare part service: business logic for spare parts and stock."""

import io
import logging
import uuid
from typing import Any, cast

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import AppException, DuplicateException, NotFoundException
from app.modules.equipment import repository as repo
from app.modules.equipment.deps import EquipmentAccessContext
from app.modules.equipment.models.spare_part import (
    SparePart,
    SparePartStock,
)
from app.modules.equipment.schemas.equipment import (
    EquipmentImportResponse,
    ImportRowError,
)
from app.modules.equipment.schemas.spare_part import (
    SparePartCreate,
    SparePartUpdate,
    StockAdjustRequest,
    StockInboundRequest,
    StockWarningResponse,
)
from app.modules.equipment.service.data_scope import verify_write_ownership

logger = logging.getLogger(__name__)


async def create_spare_part(
    db: AsyncSession,
    data: SparePartCreate,
    ctx: EquipmentAccessContext,
) -> SparePart:
    """创建备件，自动创建库存记录。department_id 自动从用户部门获取。"""
    if await repo.exists_spare_part_by_code(db, data.code):
        raise DuplicateException("备件编码", data.code)

    create_data = data.model_dump()
    # 自动归属用户部门
    if not create_data.get("department_id") and ctx.visible_department_ids:
        create_data["department_id"] = ctx.visible_department_ids[0]

    spare_part = await repo.create_spare_part(db, create_data)

    await repo.create_stock(db, {"spare_part_id": spare_part.id})

    return spare_part


async def get_spare_part_by_id(
    db: AsyncSession,
    spare_part_id: uuid.UUID,
) -> SparePart:
    """获取备件"""
    spare_part = await repo.get_spare_part_by_id(db, spare_part_id)
    if not spare_part:
        raise NotFoundException("备件", str(spare_part_id))
    return spare_part


async def get_spare_parts(
    db: AsyncSession,
    ctx: EquipmentAccessContext,
    category: str | None = None,
    keyword: str | None = None,
    is_active: bool | None = None,
    department_id: uuid.UUID | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[SparePart], int]:
    """获取备件列表，批量填充 department_name。"""
    spare_parts, total = await repo.get_spare_parts(
        db,
        ctx,
        category=category,
        keyword=keyword,
        is_active=is_active,
        department_id=department_id,
        page=page,
        page_size=page_size,
    )
    # 批量填充 department_name（缓存去重，避免 API 层直接调用 repo）
    dept_cache: dict[uuid.UUID, str] = {}
    for sp in spare_parts:
        if sp.department_id:
            if sp.department_id not in dept_cache:
                dept_info = await repo.get_department_info(db, sp.department_id)
                dept_cache[sp.department_id] = dept_info["name"] if dept_info else ""
            sp.department_name = dept_cache.get(sp.department_id, "")  # type: ignore[attr-defined]
    return spare_parts, total


async def update_spare_part(
    db: AsyncSession,
    spare_part_id: uuid.UUID,
    data: SparePartUpdate,
    ctx: EquipmentAccessContext,
) -> SparePart:
    """更新备件"""
    spare_part = await get_spare_part_by_id(db, spare_part_id)
    await verify_write_ownership(ctx, spare_part, "department_id", "department_id")

    update_data = data.model_dump(exclude_unset=True)

    if "code" in update_data:
        code_exists = await repo.exists_spare_part_by_code(
            db, update_data["code"], exclude_id=spare_part_id
        )
        if code_exists:
            raise DuplicateException("备件编码", update_data["code"])

    result = await repo.update_spare_part(db, spare_part_id, update_data)
    if not result:
        raise NotFoundException("备件", str(spare_part_id))
    return result


async def delete_spare_part(
    db: AsyncSession,
    spare_part_id: uuid.UUID,
    ctx: EquipmentAccessContext,
) -> bool:
    """删除备件"""
    spare_part = await get_spare_part_by_id(db, spare_part_id)
    await verify_write_ownership(ctx, spare_part, "department_id", "department_id")
    return await repo.delete_spare_part(db, spare_part_id)


async def get_stock_by_spare_part_id(
    db: AsyncSession,
    spare_part_id: uuid.UUID,
) -> SparePartStock:
    """获取库存记录"""
    stock = await repo.get_stock_by_spare_part_id(db, spare_part_id)
    if not stock:
        raise NotFoundException("库存记录", str(spare_part_id))
    return stock


async def inbound_stock(
    db: AsyncSession,
    spare_part_id: uuid.UUID,
    data: StockInboundRequest,
    ctx: EquipmentAccessContext | None = None,
) -> SparePartStock:
    """入库"""
    spare_part = await get_spare_part_by_id(db, spare_part_id)
    if ctx:
        await verify_write_ownership(ctx, spare_part, "department_id", "department_id")

    stock = await repo.update_stock_qty(db, spare_part_id, data.quantity)
    if not stock:
        raise NotFoundException("库存记录", str(spare_part_id))

    if data.warehouse_location is not None:
        stock.warehouse_location = data.warehouse_location

    await repo.create_transaction(
        db,
        {
            "spare_part_id": spare_part_id,
            "transaction_type": "入库",
            "quantity": data.quantity,
            "remark": data.remark,
        },
    )

    await db.flush()
    await db.refresh(stock)
    return stock


async def outbound_stock(
    db: AsyncSession,
    spare_part_id: uuid.UUID,
    quantity: int,
) -> SparePartStock:
    """出库（内部使用）"""
    stock = await get_stock_by_spare_part_id(db, spare_part_id)

    if stock.current_qty < quantity:
        raise AppException(
            message=f"库存不足，当前库存 {stock.current_qty}，出库数量 {quantity}"
        )

    stock = cast(SparePartStock, await repo.update_stock_qty(db, spare_part_id, -quantity))
    if not stock:
        raise NotFoundException("库存记录", str(spare_part_id))

    await repo.create_transaction(
        db,
        {
            "spare_part_id": spare_part_id,
            "transaction_type": "出库",
            "quantity": -quantity,
        },
    )

    await db.flush()
    await db.refresh(stock)
    return stock


async def adjust_stock(
    db: AsyncSession,
    spare_part_id: uuid.UUID,
    data: StockAdjustRequest,
    ctx: EquipmentAccessContext | None = None,
) -> SparePartStock:
    """盘点调整"""
    spare_part = await get_spare_part_by_id(db, spare_part_id)
    if ctx:
        await verify_write_ownership(ctx, spare_part, "department_id", "department_id")

    stock = await get_stock_by_spare_part_id(db, spare_part_id)

    diff = data.new_qty - stock.current_qty

    if diff != 0:
        stock = cast(SparePartStock, await repo.update_stock_qty(db, spare_part_id, diff))
        if not stock:
            raise NotFoundException("库存记录", str(spare_part_id))

        await repo.create_transaction(
            db,
            {
                "spare_part_id": spare_part_id,
                "transaction_type": "盘点调整",
                "quantity": diff,
                "remark": data.remark,
            },
        )

    await db.flush()
    await db.refresh(stock)
    return stock


async def get_stock_warnings(
    db: AsyncSession,
) -> list[StockWarningResponse]:
    """获取库存预警"""
    warnings = await repo.get_stock_warnings(db)
    result = []
    for spare_part, stock in warnings:
        result.append(
            StockWarningResponse(
                spare_part=spare_part,  # pyright: ignore[reportArgumentType]
                stock=stock,  # pyright: ignore[reportArgumentType]
                shortage=stock.safety_qty - stock.current_qty,
            )
        )
    return result


# ==================== Excel 导入 ====================

# 模板列（0-based）
SP_COL_CODE = 0          # A: 备件编码 *
SP_COL_NAME = 1           # B: 备件名称 *
SP_COL_SPECIFICATION = 2  # C: 规格型号
SP_COL_UNIT = 3           # D: 计量单位 *
SP_COL_CATEGORY = 4       # E: 备件分类
SP_COL_DEFAULT_SUPPLIER = 5  # F: 默认供应商
SP_COL_UNIT_PRICE = 6     # G: 参考单价（元）

SP_TEMPLATE_HEADERS = [
    ("备件编码 *", 18),       # A
    ("备件名称 *", 24),       # B
    ("规格型号", 22),         # C
    ("计量单位 *", 12),       # D
    ("备件分类", 16),         # E
    ("默认供应商", 22),       # F
    ("参考单价（元）", 16),    # G
]


def _sp_cell_str(row: tuple[Any, ...], col: int) -> str | None:
    val = row[col] if col < len(row) else None
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _parse_decimal(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def generate_spare_part_template_bytes() -> io.BytesIO:
    """生成备件导入模板 Excel 文件的字节流"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "备件管理"

    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws.freeze_panes = "A3"

    # 表头（第1行）
    for i, (label, width) in enumerate(SP_TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = width

    # 示例数据（第2行）
    examples = [
        "SP-001",
        "轴承密封圈",
        "Φ50×Φ34×7",
        "个",
        "机械密封",
        "XX密封件有限公司",
        "85.50",
    ]
    ws.row_dimensions[2].height = 28
    for i, example in enumerate(examples, start=1):
        cell = ws.cell(row=2, column=i, value=example)
        cell.fill = example_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        cell.font = Font(name="微软雅黑", size=9, italic=True, color="808080")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def import_spare_parts_from_excel(
    db: AsyncSession,
    file_bytes: bytes,
    ctx: EquipmentAccessContext,
) -> EquipmentImportResponse:
    """从 Excel 文件字节流导入备件"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))

    if "备件管理" not in wb.sheetnames:
        raise ValueError("Excel 中缺少「备件管理」工作表，请使用模板文件")

    ws = wb["备件管理"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    imported = 0
    skipped = 0
    errors: list[ImportRowError] = []
    warnings: list[ImportRowError] = []

    # ── 预加载系统数据 ──
    existing_result = await db.execute(
        select(SparePart.code).where(SparePart.is_deleted == False)  # noqa: E712
    )
    existing_codes: set[str] = {r[0] for r in existing_result.all()}

    # 确定导入目标部门
    if ctx.visible_department_ids:
        default_dept_id = ctx.visible_department_ids[0]
    else:
        default_dept_id = None

    for row_num, row in enumerate(rows, start=3):
        try:
            values = [_sp_cell_str(row, c) for c in range(7)]
            if not any(values):
                continue

            code = values[SP_COL_CODE]
            name = values[SP_COL_NAME]
            unit = values[SP_COL_UNIT]

            # 必填校验
            missing = []
            if not code:
                missing.append("备件编码")
            if not name:
                missing.append("备件名称")
            if not unit:
                missing.append("计量单位")
            if missing:
                warnings.append(
                    ImportRowError(
                        row=row_num,
                        message=f"缺少必填项: {', '.join(missing)}，已跳过",
                    )
                )
                skipped += 1
                continue
            assert code is not None and name is not None and unit is not None
            
            # 编码重复
            if code in existing_codes:
                errors.append(
                    ImportRowError(
                        row=row_num,
                        message=f"备件编码「{code}」已存在",
                    )
                )
                skipped += 1
                continue

            # 部门权限校验
            department_id = default_dept_id
            if not ctx.is_unrestricted:
                if not department_id:
                    errors.append(
                        ImportRowError(
                            row=row_num,
                            message="无法确定归属部门，导入失败",
                        )
                    )
                    skipped += 1
                    continue

            # 解析单价
            raw_price = row[SP_COL_UNIT_PRICE] if SP_COL_UNIT_PRICE < len(row) else None
            unit_price = _parse_decimal(raw_price)

            # 创建备件（使用 SAVEPOINT 隔离每行）
            async with db.begin_nested():
                spare_part = SparePart(
                    code=code,
                    name=name,
                    specification=values[SP_COL_SPECIFICATION],
                    unit=unit,
                    category=values[SP_COL_CATEGORY],
                    default_supplier=values[SP_COL_DEFAULT_SUPPLIER],
                    unit_price=unit_price,
                    is_active=True,
                    department_id=department_id,
                    created_by=ctx.user.id if ctx else None,
                    updated_by=ctx.user.id if ctx else None,
                )
                db.add(spare_part)
                await db.flush()

                # 创建库存记录
                db.add(SparePartStock(spare_part_id=spare_part.id))
                await db.flush()

                # 记录入库流水
                await repo.create_transaction(
                    db,
                    {
                        "spare_part_id": spare_part.id,
                        "transaction_type": "入库",
                        "quantity": 0,
                        "remark": "Excel导入创建",
                    },
                )

            existing_codes.add(code)
            imported += 1

        except Exception as e:
            errors.append(
                ImportRowError(
                    row=row_num,
                    message=f"导入异常: {e}",
                )
            )
            skipped += 1
            logger.warning("备件导入行 %d 失败: %s", row_num, e)

    return EquipmentImportResponse(
        imported=imported,
        skipped=skipped,
        errors=errors,
        warnings=warnings,
    )
