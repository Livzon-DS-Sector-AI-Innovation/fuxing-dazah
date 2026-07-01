"""Equipment service layer: business logic, validation, transaction orchestration."""

import io
import logging
import uuid
from datetime import date, datetime
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import AppException, DuplicateException, NotFoundException
from app.modules.equipment import repository as repo
from app.modules.equipment.deps import EquipmentAccessContext
from app.modules.equipment.models import Equipment, EquipmentCategory, Location
from app.modules.equipment.models.equipment import EquipmentCategoryLink
from app.modules.equipment.schemas import (
    EquipmentCategoryCreate,
    EquipmentCategoryUpdate,
    EquipmentCreate,
    EquipmentUpdate,
    LocationCreate,
    LocationUpdate,
)
from app.modules.equipment.schemas.equipment import (
    EquipmentImportResponse,
    ImportRowError,
)
from app.modules.equipment.service.data_scope import verify_write_ownership
from app.platform.identity.models import Department, User

logger = logging.getLogger(__name__)


# ==================== 设备分类 ====================
async def create_equipment_category(
    db: AsyncSession,
    data: EquipmentCategoryCreate,
    ctx: EquipmentAccessContext,
) -> EquipmentCategory:
    """创建设备分类，自动绑定部门"""
    # 获取用户所属部门 ID（用于绑定和去重校验）
    department_id = (
        ctx.visible_department_ids[0] if ctx.visible_department_ids else None
    )

    # 检查编码是否重复（部门范围内）
    if await repo.exists_category_by_code(db, data.code, department_id=department_id):
        raise DuplicateException("分类代码", data.code)

    category_data = data.model_dump()
    # 自动绑定部门
    if department_id:
        category_data["department_id"] = department_id

    return await repo.create_equipment_category(db, category_data)


async def get_equipment_category_by_id(
    db: AsyncSession,
    category_id: uuid.UUID,
) -> EquipmentCategory:
    """获取设备分类"""
    category = await repo.get_equipment_category_by_id(db, category_id)
    if not category:
        raise NotFoundException("设备分类", str(category_id))
    return category


async def get_equipment_categories(
    db: AsyncSession,
    parent_id: uuid.UUID | None = None,
    ctx: EquipmentAccessContext | None = None,
) -> list[EquipmentCategory]:
    """获取设备分类列表"""
    return await repo.get_equipment_categories(db, parent_id, ctx)


async def get_equipment_category_tree(
    db: AsyncSession,
    ctx: EquipmentAccessContext | None = None,
) -> list[EquipmentCategory]:
    """获取设备分类树形结构"""
    return await repo.get_equipment_category_tree(db, ctx)


async def update_equipment_category(
    db: AsyncSession,
    category_id: uuid.UUID,
    data: EquipmentCategoryUpdate,
    ctx: EquipmentAccessContext,
) -> EquipmentCategory:
    """更新设备分类"""
    if data.code is not None:
        # 获取该分类的部门 ID，用于同部门范围内去重
        existing = await repo.get_equipment_category_by_id(db, category_id)
        department_id = existing.department_id if existing else None
        if await repo.exists_category_by_code(
            db, data.code, exclude_id=category_id, department_id=department_id
        ):
            raise DuplicateException("分类代码", data.code)

    category = await repo.update_equipment_category(
        db, category_id, data.model_dump(exclude_unset=True)
    )
    if not category:
        raise NotFoundException("设备分类", str(category_id))
    return category


async def delete_equipment_category(
    db: AsyncSession,
    category_id: uuid.UUID,
    ctx: EquipmentAccessContext,
) -> bool:
    """删除设备分类"""
    await get_equipment_category_by_id(db, category_id)

    children = await repo.get_equipment_categories(db, parent_id=category_id, ctx=ctx)
    if children:
        raise AppException(message="该分类下存在子分类，无法删除")

    equipment_count = await repo.count_equipments_by_category(db, category_id)
    if equipment_count > 0:
        raise AppException(message="该分类下存在关联设备，无法删除")

    return await repo.delete_equipment_category(db, category_id)


# ==================== 位置管理 ====================
async def create_location(
    db: AsyncSession,
    data: LocationCreate,
    ctx: EquipmentAccessContext,
) -> Location:
    """创建位置，自动绑定部门"""
    # 获取用户所属部门 ID（用于绑定和去重校验）
    department_id = (
        ctx.visible_department_ids[0] if ctx.visible_department_ids else None
    )

    # 检查编码是否重复（部门范围内）
    if await repo.exists_location_by_code(db, data.code, department_id=department_id):
        raise DuplicateException("位置代码", data.code)

    location_data = data.model_dump()
    # 自动绑定部门
    if department_id:
        location_data["department_id"] = department_id

    return await repo.create_location(db, location_data)


async def get_location_by_id(
    db: AsyncSession,
    location_id: uuid.UUID,
) -> Location:
    """获取位置"""
    location = await repo.get_location_by_id(db, location_id)
    if not location:
        raise NotFoundException("位置", str(location_id))
    return location


async def get_locations(
    db: AsyncSession,
    parent_id: uuid.UUID | None = None,
    ctx: EquipmentAccessContext | None = None,
) -> list[Location]:
    """获取位置列表"""
    return await repo.get_locations(db, parent_id, ctx)


async def get_location_tree(
    db: AsyncSession,
    ctx: EquipmentAccessContext | None = None,
) -> list[Location]:
    """获取位置树形结构"""
    return await repo.get_location_tree(db, ctx)


async def update_location(
    db: AsyncSession,
    location_id: uuid.UUID,
    data: LocationUpdate,
    ctx: EquipmentAccessContext,
) -> Location:
    """更新位置"""
    if data.code is not None:
        # 获取该位置的部门 ID，用于同部门范围内去重
        existing = await repo.get_location_by_id(db, location_id)
        department_id = existing.department_id if existing else None
        if await repo.exists_location_by_code(
            db, data.code, exclude_id=location_id, department_id=department_id
        ):
            raise DuplicateException("位置代码", data.code)

    location = await repo.update_location(
        db, location_id, data.model_dump(exclude_unset=True)
    )
    if not location:
        raise NotFoundException("位置", str(location_id))
    return location


async def delete_location(
    db: AsyncSession,
    location_id: uuid.UUID,
    ctx: EquipmentAccessContext,
) -> bool:
    """删除位置"""
    await get_location_by_id(db, location_id)

    children = await repo.get_locations(db, parent_id=location_id, ctx=ctx)
    if children:
        raise AppException(message="该位置下存在子位置，无法删除")

    equipment_count = await repo.count_equipments_by_location(db, location_id)
    if equipment_count > 0:
        raise AppException(message="该位置下存在关联设备，无法删除")

    return await repo.delete_location(db, location_id)


# ==================== 设备管理 ====================
async def generate_equipment_no(
    db: AsyncSession,
    category_code: str,
) -> str:
    """生成设备编号"""
    max_no = await repo.get_max_equipment_no_by_category(db, category_code)
    if max_no:
        # 提取序号部分
        seq_str = max_no.split("-")[-1]
        seq = int(seq_str) + 1
    else:
        seq = 1
    return f"EQ-{category_code}-{seq:04d}"


async def create_equipment(
    db: AsyncSession,
    data: EquipmentCreate,
) -> Equipment:
    """创建设备"""
    # 校验编号唯一性
    existing = await repo.get_equipment_by_no(db, data.equipment_no)
    if existing:
        raise DuplicateException("设备编号", data.equipment_no)

    # 验证所有分类
    for cid in data.category_ids:
        await get_equipment_category_by_id(db, cid)
    await get_location_by_id(db, data.location_id)

    equipment_data = data.model_dump()

    try:
        return await repo.create_equipment(db, equipment_data)
    except IntegrityError:
        raise DuplicateException("设备编号", data.equipment_no)


async def get_equipment_by_id(
    db: AsyncSession,
    equipment_id: uuid.UUID,
) -> Equipment:
    """获取设备"""
    equipment = await repo.get_equipment_by_id(db, equipment_id)
    if not equipment:
        raise NotFoundException("设备", str(equipment_id))
    return equipment


async def get_equipments(
    db: AsyncSession,
    ctx: EquipmentAccessContext,
    category_id: uuid.UUID | None = None,
    location_id: uuid.UUID | None = None,
    department_id: uuid.UUID | None = None,
    status: str | None = None,
    keyword: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[Equipment], int]:
    """获取设备列表"""
    return await repo.get_equipments(
        db, ctx, category_id, location_id,
        department_id, status, keyword, page, page_size,
    )


async def update_equipment(
    db: AsyncSession,
    equipment_id: uuid.UUID,
    data: EquipmentUpdate,
    ctx: EquipmentAccessContext,
) -> Equipment:
    """更新设备"""
    equipment = await get_equipment_by_id(db, equipment_id)
    await verify_write_ownership(ctx, equipment, "department_id", "department_id")

    if data.category_ids is not None:
        for cid in data.category_ids:
            await get_equipment_category_by_id(db, cid)
    if data.location_id is not None:
        await get_location_by_id(db, data.location_id)

    update_data = data.model_dump(exclude_unset=True)
    return await repo.update_equipment(db, equipment_id, update_data)


async def delete_equipment(
    db: AsyncSession,
    equipment_id: uuid.UUID,
    ctx: EquipmentAccessContext,
) -> bool:
    """删除设备"""
    equipment = await get_equipment_by_id(db, equipment_id)
    await verify_write_ownership(ctx, equipment, "department_id", "department_id")
    equipment.is_deleted = True
    await db.flush()
    return True


async def get_equipment_statistics(
    db: AsyncSession, ctx: EquipmentAccessContext,
) -> dict[str, Any]:
    """获取设备统计（按数据范围过滤）"""
    return await repo.get_equipment_statistics(db, ctx)


async def get_departments_for_select(db: AsyncSession) -> list[dict[str, Any]]:
    """获取可选部门列表（含负责人），供下拉使用"""
    return await repo.get_departments_for_select(db)


# ==================== Excel 导入 ====================

# 模板列（0-based），按台账表格显示顺序
COL_EQUIPMENT_NO = 0
COL_NAME = 1
COL_CATEGORY_NAME = 2
COL_LOCATION_NAME = 3
COL_DEPARTMENT_NAME = 4
COL_RESPONSIBLE_PERSON = 5
COL_STATUS = 6
COL_IMPORTANCE = 7
COL_MODEL = 8
COL_SPECIFICATION = 9
COL_MANUFACTURER = 10
COL_SUPPLIER = 11
COL_PRODUCTION_DATE = 12
COL_COMMISSIONING_DATE = 13
COL_WARRANTY_EXPIRE = 14
COL_ASSET_VALUE = 15
COL_DEPRECIATION_YEARS = 16
COL_DESCRIPTION = 17

VALID_STATUSES = {"在用", "备用", "维修中", "停用", "报废"}
VALID_IMPORTANCE = {"高", "中", "低"}

# 模板表头
TEMPLATE_HEADERS = [
    ("设备编号 *", 20),     # A
    ("设备名称 *", 24),     # B
    ("设备分类 *", 18),     # C
    ("设备位置 *", 18),     # D
    ("归属部门 *", 18),     # E
    ("负责人 *", 14),       # F
    ("设备状态", 10),       # G
    ("重要性", 8),           # H
    ("型号", 20),            # I
    ("规格", 22),            # J
    ("制造商", 22),          # K
    ("供应商", 22),          # L
    ("出厂日期", 14),        # M
    ("投用日期", 14),        # N
    ("保修到期日", 14),      # O
    ("资产原值（元）", 16),  # P
    ("折旧年限", 10),        # Q
    ("设备描述", 30),        # R
]


def _cell_str(row, col: int) -> str | None:
    val = row[col] if col < len(row) else None
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def generate_template_bytes() -> io.BytesIO:
    """生成导入模板 Excel 文件的字节流"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备台账"

    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws.freeze_panes = "A3"

    # 表头（第1行）
    for i, (label, width) in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = width

    # 示例数据（第2行）
    examples = [
        "EQ-001", "离心泵-A-01", "离心泵", "A车间一楼", "动力部", "张三",
        "在用", "中", "IHF65-50-160", "流量25m³/h 扬程32m",
        "XX泵业有限公司", "XX机电设备公司", "2024-01-15", "2024-03-01",
        "2026-01-15", "15000", "10", "用于XX工序物料输送",
    ]
    ws.row_dimensions[2].height = 28
    for i, example in enumerate(examples, start=1):
        cell = ws.cell(row=2, column=i, value=example)
        cell.fill = example_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        cell.font = Font(name="微软雅黑", size=9, italic=True, color="808080")

    # 数据验证 — 设备状态
    dv_status = DataValidation(
        type="list",
        formula1='"在用,备用,维修中,停用,报废"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="无效状态",
        error="请选择：在用 / 备用 / 维修中 / 停用 / 报废",
    )
    dv_status.add("G3:G5002")
    ws.add_data_validation(dv_status)

    # 数据验证 — 重要性
    dv_importance = DataValidation(
        type="list",
        formula1='"高,中,低"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="无效重要性",
        error="请选择：高 / 中 / 低",
    )
    dv_importance.add("H3:H5002")
    ws.add_data_validation(dv_importance)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def import_equipments_from_excel(
    db: AsyncSession,
    file_bytes: bytes,
) -> EquipmentImportResponse:
    """从 Excel 文件字节流导入设备"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))

    if "设备台账" not in wb.sheetnames:
        raise ValueError("Excel 中缺少「设备台账」工作表，请使用模板文件")

    ws = wb["设备台账"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # 第1行表头，第2行示例，第3行起数据

    imported = 0
    skipped = 0
    errors: list[ImportRowError] = []
    warnings: list[ImportRowError] = []

    # ── 预加载系统数据 ──
    # 已有设备编号
    existing_nos_result = await db.execute(
        select(Equipment.equipment_no).where(
            Equipment.is_deleted == False,  # noqa: E712
        )
    )
    existing_nos: set[str] = {r[0] for r in existing_nos_result.all()}

    # 分类名称 → (id, name)
    cat_result = await db.execute(
        select(EquipmentCategory.id, EquipmentCategory.name).where(
            EquipmentCategory.is_deleted == False,  # noqa: E712
        )
    )
    cat_map: dict[str, uuid.UUID] = {name: cid for cid, name in cat_result.all()}

    # 位置名称 → id
    loc_result = await db.execute(
        select(Location.id, Location.name).where(
            Location.is_deleted == False,  # noqa: E712
        )
    )
    loc_map: dict[str, uuid.UUID] = {name: lid for lid, name in loc_result.all()}

    # 部门名称 → id
    dept_result = await db.execute(
        select(Department.id, Department.name).where(
            Department.is_deleted == False,  # noqa: E712
            Department.status_is_deleted.isnot(True),
        )
    )
    dept_map: dict[str, uuid.UUID] = {name: did for did, name in dept_result.all()}

    # 用户姓名 → id
    user_result = await db.execute(
        select(User.id, User.name).where(
            User.is_deleted == False,  # noqa: E712
        )
    )
    user_map: dict[str, uuid.UUID] = {name: uid for uid, name in user_result.all()}

    for row_num, row in enumerate(rows, start=3):
        try:
            # 空行跳过
            values = [_cell_str(row, c) for c in range(18)]
            if not any(values):
                continue

            eq_no = values[COL_EQUIPMENT_NO]
            name = values[COL_NAME]
            cat_name = values[COL_CATEGORY_NAME]
            loc_name = values[COL_LOCATION_NAME]
            dept_name = values[COL_DEPARTMENT_NAME]
            person_name = values[COL_RESPONSIBLE_PERSON]

            # 必填校验：缺少任意必填字段 → 跳过
            missing = []
            if not eq_no:
                missing.append("设备编号")
            if not name:
                missing.append("设备名称")
            if not cat_name:
                missing.append("设备分类")
            if not loc_name:
                missing.append("设备位置")
            if not dept_name:
                missing.append("归属部门")
            if not person_name:
                missing.append("负责人")
            if missing:
                warnings.append(ImportRowError(
                    row=row_num,
                    message=f"缺少必填项: {', '.join(missing)}，已跳过",
                ))
                skipped += 1
                continue

            assert eq_no and name and cat_name and loc_name and dept_name and person_name

            # 设备编号重复
            if eq_no in existing_nos:
                errors.append(ImportRowError(
                    row=row_num,
                    message=f"设备编号「{eq_no}」已存在",
                ))
                skipped += 1
                continue

            # 分类不存在
            category_id = cat_map.get(cat_name)
            if not category_id:
                errors.append(ImportRowError(
                    row=row_num,
                    message=f"设备分类「{cat_name}」在系统中不存在",
                ))
                skipped += 1
                continue

            # 位置不存在
            location_id = loc_map.get(loc_name)
            if not location_id:
                errors.append(ImportRowError(
                    row=row_num,
                    message=f"设备位置「{loc_name}」在系统中不存在",
                ))
                skipped += 1
                continue

            # 部门不存在
            department_id = dept_map.get(dept_name)
            if not department_id:
                errors.append(ImportRowError(
                    row=row_num,
                    message=f"归属部门「{dept_name}」在系统中不存在",
                ))
                skipped += 1
                continue

            # 负责人不存在
            responsible_person_id = user_map.get(person_name)
            if not responsible_person_id:
                errors.append(ImportRowError(
                    row=row_num,
                    message=f"负责人「{person_name}」在系统中不存在",
                ))
                skipped += 1
                continue

            # 校验状态
            status = values[COL_STATUS] or "在用"
            if status not in VALID_STATUSES:
                warnings.append(ImportRowError(
                    row=row_num,
                    message=f"设备状态「{status}」无效，已默认设为「在用」",
                ))
                status = "在用"

            # 校验重要性
            importance = values[COL_IMPORTANCE] or "低"
            if importance not in VALID_IMPORTANCE:
                warnings.append(ImportRowError(
                    row=row_num,
                    message=f"重要性「{importance}」无效，已默认设为「低」",
                ))
                importance = "低"

            # 创建装备
            raw_pd = row[COL_PRODUCTION_DATE] if COL_PRODUCTION_DATE < len(row) else None
            raw_cd = row[COL_COMMISSIONING_DATE] if COL_COMMISSIONING_DATE < len(row) else None
            raw_we = row[COL_WARRANTY_EXPIRE] if COL_WARRANTY_EXPIRE < len(row) else None
            raw_av = row[COL_ASSET_VALUE] if COL_ASSET_VALUE < len(row) else None
            raw_dy = row[COL_DEPRECIATION_YEARS] if COL_DEPRECIATION_YEARS < len(row) else None

            equipment = Equipment(
                equipment_no=eq_no,
                name=name,
                location_id=location_id,
                status=status,
                importance=importance,
                model=values[COL_MODEL],
                specification=values[COL_SPECIFICATION],
                manufacturer=values[COL_MANUFACTURER],
                supplier=values[COL_SUPPLIER],
                production_date=_parse_date(raw_pd),
                commissioning_date=_parse_date(raw_cd),
                description=values[COL_DESCRIPTION],
                warranty_expire_date=_parse_date(raw_we),
                asset_value=_parse_float(raw_av),
                depreciation_years=_parse_int(raw_dy),
                department_id=department_id,
                responsible_person_id=responsible_person_id,
            )
            db.add(equipment)
            await db.flush()

            # 关联分类
            db.add(EquipmentCategoryLink(
                equipment_id=equipment.id,
                category_id=category_id,
            ))
            await db.flush()

            existing_nos.add(eq_no)
            imported += 1

        except Exception as e:
            errors.append(ImportRowError(
                row=row_num,
                message=f"导入异常: {e}",
            ))
            skipped += 1
            await db.rollback()
            logger.warning("导入行 %d 失败: %s", row_num, e)

    return EquipmentImportResponse(
        imported=imported,
        skipped=skipped,
        errors=errors,
        warnings=warnings,
    )


def _parse_float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _parse_int(value) -> int | None:
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None
