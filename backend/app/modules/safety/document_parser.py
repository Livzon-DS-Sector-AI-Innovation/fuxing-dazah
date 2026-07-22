"""安全模块文档解析工具 — Markdown 提取 + Chapter 7 工艺阶段解析。

从 app/platform/integrations/ai/document_parser.py 迁移过来，
避免安全模块直接修改平台层代码。
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

logger = logging.getLogger(__name__)


def extract_to_markdown(file_path: str, max_chars: int = 50000) -> str:
    """Extract content as Markdown-formatted text.

    - XLSX/XLS:  tab-separated rows → Markdown table (| col | col |)
    - DOCX:      paragraphs with heading detection
    - PDF/TXT/MD: same as extract_text
    """
    from app.platform.integrations.ai.document_parser import DocumentParser

    ext = Path(file_path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        text = _extract_xlsx_to_markdown(file_path)
    elif ext == ".docx":
        text = _extract_docx_to_markdown(file_path)
    else:
        text = DocumentParser.extract_text(file_path, max_chars=999999)

    if len(text) > max_chars:
        text = text[:max_chars] + "\n\n...（文档过长，已截断）"
    return text.strip()


def _extract_xlsx_to_markdown(path: str) -> str:
    """Extract Excel content as Markdown tables (one table per sheet)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    parts: list[str] = []
    # 限制：最多 10 个 sheet，每个 sheet 最多 500 行，避免 prompt 过长
    for sheet_name in wb.sheetnames[:10]:
        ws = wb[sheet_name]
        parts.append(f"## 📊 Sheet: {sheet_name}\n")
        rows: list[list[str]] = []
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(cells)
                row_count += 1
                if row_count >= 500:
                    parts.append("*(表格超过 500 行，已截断)*\n")
                    break
        if not rows:
            parts.append("*(空表格)*\n")
            continue
        # 构建 Markdown 表格
        max_cols = max(len(r) for r in rows)
        # 表头 = 第一行
        header = rows[0]
        # 补齐列数
        header += [""] * (max_cols - len(header))
        md_rows: list[str] = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(["---"] * max_cols) + " |",
        ]
        for row in rows[1:]:
            row += [""] * (max_cols - len(row))
            md_rows.append("| " + " | ".join(row) + " |")
        parts.append("\n".join(md_rows) + "\n")
    wb.close()
    return "\n\n".join(parts)


def _extract_docx_to_markdown(path: str) -> str:
    """Extract DOCX content with basic heading detection."""
    from docx import Document

    doc = Document(path)
    lines: list[str] = []
    for p in doc.paragraphs:
        text = p.text.strip()
        if not text:
            continue
        # 检测标题样式
        if p.style and p.style.name and p.style.name.startswith("Heading"):
            level = p.style.name.replace("Heading", "").strip()
            try:
                lv = int(level)
            except ValueError:
                lv = 1
            lines.append(f"{'#' * min(lv, 4)} {text}")
        else:
            lines.append(text)
    return "\n\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════
# Chapter 7 工艺阶段解析（供批量危险源辨识使用）
# ═══════════════════════════════════════════════════════════════════════════


def get_chapter7_from_content(content: str) -> str | None:
    """从完整 9 章操规 Markdown 中提取第 7 章内容。

    定位 `# 7.` 或 `## 7.` 标题，返回该章全部内容直至下一章（`# 8.`）。
    """
    # 匹配 Chapter 7 开头（# 7. 或 ## 7. 生产工艺流程）
    ch7_start = re.search(r"^#{1,2}\s*7\.\s*生产工艺流程", content, re.MULTILINE)
    if not ch7_start:
        return None

    start_pos = ch7_start.start()

    # 匹配 Chapter 8 开头作为结束标记
    ch8_match = re.search(r"^#{1,2}\s*8\.", content[start_pos:], re.MULTILINE)
    if ch8_match:
        end_pos = start_pos + ch8_match.start()
        return content[start_pos:end_pos].strip()

    # 没有 Chapter 8 → 取到内容末尾
    return content[start_pos:].strip()


def parse_chapter7_stages(content: str) -> list[dict]:
    """解析 Chapter 7（生产工艺流程）中的工艺阶段。

    每个 `## 阶段名称` 是一个工艺阶段，可包含：
    - `### 安全要求` → 编号项列表
    - `### 操作步骤` → 编号项列表

    Args:
        content: 完整操规 Markdown（9 章）或仅 Chapter 7 内容

    Returns:
        [
            {
                "stage_name": "进料",
                "safety_items": ["1. 佩戴防毒面具", ...],
                "operation_items": ["1. 开启进料阀", ...],
                "markdown": "## 进料\\n### 安全要求\\n..."
            },
            ...
        ]
        无 Chapter 7 或无 H2 阶段时返回空列表
    """
    ch7 = get_chapter7_from_content(content)
    if not ch7:
        logger.debug("未找到 Chapter 7 生产工艺流程")
        return []

    stages: list[dict] = []

    # 按 ## 分割为各工艺阶段（跳过 Chapter 7 标题行本身）
    # 匹配 ## 标题行
    stage_pattern = re.compile(r"^##\s+(.+)$", re.MULTILINE)
    stage_matches = list(stage_pattern.finditer(ch7))

    if not stage_matches:
        logger.debug("Chapter 7 中未找到工艺阶段 (## H2)")
        return []

    for i, match in enumerate(stage_matches):
        stage_name = match.group(1).strip()
        if not stage_name:
            continue

        # 该阶段的内容范围：从当前 ## 标题到下一个 ## 标题
        section_start = match.start()
        section_end = (
            stage_matches[i + 1].start() if i + 1 < len(stage_matches) else len(ch7)
        )
        section_text = ch7[section_start:section_end].strip()

        # 解析安全要求子节
        safety_items = _parse_numbered_subsection(section_text, "安全要求")

        # 解析操作步骤子节
        operation_items = _parse_numbered_subsection(section_text, "操作步骤")

        stages.append({
            "stage_name": stage_name,
            "safety_items": safety_items,
            "operation_items": operation_items,
            "markdown": section_text,
        })

    logger.info("解析 Chapter 7: %d 个工艺阶段", len(stages))
    return stages


def _parse_numbered_subsection(section_text: str, subsection_name: str) -> list[str]:
    """从 Markdown 节选中提取 `### subsection_name` 下的编号项。

    匹配 `### 安全要求` 或 `### 操作步骤` 下方以 `1.` / `2.` / `-` 等开头的项。
    """
    # 匹配 ### subsection_name 到下一个 ### 或文本末尾
    sub_pattern = re.compile(
        r"^###\s+" + re.escape(subsection_name) + r"\s*\n(.*?)(?=\n###\s|\Z)",
        re.MULTILINE | re.DOTALL,
    )
    sub_match = sub_pattern.search(section_text)
    if not sub_match:
        return []

    sub_content = sub_match.group(1)

    # 提取编号项：以数字+点号开头（如 "1. xxx"）或 "-" 开头
    items: list[str] = []
    for line in sub_content.strip().split("\n"):
        stripped = line.strip()
        if not stripped:
            continue
        # 匹配: 1. / 2. / (1) / - 等
        if re.match(r"^(\d+[\.\、\)]\s*|[-•]\s+)", stripped):
            items.append(stripped)

    return items
