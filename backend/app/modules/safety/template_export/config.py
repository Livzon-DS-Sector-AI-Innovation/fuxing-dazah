"""
Excel 标准化输出 — 模板配置与自动检测。

定义模板结构描述符、列映射、风险颜色规则、页面设置。
提供 ``TemplateInspector`` 自动检测任意 Excel 模板的结构，
只需一个模板文件即可生成 ``TemplateConfig``。
"""

from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path

# ═══════════════════════════════════════════════════════════════════════════════
# 默认列 → DB 字段映射（危险源辨识模板）
# ═══════════════════════════════════════════════════════════════════════════════
# Keys are lowercase column letters; values are DB field names.
# Fields that don't exist in DB yet are mapped to "" and left blank.

HAZARD_COLUMN_MAPPING: dict[str, str] = {
    "a": "hazard_id_no",
    "b": "position",
    "c": "production_step",
    "d": "specific_activity",
    "e": "unsafe_behavior",
    "f": "possible_accident",
    "g": "existing_engineering_controls",
    "h": "existing_management_controls",
    "i": "",   # 培训教育措施·现有 — no DB field yet
    "j": "existing_ppe",
    "k": "existing_emergency_measures",
    "l": "l_inherent",
    "m": "e_inherent",
    "n": "d_inherent",
    "o": "inherent_risk_label",
    "p": "recommendation_content",
    "q": "recommendation_content",
    "r": "recommendation_content",
    "s": "recommendation_content",
    "t": "recommendation_content",
    "u": "control_level",
    "v": "department",
    "w": "responsible_person",
    "x": "",   # 检查频次 — no DB field yet
}

# Columns whose values should be cast to float (risk scores)
NUMERIC_COLUMNS = frozenset({"l", "m", "n"})


# ═══════════════════════════════════════════════════════════════════════════════
# 风险等级 → 字体颜色
# ═══════════════════════════════════════════════════════════════════════════════

RISK_LABEL_COLORS: dict[str, str] = {
    "一级": "FF0000",   # red
    "重大": "FF0000",
    "level_1": "FF0000",
    "二级": "FF8C00",   # orange
    "较大": "FF8C00",
    "level_2": "FF8C00",
    "三级": "0000FF",   # blue
    "一般": "0000FF",
    "level_3": "0000FF",
    "四级": "008000",   # green
    "level_4": "008000",
}


# ═══════════════════════════════════════════════════════════════════════════════
# 页面 / 打印设置
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class PageSetup:
    """openpyxl 页面设置，影响 PDF 导出效果。"""

    orientation: str = "landscape"   # "portrait" | "landscape"
    paper_size: int = 8             # 8=A3 (297×420mm), 9=A4, 13=B4, …
    fit_to_width: int = 1           # 0=不缩放, 1=适应1页宽
    fit_to_height: int = 0          # 0=自动（按需分页）


# ═══════════════════════════════════════════════════════════════════════════════
# 模板结构描述符
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class TemplateConfig:
    """描述 Excel 模板的结构，驱动填充引擎。

    所有行号均为 1-based（与 Excel 行号一致）。

    通用字段说明
    -------------
    title_row:
        ``***`` 占位符所在行号。设为 0 则跳过标题替换。
    header_row_count:
        表头行数（标题行 + 签名行 + 列标行等），数据从下一行开始。
    sample_row:
        样式来源行 — 所有数据行的字体/填充/对齐/边框克隆自此行。
    total_columns:
        总列数（A=1 到 X=24）。
    title_placeholder:
        标题中需要替换的文本片段，如 ``"***"``。替换为空字符串则不替换。
    title_resolver:
        从数据列表中提取标题替换文本的回调函数。
        默认用出现最多的 ``department`` 值。不需要标题替换时设为 ``None``。
    column_mapping:
        ``{列字母: DB字段名}`` 映射。DB 字段为 ``""`` 的列留空。
    numeric_columns:
        需要转为 float 的列字母集合。
    risk_label_column:
        需要风险着色的列字母（如 ``"o"``），不需要时设为 ``""``。
    risk_label_colors:
        风险标签关键字 → RRGGBB 颜色映射。
    page_setup:
        PDF 导出的页面设置（方向/纸张/缩放）。
    sequence_column:
        序号列（1=A, 2=B, …），自动填入 1, 2, 3…。设为 0 禁用。
    """

    # ── 结构行（1-based） ──
    title_row: int = 1
    header_row_count: int = 6
    sample_row: int = 7
    total_columns: int = 24

    # ── 标题占位符 ──
    title_placeholder: str = "***"
    title_resolver: Callable[[list[dict]], str] | None = None

    # ── 列映射 ──
    column_mapping: dict[str, str] = field(
        default_factory=lambda: HAZARD_COLUMN_MAPPING
    )
    numeric_columns: frozenset = NUMERIC_COLUMNS

    # ── 风险着色 ──
    risk_label_column: str = "o"
    risk_label_colors: dict[str, str] = field(
        default_factory=lambda: RISK_LABEL_COLORS
    )

    # ── 页面设置 ──
    page_setup: PageSetup = field(default_factory=PageSetup)

    # ── 序号列 ──
    sequence_column: int = 1   # 1=A, 0=disabled


# ═══════════════════════════════════════════════════════════════════════════════
# 预置配置
# ═══════════════════════════════════════════════════════════════════════════════

def _default_title_resolver(data: list[dict]) -> str:
    """从记录中取出现最多的部门名作为标题替换文本。"""
    from collections import Counter
    depts = [r.get("department", "") for r in data if r.get("department")]
    if not depts:
        return ""
    return Counter(depts).most_common(1)[0][0]


HAZARD_TEMPLATE_CONFIG = TemplateConfig(
    title_row=1,
    header_row_count=6,
    sample_row=7,
    total_columns=24,
    title_placeholder="***",
    title_resolver=_default_title_resolver,
    column_mapping=HAZARD_COLUMN_MAPPING,
    numeric_columns=NUMERIC_COLUMNS,
    risk_label_column="o",
    risk_label_colors=RISK_LABEL_COLORS,
    page_setup=PageSetup(
        orientation="landscape",
        paper_size=8,       # A3
        fit_to_width=1,
        fit_to_height=0,
    ),
    sequence_column=1,
)


# ═══════════════════════════════════════════════════════════════════════════════
# 模板自动检测器
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class InspectionResult:
    """``TemplateInspector.inspect()`` 的返回结果。"""

    config: TemplateConfig
    merged_cells: list[dict]        # 合并单元格信息
    has_column_letters_row: bool    # 是否存在 a, b, c, … 列字母行
    column_letters_row: int         # 列字母行号，0 表示不存在
    title_cell_ref: str             # 标题单元格引用（如 "A1"）
    empty_data_rows: int            # 模板中预留给数据的空行数


class TemplateInspector:
    """自动检测 Excel 模板结构并生成 ``TemplateConfig``。

    检测策略（按优先级）：
    1. **样本行**：最后一个同时有内容和格式（边框）的行
    2. **表头行数**：样本行之前的所有行
    3. **总列数**：样本行的最大非空列号
    4. **列字母行**：是否有一行的内容为 ``a, b, c, …``
    5. **标题占位符**：首行是否包含 ``***``
    6. **合并单元格**：模板中所有合并区域

    Usage::

        inspector = TemplateInspector()
        result = inspector.inspect("模板.xlsx")
        config = result.config
        # 按需微调 config，然后传给 ExcelTemplateFiller
    """

    # 检测列字母行的阈值：一行中连续单字母单元格 >= 这个数就判定为列字母行
    _COLUMN_LETTER_THRESHOLD = 10

    # 常见列字母序列
    _COLUMN_LETTERS = tuple(
        chr(ord("a") + i) for i in range(26)
    )  # a-z
    _COLUMN_LETTERS_UPPER = tuple(chr(ord("A") + i) for i in range(26))

    def inspect(
        self,
        template_path: str | Path,
        *,
        column_mapping: dict[str, str] | None = None,
    ) -> InspectionResult:
        """检测模板并返回 ``InspectionResult``。

        *column_mapping* 可选 — 提供则优先使用，否则生成空映射。
        """
        import openpyxl

        wb = openpyxl.load_workbook(str(template_path))
        ws = wb.active

        total_cols = self._detect_total_columns(ws)
        sample_row = self._detect_sample_row(ws, total_cols)
        col_letters_row, has_col_letters = self._detect_column_letters_row(
            ws, total_cols, sample_row
        )

        # 表头行 = 样本行之上的所有行
        header_count = sample_row - 1

        # 标题行检测
        title_row, title_placeholder = self._detect_title(ws, header_count)

        # 合并单元格
        merged = self._capture_merged_cells(ws)

        # 生成列映射
        if column_mapping is None:
            column_mapping = self._build_empty_mapping(total_cols)

        config = TemplateConfig(
            title_row=title_row,
            header_row_count=header_count,
            sample_row=sample_row,
            total_columns=total_cols,
            title_placeholder=title_placeholder,
            title_resolver=_default_title_resolver if title_placeholder else None,
            column_mapping=column_mapping,
            numeric_columns=frozenset(),   # 调用方自行设置
            risk_label_column="",          # 调用方自行设置
            risk_label_colors=RISK_LABEL_COLORS,
            page_setup=PageSetup(),        # 调用方自行设置
            sequence_column=1,
        )

        wb.close()
        return InspectionResult(
            config=config,
            merged_cells=merged,
            has_column_letters_row=has_col_letters,
            column_letters_row=col_letters_row,
            title_cell_ref=f"{openpyxl.utils.get_column_letter(1)}{title_row}"
            if title_row else "",
            empty_data_rows=0,  # 需调用方根据实际需求设置
        )

    # ── 检测方法 ──────────────────────────────────────────────────────────

    @staticmethod
    def _detect_total_columns(ws) -> int:
        """扫描所有行找出最大非空列号。"""
        max_col = 0
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row or 50, 50),
            max_col=min(ws.max_column or 30, 100),
        ):
            for cell in row:
                if cell.value is not None:
                    max_col = max(max_col, cell.column)
        return max(1, max_col)

    @staticmethod
    def _detect_sample_row(ws, total_cols: int) -> int:
        """找最后一个同时有内容和边框的「格式行」作为样本行。

        从下往上扫描，第一行满足「至少 3 列有边框且有空列有内容」即为样本行。
        """
        from openpyxl.styles import Border

        no_border = Border()
        max_row = min(ws.max_row or 100, 100)

        for row_idx in range(max_row, 0, -1):
            has_content = False
            border_count = 0
            for c in range(1, total_cols + 1):
                cell = ws.cell(row_idx, c)
                if cell.value is not None:
                    has_content = True
                if cell.border and cell.border != no_border:
                    # 检查是否有任一边框线
                    b = cell.border
                    if any(
                        getattr(b, side, None) and getattr(b, side).style
                        for side in ("left", "right", "top", "bottom")
                    ):
                        border_count += 1

            if has_content and border_count >= 3:
                return row_idx

        # 降级：返回最后有内容的行
        return max_row

    @classmethod
    def _detect_column_letters_row(
        cls, ws, total_cols: int, sample_row: int
    ) -> tuple[int, bool]:
        """检测样本行上方是否存在列字母行（a, b, c, … 或 A, B, C, …）。"""
        for row_idx in range(sample_row - 1, 0, -1):
            letters = []
            for c in range(1, min(total_cols + 1, 27)):
                val = ws.cell(row_idx, c).value
                if isinstance(val, str):
                    letters.append(val.strip().lower())
            if len(letters) >= cls._COLUMN_LETTER_THRESHOLD:
                # 检查是否是连续字母序列
                matches = 0
                expected_idx = 0
                for letter in letters:
                    exp_lower = cls._COLUMN_LETTERS[expected_idx] if expected_idx < 26 else ""
                    exp_upper = cls._COLUMN_LETTERS_UPPER[expected_idx] if expected_idx < 26 else ""
                    if letter == exp_lower or letter == exp_upper:
                        matches += 1
                        expected_idx += 1
                    elif expected_idx > 0:
                        break
                if matches >= cls._COLUMN_LETTER_THRESHOLD:
                    return row_idx, True
        return 0, False

    @staticmethod
    def _detect_title(ws, header_count: int) -> tuple[int, str]:
        """检测标题行：在表头区域查找包含 ``***`` 的行。"""
        for row_idx in range(1, header_count + 1):
            for c in range(1, min(ws.max_column or 10, 10)):
                val = ws.cell(row_idx, c).value
                if isinstance(val, str) and "***" in val:
                    return row_idx, "***"
        # 无占位符 —— 首行作为标题行但不做替换
        return 1, ""

    @staticmethod
    def _capture_merged_cells(ws) -> list[dict]:
        """捕获所有合并单元格的范围信息。"""
        result = []
        for merged_range in ws.merged_cells.ranges:
            result.append({
                "range": str(merged_range),
                "min_row": merged_range.min_row,
                "max_row": merged_range.max_row,
                "min_col": merged_range.min_col,
                "max_col": merged_range.max_col,
            })
        return result

    @staticmethod
    def _build_empty_mapping(total_cols: int) -> dict[str, str]:
        """根据列数生成空的列映射。"""
        from openpyxl.utils import get_column_letter

        return {
            get_column_letter(c).lower(): "" for c in range(1, total_cols + 1)
        }

    # ── 便捷方法 ──────────────────────────────────────────────────────────

    def build_config(
        self,
        template_path: str | Path,
        *,
        column_mapping: dict[str, str] | None = None,
        **overrides,
    ) -> TemplateConfig:
        """检测模板并返回可直接使用的 ``TemplateConfig``。

        ``overrides`` 可以覆盖自动检测的任何字段：
        ``title_row``, ``header_row_count``, ``sample_row``, ``total_columns``,
        ``numeric_columns``, ``risk_label_column``, ``risk_label_colors``,
        ``page_setup``, ``sequence_column``, ``title_resolver``。
        """
        result = self.inspect(template_path, column_mapping=column_mapping)
        config = result.config

        # 应用覆盖
        for key, value in overrides.items():
            if hasattr(config, key):
                setattr(config, key, value)

        # PageSetup 特殊处理
        if "page_setup" in overrides:
            ps = overrides["page_setup"]
            if isinstance(ps, dict):
                current = config.page_setup
                for k, v in ps.items():
                    if hasattr(current, k):
                        setattr(current, k, v)

        return config
