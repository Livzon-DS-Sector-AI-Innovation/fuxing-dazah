"""Meter HTTP API — 路由、入参、依赖注入、响应."""

from __future__ import annotations

import logging
from datetime import date
from typing import Any
from uuid import UUID

from fastapi import Depends, File, Form, Query, Request, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.exceptions import NotFoundException
from app.core.response import paginated_response, success_response
from app.modules.meter import repository as repo
from app.modules.meter import service
from app.modules.meter.ai_service import extract_and_update_date, get_meter_ai_config
from app.modules.meter.schemas import (
    BatchCreateRequest,
    BatchCreateResult,
    BatchDeleteRequest,
    BatchExtractRequest,
    BatchExtractResponse,
    DateStatsResponse,
    DepartmentCreate,
    DepartmentResponse,
    DepartmentUpdate,
    ExportReportRequest,
    ExtractDateResponse,
    FileMatchItem,
    FileMatchRequest,
    GasDetectorBatchCreateRequest,
    GasDetectorCreate,
    GasDetectorFilter,
    GasDetectorFilterOptions,
    GasDetectorListResponse,
    GasDetectorResponse,
    GasDetectorUpdate,
    InstrumentCreate,
    InstrumentFilter,
    InstrumentFilterOptions,
    InstrumentListResponse,
    InstrumentResponse,
    InstrumentUpdate,
    LedgerImportResult,
    MeterOverviewResponse,
    MeterSettingsResponse,
    MeterSettingsUpdate,
    PersonnelCandidate,
    ReportItem,
    ReportResponse,
    _normalize_department,
)
from app.shared.module_api import create_module_router
from app.shared.module_registry import MODULES_BY_CODE

logger = logging.getLogger(__name__)

# ── 为所有 meter API 添加请求日志 ──

async def _log_meter_request(request: Request) -> None:
    """记录每个 meter API 请求的方法和路径，便于排查问题。"""
    logger.info(
        "[meter] %s %s | client=%s",
        request.method, request.url.path,
        request.client.host if request.client else "unknown",
    )

router = create_module_router(MODULES_BY_CODE["meter"])
# 注入请求日志依赖：所有 meter 路由执行前会自动调用 _log_meter_request
router.dependencies.append(Depends(_log_meter_request))


@router.get("/overview", summary="仪表总览统计")
async def get_meter_overview(
    source: str = Query(default="instrument", pattern="^(instrument|gas_detector)$", description="数据源"),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    stats = await service.get_meter_overview(db, source)
    return success_response(MeterOverviewResponse(**stats).model_dump(mode="json"))


# ═══════════════════════════════════════════
# 标准计量器具
# ═══════════════════════════════════════════


@router.get("/instruments", summary="标准计量器具列表")
async def list_instruments(
    filters: InstrumentFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    records, total = await service.list_instruments(db, filters)

    # 批量查询报告数量
    report_counts = await repo.count_reports_by_instrument_ids(db, [r.id for r in records])

    items: list[dict[str, Any]] = []
    for r in records:
        anomaly = r.anomaly_flags or {}
        items.append(
            InstrumentListResponse(
                id=str(r.id),
                department=r.department,
                asset_number=r.asset_number,
                instrument_name=r.instrument_name,
                model_spec=r.model_spec,
                measurement_range=r.measurement_range,
                accuracy_grade=r.accuracy_grade,
                serial_number=r.serial_number,
                calibration_cycle_months=r.calibration_cycle_months,
                color_marking=r.color_marking,
                location=r.location,
                manufacturer=r.manufacturer,
                status=service.compute_status(r.status, r.next_calibration_date),
                calibration_date=r.calibration_date,
                calibration_unit=r.calibration_unit,
                calibration_result=r.calibration_result,
                next_calibration_date=r.next_calibration_date,
                has_anomaly=bool(anomaly),
                report_count=report_counts.get(r.id, 0),
                remark=r.remark,
                updated_at=r.updated_at,
            ).model_dump(mode="json")
        )

    return paginated_response(
        data=items,
        page=filters.page,
        page_size=filters.page_size,
        total=total,
    )


@router.get("/instruments/filter-options", summary="获取标准计量器具筛选选项")
async def get_instrument_filter_options(
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    options = await service.get_instrument_filter_options(db)
    if "department" in options:
        options["department"] = [d for d in (_normalize_department(x) for x in options["department"]) if d is not None]
    return success_response(InstrumentFilterOptions(**options).model_dump(mode="json"))


@router.get("/instruments/date-stats", summary="标准计量器具日期聚合统计")
async def get_instrument_date_stats(
    field: str = Query(default="calibration_date", pattern="^(calibration_date|next_calibration_date)$", description="统计的日期字段"),
    filters: InstrumentFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    stats = await service.get_instrument_date_stats(db, filters, field)
    return success_response(DateStatsResponse(**stats).model_dump(mode="json"))


@router.get("/instruments/export", summary="导出标准计量器具为 CSV")
async def export_instruments(
    filters: InstrumentFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    import csv
    import io

    # 导出全量数据（最多 20000 条），绕过 page_size 上限
    object.__setattr__(filters, 'page_size', 20000)
    records, _ = await service.list_instruments(db, filters)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "资产编号", "器具名称", "型号规格", "测量范围", "精度等级",
        "出厂编号", "检定周期(月)", "使用地点", "制造商", "状态",
        "检定日期", "检定单位", "检定结论", "下次检定日期", "部门",
    ])
    for r in records:
        writer.writerow([
            r.asset_number, r.instrument_name, r.model_spec, r.measurement_range,
            r.accuracy_grade, r.serial_number, r.calibration_cycle_months,
            r.location, r.manufacturer, service.compute_status(r.status, r.next_calibration_date),
            r.calibration_date.isoformat() if r.calibration_date else "",
            r.calibration_unit, r.calibration_result,
            r.next_calibration_date.isoformat() if r.next_calibration_date else "",
            _normalize_department(r.department),
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=instruments_export.csv"},
    )


@router.get("/instruments/export-excel", summary="导出标准计量器具为 Excel")
async def export_instruments_excel(
    filters: InstrumentFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    import io as io_mod

    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.styles import (  # type: ignore[import-untyped]
        Alignment,
        Font,
        PatternFill,
    )

    # 导出全量数据，绕过 page_size 上限
    object.__setattr__(filters, 'page_size', 100000)
    records, _ = await service.list_instruments(db, filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "标准计量器具台账"

    headers = [
        "资产编号", "器具名称", "型号规格", "测量范围", "精度等级",
        "出厂编号", "检定周期(月)", "使用地点", "制造商", "状态",
        "检定日期", "检定单位", "检定结论", "下次检定日期", "部门",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, r in enumerate(records, 2):
        values = [
            r.asset_number, r.instrument_name, r.model_spec, r.measurement_range,
            r.accuracy_grade, r.serial_number, r.calibration_cycle_months,
            r.location, r.manufacturer, service.compute_status(r.status, r.next_calibration_date),
            r.calibration_date.isoformat() if r.calibration_date else "",
            r.calibration_unit, r.calibration_result,
            r.next_calibration_date.isoformat() if r.next_calibration_date else "",
            _normalize_department(r.department),
        ]
        for col_idx, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

    buf = io_mod.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=instruments_export.xlsx"},
    )


@router.get("/instruments/{instrument_id}", summary="标准计量器具详情")
async def get_instrument(
    instrument_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    obj = await service.get_instrument(db, instrument_id)
    reports = _build_report_items(obj.reports if hasattr(obj, 'reports') else [])
    return success_response(
        InstrumentResponse(
            id=str(obj.id),
            asset_number=obj.asset_number,
            instrument_name=obj.instrument_name,
            model_spec=obj.model_spec,
            measurement_range=obj.measurement_range,
            accuracy_grade=obj.accuracy_grade,
            serial_number=obj.serial_number,
            calibration_cycle_months=obj.calibration_cycle_months,
            location=obj.location,
            manufacturer=obj.manufacturer,
            status=service.compute_status(obj.status, obj.next_calibration_date),
            color_marking=obj.color_marking,
            calibration_date=obj.calibration_date,
            calibration_unit=obj.calibration_unit,
            calibration_result=obj.calibration_result,
            next_calibration_date=obj.next_calibration_date,
            department=obj.department,
            sheet_name=obj.sheet_name,
            anomaly_flags=obj.anomaly_flags,
            is_deleted=obj.is_deleted,
            created_at=obj.created_at,
            updated_at=obj.updated_at,
            reports=reports,
        ).model_dump(mode="json")
    )


@router.post("/instruments", summary="新增标准计量器具")
async def create_instrument(
    data: InstrumentCreate,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    record = await service.create_instrument(db, data)
    return success_response(
        InstrumentResponse(
            id=str(record.id),
            asset_number=record.asset_number,
            instrument_name=record.instrument_name,
            model_spec=record.model_spec,
            measurement_range=record.measurement_range,
            accuracy_grade=record.accuracy_grade,
            serial_number=record.serial_number,
            calibration_cycle_months=record.calibration_cycle_months,
            location=record.location,
            manufacturer=record.manufacturer,
            status=service.compute_status(record.status, record.next_calibration_date),
            color_marking=record.color_marking,
            calibration_date=record.calibration_date,
            calibration_unit=record.calibration_unit,
            calibration_result=record.calibration_result,
            next_calibration_date=record.next_calibration_date,
            department=record.department,
            sheet_name=record.sheet_name,
            anomaly_flags=record.anomaly_flags,
            is_deleted=record.is_deleted,
            created_at=record.created_at,
            updated_at=record.updated_at,
            reports=[],
        ).model_dump(mode="json"),
        status_code=201,
    )


@router.post("/instruments/batch", summary="批量新增标准计量器具（单次最多 200 条）")
async def batch_create_instruments(
    body: dict[str, Any],
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    req = BatchCreateRequest(**body)
    result = await service.batch_create_instruments(
        db, [item.model_dump() for item in req.items]
    )
    return success_response(
        BatchCreateResult(**result).model_dump(mode="json"),
        status_code=201,
    )


@router.post("/gas-detectors/batch", summary="批量新增有毒有害可燃探测器（单次最多 200 条）")
async def batch_create_gas_detectors(
    body: dict[str, Any],
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    req = GasDetectorBatchCreateRequest(**body)
    result = await service.batch_create_gas_detectors(
        db, [item.model_dump() for item in req.items]
    )
    return success_response(
        BatchCreateResult(**result).model_dump(mode="json"),
        status_code=201,
    )


# ═══════════════════════════════════════════
# Excel 台账导入
# ═══════════════════════════════════════════


@router.post("/instruments/import-ledger", summary="导入标准计量器具台账Excel（全量替换）")
async def import_instrument_ledger(
    file: UploadFile = File(..., description="Excel 文件 (.et 或 .xlsx)"),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    """上传计量器具台账 Excel，全量替换现有标准计量器具数据。

    支持 .et (WPS) 和 .xlsx 格式，文件限制 50MB。
    处理所有 sheet（跳过探测器 sheet）。
    """
    max_size = 50 * 1024 * 1024
    filename = file.filename or "unknown"

    try:
        file_data = await file.read()
        if len(file_data) > max_size:
            return JSONResponse(
                status_code=400,
                content={"code": 400, "message": "文件大小超过 50MB 限制"},
            )

        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in ("et", "xlsx", "xls"):
            return JSONResponse(
                status_code=400,
                content={"code": 400, "message": "不支持的文件格式，请上传 .et 或 .xlsx 文件"},
            )

        result = await service.import_instrument_ledger(db, file_data, filename)
    except ValueError as e:
        return JSONResponse(
            status_code=400,
            content={"code": 400, "message": str(e)},
        )
    except Exception as e:
        logger.exception("import_instrument_ledger failed")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"导入失败: {e}"},
        )

    return success_response(
        LedgerImportResult(**result).model_dump(mode="json"),
        status_code=201,
    )


@router.post("/gas-detectors/import-ledger", summary="导入有毒有害探测器台账Excel（全量替换）")
async def import_gas_detector_ledger(
    file: UploadFile = File(..., description="Excel 文件 (.et 或 .xlsx)"),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    """上传探测器台账 Excel，全量替换现有探测器数据。

    仅处理探测器 sheet（含"可燃""有毒""探测器"等关键词的 sheet 或 Sheet 0）。
    支持 .et (WPS) 和 .xlsx 格式，文件限制 50MB。
    """
    max_size = 50 * 1024 * 1024
    filename = file.filename or "unknown"

    try:
        file_data = await file.read()
        if len(file_data) > max_size:
            return JSONResponse(
                status_code=400,
                content={"code": 400, "message": "文件大小超过 50MB 限制"},
            )

        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in ("et", "xlsx", "xls"):
            return JSONResponse(
                status_code=400,
                content={"code": 400, "message": "不支持的文件格式，请上传 .et 或 .xlsx 文件"},
            )

        result = await service.import_gas_detector_ledger(db, file_data, filename)
    except ValueError as e:
        return JSONResponse(
            status_code=400,
            content={"code": 400, "message": str(e)},
        )
    except Exception as e:
        logger.exception("import_gas_detector_ledger failed")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"导入失败: {e}"},
        )

    return success_response(
        LedgerImportResult(**result).model_dump(mode="json"),
        status_code=201,
    )


@router.put("/instruments/{instrument_id}", summary="更新标准计量器具")
async def update_instrument(
    instrument_id: UUID,
    data: InstrumentUpdate,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    record = await service.update_instrument(db, instrument_id, data)
    reports = _build_report_items(record.reports if hasattr(record, 'reports') else [])
    return success_response(
        InstrumentResponse(
            id=str(record.id),
            asset_number=record.asset_number,
            instrument_name=record.instrument_name,
            model_spec=record.model_spec,
            measurement_range=record.measurement_range,
            accuracy_grade=record.accuracy_grade,
            serial_number=record.serial_number,
            calibration_cycle_months=record.calibration_cycle_months,
            location=record.location,
            manufacturer=record.manufacturer,
            status=service.compute_status(record.status, record.next_calibration_date),
            color_marking=record.color_marking,
            calibration_date=record.calibration_date,
            calibration_unit=record.calibration_unit,
            calibration_result=record.calibration_result,
            next_calibration_date=record.next_calibration_date,
            department=record.department,
            sheet_name=record.sheet_name,
            anomaly_flags=record.anomaly_flags,
            is_deleted=record.is_deleted,
            created_at=record.created_at,
            updated_at=record.updated_at,
            reports=reports,
        ).model_dump(mode="json")
    )


@router.delete("/instruments/{instrument_id}", summary="删除标准计量器具（软删除）")
async def delete_instrument(
    instrument_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    await service.delete_instrument(db, instrument_id)
    return success_response(message="删除成功")


@router.post("/instruments/batch-delete", summary="批量删除标准计量器具（软删除）")
async def batch_delete_instruments(
    body: BatchDeleteRequest,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    ids = [UUID(i) for i in body.ids]
    deleted_count = await service.batch_delete_instruments(db, ids)
    return success_response({"deleted_count": deleted_count}, message=f"成功删除 {deleted_count} 条记录")


@router.get("/instruments/ids", summary="获取筛选条件下所有标准计量器具 ID（用于跨页全选）")
async def get_instrument_ids(
    filters: InstrumentFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    ids = await service.get_all_instrument_ids(db, filters)
    return success_response([str(i) for i in ids])


@router.post("/instruments/export-reports", summary="批量导出标准计量器具最新报告 ZIP（单次最多 200 份）")
async def export_instrument_reports(
    body: ExportReportRequest,
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    ids = [UUID(i) for i in body.ids]
    zip_data, filename, count = await service.export_instrument_reports(db, ids)
    return StreamingResponse(
        iter([zip_data]),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Report-Count": str(count),
        },
    )


@router.get("/departments/instruments", summary="获取标准计量器具部门列表")
async def list_instrument_departments(
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    depts = await service.get_instrument_departments(db)
    return success_response([_normalize_department(d) for d in depts])


# ═══════════════════════════════════════════
# 有毒有害可燃探测器
# ═══════════════════════════════════════════


@router.get("/gas-detectors", summary="有毒有害可燃探测器列表")
async def list_gas_detectors(
    filters: GasDetectorFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    records, total = await service.list_gas_detectors(db, filters)

    # 批量查询报告数量
    report_counts = await repo.count_reports_by_gas_detector_ids(db, [r.id for r in records])

    items: list[dict[str, Any]] = []
    for r in records:
        anomaly = r.anomaly_flags or {}
        items.append(
            GasDetectorListResponse(
                id=str(r.id),
                department=r.department,
                instrument_name=r.instrument_name,
                detection_model=r.detection_model,
                measurement_range=r.measurement_range,
                product_number=r.product_number,
                installation_type=r.installation_type,
                installation_location=r.installation_location,
                medium=r.medium,
                calibration_factor=r.calibration_factor,
                manufacturer_supplier=r.manufacturer_supplier,
                manufacturer=r.manufacturer,
                status=service.compute_status(r.status, r.next_calibration_date),
                calibration_date=r.calibration_date,
                next_calibration_date=r.next_calibration_date,
                detection_unit=r.detection_unit,
                calibration_result=r.calibration_result,
                has_anomaly=bool(anomaly),
                report_count=report_counts.get(r.id, 0),
                remark=r.remark,
                updated_at=r.updated_at,
            ).model_dump(mode="json")
        )

    return paginated_response(
        data=items,
        page=filters.page,
        page_size=filters.page_size,
        total=total,
    )


@router.get("/gas-detectors/filter-options", summary="获取有毒有害可燃探测器筛选选项")
async def get_gas_detector_filter_options(
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    options = await service.get_gas_detector_filter_options(db)
    if "department" in options:
        options["department"] = [d for d in (_normalize_department(x) for x in options["department"]) if d is not None]
    return success_response(GasDetectorFilterOptions(**options).model_dump(mode="json"))


@router.get("/gas-detectors/date-stats", summary="有毒有害可燃探测器日期聚合统计")
async def get_gas_detector_date_stats(
    field: str = Query(default="calibration_date", pattern="^(calibration_date|next_calibration_date)$", description="统计的日期字段"),
    filters: GasDetectorFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    stats = await service.get_gas_detector_date_stats(db, filters, field)
    return success_response(DateStatsResponse(**stats).model_dump(mode="json"))


@router.get("/gas-detectors/export-excel", summary="导出有毒有害可燃探测器为 Excel")
async def export_gas_detectors_excel(
    filters: GasDetectorFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    import io as io_mod

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    # 导出全量数据，绕过 page_size 上限
    object.__setattr__(filters, 'page_size', 100000)
    records, _ = await service.list_gas_detectors(db, filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "有毒有害可燃探测器台账"

    headers = [
        "部门", "器具名称", "检测型号", "量程", "产品编号",
        "安装方式", "安装位置", "使用介质", "标定系数", "制造商/供应商",
        "检定时间", "检测单位", "下次检定时间", "检定结论", "制造单位",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, r in enumerate(records, 2):
        values = [
            _normalize_department(r.department), r.instrument_name, r.detection_model,
            r.measurement_range, r.product_number, r.installation_type,
            r.installation_location, r.medium, r.calibration_factor,
            r.manufacturer_supplier, r.calibration_date.isoformat() if r.calibration_date else "",
            r.detection_unit, r.next_calibration_date.isoformat() if r.next_calibration_date else "",
            r.calibration_result, r.manufacturer,
        ]
        for col_idx, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

    buf = io_mod.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=gas_detectors_export.xlsx"},
    )


@router.get("/gas-detectors/{detector_id}", summary="有毒有害可燃探测器详情")
async def get_gas_detector(
    detector_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    obj = await service.get_gas_detector(db, detector_id)
    reports = _build_report_items(obj.reports if hasattr(obj, 'reports') else [])
    return success_response(
        GasDetectorResponse(
            id=str(obj.id),
            instrument_name=obj.instrument_name,
            detection_model=obj.detection_model,
            measurement_range=obj.measurement_range,
            product_number=obj.product_number,
            installation_type=obj.installation_type,
            installation_location=obj.installation_location,
            medium=obj.medium,
            calibration_factor=obj.calibration_factor,
            manufacturer_supplier=obj.manufacturer_supplier,
            calibration_date=obj.calibration_date,
            calibration_result=obj.calibration_result,
            detection_unit=obj.detection_unit,
            next_calibration_date=obj.next_calibration_date,
            manufacturer=obj.manufacturer,
            status=service.compute_status(obj.status, obj.next_calibration_date),
            department=obj.department,
            sheet_name=obj.sheet_name,
            anomaly_flags=obj.anomaly_flags,
            is_deleted=obj.is_deleted,
            created_at=obj.created_at,
            updated_at=obj.updated_at,
            reports=reports,
        ).model_dump(mode="json")
    )


@router.post("/gas-detectors", summary="新增有毒有害可燃探测器")
async def create_gas_detector(
    data: GasDetectorCreate,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    record = await service.create_gas_detector(db, data)
    return success_response(
        GasDetectorResponse(
            id=str(record.id),
            instrument_name=record.instrument_name,
            detection_model=record.detection_model,
            measurement_range=record.measurement_range,
            product_number=record.product_number,
            installation_type=record.installation_type,
            installation_location=record.installation_location,
            medium=record.medium,
            calibration_factor=record.calibration_factor,
            manufacturer_supplier=record.manufacturer_supplier,
            calibration_date=record.calibration_date,
            calibration_result=record.calibration_result,
            detection_unit=record.detection_unit,
            next_calibration_date=record.next_calibration_date,
            manufacturer=record.manufacturer,
            status=service.compute_status(record.status, record.next_calibration_date),
            department=record.department,
            sheet_name=record.sheet_name,
            anomaly_flags=record.anomaly_flags,
            is_deleted=record.is_deleted,
            created_at=record.created_at,
            updated_at=record.updated_at,
            reports=[],
        ).model_dump(mode="json"),
        status_code=201,
    )


@router.put("/gas-detectors/{detector_id}", summary="更新有毒有害可燃探测器")
async def update_gas_detector(
    detector_id: UUID,
    data: GasDetectorUpdate,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    record = await service.update_gas_detector(db, detector_id, data)
    reports = _build_report_items(record.reports if hasattr(record, 'reports') else [])
    return success_response(
        GasDetectorResponse(
            id=str(record.id),
            instrument_name=record.instrument_name,
            detection_model=record.detection_model,
            measurement_range=record.measurement_range,
            product_number=record.product_number,
            installation_type=record.installation_type,
            installation_location=record.installation_location,
            medium=record.medium,
            calibration_factor=record.calibration_factor,
            manufacturer_supplier=record.manufacturer_supplier,
            calibration_date=record.calibration_date,
            calibration_result=record.calibration_result,
            detection_unit=record.detection_unit,
            next_calibration_date=record.next_calibration_date,
            manufacturer=record.manufacturer,
            status=service.compute_status(record.status, record.next_calibration_date),
            department=record.department,
            sheet_name=record.sheet_name,
            anomaly_flags=record.anomaly_flags,
            is_deleted=record.is_deleted,
            created_at=record.created_at,
            updated_at=record.updated_at,
            reports=reports,
        ).model_dump(mode="json")
    )


@router.delete("/gas-detectors/{detector_id}", summary="删除有毒有害可燃探测器（软删除）")
async def delete_gas_detector(
    detector_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    await service.delete_gas_detector(db, detector_id)
    return success_response(message="删除成功")


@router.post("/gas-detectors/batch-delete", summary="批量删除有毒有害可燃探测器（软删除）")
async def batch_delete_gas_detectors(
    body: BatchDeleteRequest,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    ids = [UUID(i) for i in body.ids]
    deleted_count = await service.batch_delete_gas_detectors(db, ids)
    return success_response({"deleted_count": deleted_count}, message=f"成功删除 {deleted_count} 条记录")


@router.get("/gas-detectors/ids", summary="获取筛选条件下所有探测器 ID（用于跨页全选）")
async def get_gas_detector_ids(
    filters: GasDetectorFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    ids = await service.get_all_gas_detector_ids(db, filters)
    return success_response([str(i) for i in ids])


@router.post("/gas-detectors/export-reports", summary="批量导出探测器最新报告 ZIP（单次最多 200 份）")
async def export_gas_detector_reports(
    body: ExportReportRequest,
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    ids = [UUID(i) for i in body.ids]
    zip_data, filename, count = await service.export_gas_detector_reports(db, ids)
    return StreamingResponse(
        iter([zip_data]),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Report-Count": str(count),
        },
    )


@router.get("/departments/gas-detectors", summary="获取有毒有害可燃探测器部门列表")
async def list_gas_detector_departments(
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    depts = await service.get_gas_detector_departments(db)
    return success_response([_normalize_department(d) for d in depts])


# ═══════════════════════════════════════════
# 检测报告
# ═══════════════════════════════════════════


@router.post("/reports", summary="上传检测报告")
async def upload_report(
    file: UploadFile = File(..., description="检测报告文件（最大50MB）"),
    instrument_id: UUID | None = Form(default=None, description="标准计量器具 ID"),
    gas_detector_id: UUID | None = Form(default=None, description="探测器 ID"),
    report_date: date | None = Form(default=None, description="报告日期"),
    remark: str | None = Form(default=None, description="备注"),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    report = await service.upload_report(
        db,
        file=file,
        instrument_id=instrument_id,
        gas_detector_id=gas_detector_id,
        report_date=report_date,
        remark=remark,
    )
    return success_response(
        ReportResponse(
            id=str(report.id),
            instrument_id=str(report.instrument_id) if report.instrument_id else None,
            gas_detector_id=str(report.gas_detector_id) if report.gas_detector_id else None,
            file_name=report.file_name,
            file_size=report.file_size,
            content_type=report.content_type,
            report_date=report.report_date,
            remark=report.remark,
            uploaded_at=report.created_at,
        ).model_dump(mode="json"),
        status_code=201,
    )


@router.get("/reports/{report_id}", summary="获取检测报告元数据")
async def get_report(
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    report = await service.get_report(db, report_id)
    return success_response(
        ReportResponse(
            id=str(report.id),
            instrument_id=str(report.instrument_id) if report.instrument_id else None,
            gas_detector_id=str(report.gas_detector_id) if report.gas_detector_id else None,
            file_name=report.file_name,
            file_size=report.file_size,
            content_type=report.content_type,
            report_date=report.report_date,
            remark=report.remark,
            uploaded_at=report.created_at,
        ).model_dump(mode="json")
    )


@router.get("/reports/{report_id}/download", summary="下载检测报告文件")
async def download_report(
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    report = await service.get_report(db, report_id)
    result = await service.download_report_data(report)
    if result is None:
        raise NotFoundException("检测报告文件", str(report_id))
    data, content_type = result
    filename = report.file_name.encode("ascii", "ignore").decode() or "report"
    return StreamingResponse(
        iter([data]),
        media_type=content_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(data)),
        },
    )


@router.get("/reports/{report_id}/preview", summary="在线预览检测报告文件")
async def preview_report(
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    report = await service.get_report(db, report_id)
    result = await service.download_report_data(report)
    if result is None:
        raise NotFoundException("检测报告文件", str(report_id))
    data, content_type = result
    filename = report.file_name.encode("ascii", "ignore").decode() or "report"
    return StreamingResponse(
        iter([data]),
        media_type=content_type,
        headers={
            "Content-Disposition": f'inline; filename="{filename}"',
            "Content-Length": str(len(data)),
        },
    )


@router.delete("/reports/{report_id}", summary="删除检测报告（软删除）")
async def delete_report(
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    await service.delete_report(db, report_id)
    return success_response(message="删除成功")


@router.get("/instruments/{instrument_id}/reports", summary="获取标准计量器具的报告列表")
async def list_instrument_reports(
    instrument_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    reports = await service.list_instrument_reports(db, instrument_id)
    items = _build_report_items(reports)
    return success_response(items)


@router.get("/gas-detectors/{detector_id}/reports", summary="获取探测器的报告列表")
async def list_gas_detector_reports(
    detector_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    reports = await service.list_gas_detector_reports(db, detector_id)
    items = _build_report_items(reports)
    return success_response(items)


@router.post("/reports/match", summary="批量匹配文件名到仪表（单次最多 200 个文件）")
async def match_files(
    body: FileMatchRequest,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    results = await service.match_filenames(db, body.filenames)
    items = [FileMatchItem(**r).model_dump(mode="json") for r in results]
    return success_response(items)


@router.post("/reports/batch", summary="批量上传检测报告（单次最多 200 份）")
async def batch_upload_reports(
    files: list[UploadFile] = File(..., max_length=200, description="报告文件列表（最多200份）"),
    items_json: str = Form(..., description="JSON: [{filename, instrument_id, gas_detector_id, report_date, remark}]"),
    report_date: date | None = Form(default=None, description="报告日期（统一）"),
    remark: str | None = Form(default=None, description="备注（统一）"),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    import json
    try:
        items = json.loads(items_json)
    except json.JSONDecodeError:
        return JSONResponse(status_code=400, content={"code": 400, "message": "items_json JSON 格式错误"})

    file_list = []
    for f in files:
        data = await f.read()
        file_list.append((f.filename or "unknown", data, f.content_type or "application/octet-stream"))

    result = await service.batch_upload_reports(db, file_list, items, report_date=report_date, remark=remark)
    return success_response(result, status_code=201 if result["success"] > 0 else 200)


@router.post("/reports/batch-extract-dates", summary="批量 AI 识别报告日期并更新仪表（单次最多 200 份）")
async def batch_extract_dates(
    body: BatchExtractRequest,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    report_ids = [UUID(rid) for rid in body.report_ids]
    result = await service.batch_extract_dates(db, report_ids)
    return success_response(
        BatchExtractResponse(**result).model_dump(mode="json")
    )


@router.post("/reports/batch-extract-dates/stream", summary="批量 AI 识别报告日期 SSE 流式（单次最多 200 份，支持中断）")
async def batch_extract_dates_stream(
    body: BatchExtractRequest,
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """逐份 AI 识别并推送 SSE 进度事件。

    事件类型：
    - start: 任务开始，含 total
    - progress: 开始处理一份报告，含 current/total/report_id/file_name
    - result: 单份报告处理结果，含 status (success/failed)
    - error: 全局错误（如未配置 AI）
    - complete: 全部完成或被中断，含 success/failed/interrupted
    """
    import json
    from collections.abc import AsyncGenerator

    def _sse(event: str, data: dict[str, Any]) -> str:
        return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"

    report_ids = [UUID(rid) for rid in body.report_ids]
    interrupted = False

    async def event_stream() -> AsyncGenerator[str, None]:
        nonlocal interrupted
        async for chunk in service.batch_extract_dates_stream(db, report_ids):
            yield chunk
            if await request.is_disconnected():
                interrupted = True
                break

        if interrupted:
            yield _sse("complete", {
                "total": len(report_ids), "success": 0, "failed": 0,
                "interrupted": True,
            })

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


# ═══════════════════════════════════════════
# 检定到期提醒
# ═══════════════════════════════════════════


@router.get("/calibration/alerts", summary="检定到期提醒")
async def calibration_alerts(
    days_before: int = Query(default=30, ge=0, le=365, description="0=截止今天(含超期), >0=未来N天内到期"),
    department: str | None = Query(default=None, description="部门筛选"),
    source: str | None = Query(default=None, pattern="^(instrument|gas_detector)$", description="数据源筛选"),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    alerts = await service.get_calibration_alerts(db, days_before=days_before, department=department, source=source)
    return success_response(alerts)


@router.get("/calibration/alerts/export-excel", summary="导出检定到期提醒为 Excel")
async def export_calibration_alerts_excel(
    days_before: int = Query(default=30, ge=0, le=365, description="0=截止今天(含超期), >0=未来N天内到期"),
    department: str | None = Query(default=None, description="部门筛选"),
    source: str | None = Query(default=None, pattern="^(instrument|gas_detector)$", description="数据源筛选"),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    import io as io_mod

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    alerts = await service.get_calibration_alerts(db, days_before=days_before, department=department, source=source)

    wb = Workbook()
    ws = wb.active
    ws.title = "检定到期提醒"

    headers = ["来源", "编号", "名称", "位置", "部门", "下次检定日期", "距到期天数"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, a in enumerate(alerts, 2):
        source_label = "计量器具" if a["source"] == "instrument" else "探测器"
        days_text = ""
        if a["days_until_due"] is not None:
            d = a["days_until_due"]
            if d < 0:
                days_text = f"已过期 {abs(d)} 天"
            elif d == 0:
                days_text = "今天到期"
            else:
                days_text = f"{d} 天"
        values = [
            source_label,
            a.get("serial_number", ""),
            a.get("instrument_name", ""),
            a.get("location", ""),
            a.get("department", ""),
            a["next_calibration_date"].isoformat() if a.get("next_calibration_date") else "",
            days_text,
        ]
        for col_idx, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, len(alerts) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

    buf = io_mod.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=calibration_alerts_export.xlsx"},
    )


# ═══════════════════════════════════════════
# AI 日期提取（配置从环境变量 METER_AI_* 读取）
# ═══════════════════════════════════════════


@router.post("/reports/{report_id}/extract-date", summary="从报告中提取校准日期")
async def extract_date(
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    import logging

    logger = logging.getLogger(__name__)

    try:
        config = get_meter_ai_config()
        if not config:
            return JSONResponse(
                status_code=400,
                content={"code": 400, "message": "请先配置环境变量 METER_AI_BASE_URL / METER_AI_API_KEY / METER_AI_MODEL"},
            )

        report = await service.get_report(db, report_id)
        result = await service.download_report_data(report)
        if result is None:
            raise NotFoundException("报告文件", str(report_id))
        pdf_data, _ = result

        # 获取关联仪表的检定周期
        calibration_cycle = None
        if report.instrument_id:
            inst = await service.get_instrument(db, report.instrument_id)
            calibration_cycle = inst.calibration_cycle_months
        # 探测器无 calibration_cycle_months 字段，保持 None

        ai_result = await extract_and_update_date(
            pdf_data, config["api_url"], config["api_key"], config["model"], calibration_cycle
        )

        if ai_result["success"]:
            # 回写数据库（asyncpg 需要 Python date 对象，不能传字符串）
            updates: dict[str, Any] = {"calibration_date": date.fromisoformat(ai_result["calibration_date"])}
            if ai_result.get("next_calibration_date"):
                updates["next_calibration_date"] = date.fromisoformat(ai_result["next_calibration_date"])
            if report.instrument_id:
                await repo.update_instrument(db, report.instrument_id, updates)
            elif report.gas_detector_id:
                await repo.update_gas_detector(db, report.gas_detector_id, updates)
            await db.commit()

        return success_response(
            ExtractDateResponse(**ai_result).model_dump(mode="json")
        )
    except Exception:
        logger.exception("extract_date 未捕获异常")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": "服务器内部错误"},
        )


# ═══════════════════════════════════════════
# 辅助函数
# ═══════════════════════════════════════════


def _build_report_items(reports: list[Any]) -> list[dict[str, Any]]:
    """将 ORM report 对象转换为 ReportItem 字典列表。"""
    items: list[dict[str, Any]] = []
    for r in reports:
        items.append(
            ReportItem(
                id=str(r.id),
                file_name=r.file_name,
                file_size=r.file_size,
                content_type=r.content_type,
                report_date=r.report_date,
                remark=r.remark,
                uploaded_at=r.created_at,
                download_url=f"./reports/{r.id}/download",
            ).model_dump(mode="json")
        )
    return items


# ═══════════════════════════════════════════
# 部门管理
# ═══════════════════════════════════════════


@router.get("/departments", summary="部门列表")
async def list_departments(
    source: str | None = Query(default=None, pattern="^(instrument|gas_detector)$", description="来源筛选"),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    result = await service.list_departments(db, source=source)
    return success_response(
        [DepartmentResponse(**r).model_dump(mode="json") for r in result]
    )


@router.post("/departments", summary="新增部门")
async def create_department(
    data: DepartmentCreate,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    dept = await service.create_department(db, data)
    return success_response(
        DepartmentResponse(
            id=str(dept.id),
            source=dept.source,
            name=dept.name,
            heads=dept.heads or [],
            auto_notify_enabled=dept.auto_notify_enabled,
            created_at=dept.created_at,
            updated_at=dept.updated_at,
        ).model_dump(mode="json"),
        status_code=201,
    )


@router.put("/departments/{dept_id}", summary="更新部门（联动更新表中记录）")
async def update_department(
    dept_id: UUID,
    data: DepartmentUpdate,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    dept = await service.update_department(db, dept_id, data)
    return success_response(
        DepartmentResponse(
            id=str(dept.id),
            source=dept.source,
            name=dept.name,
            heads=dept.heads or [],
            auto_notify_enabled=dept.auto_notify_enabled,
            created_at=dept.created_at,
            updated_at=dept.updated_at,
        ).model_dump(mode="json")
    )


@router.put("/departments/{dept_id}/auto-notify", summary="切换部门自动提醒开关")
async def toggle_department_auto_notify(
    dept_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    dept = await repo.get_department_by_id(db, dept_id)
    if dept is None:
        raise NotFoundException("部门", str(dept_id))
    new_state = not dept.auto_notify_enabled
    updated = await service.update_department(
        db, dept_id,
        DepartmentUpdate(name=dept.name, auto_notify_enabled=new_state)
    )
    return success_response(
        DepartmentResponse(
            id=str(updated.id),
            source=updated.source,
            name=updated.name,
            heads=updated.heads or [],
            auto_notify_enabled=updated.auto_notify_enabled,
            created_at=updated.created_at,
            updated_at=updated.updated_at,
        ).model_dump(mode="json")
    )


@router.get("/departments/personnel-candidates", summary="获取可选负责人列表（从 identity.users 查询）")
async def get_personnel_candidates(
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    """从平台 identity.users 查询所有用户，作为负责人候选人列表。"""
    candidates = await service.get_personnel_candidates(db)
    return success_response(
        [PersonnelCandidate(**c).model_dump(mode="json") for c in candidates]
    )


@router.delete("/departments/{dept_id}", summary="删除部门")
async def delete_department(
    dept_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    await service.delete_department(db, dept_id)
    return success_response(message="删除成功")


# ═══════════════════════════════════════════
# 全局设置
# ═══════════════════════════════════════════


@router.get("/settings", summary="获取全局设置（提醒时间）")
async def get_settings(
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    result = await service.get_meter_settings(db)
    return success_response(MeterSettingsResponse(**result).model_dump(mode="json"))


@router.put("/settings", summary="更新全局设置（提醒时间）")
async def update_settings(
    data: MeterSettingsUpdate,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    try:
        result = await service.update_meter_settings(db, data.notify_time)
    except ValueError as e:
        return JSONResponse(
            status_code=400,
            content={"code": 400, "message": str(e)},
        )
    await db.commit()
    return success_response(MeterSettingsResponse(**result).model_dump(mode="json"))
