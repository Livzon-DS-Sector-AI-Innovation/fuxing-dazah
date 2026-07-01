# ruff: noqa: E402, E501
"""
生成设备台账导入 Excel 模板。

用法：在 dazah-backend 目录下执行
    uv run python -X utf8 scripts/import_data/generate_template.py
"""

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = PROJECT_ROOT / "assets" / "设备台账导入模板.xlsx"

HEADERS: list[dict] = [
    {"col": 1,  "field": "name",              "label": "设备名称 *",
     "width": 24, "example": "离心泵-A-01",     "note": "必填"},
    {"col": 2,  "field": "category_name",     "label": "设备分类 *",
     "width": 18, "example": "离心泵",           "note": "必填，填分类名称，系统不存在则自动创建"},
    {"col": 3,  "field": "location_name",     "label": "设备位置 *",
     "width": 18, "example": "A车间一楼",        "note": "必填，填位置名称，系统不存在则自动创建"},
    {"col": 4,  "field": "department_name",   "label": "归属部门 *",
     "width": 18, "example": "动力部",           "note": "必填，填部门名称，须与系统一致"},
    {"col": 5,  "field": "status",            "label": "设备状态",
     "width": 10, "example": "在用",             "note": "在用 / 备用 / 维修中 / 停用 / 报废，不填默认'在用'"},
    {"col": 6,  "field": "model",             "label": "设备型号",
     "width": 20, "example": "IHF65-50-160",    "note": "选填"},
    {"col": 7,  "field": "specification",     "label": "设备规格",
     "width": 22, "example": "流量25m³/h 扬程32m","note": "选填"},
    {"col": 8,  "field": "manufacturer",      "label": "制造商",
     "width": 22, "example": "XX泵业有限公司",   "note": "选填"},
    {"col": 9,  "field": "supplier",          "label": "供应商",
     "width": 22, "example": "XX机电设备公司",   "note": "选填"},
    {"col": 10, "field": "production_date",   "label": "出厂日期",
     "width": 14, "example": "2024-01-15",       "note": "选填，格式 YYYY-MM-DD"},
    {"col": 11, "field": "commissioning_date","label": "投用日期",
     "width": 14, "example": "2024-03-01",       "note": "选填，格式 YYYY-MM-DD"},
    {"col": 12, "field": "description",       "label": "设备描述",
     "width": 30, "example": "用于XX工序物料输送","note": "选填"},
    {"col": 13, "field": "importance",        "label": "重要性",
     "width": 8,  "example": "中",               "note": "高 / 中 / 低，不填默认'低'"},
    {"col": 14, "field": "warranty_expire",   "label": "保修到期日",
     "width": 14, "example": "2026-01-15",       "note": "选填，格式 YYYY-MM-DD"},
    {"col": 15, "field": "asset_value",       "label": "资产原值（元）",
     "width": 16, "example": "15000.00",         "note": "选填，数字"},
    {"col": 16, "field": "depreciation_years","label": "折旧年限",
     "width": 10, "example": "10",               "note": "选填，整数"},
    {"col": 17, "field": "responsible_person","label": "负责人",
     "width": 14, "example": "张三",              "note": "选填，填姓名，须与系统用户一致"},
]


def create_template():
    wb = openpyxl.Workbook()

    ws_guide = wb.active
    ws_guide.title = "填写说明"
    _write_guide(ws_guide)

    ws_data = wb.create_sheet("设备台账")
    _write_data_sheet(ws_data)

    wb.save(OUTPUT_PATH)
    print(f"[+] 模板已生成: {OUTPUT_PATH}")


def _write_guide(ws):
    title_font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    section_font = Font(name="微软雅黑", size=11, bold=True, color="2E75B6")
    body_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 62

    row = 1
    ws.merge_cells("B1:C1")
    ws.cell(row=row, column=2, value="设备台账导入模板 — 填写说明").font = title_font
    row += 2

    ws.cell(row=row, column=2, value="一、基本要求").font = section_font
    row += 1
    tips = [
        "1. 请在「设备台账」工作表中逐行填写设备信息。",
        "2. 第一行是表头，第二行是示例数据，请从第三行开始填写实际数据。",
        "3. 带 * 号的列为必填项，其余可选填。",
        "4. 设备编号由系统自动生成，无需手动填写。",
        "5. 「设备分类」「设备位置」直接填中文名称即可，系统不存在会自动创建。",
        "6. 自动创建的分类/位置，编号会暂设为临时值，请后续到系统界面补填正式编号。",
        "7. 「归属部门」必须与系统中已有部门名称一致，否则导入失败。",
        "8. 不要修改表头名称、列顺序或删除列，否则导入会失败。",
    ]
    for tip in tips:
        ws.cell(row=row, column=2, value=tip).font = body_font
        row += 1
    row += 1

    ws.cell(row=row, column=2, value="二、字段说明").font = section_font
    row += 1

    for c, label in enumerate(["列", "字段", "填写说明"], start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    for h in HEADERS:
        col_letter = get_column_letter(h["col"])
        cells = [(1, col_letter), (2, h["label"]), (3, h["note"])]
        for c, val in cells:
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="三、注意事项").font = section_font
    row += 1
    notes = [
        "1. 设备编号由系统按 EQ-IMP-日期-序号 格式自动生成，无需填写。",
        "2. 设备分类和位置：填名称即可。如系统中不存在，导入时会自动创建并归属到设备所在部门。",
        "3. 自动创建的分类和位置编号为临时值（IMP_CAT_xxx / IMP_LOC_xxx），请导入后到系统界面补填正式编号。",
        "4. 归属部门名称必须与系统中已有部门完全一致，否则该行导入失败。",
        "5. 日期格式统一为 YYYY-MM-DD，例如 2024-01-15。",
        "6. 数字字段不要带千分位分隔符，直接写 15000.00。",
        "7. 导入失败的行会在日志中打印，修正后重新导入即可（已成功的会自动跳过）。",
    ]
    for note in notes:
        ws.cell(row=row, column=2, value=note).font = body_font
        row += 1


def _write_data_sheet(ws):
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws.freeze_panes = "A3"

    # 表头（第1行）
    for h in HEADERS:
        col_letter = get_column_letter(h["col"])
        cell = ws.cell(row=1, column=h["col"], value=h["label"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[col_letter].width = h["width"]

    # 示例（第2行）
    ws.row_dimensions[2].height = 28
    for i, h in enumerate(HEADERS):
        cell = ws.cell(row=2, column=i + 1, value=h["example"])
        cell.fill = example_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        cell.font = Font(name="微软雅黑", size=9, italic=True, color="808080")

    # 数据验证 — 设备状态
    dv_status = DataValidation(
        type="list",
        formula1='"在用,备用,维修中,停用,报废"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="无效状态",
        error="请选择：在用 / 备用 / 维修中 / 停用 / 报废",
    )
    dv_status.add("E3:E5002")
    ws.add_data_validation(dv_status)

    # 数据验证 — 重要性
    dv_importance = DataValidation(
        type="list",
        formula1='"高,中,低"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="无效重要性",
        error="请选择：高 / 中 / 低",
    )
    dv_importance.add("M3:M5002")
    ws.add_data_validation(dv_importance)


if __name__ == "__main__":
    create_template()
