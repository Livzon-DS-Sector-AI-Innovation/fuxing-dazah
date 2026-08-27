"""HR business workflows live here."""

import asyncio
import logging
from datetime import UTC, date, datetime
from uuid import UUID

from sqlalchemy import func, or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import DuplicateException, NotFoundException
from app.modules.hr.models import (
    AnnualTrainingPlan,
    AnnualTrainingPlanItem,
    Candidate,
    CandidateAiEvaluation,
    CandidateAnalysisReport,
    CandidateReview,
    CandidateStatusLog,
    DepartureRecord,
    DeptTrainingPersonnel,
    Employee,
    HrDepartment,
    Interview,
    JobRequirement,
    MonthlyPerformanceEvaluation,
    OffboardingRecord,
    OnboardingRecord,
    OnboardingTask,
    PositionTraining,
    Team,
    TrainingLedger,
    TrainingLedgerPage,
)
from app.modules.hr.repository import (
    AnnualTrainingPlanItemRepository,
    AnnualTrainingPlanRepository,
    CandidateAiEvaluationRepository,
    CandidateAnalysisReportRepository,
    CandidateRepository,
    CandidateReviewRepository,
    CandidateStatusLogRepository,
    DepartmentRepository,
    DepartureRecordRepository,
    EmployeeRepository,
    InterviewRepository,
    JobRequirementRepository,
    OffboardingRecordRepository,
    OnboardingRecordRepository,
    PerformanceEvaluationRepository,
    TeamRepository,
    TrainingLedgerPageRepository,
    TrainingLedgerRepository,
)
from app.modules.hr.schemas import (
    AnnualTrainingPlanCreate,
    AnnualTrainingPlanItemBatchUpdate,
    AnnualTrainingPlanUpdate,
    CandidateCreate,
    CandidateUpdate,
    DepartmentCreate,
    DepartmentUpdate,
    DepartureRecordCreate,
    DepartureRecordUpdate,
    EmployeeCreate,
    EmployeeUpdate,
    InterviewCreate,
    InterviewUpdate,
    JobRequirementCreate,
    JobRequirementUpdate,
    OffboardingRecordCreate,
    OffboardingRecordUpdate,
    PerformanceEvaluationCreate,
    PerformanceEvaluationUpdate,
    PerformanceListParams,
    TeamCreate,
    TeamUpdate,
    TrainingLedgerCreate,
    TrainingLedgerUpdate,
)

logger = logging.getLogger(__name__)

# ─── Services ───

class EmployeeService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = EmployeeRepository(session)

    async def get_employee(self, employee_id: UUID) -> Employee:
        employee = await self.repo.get_by_id(employee_id)
        if not employee:
            raise NotFoundException("员工", str(employee_id))
        return employee

    async def get_employee_by_number(self, employee_number: str) -> Employee:
        employee = await self.repo.get_by_employee_number(employee_number)
        if not employee:
            raise NotFoundException("员工", employee_number)
        return employee

    async def create_employee(self, data: EmployeeCreate) -> Employee:
        existing = await self.repo.get_by_employee_number(
            data.employee_number, include_deleted=True
        )
        if existing:
            raise DuplicateException("工号", data.employee_number)

        employee = Employee(**data.model_dump())
        if not employee.department:
            employee.department = "未分类"
        employee.status = "在职"

        result = await self.repo.create(employee)

        # 同时创建入职台账记录
        await self._create_onboarding_record(result)

        # 自动根据岗位创建培训台账
        await self._sync_employee_training(result)

        return result

    async def _sync_employee_training(self, employee: Employee) -> None:
        """为新员工自动创建培训台账页面和岗位关联的培训记录。"""
        from app.modules.hr.models import TrainingLedgerPage
        # 创建台账页面
        q = select(TrainingLedgerPage).where(
            TrainingLedgerPage.employee_number == employee.employee_number,
            TrainingLedgerPage.is_deleted == False,  # noqa: E712
        )
        r = await self.repo.session.execute(q)
        if not r.scalar_one_or_none():
            self.repo.session.add(TrainingLedgerPage(
                employee_number=employee.employee_number,
                employee_name=employee.name,
            ))

        # 根据岗位导入培训内容（兼容带部门前缀和不带前缀的岗位名）
        if employee.position and employee.department:
            pt_q = select(PositionTraining).where(
                PositionTraining.is_deleted == False,  # noqa: E712
                PositionTraining.department == employee.department,
                or_(
                    PositionTraining.position_name == employee.position,
                    PositionTraining.position_name.endswith(employee.position),
                ),
            )
            pts = (await self.repo.session.execute(pt_q)).scalars().all()
            for pt in pts:
                exist_q = select(TrainingLedger).where(
                    TrainingLedger.employee_number == employee.employee_number,
                    TrainingLedger.training_subject == pt.training_category,
                    TrainingLedger.is_deleted == False,  # noqa: E712
                )
                ext = await self.repo.session.execute(exist_q)
                if not ext.scalar_one_or_none():
                    self.repo.session.add(TrainingLedger(
                        employee_number=employee.employee_number,
                        training_subject=pt.training_category,
                        training_method=pt.training_method,
                        trainer=pt.trainer,
                        training_date=employee.hire_date or date.today(),
                    ))

    async def _create_onboarding_record(self, employee: Employee) -> None:
        """为新员工创建入职台账记录。"""
        from app.modules.hr.models import OnboardingRecord
        record = OnboardingRecord(
            employee_number=employee.employee_number,
            name=employee.name,
            domain_account=employee.domain_account,
            department=employee.department,
            team=employee.team,
            position=employee.position,
            job_category=employee.job_category,
            status_category=employee.status_category,
            is_employed="是",
            hire_date=employee.hire_date,
            factory_entry_date=employee.factory_entry_date,
            education=employee.education,
            school=employee.school,
            major=employee.major,
            phone=employee.phone,
        )
        self.repo.session.add(record)
        await self.repo.session.flush()

    # ── Excel 列名 → 模型字段名 映射 ──
    _UPLOAD_COLUMN_MAP: dict[str, str] = {
        "工号": "employee_number",
        "姓名": "name",
        "域账号": "domain_account",
        "体现部门": "department",
        "部门": "actual_department",
        "班组": "team",
        "职位": "position",
        "岗位": "job_category",
        "体现岗位": "position",
        "职务": "duty",
        "岗位类别": "job_category",
        "级别": "level",
        "职级": "level",
        "兼任部门": "concurrent_departments",
        "资格": "qualifications",
        "证书": "certificates",
        "资格类型": "qualification_type",
        "性别": "gender",
        "籍贯": "native_place",
        "政治面貌": "political_status",
        "婚姻状况": "marital_status",
        "户籍类型": "household_type",
        "用工性质": "status_category",
        "员工性质": "status_category",
        "人员状态": "status",
        "出生年份": "birth_year",
        "出生月份": "birth_month",
        "出生日": "birth_day",
        "出生年月": "_birth_date",
        "年龄": "age",
        "参加工作时间": "work_start_date",
        "进厂时间": "factory_entry_date",
        "进丽珠时间": "livo_entry_date",
        "入职日期": "hire_date",
        "毕业时间": "graduation_date",
        "工龄": "work_years",
        "厂龄": "factory_tenure",
        "公司工龄": "company_tenure",
        "学历": "education",
        "分类": "classification",
        "毕业学校": "school",
        "毕业院校": "school",
        "专业": "major",
        "品种": "variety",
        "兼任品种": "concurrent_variety",
        "身份证号": "id_card",
        "身份证号码": "id_card",
        "身份证有效期": "id_card_expiry",
        "身份证地址": "id_card_address",
        "现居住地址": "current_address",
        "合同类型": "contract_type",
        "合同开始日期": "contract_start_date",
        "合同结束日期": "contract_end_date",
        "合同开始日期2": "contract_start_2",
        "合同结束日期2": "contract_end_2",
        "合同开始日期3": "contract_start_3",
        "合同结束日期3": "contract_end_3",
        "合同开始日期4": "contract_start_4",
        "合同结束日期4": "contract_end_4",
        "手机": "phone",
        "邮箱": "email",
        "紧急联系人": "emergency_contact_name",
        "紧急联系人电话": "emergency_contact_phone",
        "紧急联系人关系": "emergency_contact_relation",
        "银行卡号": "bank_account",
        "培训编号": "training_id",
        "调动历史": "transfer_history",
        "备注": "remarks",
        # Excel 扩展字段
        "部门管理者": "dept_manager",
        "额外管理者": "additional_manager",
        "报表用职级": "report_grade",
        "部门负责人/一级培训师": "dept_head_trainer",
        "入职安全培训日期": "safety_training_date",
        "入职安全培训成绩": "safety_training_score",
        "企业文化培训日期": "culture_training_date",
        "GMP基础培训时间": "gmp_training_date",
        "离职时间": "departure_date",
    }

    # Excel 上传时，以下字段全部以 Excel 为准（Excel 中未出现的字段置空）
    _EXCEL_RESETTABLE_FIELDS: set[str] = {
        "name", "domain_account", "department", "actual_department", "team", "position",
        "job_category", "level", "concurrent_departments",
        "qualifications", "qualification_type",
        "gender", "native_place", "political_status", "marital_status",
        "household_type", "status_category",
        "birth_year", "birth_month", "birth_day", "age",
        "work_start_date", "factory_entry_date", "livo_entry_date",
        "hire_date", "graduation_date",
        "work_years", "factory_tenure", "company_tenure",
        "education", "classification", "school", "major", "variety",
        "id_card", "id_card_expiry", "id_card_address", "current_address",
        "contract_type", "contract_start_date", "contract_end_date",
        "contract_start_2", "contract_end_2",
        "contract_start_3", "contract_end_3",
        "contract_start_4", "contract_end_4",
        "phone", "email",
        "emergency_contact_name", "emergency_contact_phone",
        "emergency_contact_relation",
        "bank_account", "training_id", "transfer_history", "remarks",
        "status", "certificates", "concurrent_variety",
        # Excel 扩展字段
        "duty", "dept_manager", "additional_manager", "report_grade",
        "dept_head_trainer", "safety_training_date", "safety_training_score",
        "culture_training_date", "gmp_training_date", "departure_date",
    }

    _DATE_FIELDS: set[str] = {
        "work_start_date", "factory_entry_date", "livo_entry_date",
        "hire_date", "graduation_date", "id_card_expiry",
        "contract_start_date", "contract_end_date",
        "contract_start_2", "contract_end_2",
        "contract_start_3", "contract_end_3",
        "contract_start_4", "contract_end_4",
        "safety_training_date", "culture_training_date",
        "gmp_training_date", "departure_date",
    }

    _INT_FIELDS: set[str] = {
        "birth_year", "birth_month", "birth_day", "age", "work_years",
    }

    # 常见日期格式（按优先级从高到低排序）
    _DATE_FORMATS: list[str] = [
        "%Y-%m-%d",       # 2024-01-15
        "%Y/%m/%d",       # 2024/01/15
        "%Y.%m.%d",       # 2024.01.15
        "%Y年%m月%d日",    # 2024年01月15日
        "%Y%m%d",         # 20240115
        "%Y-%m-%d %H:%M:%S",       # 2024-01-15 00:00:00
        "%Y/%m/%d %H:%M:%S",       # 2024/01/15 00:00:00
        "%d/%m/%Y",       # 15/01/2024
        "%m/%d/%Y",       # 01/15/2024
    ]

    @staticmethod
    def _parse_date_value(val: object) -> date | None:
        """将各种格式的日期值统一转换为 date 对象。返回 None 表示无法解析。"""
        from datetime import date as date_cls
        from datetime import datetime as datetime_cls

        if isinstance(val, datetime_cls):
            return val.date()
        if isinstance(val, date_cls):
            return val
        if isinstance(val, (int, float)):
            # Excel 日期序列号（以 1899-12-30 为第 0 天）
            try:
                from datetime import timedelta
                excel_epoch = date_cls(1899, 12, 30)
                return excel_epoch + timedelta(days=int(val))
            except (ValueError, OverflowError):
                return None
        if isinstance(val, str):
            val = val.strip()
            if not val:
                return None
            for fmt in EmployeeService._DATE_FORMATS:
                try:
                    return datetime_cls.strptime(val, fmt).date()
                except ValueError:
                    continue
        return None

    async def upload_employees(
        self, file_bytes: bytes, allowed_departments: set[str] | None = None
    ) -> dict:
        """从 Excel 文件批量导入员工，按工号 upsert。返回 {created, updated, skipped, errors}。

        每行使用独立的 SQL savepoint，单行失败通过 ROLLBACK TO SAVEPOINT 恢复事务，
        确保失败行不影响其他行。allowed_departments 非空时仅导入该范围内的部门行。
        """
        from io import BytesIO

        from openpyxl import load_workbook

        wb = await asyncio.to_thread(load_workbook, BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("文件为空")

        header = [str(c).strip() if c else "" for c in rows[0]]
        # 建立列索引
        col_map: dict[int, str] = {}
        for idx, col_name in enumerate(header):
            field = self._UPLOAD_COLUMN_MAP.get(col_name)
            if field:
                col_map[idx] = field

        if "employee_number" not in col_map.values():
            raise ValueError("缺少「工号」列，无法导入")

        created = 0
        updated = 0
        skipped = 0
        errors: list[str] = []

        # 获取底层连接，用于手动管理 savepoint
        conn = await self.repo.session.connection()
        sp_id = 0

        for row_idx, row in enumerate(rows[1:], start=2):
            if all(c is None for c in row):
                continue

            sp_id += 1
            sp_name = f"row_{sp_id}"

            try:
                await conn.execute(text(f"SAVEPOINT {sp_name}"))

                data: dict = {}
                for col_idx, field_name in col_map.items():
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is None or (isinstance(val, str) and val.strip() == ""):
                        continue
                    if isinstance(val, str):
                        val = val.strip()
                    if field_name in self._DATE_FIELDS:
                        parsed = self._parse_date_value(val)
                        if parsed is None:
                            continue
                        val = parsed
                    elif field_name in self._INT_FIELDS:
                        try:
                            val = int(float(str(val)))
                        except (ValueError, TypeError):
                            continue
                    elif field_name == "qualifications":
                        vals = [v.strip() for v in str(val).split(",") if v.strip()]
                        import json as _json
                        val = _json.dumps(vals, ensure_ascii=False)
                    elif field_name == "_birth_date":
                        # 出生年月：datetime 对象 → 拆分为 birth_year / birth_month / birth_day
                        if isinstance(val, (int, float)):
                            parsed_date = self._parse_date_value(val)
                        else:
                            parsed_date = self._parse_date_value(str(val))
                        if parsed_date:
                            data["birth_year"] = parsed_date.year
                            data["birth_month"] = parsed_date.month
                            data["birth_day"] = parsed_date.day
                        continue  # 不把 _birth_date 本身写入 data
                    elif field_name == "status":
                        # 规范化人员状态：'—'/空/未知值 → 不写入（更新保留原状态，
                        # 新员工由 upsert INSERT 分支默认「在职」）
                        val_str = str(val).strip()
                        if val_str in ("—", "——", "-", "", "无"):
                            continue
                        if val_str in ("在职", "离职", "待审批", "病假", "产假", "产假复岗"):
                            val = val_str
                        else:
                            continue
                    data[field_name] = val

                data.setdefault("position", "")
                data.setdefault("department", "未分类")
                data.setdefault("name", "")

                if "employee_number" not in data:
                    raise ValueError("缺少工号")

                if "hire_date" not in data:
                    data["hire_date"] = date.today()
                # 状态列缺失/为空时不再默认「在职」：已有员工保留原状态（病假/产假不被覆盖），
                # 新员工由 upsert INSERT 分支默认「在职」

                # ── Excel 没有的字段全部置空，确保完全以 Excel 为准 ──
                for field in self._EXCEL_RESETTABLE_FIELDS:
                    if field not in data:
                        data[field] = None

                # 数据范围受限时仅导入授权部门行
                if (
                    allowed_departments is not None
                    and data.get("department") not in allowed_departments
                ):
                    skipped += 1
                    await conn.execute(text(f"RELEASE SAVEPOINT {sp_name}"))
                    continue

                is_new = await self.repo.upsert_by_employee_number(data)
                if is_new:
                    created += 1
                else:
                    updated += 1

                await conn.execute(text(f"RELEASE SAVEPOINT {sp_name}"))
            except Exception as e:
                # ROLLBACK TO SAVEPOINT 是 Postgres 恢复事务的唯一方式
                await conn.execute(text(f"ROLLBACK TO SAVEPOINT {sp_name}"))
                await conn.execute(text(f"RELEASE SAVEPOINT {sp_name}"))
                errors.append(f"第{row_idx}行: {e}")

        # ── 上传完成后自动同步部门表 ──
        if created > 0 or updated > 0:
            await self._sync_departments_from_employees()

        return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}

    async def upload_trainers(self, file_bytes: bytes, allowed_departments: set[str] | None = None) -> dict:
        """上传内训师 Excel，按姓名+部门 upsert。allowed_departments 非空时仅导入该范围内的部门。"""
        from io import BytesIO

        from openpyxl import load_workbook

        from app.modules.hr.models import HrTrainer

        wb = await asyncio.to_thread(load_workbook, BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("文件为空")
        header = [str(c).strip() if c else "" for c in rows[0]]
        col_map = {}
        col_name_map = {"姓名": "name", "部门": "department", "可培训部门": "trainable_departments",
                        "资格范围": "qualification_scope", "认证日期": "certification_date",
                        "确认日期": "confirmation_date", "确认提醒": "confirmation_reminder",
                        "是否一级培训师": "is_level1", "培训管理员": "admin", "备注": "remarks"}
        for idx, h in enumerate(header):
            if h in col_name_map:
                col_map[idx] = col_name_map[h]

        from sqlalchemy import text as sql_text
        conn = await self.repo.session.connection()
        created, updated, skipped, errors = 0, 0, 0, []
        for row_idx, row in enumerate(rows[1:], start=2):
            if all(c is None for c in row): continue
            sp_name = f"trainer_row_{row_idx}"
            try:
                await conn.execute(sql_text(f"SAVEPOINT {sp_name}"))
                data = {}
                for ci, fn in col_map.items():
                    v = row[ci] if ci < len(row) else None
                    if v is None or (isinstance(v, str) and v.strip() == ""): continue
                    if isinstance(v, str): v = v.strip()
                    if fn in ("certification_date", "confirmation_date", "confirmation_reminder"):
                        v = self._parse_date_value(v)
                    data[fn] = v
                if "name" not in data or "department" not in data:
                    errors.append(f"第{row_idx}行: 缺少姓名或部门"); continue
                if allowed_departments is not None and data.get("department") not in allowed_departments:
                    skipped += 1
                    continue

                q = select(HrTrainer).where(
                    HrTrainer.name == data["name"],
                    HrTrainer.department == data["department"],
                    HrTrainer.is_deleted == False)  # noqa: E712
                r = await self.repo.session.execute(q)
                existing = r.scalar_one_or_none()
                if existing:
                    for k, v in data.items(): setattr(existing, k, v)
                    updated += 1
                else:
                    self.repo.session.add(HrTrainer(**data))
                    created += 1
                await self.repo.session.flush()
                await conn.execute(sql_text(f"RELEASE SAVEPOINT {sp_name}"))
            except Exception as e:
                # 单行失败只回滚本行，不毒化整个导入事务
                await conn.execute(sql_text(f"ROLLBACK TO SAVEPOINT {sp_name}"))
                await conn.execute(sql_text(f"RELEASE SAVEPOINT {sp_name}"))
                errors.append(f"第{row_idx}行: {e}")
        return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}

    async def upload_sop_catalog(self, file_bytes: bytes) -> dict:
        """上传 SOP 目录 Excel，按 SOP 编号 upsert。"""
        from io import BytesIO

        from openpyxl import load_workbook

        from app.modules.hr.models import PositionTraining, SopCatalog

        wb = await asyncio.to_thread(load_workbook, BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: raise ValueError("文件为空")
        header = [str(c).strip() if c else "" for c in rows[0]]
        col_map = {}
        col_name_map = {"SOP编号": "sop_number", "文件名称": "file_name", "培训类别": "category",
                        "部门": "department", "所属部门": "department", "岗位": "position_name",
                        "体现部门": "department", "SOP名称": "file_name", "文件名": "file_name",
                        "类别": "category", "分类": "category", "培训分类": "category",
                        "职位": "position_name", "岗位名称": "position_name", "培训内容": "file_name"}
        for idx, h in enumerate(header):
            if h in col_name_map: col_map[idx] = col_name_map[h]

        from sqlalchemy import text as sql_text
        conn = await self.repo.session.connection()
        created, updated, errors = 0, 0, []
        for row_idx, row in enumerate(rows[1:], start=2):
            if all(c is None for c in row): continue
            sp_name = f"sop_row_{row_idx}"
            try:
                await conn.execute(sql_text(f"SAVEPOINT {sp_name}"))
                data = {}
                for ci, fn in col_map.items():
                    v = row[ci] if ci < len(row) else None
                    if v is None or (isinstance(v, str) and v.strip() == ""): continue
                    if isinstance(v, str): v = v.strip()
                    data[fn] = v
                if "file_name" not in data:
                    errors.append(f"第{row_idx}行: 缺少文件名称"); continue

                # 培训内容可能包含逗号分隔的多个条目，拆分为独立记录
                raw_names = str(data["file_name"])
                split_names = [n.strip() for n in raw_names.replace("，", ",").split(",") if n.strip()]
                # 没有培训类别时，用拆分后的值作为类别
                has_category = "category" in data

                for name in split_names:
                    row_data = {**data, "file_name": name}
                    if not has_category:
                        row_data["category"] = name
                    # 去重：有 SOP 编号按编号+文件名匹配，否则按部门+岗位+文件名匹配
                    if row_data.get("sop_number"):
                        q = select(SopCatalog).where(
                            SopCatalog.sop_number == row_data["sop_number"],
                            SopCatalog.file_name == name,
                            SopCatalog.is_deleted == False)  # noqa: E712
                    else:
                        q = select(SopCatalog).where(
                            SopCatalog.file_name == name,
                            SopCatalog.department == row_data.get("department"),
                            SopCatalog.position_name == row_data.get("position_name"),
                            SopCatalog.is_deleted == False)  # noqa: E712
                    r = await self.repo.session.execute(q)
                    existing = r.scalar_one_or_none()
                    if existing:
                        for k, v in row_data.items():
                            if v is not None:
                                setattr(existing, k, v)
                        updated += 1
                    else:
                        self.repo.session.add(SopCatalog(**row_data))
                        created += 1
                    await self.repo.session.flush()

                    # 同步写入岗位培训关联表（用于入职培训自动匹配），同样去重
                    if row_data.get("position_name") and row_data.get("department"):
                        pt_category = row_data.get("category") or name
                        pt_q = select(PositionTraining).where(
                            PositionTraining.position_name == row_data["position_name"],
                            PositionTraining.department == row_data["department"],
                            PositionTraining.training_category == pt_category,
                            PositionTraining.is_deleted == False)  # noqa: E712
                        pt_r = await self.repo.session.execute(pt_q)
                        pt_existing = pt_r.scalar_one_or_none()
                        if pt_existing:
                            pt_existing.file_name = name
                            pt_existing.sop_number = row_data.get("sop_number")
                            pt_existing.updated_at = func.now()
                        else:
                            self.repo.session.add(PositionTraining(
                                position_name=row_data["position_name"],
                                department=row_data["department"],
                                training_category=pt_category,
                                sop_number=row_data.get("sop_number"),
                                file_name=name,
                            ))
                await conn.execute(sql_text(f"RELEASE SAVEPOINT {sp_name}"))
            except Exception as e:
                await conn.execute(sql_text(f"ROLLBACK TO SAVEPOINT {sp_name}"))
                await conn.execute(sql_text(f"RELEASE SAVEPOINT {sp_name}"))
                errors.append(f"第{row_idx}行: {e}")
        return {"created": created, "updated": updated, "errors": errors}

    async def _sync_departments_from_employees(self) -> int:
        """从员工表同步部门：激活已删除的 + 新增不存在的。"""
        from sqlalchemy import text as sql_text
        conn = await self.repo.session.connection()

        # 1. 激活已存在但被软删除的部门
        r1 = await conn.execute(sql_text("""
            UPDATE hr.departments SET is_deleted = false, updated_at = now()
            WHERE code IN (SELECT DISTINCT department FROM hr.employees WHERE is_deleted = false AND department != '')
              AND is_deleted = true
        """))

        # 2. 新增员工表中有但部门表中不存在的部门
        r2 = await conn.execute(sql_text("""
            INSERT INTO hr.departments (id, name, code, created_at)
            SELECT gen_random_uuid(), e.department, e.department, now()
            FROM (SELECT DISTINCT department FROM hr.employees WHERE is_deleted = false AND department != '') e
            WHERE NOT EXISTS (SELECT 1 FROM hr.departments d WHERE d.code = e.department)
        """))
        return (r1.rowcount or 0) + (r2.rowcount or 0)

    async def approve_employee(self, employee_number: str) -> Employee:
        employee = await self.repo.get_by_employee_number(employee_number)
        if not employee:
            raise NotFoundException("员工", employee_number)
        if employee.status != "待审批":
            raise DuplicateException("审批", "该员工已审批完成")

        employee.status = "在职"
        result = await self.repo.update(employee)
        return result

    async def update_employee(self, employee_id: UUID, data: EmployeeUpdate) -> Employee:
        employee = await self.get_employee(employee_id)
        update_data = data.model_dump(exclude_unset=True)

        if "employee_number" in update_data:
            existing = await self.repo.get_by_employee_number(
                update_data["employee_number"], include_deleted=True
            )
            if existing and existing.id != employee_id:
                raise DuplicateException("工号", update_data["employee_number"])

        for field, value in update_data.items():
            setattr(employee, field, value)

        result = await self.repo.update(employee)
        return result

    async def delete_employee(self, employee_id: UUID) -> None:
        employee = await self.repo.get_by_id(employee_id)
        if not employee:
            raise NotFoundException("员工", str(employee_id))
        employee.is_deleted = True
        await self.repo.update(employee)

    async def list_employees(
        self,
        *,
        department: str | None = None,
        status: str | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
        sort_by: str = "department",
        sort_order: str = "asc",
        include_uncategorized: bool = False,
    ) -> tuple[list[Employee], int]:
        return await self.repo.list_employees(
            department=department,
            include_uncategorized=include_uncategorized,
            status=status,
            keyword=keyword,
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            sort_order=sort_order,
        )

    # ── 年度培训计划上传 ──

    _PLAN_COLUMN_MAP: dict[str, str] = {
        "年度": "year",
        "年份": "year",
        "部门": "department",
        "体现部门": "department",
        "品种": "variety",
        "月份": "month",
        "培训时间": "month",
        "培训人数": "trainee_count",
        "课时": "duration_hours",
        "培训内容及使用教材": "content_and_textbook",
        "培训内容": "content_and_textbook",
        "培训对象": "target_audience",
        "参加岗位/参加人数": "position_and_count",
        "授课单位及授课人": "position_and_count",
        "授课单位/培训师": "position_and_count",
        "培训方式": "training_method",
        "考核方式": "assessment_method",
        "培训地点": "location",
        "注意事项": "notes",
        "培训学时": "training_hours",
        "培训跟踪": "tracking_status",
        "确认者": "confirmer",
        "确认人": "confirmer",
        "确认人/日期": "_confirmer_date",
        "确认日期": "confirm_date",
        "实施日期": "confirm_date",
        "提醒实施": "confirm_date",
        "培训季度及课时": "_quarter_hours",
        "部门管理员": "confirmer",
        "备注": "remarks",
    }

    @staticmethod
    def _parse_combined_quarter_hours(raw: str) -> dict:
        """解析导出格式的"培训季度及课时"列：'1月\n8课时' → {month, duration_hours}"""
        result: dict = {}
        if not raw or not raw.strip():
            return result
        parts = [p.strip() for p in raw.replace("\n", " ").split() if p.strip()]
        for part in parts:
            if part.endswith("课时") or part.endswith("学时"):
                try:
                    result["duration_hours"] = float(part.rstrip("课时学时"))
                except ValueError:
                    pass
            elif "月" in part:
                result["month"] = part
            else:
                # 可能是纯数字月份
                try:
                    int(part)
                    result["month"] = f"{part}月"
                except ValueError:
                    pass
        return result

    @staticmethod
    def _parse_combined_confirmer_date(raw: str) -> dict:
        """解析导出格式的"确认人/日期"列：'张三 / 2024-01-15' → {confirmer, confirm_date}"""
        result: dict = {}
        if not raw or not raw.strip():
            return result
        if " / " in raw:
            parts = raw.split(" / ", 1)
            name = parts[0].strip()
            if name:
                result["confirmer"] = name
            date_str = parts[1].strip()
            if date_str:
                parsed = EmployeeService._parse_date_value(date_str)
                if parsed:
                    result["confirm_date"] = parsed
        else:
            # 可能只是确认人名字
            result["confirmer"] = raw.strip()
        return result

    async def upload_annual_plan(self, file_bytes: bytes) -> dict:
        """从 Excel 批量导入年度培训计划，按年度+部门自动分类。

        支持两种 Excel 格式：
        1. 系统导出格式：标题行 → 部门行 → 表头行 → 数据行
        2. 自定义格式：第一行直接为表头，需包含「年度」「部门」列

        每行使用独立 SAVEPOINT，单行失败不影响其他行。
        """
        from io import BytesIO

        from openpyxl import load_workbook

        from app.modules.hr.models import AnnualTrainingPlan, AnnualTrainingPlanItem

        wb = await asyncio.to_thread(load_workbook, BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("文件为空")

        # ── 智能识别表头行：跳过标题行（如 "2026 年培训计划"、"部门：xxx"）──
        header_row_idx = 0
        col_map: dict[int, str] = {}
        for i, row in enumerate(rows):
            candidate = [str(c).strip() if c else "" for c in row]
            mapping = {idx: self._PLAN_COLUMN_MAP[h] for idx, h in enumerate(candidate) if h in self._PLAN_COLUMN_MAP}
            if len(mapping) >= 2:  # 至少有 2 列能匹配才认为是表头行
                header_row_idx = i
                col_map = mapping
                break

        if not col_map:
            raise ValueError(
                "未识别到有效的表头列。支持的列名：年度/年份、部门、月份、培训人数、课时、"
                "培训内容及使用教材、培训对象、参加岗位/参加人数、授课单位及授课人、"
                "培训方式/考核方式、培训跟踪、确认人/确认者/确认人及日期、备注、培训季度及课时"
            )

        has_year_col = "year" in col_map.values()
        has_dept_col = "department" in col_map.values()

        # 如果表头行不含「年度」或「部门」列，尝试从标题行提取缺失的字段
        plan_year: int | None = None
        plan_dept: str | None = None
        if not has_year_col or not has_dept_col:
            for i in range(header_row_idx):
                cell_text = " ".join(
                    [str(c).strip() for c in rows[i] if c is not None]
                ).strip()
                import re
                if not has_year_col:
                    year_match = re.search(r"(\d{4})\s*年", cell_text)
                    if year_match:
                        plan_year = int(year_match.group(1))
                if not has_dept_col:
                    dept_match = re.search(r"部门[：:]\s*(.+)", cell_text)
                    if dept_match:
                        plan_dept = dept_match.group(1).strip()
            if not has_year_col and plan_year is None:
                plan_year = date.today().year
            if not has_dept_col and plan_dept is None:
                raise ValueError(
                    "Excel 表头缺少「部门」列，且标题行未找到「部门：xxx」。"
                    "请在 Excel 中添加「部门」列，或在标题行写明「部门：xxx」"
                )

        # ── 逐行解析数据（用 begin_nested 做行级 SAVEPOINT 隔离）──
        plan_cache: dict[tuple, AnnualTrainingPlan] = {}
        created, updated, errors = 0, 0, []
        skip_fields = {"year", "department", "_quarter_hours", "_confirmer_date", "variety"}

        data_start = header_row_idx + 1
        for row_idx, row in enumerate(rows[data_start:], start=data_start + 1):
            if all(c is None for c in row):
                continue

            # ── 先解析数据（无 DB 操作），做预校验 ──
            data: dict = {}
            for ci, fn in col_map.items():
                v = row[ci] if ci < len(row) else None
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    continue
                if isinstance(v, str):
                    v = v.strip()

                # 处理组合列
                if fn == "_quarter_hours":
                    combined = self._parse_combined_quarter_hours(str(v))
                    data.update(combined)
                    continue
                if fn == "_confirmer_date":
                    combined = self._parse_combined_confirmer_date(str(v))
                    data.update(combined)
                    continue

                if fn == "month" and isinstance(v, str) and len(v) > 16:
                    v = v[:16]
                if fn == "year":
                    v = int(float(str(v)))
                if fn in ("trainee_count",):
                    v = int(float(str(v)))
                if fn in ("duration_hours", "training_hours"):
                    try:
                        s = str(v).strip().rstrip('hH小时课时')
                        v = float(s)
                    except (ValueError, TypeError):
                        continue
                if fn == "confirm_date":
                    parsed = self._parse_date_value(v)
                    if parsed:
                        v = parsed
                    else:
                        continue
                data[fn] = v

            # 预校验：必须有部门和有效数据
            dept = data.get("department", "") or plan_dept or ""
            year = data.get("year") or plan_year or date.today().year
            if not dept:
                errors.append(f"第{row_idx}行: 缺少部门")
                continue

            item_data = {k: v for k, v in data.items() if k not in skip_fields}
            if not item_data:
                errors.append(f"第{row_idx}行: 无有效数据")
                continue

            cache_key = (int(year), str(dept))

            # ── DB 操作放在嵌套事务内，单行失败不影响其他行 ──
            try:
                nested = await self.repo.session.begin_nested()

                # 找或创建年度计划
                plan = plan_cache.get(cache_key)
                if not plan:
                    q = select(AnnualTrainingPlan).where(
                        AnnualTrainingPlan.year == year,
                        AnnualTrainingPlan.department == dept,
                        AnnualTrainingPlan.is_deleted == False,  # noqa: E712
                    )
                    r = await self.repo.session.execute(q)
                    plan = r.scalar_one_or_none()
                    if not plan:
                        plan = AnnualTrainingPlan(year=int(year), department=str(dept), status="草稿")
                        self.repo.session.add(plan)
                        await self.repo.session.flush()
                    plan_cache[cache_key] = plan

                # 添加计划项（去重：同计划+同内容+同月份视为重复，跳过）
                content = item_data.get("content_and_textbook") or ""
                month_val = item_data.get("month") or ""
                existing_item = (await self.repo.session.execute(
                    select(AnnualTrainingPlanItem).where(
                        AnnualTrainingPlanItem.plan_id == plan.id,
                        AnnualTrainingPlanItem.content_and_textbook == content,
                        AnnualTrainingPlanItem.month == month_val,
                        AnnualTrainingPlanItem.is_deleted == False,  # noqa: E712
                    )
                )).scalar_one_or_none()
                if existing_item:
                    for k, v in item_data.items():
                        setattr(existing_item, k, v)
                    updated += 1
                else:
                    item = AnnualTrainingPlanItem(plan_id=plan.id, **item_data)
                    self.repo.session.add(item)
                    created += 1
                await self.repo.session.flush()

                await nested.commit()
            except Exception as e:
                await nested.rollback()
                # 清除可能已失效的缓存（plan 在本次嵌套事务中新建但被回滚了）
                plan_cache.pop(cache_key, None)
                errors.append(f"第{row_idx}行: {e}")

        return {"created": created, "updated": updated, "errors": errors}

class DepartmentService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = DepartmentRepository(session)

    async def get_department(self, department_id: UUID) -> HrDepartment:
        department = await self.repo.get_by_id(department_id)
        if not department:
            raise NotFoundException("部门", str(department_id))
        return department

    async def create_department(self, data: DepartmentCreate) -> HrDepartment:
        existing = await self.repo.get_by_code(data.code)
        if existing:
            raise DuplicateException("部门编码", data.code)

        department = HrDepartment(**data.model_dump())
        result = await self.repo.create(department)
        return result

    async def update_department(self, department_id: UUID, data: DepartmentUpdate) -> HrDepartment:
        department = await self.get_department(department_id)
        update_data = data.model_dump(exclude_unset=True)

        if "code" in update_data:
            existing = await self.repo.get_by_code(update_data["code"])
            if existing and existing.id != department_id:
                raise DuplicateException("部门编码", update_data["code"])

        for field, value in update_data.items():
            setattr(department, field, value)

        result = await self.repo.update(department)
        return result

    async def delete_department(self, department_id: UUID) -> None:
        await self.repo.session.execute(text("DELETE FROM hr.departments WHERE id = :id"), {"id": department_id})

    async def list_departments(
        self,
        *,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[HrDepartment], int]:
        departments, total = await self.repo.list_departments(
            keyword=keyword,
            page=page,
            page_size=page_size,
        )
        # Attach employee count to each department
        from sqlalchemy import func, select

        from app.modules.hr.models import Employee
        for dept in departments:
            count = await self.repo.session.scalar(
                select(func.count()).select_from(Employee).where(
                    Employee.department == dept.name,
                    Employee.is_deleted.is_(False),
                    Employee.status != "离职",
                )
            )
            dept.employee_count = count or 0

        return departments, total

class TeamService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = TeamRepository(session)
        self.department_repo = DepartmentRepository(session)

    async def get_team(self, team_id: UUID) -> Team:
        team = await self.repo.get_by_id(team_id)
        if not team:
            raise NotFoundException("班组", str(team_id))
        return team

    async def create_team(self, data: TeamCreate) -> Team:
        department = await self.department_repo.get_by_id(data.department_id)
        if not department:
            raise NotFoundException("部门", str(data.department_id))

        team = Team(**data.model_dump())
        result = await self.repo.create(team)
        return result

    async def update_team(self, team_id: UUID, data: TeamUpdate) -> Team:
        team = await self.get_team(team_id)
        update_data = data.model_dump(exclude_unset=True)

        if "department_id" in update_data:
            department = await self.department_repo.get_by_id(update_data["department_id"])
            if not department:
                raise NotFoundException("部门", str(update_data["department_id"]))

        for field, value in update_data.items():
            setattr(team, field, value)

        result = await self.repo.update(team)
        return result

    async def delete_team(self, team_id: UUID) -> None:
        await self.repo.session.execute(text("DELETE FROM hr.teams WHERE id = :id"), {"id": team_id})

    async def list_teams(
        self,
        *,
        department_id: UUID | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[Team], int]:
        return await self.repo.list_teams(
            department_id=department_id,
            keyword=keyword,
            page=page,
            page_size=page_size,
        )

class OffboardingRecordService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = OffboardingRecordRepository(session)
        self.employee_repo = EmployeeRepository(session)

    async def get_record(self, record_id: UUID) -> OffboardingRecord:
        record = await self.repo.get_by_id(record_id)
        if not record:
            raise NotFoundException("离职记录", str(record_id))
        return record

    async def create_record(self, data: OffboardingRecordCreate) -> OffboardingRecord:
        employee = await self.employee_repo.get_by_id(data.employee_id)
        if not employee:
            raise NotFoundException("员工", str(data.employee_id))

        record = OffboardingRecord(**data.model_dump())
        record = await self.repo.create(record)

        # 自动将员工状态更新为离职
        employee.status = "离职"
        await self.employee_repo.update(employee)

        return record

    async def update_record(self, record_id: UUID, data: OffboardingRecordUpdate) -> OffboardingRecord:
        record = await self.get_record(record_id)
        update_data = data.model_dump(exclude_unset=True)

        for field, value in update_data.items():
            setattr(record, field, value)

        result = await self.repo.update(record)

        return result

    async def delete_record(self, record_id: UUID) -> None:
        # 删除离职记录时恢复员工在职状态（与 departure 流程对称）
        record = await self.get_record(record_id)
        if record and record.employee_id:
            employee = await self.employee_repo.get_by_id(record.employee_id)
            if employee and employee.status == "离职":
                employee.status = "在职"
                await self.employee_repo.update(employee)
        await self.repo.session.execute(text("DELETE FROM hr.offboarding_records WHERE id = :id"), {"id": record_id})

    async def list_records(
        self,
        *,
        employee_id: UUID | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[OffboardingRecord], int]:
        return await self.repo.list_records(
            employee_id=employee_id,
            keyword=keyword,
            page=page,
            page_size=page_size,
        )

class OnboardingRecordService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = OnboardingRecordRepository(session)

    async def get_record(self, record_id: UUID) -> OnboardingRecord:
        record = await self.repo.get_by_id(record_id)
        if not record:
            raise NotFoundException("入职记录", str(record_id))
        return record

    async def list_records(
        self,
        *,
        department: str | None = None,
        position: str | None = None,
        is_employed: str | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
        sort_by: str = "hire_date",
        sort_order: str = "desc",
        days: int = 7,
    ) -> tuple[list[OnboardingRecord], int]:
        if days > 0:
            await self._cleanup_old_records(days)
        return await self.repo.list_records(
            department=department,
            position=position,
            is_employed=is_employed,
            keyword=keyword,
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            sort_order=sort_order,
            days=days,
        )

    async def delete_record(self, record_id: UUID) -> None:
        await self.repo.session.execute(text("DELETE FROM hr.onboarding_records WHERE id = :id"), {"id": record_id})

    async def _cleanup_old_records(self, days: int) -> None:
        """软删除超过 N 天的入职台账记录。"""
        from sqlalchemy import text as sql_text
        await self.repo.session.execute(sql_text("""
            UPDATE hr.onboarding_records SET is_deleted = true, updated_at = now()
            WHERE is_deleted = false AND created_at < now() - make_interval(days => :days)
        """), {"days": days})

class DepartureRecordService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = DepartureRecordRepository(session)
        self.employee_repo = EmployeeRepository(session)

    async def get_record(self, record_id: UUID) -> DepartureRecord:
        record = await self.repo.get_by_id(record_id)
        if not record:
            raise NotFoundException("离职台账记录", str(record_id))
        return record

    async def list_records(
        self,
        *,
        department: str | None = None,
        offboarding_type: str | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
        sort_by: str = "offboarding_date",
        sort_order: str = "desc",
    ) -> tuple[list[DepartureRecord], int]:
        return await self.repo.list_records(
            department=department,
            offboarding_type=offboarding_type,
            keyword=keyword,
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            sort_order=sort_order,
        )

    async def create_record(self, data: DepartureRecordCreate) -> DepartureRecord:
        record = DepartureRecord(**data.model_dump())
        record = await self.repo.create(record)

        # 自动将对应员工状态更新为离职（按姓名+部门匹配）
        employee = await self.employee_repo.get_by_name_and_department(data.name, data.department)
        if employee and employee.status != "离职":
            employee.status = "离职"
            await self.employee_repo.update(employee)

        # 同步更新入职台账为不在职，确保所有列表过滤生效
        from app.modules.hr.models import OnboardingRecord
        onboarding_q = select(OnboardingRecord).where(
            OnboardingRecord.name == data.name,
            OnboardingRecord.department == data.department,
            OnboardingRecord.is_deleted == False,  # noqa: E712
        )
        onboarding_r = await self.repo.session.execute(onboarding_q)
        for onboarding in onboarding_r.scalars().all():
            if onboarding.is_employed != "否":
                onboarding.is_employed = "否"
                self.repo.session.add(onboarding)

        await self.repo.session.flush()
        return record

    async def update_record(self, record_id: UUID, data: DepartureRecordUpdate) -> DepartureRecord:
        record = await self.get_record(record_id)
        update_data = data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(record, field, value)
        return await self.repo.update(record)

    async def delete_record(self, record_id: UUID) -> None:
        # 恢复入职台账的在职状态和员工状态
        record = await self.get_record(record_id)
        from app.modules.hr.models import OnboardingRecord
        onboarding_q = select(OnboardingRecord).where(
            OnboardingRecord.name == record.name,
            OnboardingRecord.department == record.department,
            OnboardingRecord.is_deleted == False,  # noqa: E712
        )
        onboarding_r = await self.repo.session.execute(onboarding_q)
        for onboarding in onboarding_r.scalars().all():
            onboarding.is_employed = "是"
            self.repo.session.add(onboarding)

        employee = await self.employee_repo.get_by_name_and_department(record.name, record.department)
        if employee and employee.status == "离职":
            employee.status = "在职"
            await self.employee_repo.update(employee)

        await self.repo.session.execute(text("DELETE FROM hr.departure_records WHERE id = :id"), {"id": record_id})

class TrainingLedgerService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = TrainingLedgerRepository(session)

    async def get_record(self, record_id: UUID) -> TrainingLedger:
        record = await self.repo.get_by_id(record_id)
        if not record:
            raise NotFoundException("培训台账记录", str(record_id))
        return record

    async def create_record(self, data: TrainingLedgerCreate) -> TrainingLedger:
        # 去重：同员工+同日期+同主题视为重复
        existing = await self.repo.session.execute(
            select(TrainingLedger).where(
                TrainingLedger.employee_number == data.employee_number,
                TrainingLedger.training_date == data.training_date,
                TrainingLedger.training_subject == data.training_subject,
                TrainingLedger.is_deleted == False,  # noqa: E712
            )
        )
        dup = existing.scalar_one_or_none()
        if dup:
            return dup
        record = TrainingLedger(**data.model_dump())
        return await self.repo.create(record)

    async def update_record(
        self, record_id: UUID, data: TrainingLedgerUpdate
    ) -> TrainingLedger:
        record = await self.get_record(record_id)
        update_data = data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(record, field, value)
        return await self.repo.update(record)

    async def delete_record(self, record_id: UUID) -> None:
        record = await self.get_record(record_id)
        record.is_deleted = True
        await self.repo.session.flush()

    async def list_records(
        self,
        *,
        employee_number: str | None = None,
        date_from: date | None = None,
        date_to: date | None = None,
        page: int = 1,
        page_size: int = 20,
        sort_by: str = "training_date",
        sort_order: str = "asc",
    ) -> tuple[list[TrainingLedger], int]:
        # 排除已离职员工的培训记录
        from app.modules.hr.models import OnboardingRecord
        departed_subq = select(OnboardingRecord.employee_number).where(
            OnboardingRecord.is_deleted == False,  # noqa: E712
            OnboardingRecord.is_employed == "否",
        )
        return await self.repo.list_records(
            employee_number=employee_number,
            date_from=date_from,
            date_to=date_to,
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            sort_order=sort_order,
            exclude_employee_numbers=departed_subq,
        )

    async def create_from_notification(
        self,
        *,
        employee_number: str,
        training_date: date,
        training_subject: str,
        training_method: str | None,
        trainer: str | None,
        source_id: str | None = None,
    ) -> TrainingLedger | None:
        """当培训通知包含特定员工时，自动创建培训台账记录。"""
        if source_id:
            existing = await self.repo.get_by_source("notification", source_id)
            if existing:
                return existing

        record = TrainingLedger(
            employee_number=employee_number,
            training_date=training_date,
            training_subject=training_subject,
            training_method=training_method,
            trainer=trainer,
            source_type="notification",
            source_id=source_id,
        )
        return await self.repo.create(record)

class TrainingLedgerPageService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = TrainingLedgerPageRepository(session)

    async def list_pages(self) -> list[TrainingLedgerPage]:
        return await self.repo.list_pages()

    async def list_pages_with_department(self) -> list[tuple[TrainingLedgerPage, str | None]]:
        return await self.repo.list_pages_with_department()

    async def create_page(self, data) -> TrainingLedgerPage:
        existing = await self.repo.get_by_employee_number(data.employee_number)
        if existing:
            raise DuplicateException("培训台账页面", data.employee_number)
        page = TrainingLedgerPage(**data.model_dump())
        return await self.repo.create(page)

class AnnualTrainingPlanService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = AnnualTrainingPlanRepository(session)
        self.item_repo = AnnualTrainingPlanItemRepository(session)

    async def get_plan(self, plan_id: UUID) -> AnnualTrainingPlan:
        plan = await self.repo.get_by_id(plan_id)
        if not plan:
            raise NotFoundException("年度培训计划", str(plan_id))
        return plan

    async def create_plan(self, data: AnnualTrainingPlanCreate) -> AnnualTrainingPlan:
        existing = await self.repo.get_by_year_and_department(data.year, data.department)
        if existing:
            raise DuplicateException("年度培训计划", f"{data.year}年-{data.department}")
        plan = AnnualTrainingPlan(**data.model_dump())
        return await self.repo.create(plan)

    async def update_plan(self, plan_id: UUID, data: AnnualTrainingPlanUpdate) -> AnnualTrainingPlan:
        plan = await self.get_plan(plan_id)
        update_data = data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(plan, field, value)
        return await self.repo.update(plan)

    async def delete_plan(self, plan_id: UUID) -> None:
        await self.repo.session.execute(text("DELETE FROM hr.annual_training_plans WHERE id = :id"), {"id": plan_id})

    async def list_plans(
        self,
        *,
        year: int | None = None,
        department: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[AnnualTrainingPlan], int]:
        return await self.repo.list_plans(
            year=year,
            department=department,
            page=page,
            page_size=page_size,
        )

class AnnualTrainingPlanItemService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = AnnualTrainingPlanItemRepository(session)
        self.plan_repo = AnnualTrainingPlanRepository(session)

    async def list_items(self, plan_id: UUID) -> list[AnnualTrainingPlanItem]:
        return await self.repo.list_items(plan_id)

    async def batch_update_items(
        self, plan_id: UUID, data: AnnualTrainingPlanItemBatchUpdate
    ) -> list[AnnualTrainingPlanItem]:
        plan = await self.plan_repo.get_by_id(plan_id)
        if not plan:
            raise NotFoundException("年度培训计划", str(plan_id))

        # 删除旧明细
        await self.repo.delete_by_plan_id(plan_id)

        # 创建新明细
        results: list[AnnualTrainingPlanItem] = []
        for idx, item_data in enumerate(data.items):
            item = AnnualTrainingPlanItem(
                plan_id=plan_id,
                sort_order=idx,
                **item_data.model_dump(exclude={"sort_order"}),
            )
            created = await self.repo.create(item)
            results.append(created)
        return results


# ─── Recruitment Services ───

# 候选人状态流转规则
_CANDIDATE_TRANSITIONS: dict[str, set[str]] = {
    "待筛选": {"已筛选", "已拒绝"},
    "已筛选": {"待部门审核", "已拒绝"},
    "待部门审核": {"面试中", "已拒绝"},
    "面试中": {"已面试", "已拒绝"},
    "已面试": {"录用中", "已拒绝"},
    "录用中": {"已录用", "已拒绝"},
    "已录用": {"待入职审批", "已拒绝"},
    "待入职审批": {"已录用", "已入职", "已拒绝"},
    "已拒绝": set(),
}


class JobRequirementService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = JobRequirementRepository(session)

    async def list_all(self, *, status: str | None = None) -> list[JobRequirement]:
        return await self.repo.list_all(status=status)

    async def get(self, req_id: UUID) -> JobRequirement:
        req = await self.repo.get_by_id(req_id)
        if not req:
            raise NotFoundException("岗位需求", str(req_id))
        return req

    async def create(self, data: JobRequirementCreate) -> JobRequirement:
        req = JobRequirement(**data.model_dump())
        return await self.repo.create(req)

    async def update(self, req_id: UUID, data: JobRequirementUpdate) -> JobRequirement:
        req = await self.get(req_id)
        for k, v in data.model_dump(exclude_unset=True).items():
            setattr(req, k, v)
        return await self.repo.update(req)

    async def delete(self, req_id: UUID) -> None:
        req = await self.get(req_id)
        await self.repo.soft_delete(req.id)


class CandidateService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = CandidateRepository(session)
        self.log_repo = CandidateStatusLogRepository(session)

    async def list_all(
        self, *, job_requirement_id: UUID | None = None, status: str | None = None,
        keyword: str | None = None, candidate_type: str | None = None,
        page: int = 1, page_size: int = 100,
    ) -> tuple[list[Candidate], int]:
        return await self.repo.list_all(
            job_requirement_id=job_requirement_id, status=status,
            keyword=keyword, candidate_type=candidate_type,
            page=page, page_size=page_size,
        )

    async def get(self, candidate_id: UUID) -> Candidate:
        c = await self.repo.get_by_id(candidate_id)
        if not c:
            raise NotFoundException("候选人", str(candidate_id))
        return c

    async def create(self, data: CandidateCreate) -> Candidate:
        c = Candidate(**data.model_dump())
        result = await self.repo.create(c)
        # 记录初始状态
        self.repo.session.add(CandidateStatusLog(
            candidate_id=result.id, from_status=None, to_status=result.status, remark="创建候选人"
        ))
        await self.repo.session.flush()
        return result

    async def update(self, candidate_id: UUID, data: CandidateUpdate) -> Candidate:
        c = await self.get(candidate_id)
        for k, v in data.model_dump(exclude_unset=True).items():
            setattr(c, k, v)
        return await self.repo.update(c)

    async def delete(self, candidate_id: UUID) -> None:
        c = await self.get(candidate_id)
        await self.repo.soft_delete(c.id)

    # ─── Excel 批量导入 ───

    _CANDIDATE_UPLOAD_COLUMN_MAP: dict[str, str] = {
        "姓名": "name",
        "手机": "phone",
        "邮箱": "email",
        "应聘岗位": "position",
        "部门": "department",
        "性别": "gender",
        "学校": "school",
        "学历": "education",
        "专业": "major",
        "毕业时间": "graduation_date",
        "推荐等级": "recommendation_level",
        "候选人类型": "candidate_type",
        "简历来源": "source",
        "期望薪资": "expected_salary",
        "当前公司": "current_company",
        "工作年限": "work_years",
        "备注": "notes",
    }

    _CANDIDATE_DATE_FIELDS: set[str] = {"graduation_date"}
    _CANDIDATE_INT_FIELDS: set[str] = {"work_years"}

    async def upload_candidates(self, file_bytes: bytes) -> dict:
        """从 Excel 批量导入候选人，按姓名+手机/邮箱 upsert。"""
        from io import BytesIO

        from openpyxl import load_workbook
        from sqlalchemy import text as sa_text

        wb = await asyncio.to_thread(load_workbook, BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("文件为空")

        header = [str(c).strip() if c else "" for c in rows[0]]
        col_map: dict[int, str] = {}
        for idx, col_name in enumerate(header):
            field = self._CANDIDATE_UPLOAD_COLUMN_MAP.get(col_name)
            if field:
                col_map[idx] = field

        if "name" not in col_map.values():
            raise ValueError("缺少「姓名」列，无法导入")

        created = 0
        updated = 0
        errors: list[str] = []

        conn = await self.repo.session.connection()
        sp_id = 0

        for row_idx, row in enumerate(rows[1:], start=2):
            if all(c is None for c in row):
                continue
            sp_id += 1
            sp_name = f"cand_row_{sp_id}"
            try:
                await conn.execute(sa_text(f"SAVEPOINT {sp_name}"))
                data: dict = {}
                name = ""
                phone = None
                email = None
                for col_idx, field_name in col_map.items():
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is None or (isinstance(val, str) and val.strip() == ""):
                        continue
                    if isinstance(val, str):
                        val = val.strip()
                    if field_name == "name":
                        name = str(val)
                    elif field_name == "phone":
                        phone = str(val)
                    elif field_name == "email":
                        email = str(val)
                    if field_name in self._CANDIDATE_DATE_FIELDS:
                        parsed = self._parse_date_value(val)
                        if parsed:
                            val = parsed
                        else:
                            continue
                    elif field_name in self._CANDIDATE_INT_FIELDS:
                        try:
                            val = int(float(str(val)))
                        except (ValueError, TypeError):
                            continue
                    data[field_name] = val

                if not name:
                    raise ValueError("姓名为空")
                data.setdefault("status", "待筛选")
                data.setdefault("candidate_type", "职能")

                await self.repo.upsert(name, phone, email, data)
                # 如果是新候选人也记录初始状态日志（upsert 内部 create 不会记 log）
                created += 1  # 简化：不做精确 is_new 判断，统一计数
                await conn.execute(sa_text(f"RELEASE SAVEPOINT {sp_name}"))
            except Exception as e:
                await conn.execute(sa_text(f"ROLLBACK TO SAVEPOINT {sp_name}"))
                try:
                    await conn.execute(sa_text(f"RELEASE SAVEPOINT {sp_name}"))
                except Exception:
                    pass
                errors.append(f"第{row_idx}行: {e}")

        return {"created": created, "updated": updated, "errors": errors}

    @staticmethod
    def _parse_date_value(val) -> date | None:
        """日期解析，复用 EmployeeService 模式。"""
        from datetime import datetime as _dt
        if isinstance(val, (int, float)):
            try:
                return date.fromordinal(date(1900, 1, 1).toordinal() + int(val) - 2)
            except Exception:
                return None
        if isinstance(val, _dt):
            return val.date()
        if isinstance(val, date):
            return val
        import re as _re
        s = str(val).strip()
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"]:
            try:
                return _dt.strptime(s, fmt).date()
            except ValueError:
                continue
        m = _re.match(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", s)
        if m:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        return None

    async def transition_status(self, candidate_id: UUID, new_status: str, remark: str | None = None) -> Candidate:
        c = await self.get(candidate_id)
        allowed = _CANDIDATE_TRANSITIONS.get(c.status, set())
        if new_status not in allowed:
            raise ValueError(f"不允许从「{c.status}」变更为「{new_status}」")
        old_status = c.status
        c.status = new_status
        result = await self.repo.update(c)
        self.repo.session.add(CandidateStatusLog(
            candidate_id=result.id, from_status=old_status, to_status=new_status, remark=remark
        ))
        await self.repo.session.flush()
        return result

    async def get_status_logs(self, candidate_id: UUID) -> list[CandidateStatusLog]:
        return await self.log_repo.list_by_candidate(candidate_id)

    async def onboard(self, candidate_id: UUID) -> tuple[Candidate, OnboardingRecord, str]:
        """候选人入职：校验状态、生成工号、创建入职记录、更新岗位需求 hired_count。"""
        import uuid as _uuid
        from datetime import date as date_type

        from app.modules.hr.models import OnboardingRecord
        from app.modules.hr.repository import (
            CandidateRepository,
            JobRequirementRepository,
        )
        from app.modules.hr.schemas import CandidateUpdate

        # SELECT FOR UPDATE 防止并发入职
        candidate_repo = CandidateRepository(self.repo.session)
        c = await candidate_repo.get_by_id_for_update(candidate_id)
        if not c:
            raise NotFoundException("候选人", str(candidate_id))
        if c.status != "待入职审批":
            raise ValueError("候选人状态必须为「待入职审批」才能入职（请先通过入职审批）")

        # 生成工号
        emp_no = f"ZP{date_type.today().strftime('%y%m%d')}{str(_uuid.uuid4())[:4].upper()}"

        onboarding = OnboardingRecord(
            employee_number=emp_no,
            name=c.name or "",
            department=c.department or "未分配",
            position=c.position or "未分配",
            hire_date=date_type.today(),
            phone=c.phone,
            email=c.email,
            education=c.education,
            school=c.school,
            major=c.major,
            source="recruitment",
        )
        self.repo.session.add(onboarding)
        await self.repo.session.flush()

        # 自动创建入职子任务
        task_types = [
            ("体检确认", 0), ("资料审核", 1), ("合同签署", 2), ("入职培训", 3),
        ]
        for task_type, sort_order in task_types:
            task = OnboardingTask(
                candidate_id=candidate_id,
                task_type=task_type,
                sort_order=sort_order,
                status="待完成",
            )
            self.repo.session.add(task)
        await self.repo.session.flush()

        # 更新候选人状态
        await self.update(candidate_id, CandidateUpdate(status="已入职", offer_status="已接受"))

        # 更新岗位需求 hired_count
        if c.job_requirement_id:
            jd_repo = JobRequirementRepository(self.repo.session)
            await jd_repo.increment_hired_count(c.job_requirement_id)

        # UPDATE 后 re-fetch，确保返回最新状态
        c = await candidate_repo.get_by_id(candidate_id) or c

        return c, onboarding, emp_no


class InterviewService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = InterviewRepository(session)

    async def list_by_candidate(self, candidate_id: UUID) -> list[Interview]:
        return await self.repo.list_by_candidate(candidate_id)

    async def get(self, interview_id: UUID) -> Interview:
        iv = await self.repo.get_by_id(interview_id)
        if not iv:
            raise NotFoundException("面试记录", str(interview_id))
        return iv

    async def create(self, data: InterviewCreate) -> Interview:
        create_calendar = data.create_calendar_event
        iv = Interview(**data.model_dump(exclude={"create_calendar_event"}))
        result = await self.repo.create(iv)

        if create_calendar and data.interview_date:
            try:
                candidate_repo = CandidateRepository(self.repo.session)
                candidate = await candidate_repo.get_by_id(data.candidate_id)
                candidate_name = candidate.name if candidate else "候选人"
                position = candidate.position if candidate else ""

                from app.modules.hr.feishu_calendar import FeishuCalendarService
                calendar = FeishuCalendarService()
                event_id = await calendar.create_interview_event(
                    interview_id=result.id,
                    candidate_name=candidate_name,
                    position=position,
                    interview_date_val=data.interview_date,
                    interviewer_name=data.interviewer or "未指定",
                    location=data.location or "未指定",
                    interview_type=data.interview_type,
                )
                result.calendar_event_id = event_id
                result = await self.repo.update(result)
            except Exception as e:
                logger.warning("创建日历事件失败（面试已保存）: %s", e)

        return result

    async def update(self, interview_id: UUID, data: InterviewUpdate) -> Interview:
        iv = await self.get(interview_id)
        old_date = iv.interview_date
        old_location = iv.location
        old_interviewer = iv.interviewer

        for k, v in data.model_dump(exclude_unset=True).items():
            setattr(iv, k, v)
        result = await self.repo.update(iv)

        if result.calendar_event_id and (
            old_date != result.interview_date
            or old_location != result.location
            or old_interviewer != result.interviewer
        ):
            try:
                from app.modules.hr.feishu_calendar import FeishuCalendarService
                calendar = FeishuCalendarService()
                await calendar.update_interview_event(
                    result.calendar_event_id,
                    interview_date_val=result.interview_date,
                    location=result.location,
                    interviewer_name=result.interviewer,
                )
            except Exception as e:
                logger.warning("同步日历事件失败（面试已更新）: %s", e)

        return result

    async def delete(self, interview_id: UUID) -> None:
        iv = await self.get(interview_id)
        if iv.calendar_event_id:
            try:
                from app.modules.hr.feishu_calendar import FeishuCalendarService
                calendar = FeishuCalendarService()
                await calendar.delete_interview_event(iv.calendar_event_id)
            except Exception as e:
                logger.warning("删除日历事件失败: %s", e)
        await self.repo.soft_delete(iv.id)


class AiEvaluationService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = CandidateAiEvaluationRepository(session)

    async def get_by_interview(self, interview_id: UUID) -> CandidateAiEvaluation | None:
        return await self.repo.get_by_interview(interview_id)

    async def evaluate(self, interview_id: UUID) -> CandidateAiEvaluation:
        """对面试逐字稿进行AI评估"""
        interview_repo = InterviewRepository(self.repo.session)
        interview = await interview_repo.get_by_id(interview_id)
        if not interview:
            raise NotFoundException("面试记录", str(interview_id))
        if not interview.transcript_text or not interview.transcript_text.strip():
            raise ValueError("请先填写面试逐字稿")

        candidate_repo = CandidateRepository(self.repo.session)
        candidate = await candidate_repo.get_by_id(interview.candidate_id)
        if not candidate:
            raise NotFoundException("候选人", str(interview.candidate_id))

        # 获取JD文本
        jd_text = ""
        if interview.job_requirement_id:
            jd_repo = JobRequirementRepository(self.repo.session)
            jd = await jd_repo.get_by_id(interview.job_requirement_id)
            if jd and jd.requirements:
                jd_text = jd.requirements

        # 组装简历文本
        resume_parts = [
            f"姓名：{candidate.name}",
            f"学校：{candidate.school or '未知'}",
            f"学历：{candidate.education or '未知'}",
            f"专业：{candidate.major or '未知'}",
        ]
        if candidate.current_company:
            resume_parts.append(f"当前公司：{candidate.current_company}")
        if candidate.work_years is not None:
            resume_parts.append(f"工作年限：{candidate.work_years}年")
        resume_text = "\n".join(resume_parts)

        # 调用LLM评估
        from app.modules.hr.ai_service import AiChatService
        prompt = self._build_evaluation_prompt(jd_text, resume_text, interview.transcript_text)
        result = await AiChatService.call_json(prompt)

        # 删除旧评估
        existing = await self.repo.get_by_interview(interview_id)
        if existing:
            existing.is_deleted = True
            await self.repo.session.flush()

        eval_obj = CandidateAiEvaluation(
            candidate_id=interview.candidate_id,
            job_requirement_id=interview.job_requirement_id,
            interview_id=interview_id,
            jd_match_score=result.get("jd_match_score"),
            professional_score=result.get("professional_score"),
            communication_score=result.get("communication_score"),
            learning_score=result.get("learning_score"),
            stability_score=result.get("stability_score"),
            overall_score=result.get("overall_score"),
            strengths=result.get("strengths"),
            weaknesses=result.get("weaknesses"),
            ai_summary=result.get("ai_summary"),
            risk_flags=result.get("risk_flags"),
            jd_text_snapshot=jd_text,
            transcript_snapshot=interview.transcript_text,
            model_version="deepseek-v3",
        )
        eval_obj = await self.repo.create(eval_obj)

        # 回写候选人 match_report
        c = await candidate_repo.get_by_id(interview.candidate_id)
        if c:
            c.match_report = result.get("ai_summary")
            await candidate_repo.update(c)

        return eval_obj

    @staticmethod
    def _build_evaluation_prompt(jd_text: str, resume_text: str, transcript: str) -> str:
        return f"""你是一位专业的制药行业招聘评估专家。请根据以下信息对候选人进行综合评价。

## 岗位JD
{jd_text or "（未提供JD）"}

## 候选人简历
{resume_text}

## 面试逐字稿
{transcript}

请严格按照以下JSON格式输出评估结果（不要输出其他内容）：
{{
  "jd_match_score": <1-10的数字，JD匹配度>,
  "professional_score": <1-10的数字，专业能力>,
  "communication_score": <1-10的数字，沟通表达>,
  "learning_score": <1-10的数字，学习能力>,
  "stability_score": <1-10的数字，稳定性评估>,
  "overall_score": <1-10的数字，综合评分>,
  "strengths": "候选人优势（3-5条，每条30字以内，用\\n分隔）",
  "weaknesses": "候选人不足（2-3条，每条30字以内，用\\n分隔）",
  "ai_summary": "综合评价200-300字，包括JD匹配度分析、专业能力评价、软性素质评价、录用建议",
  "risk_flags": "风险提示，无则填'无'"
}}

评分标准：8-10优秀/强烈推荐，6-7良好/可考虑，4-5一般/需对比，1-3不推荐。
评分必须基于面试逐字稿中的实际表现，不要假设简历内容即为真实能力。"""


# ─── Candidate Review Service ───


class CandidateReviewService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = CandidateReviewRepository(session)
        self.candidate_repo = CandidateRepository(session)
        self.jd_repo = JobRequirementRepository(session)

    async def list_pending(self, *, reviewer: str | None = None) -> list[dict]:
        reviews = await self.repo.list_pending(reviewer=reviewer)
        # 批量预取候选人 + 岗位需求，避免 N+1
        candidate_ids = [rv.candidate_id for rv in reviews]
        jd_ids = [rv.job_requirement_id for rv in reviews if rv.job_requirement_id]
        candidate_map = {}
        jd_map = {}
        if candidate_ids:
            candidates = await self.candidate_repo.get_by_ids(candidate_ids)
            candidate_map = {c.id: c for c in candidates}
        if jd_ids:
            jds = await self.jd_repo.get_by_ids(jd_ids)
            jd_map = {j.id: j for j in jds}
        result = []
        for rv in reviews:
            c = candidate_map.get(rv.candidate_id)
            if not c:
                continue
            result.append({
                "review": rv,
                "candidate": c,
                "job_requirement": jd_map.get(rv.job_requirement_id) if rv.job_requirement_id else None,
            })
        return result

    async def push(self, candidate_id: UUID, pushed_by: str, push_note: str | None = None, reviewer: str | None = None) -> CandidateReview:
        c = await self.candidate_repo.get_by_id(candidate_id)
        if not c:
            raise NotFoundException("候选人", str(candidate_id))
        if not c.job_requirement_id:
            raise ValueError("该候选人未关联岗位需求，无法推送审核")

        jd = await self.jd_repo.get_by_id(c.job_requirement_id)
        # 优先用传入的 reviewer，其次用岗位负责人，最后报错
        final_reviewer = reviewer or (jd.owner if jd and jd.owner else None)
        if not final_reviewer:
            raise ValueError("请填写审核人姓名，或为岗位需求设置负责人")

        # 候选人状态前置校验：已入职/已录用/已拒绝 不可推送审核
        if c.status in ("已录用", "已入职", "已拒绝"):
            raise ValueError(f"候选人状态为「{c.status}」，无法推送审核")

        rv = CandidateReview(
            candidate_id=candidate_id,
            job_requirement_id=c.job_requirement_id,
            pushed_by=pushed_by,
            push_note=push_note,
            reviewer=final_reviewer,
            status="待审核",
        )
        result = await self.repo.create(rv)

        # 候选人状态流转
        old_status = c.status
        c.status = "待部门审核"
        await self.candidate_repo.update(c)
        self.candidate_repo.session.add(CandidateStatusLog(
            candidate_id=candidate_id,
            from_status=old_status,
            to_status="待部门审核",
            remark=f"推送至{final_reviewer}审核",
        ))

        # 发送飞书消息卡片给审核人
        await self._send_review_card(result, c, jd, push_note)

        return result

    async def find_pending_review_id(self, candidate_id: UUID) -> UUID:
        """按候选人查找待审核记录的ID（用于审核决策时自动查找）。"""
        rv = await self.repo.get_by_candidate(candidate_id)
        if not rv:
            raise ValueError("未找到审核记录")
        if rv.status != "待审核":
            raise ValueError("该审核已处理")
        return rv.id

    async def decide(self, review_id: UUID, decision: str, review_comment: str | None = None) -> CandidateReview:
        rv = await self.repo.get_by_id(review_id)
        if not rv:
            raise NotFoundException("审核记录", str(review_id))
        if rv.status != "待审核":
            raise ValueError("该审核已处理")

        # 先校验候选人状态，再落库审核结果——避免校验失败时审核记录已被
        # flush 且随路由 commit（候选人卡死、审核无法重做）
        is_onboarding = rv.review_type == "入职审批"
        c = await self.candidate_repo.get_by_id(rv.candidate_id)
        new_status: str | None = None
        if c:
            if is_onboarding:
                if c.status != "待入职审批":
                    raise ValueError(f"候选人状态为「{c.status}」，与入职审批不匹配")
                new_status = "已入职" if decision == "已同意" else "已录用"
            else:
                if c.status not in ("待部门审核", "面试中"):
                    raise ValueError(f"候选人状态为「{c.status}」，与当前审核不匹配，无法更新状态")
                new_status = "面试中" if decision == "已同意" else "已拒绝"

        rv.status = decision
        rv.review_comment = review_comment
        rv.reviewed_at = datetime.now()
        result = await self.repo.update(rv)

        # 候选人状态更新
        if c and new_status:
            old_status = c.status
            c.status = new_status
            await self.candidate_repo.update(c)
            self.candidate_repo.session.add(CandidateStatusLog(
                candidate_id=c.id,
                from_status=old_status,
                to_status=new_status,
                remark=f"审核人{decision}" + (f"：{review_comment}" if review_comment else ""),
            ))

        # 通知HR审核结果
        if rv.pushed_by:
            await self._send_decision_notification(result, c, decision, review_comment)

        return result

    async def _send_review_card(self, review: CandidateReview, candidate: Candidate, jd: JobRequirement | None, push_note: str | None) -> None:
        """发送飞书消息卡片给审核人"""
        try:
            from app.modules.hr.feishu_review_service import send_review_card
            await send_review_card(review, candidate, jd, push_note)
        except Exception as e:
            logger.warning(f"发送飞书审核卡片失败: {e}")

    async def push_onboarding(self, candidate_id: UUID, pushed_by: str, push_note: str | None = None) -> CandidateReview:
        """发起入职审批"""
        c = await self.candidate_repo.get_by_id(candidate_id)
        if not c:
            raise NotFoundException("候选人", str(candidate_id))
        if c.status != "已录用":
            raise ValueError(f"候选人状态必须为「已录用」才能发起入职审批，当前状态：「{c.status}」")

        reviewer = pushed_by  # 入职审批由发起人自行指定审批人，或使用部门负责人
        if c.job_requirement_id:
            jd = await self.jd_repo.get_by_id(c.job_requirement_id)
            if jd and jd.owner:
                reviewer = jd.owner

        rv = CandidateReview(
            candidate_id=candidate_id,
            job_requirement_id=c.job_requirement_id,
            pushed_by=pushed_by,
            push_note=push_note,
            reviewer=reviewer,
            status="待审核",
            review_type="入职审批",
        )
        result = await self.repo.create(rv)

        old_status = c.status
        c.status = "待入职审批"
        await self.candidate_repo.update(c)
        self.candidate_repo.session.add(CandidateStatusLog(
            candidate_id=candidate_id,
            from_status=old_status,
            to_status="待入职审批",
            remark=f"发起入职审批，审批人：{reviewer}",
        ))

        return result

    async def _send_decision_notification(self, review: CandidateReview, candidate: Candidate | None, decision: str, comment: str | None) -> None:
        """通知HR审核结果"""
        try:
            from app.modules.hr.feishu_review_service import send_decision_notification
            await send_decision_notification(review, candidate, decision, comment)
        except Exception as e:
            logger.warning(f"发送飞书审核通知失败: {e}")


# ─── 月度绩效考核 ───


class PerformanceEvaluationService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db
        self.repo = PerformanceEvaluationRepository(db)

    async def list_evaluations(
        self, params: PerformanceListParams, departments: frozenset[str] | None = None,
    ) -> tuple[list[dict], int]:
        evals, total = await self.repo.list_evaluations(
            month=params.evaluation_month,
            department=params.department,
            status=params.status,
            departments=departments,
            page=params.page,
            page_size=params.page_size,
        )
        items = []
        for e in evals:
            items.append(self._to_response_dict(e))
        return items, total

    async def get_evaluation(self, evaluation_id: UUID) -> dict | None:
        e = await self.repo.get_by_id(evaluation_id)
        if e is None:
            return None
        return self._to_response_dict(e)

    async def create_evaluation(self, data: PerformanceEvaluationCreate) -> dict:
        existing = await self.repo.get_by_month_dept(data.evaluation_month, data.department)
        if existing:
            raise ValueError(f"部门 {data.department} 在 {data.evaluation_month} 已有考核记录")

        # 如果未提供部门负责人，从部门培训人员表拉取
        dept_head = data.department_head
        if not dept_head:
            dtp = (await self.db.execute(
                select(DeptTrainingPersonnel).where(
                    DeptTrainingPersonnel.department == data.department,
                    DeptTrainingPersonnel.is_deleted == False,  # noqa: E712
                )
            )).scalars().first()
            if dtp and dtp.department_head:
                dept_head = dtp.department_head

        evaluation = MonthlyPerformanceEvaluation(
            department=data.department,
            department_head=dept_head or "",
            evaluator_leader=data.evaluator_leader,
            evaluation_month=data.evaluation_month,
            headcount=data.headcount,
        )
        evaluation = await self.repo.create(evaluation)

        if data.items:
            await self.repo.upsert_items(evaluation.id, [
                it.model_dump() for it in data.items
            ])
        evaluation = await self.repo.get_by_id(evaluation.id)
        assert evaluation is not None
        return self._to_response_dict(evaluation)

    async def update_evaluation(self, evaluation_id: UUID, data: PerformanceEvaluationUpdate) -> dict:
        e = await self.repo.get_by_id(evaluation_id)
        if e is None:
            raise ValueError("考核记录不存在")
        if e.status not in ("draft",):
            raise ValueError("仅草稿状态可编辑")

        if data.department_head is not None:
            e.department_head = data.department_head
        if data.evaluator_leader is not None:
            e.evaluator_leader = data.evaluator_leader
        if data.headcount is not None:
            e.headcount = data.headcount
        e = await self.repo.update(e)

        if data.items is not None:
            await self.repo.upsert_items(e.id, [
                it.model_dump() for it in data.items
            ])
        e = await self.repo.get_by_id(evaluation_id)
        assert e is not None
        return self._to_response_dict(e)

    async def submit_self(self, evaluation_id: UUID, user_name: str) -> dict:
        e = await self.repo.get_by_id(evaluation_id)
        if e is None:
            raise ValueError("考核记录不存在")
        if e.status != "draft":
            raise ValueError("仅草稿状态可提交自评")
        if user_name and e.department_head and user_name != e.department_head:
            raise PermissionError("仅部门负责人可提交自评")
        e.status = "self_submitted"
        e.self_submitted_at = datetime.now(UTC)
        await self.repo.update(e)
        if _is_notify_enabled():
            await self._notify_leader(e)
        return self._to_response_dict(
            await self.repo.get_by_id(evaluation_id)  # type: ignore[arg-type]
        )

    async def submit_leader(self, evaluation_id: UUID, user_name: str | None = None) -> dict:
        e = await self.repo.get_by_id(evaluation_id)
        if e is None:
            raise ValueError("考核记录不存在")
        if e.status not in ("draft", "self_submitted"):
            raise ValueError("仅草稿或已自评状态可提交领导评分")
        if user_name and e.evaluator_leader and user_name != e.evaluator_leader:
            raise PermissionError("仅分管领导可提交领导评分")
        e.status = "leader_scored"
        e.leader_submitted_at = datetime.now(UTC)
        await self.repo.update(e)
        if _is_notify_enabled():
            await self._notify_dept_head(e)
        return self._to_response_dict(
            await self.repo.get_by_id(evaluation_id)  # type: ignore[arg-type]
        )

    async def auto_create_for_month(self, month: str, departments: list[str]) -> list[dict]:
        created: list[dict] = []
        for dept in departments:
            existing = await self.repo.get_by_month_dept(month, dept)
            if existing:
                continue
            dtp = (await self.db.execute(
                select(DeptTrainingPersonnel).where(
                    DeptTrainingPersonnel.department == dept,
                    DeptTrainingPersonnel.is_deleted == False,  # noqa: E712
                )
            )).scalars().first()
            dept_head = dtp.department_head if dtp and dtp.department_head else ""
            evaluation = MonthlyPerformanceEvaluation(
                department=dept,
                department_head=dept_head,
                evaluation_month=month,
            )
            await self.repo.create(evaluation)
            # 为每个启用的考核项目预填空评分记录
            active_cats = await self.repo.list_categories()
            for cat in active_cats:
                if cat.is_active:
                    await self.repo.upsert_category_score(
                        evaluation.id, cat.id, None, None, cat.weight,
                    )
            # 添加默认考核项目
            default_items = [
                {"category": "key_work", "indicator": "月度重点工作1", "weight": 20, "sort_order": 0},
                {"category": "key_work", "indicator": "月度重点工作2", "weight": 20, "sort_order": 1},
                {"category": "key_work", "indicator": "月度重点工作3", "weight": 20, "sort_order": 2},
                {"category": "routine_work", "indicator": "月度常规工作", "weight": 30, "sort_order": 3},
                {"category": "reward_penalty", "indicator": "奖惩项目", "standard": "造成重大事故可否决绩效总分；获得省级以上奖项可加分", "weight": 10, "sort_order": 4},
            ]
            await self.repo.upsert_items(evaluation.id, default_items)
            # re-fetch 避免 MissingGreenlet（items 是 lazy relationship）
            evaluation = await self.repo.get_by_id(evaluation.id)
            assert evaluation is not None
            # 通知部门负责人（测试阶段默认关闭）
            if dept_head and _is_notify_enabled():
                await self._notify_self_required(evaluation)
            created.append(self._to_response_dict(evaluation))
        # 通知所有项目负责人（仅 PERFORMANCE_NOTIFY=true 时发送）
        if created and _is_notify_enabled():
            await self._notify_evaluators(month, created)
        return created

    async def _notify_evaluators(self, month: str, evaluations: list[dict]) -> None:
        """通知各项目负责人：本月有 N 个部门需要评分"""
        cats = await self.repo.list_categories()
        # evaluator → list of departments
        evaluator_depts: dict[str, list[str]] = {}
        for cat in cats:
            if not cat.is_active or not cat.evaluator:
                continue
            evaluator_depts.setdefault(cat.evaluator, []).extend(
                [e["department"] for e in evaluations]
            )
        for evaluator, depts in evaluator_depts.items():
            unique_depts = list(dict.fromkeys(depts))  # 去重保序
            try:
                from app.modules.hr.feishu_review_service import (
                    _lookup_open_id,
                    _send_card,
                )
                open_id = await _lookup_open_id(evaluator)
                if open_id:
                    card = {
                        "config": {"wide_screen_mode": True},
                        "header": {"title": {"tag": "plain_text", "content": f"📋 {month} 月度考核 - 待评分"}, "template": "blue"},
                        "elements": [
                            {"tag": "markdown", "content": f"您负责的考核项目需要为 **{len(unique_depts)}** 个部门评分：\n\n" + "\n".join(f"• {d}" for d in unique_depts[:10]) + ("\n..." if len(unique_depts) > 10 else "")},
                            {"tag": "hr"},
                            {"tag": "action", "actions": [{"tag": "button", "text": {"tag": "plain_text", "content": "前往评分"}, "type": "primary", "url": f"{_get_base_url()}/hr/performance/score?month={month}"}]},
                        ],
                    }
                    await _send_card(open_id, card)
            except Exception as exc:
                logger.warning(f"通知项目负责人 {evaluator} 失败: {exc}")

    async def _notify_self_required(self, evaluation: MonthlyPerformanceEvaluation) -> None:
        try:
            from app.modules.hr.feishu_review_service import _lookup_open_id, _send_card
            open_id = await _lookup_open_id(evaluation.department_head)
            if open_id:
                card = {
                    "config": {"wide_screen_mode": True},
                    "header": {"title": {"tag": "plain_text", "content": "📋 月度绩效考核 - 待自评"}, "template": "blue"},
                    "elements": [
                        {"tag": "markdown", "content": f"**{evaluation.evaluation_month}** 月度绩效考核已生成，请及时完成自评。\n\n部门：{evaluation.department}\n负责人：{evaluation.department_head}"},
                        {"tag": "hr"},
                        {"tag": "action", "actions": [{"tag": "button", "text": {"tag": "plain_text", "content": "前往填写自评"}, "type": "primary", "url": f"{_get_base_url()}/hr/performance/{evaluation.id}"}]},
                    ],
                }
                await _send_card(open_id, card)
        except Exception as exc:
            logger.warning(f"发送绩效考核自评提醒失败: {exc}")

    async def _notify_leader(self, evaluation: MonthlyPerformanceEvaluation) -> None:
        if not evaluation.evaluator_leader:
            return
        try:
            from app.modules.hr.feishu_review_service import _lookup_open_id, _send_card
            open_id = await _lookup_open_id(evaluation.evaluator_leader)
            if open_id:
                card = {
                    "config": {"wide_screen_mode": True},
                    "header": {"title": {"tag": "plain_text", "content": "📋 月度绩效考核 - 待领导评分"}, "template": "blue"},
                    "elements": [
                        {"tag": "markdown", "content": f"**{evaluation.evaluation_month}** 部门负责人已完成自评，请进行分管领导评分。\n\n部门：{evaluation.department}\n负责人：{evaluation.department_head}"},
                        {"tag": "hr"},
                        {"tag": "action", "actions": [{"tag": "button", "text": {"tag": "plain_text", "content": "前往评分"}, "type": "primary", "url": f"{_get_base_url()}/hr/performance/{evaluation.id}"}]},
                    ],
                }
                await _send_card(open_id, card)
        except Exception as exc:
            logger.warning(f"发送绩效考核领导评分提醒失败: {exc}")

    async def _notify_dept_head(self, evaluation: MonthlyPerformanceEvaluation) -> None:
        try:
            from app.modules.hr.feishu_review_service import _lookup_open_id, _send_card
            open_id = await _lookup_open_id(evaluation.department_head)
            if open_id:
                card = {
                    "config": {"wide_screen_mode": True},
                    "header": {"title": {"tag": "plain_text", "content": "📋 月度绩效考核 - 领导已评分"}, "template": "green"},
                    "elements": [
                        {"tag": "markdown", "content": f"**{evaluation.evaluation_month}** 绩效考核领导评分已完成，请查看结果。\n\n部门：{evaluation.department}"},
                        {"tag": "hr"},
                        {"tag": "action", "actions": [{"tag": "button", "text": {"tag": "plain_text", "content": "查看考核结果"}, "type": "primary", "url": f"{_get_base_url()}/hr/performance/{evaluation.id}"}]},
                    ],
                }
                await _send_card(open_id, card)
        except Exception as exc:
            logger.warning(f"发送绩效考核结果通知失败: {exc}")

    @staticmethod
    def _to_response_dict(e: MonthlyPerformanceEvaluation) -> dict:
        sorted_items = sorted(e.items or [], key=lambda x: x.sort_order)
        return {
            "id": str(e.id),
            "department": e.department,
            "department_head": e.department_head,
            "evaluator_leader": e.evaluator_leader,
            "evaluation_month": e.evaluation_month,
            "headcount": e.headcount,
            "status": e.status,
            "self_submitted_at": e.self_submitted_at.isoformat() if e.self_submitted_at else None,
            "leader_submitted_at": e.leader_submitted_at.isoformat() if e.leader_submitted_at else None,
            "created_at": e.created_at.isoformat() if e.created_at else None,
            "updated_at": e.updated_at.isoformat() if e.updated_at else None,
            "items": [
                {
                    "id": str(it.id),
                    "category": it.category,
                    "indicator": it.indicator,
                    "standard": it.standard,
                    "weight": it.weight,
                    "self_score": it.self_score,
                    "leader_score": it.leader_score,
                    "final_score": it.final_score,
                    "completion": it.completion,
                    "sort_order": it.sort_order,
                }
                for it in sorted_items
            ],
        }


def _get_base_url() -> str:
    from app.core.config import get_settings
    return getattr(get_settings(), "APP_BASE_URL", "http://localhost:3000")


class OnboardingTaskService:
    """入职子任务管理"""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def list_by_candidate(self, candidate_id: UUID) -> list[OnboardingTask]:
        from sqlalchemy import select as _select
        result = await self.session.execute(
            _select(OnboardingTask)
            .where(OnboardingTask.candidate_id == candidate_id, OnboardingTask.is_deleted == False)  # noqa: E712
            .order_by(OnboardingTask.sort_order)
        )
        return list(result.scalars().all())

    async def update(self, task_id: UUID, data: dict) -> OnboardingTask:
        from app.core.exceptions import NotFoundException as _NF
        result = await self.session.execute(
            select(OnboardingTask).where(OnboardingTask.id == task_id, OnboardingTask.is_deleted == False)  # noqa: E712
        )
        task = result.scalar_one_or_none()
        if not task:
            raise _NF("入职任务", str(task_id))
        for k, v in data.items():
            setattr(task, k, v)
        if data.get("status") == "已完成" and not task.completed_at:
            task.completed_at = datetime.now()
        await self.session.flush()
        return task


def _is_notify_enabled() -> bool:
    """是否启用绩效飞书通知（默认关闭，设置 HR_PERFORMANCE_NOTIFY=true 开启）"""
    from app.core.config import get_settings

    return get_settings().HR_PERFORMANCE_NOTIFY


# ─── 候选人胜任度多维分析报告 ───

_ANALYSIS_SYSTEM_PROMPT = """你是企业人力资源部的资深招聘评估专家。基于候选人简历、岗位JD和面试记录，输出结构化的胜任度多维分析报告。评估必须客观、具体，引用事实依据，评分严格（不及格就直说不及格，合格就直说合格）。"""

_ANALYSIS_PROMPT_TEMPLATE = """请对候选人进行多维度胜任度分析，参照以下报告结构输出 JSON：

{{
  "dimensions": [
    {{"name": "学历专业匹配度", "score": 0-100, "star": 1-5, "assessment": "该维度评价"}},
    {{"name": "英语能力匹配度", "score": 0-100, "star": 1-5, "assessment": "该维度评价"}},
    {{"name": "工作经验匹配度", "score": 0-100, "star": 1-5, "assessment": "该维度评价"}},
    {{"name": "专业技能匹配度", "score": 0-100, "star": 1-5, "assessment": "该维度评价"}},
    {{"name": "软素质匹配度", "score": 0-100, "star": 1-5, "assessment": "该维度评价"}},
    {{"name": "稳定性评估", "score": 0-100, "star": 1-5, "assessment": "该维度评价"}},
    {{"name": "薪资匹配度", "score": 0-100, "star": 1-5, "assessment": "该维度评价"}}
  ],
  "strengths": ["核心优势1", "核心优势2"],
  "risks": ["潜在风险1", "潜在风险2"],
  "total_score": 0-100,
  "recommend_level": "强烈推荐/推荐/待定/不推荐 之一",
  "interview_suggestions": ["面试考察重点建议1", "建议2"],
  "training_suggestions": ["录用后培养建议1", "建议2"],
  "summary": "总体结论一段话"
}}

【候选人简历】
{resume_text}

【岗位JD】
{jd_text}

【面试记录】
{interview_text}
"""


class CandidateAnalysisService:
    """候选人胜任度多维分析报告：面试记录提交后自动生成。"""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.repo = CandidateAnalysisReportRepository(session)

    async def list_by_candidate(self, candidate_id: UUID) -> list[CandidateAnalysisReport]:
        return await self.repo.list_by_candidate(candidate_id)

    async def get_by_interview(self, interview_id: UUID) -> CandidateAnalysisReport | None:
        return await self.repo.get_by_interview(interview_id)

    async def generate(
        self, candidate_id: UUID, interview_id: UUID
    ) -> CandidateAnalysisReport:
        """基于简历+JD+面试记录生成报告；面试建议联动写入面试备注。"""
        from datetime import datetime as _dt

        interview = await InterviewRepository(self.session).get_by_id(interview_id)
        if not interview:
            raise NotFoundException("面试记录", str(interview_id))
        interview_text = (interview.transcript_text or "").strip() or (interview.notes or "").strip()
        if not interview_text:
            raise ValueError("请先填写面试记录（逐字稿或备注）")

        candidate = await CandidateRepository(self.session).get_by_id(candidate_id)
        if not candidate:
            raise NotFoundException("候选人", str(candidate_id))

        jd_text = ""
        if interview.job_requirement_id:
            jd = await JobRequirementRepository(self.session).get_by_id(interview.job_requirement_id)
            if jd:
                jd_parts = []
                if jd.duties:
                    jd_parts.append(f"岗位职责：{jd.duties}")
                if jd.requirements:
                    jd_parts.append(f"任职要求：{jd.requirements}")
                jd_text = "\n".join(jd_parts)

        resume_parts = [
            f"姓名：{candidate.name}",
            f"学历：{candidate.education or '未知'}",
            f"学校：{candidate.school or '未知'}",
            f"专业：{candidate.major or '未知'}",
        ]
        if candidate.current_company:
            resume_parts.append(f"当前公司：{candidate.current_company}")
        if candidate.work_years is not None:
            resume_parts.append(f"工作年限：{candidate.work_years}年")
        if candidate.resume_url:
            resume_parts.append(f"简历文件：{candidate.resume_url}")
        resume_text = "\n".join(resume_parts)

        from app.modules.hr.ai_service import AiChatService

        prompt = _ANALYSIS_PROMPT_TEMPLATE.format(
            resume_text=resume_text,
            jd_text=jd_text or "（无岗位JD）",
            interview_text=interview_text[:6000],
        )
        result = await AiChatService.call_json(prompt, system_prompt=_ANALYSIS_SYSTEM_PROMPT)
        if not isinstance(result, dict) or not result.get("dimensions"):
            raise ValueError("AI 返回结构异常，请重试")

        # 同面试记录旧报告软删
        existing = await self.repo.get_by_interview(interview_id)
        if existing:
            existing.is_deleted = True
            await self.repo.session.flush()

        report = await self.repo.create(
            CandidateAnalysisReport(
                candidate_id=candidate_id,
                job_requirement_id=interview.job_requirement_id,
                interview_id=interview_id,
                dimensions=result.get("dimensions"),
                strengths=result.get("strengths"),
                risks=result.get("risks"),
                total_score=result.get("total_score"),
                recommend_level=result.get("recommend_level"),
                interview_suggestions=result.get("interview_suggestions"),
                training_suggestions=result.get("training_suggestions"),
                raw_text=str(result.get("summary") or ""),
                model_version="deepseek-chat",
                generated_at=_dt.now().astimezone(),
            )
        )

        # 联动：面试建议写入面试备注，HR 面试流程中直接可见
        suggestions = result.get("interview_suggestions") or []
        if suggestions:
            prefix = "【AI面试建议】"
            notes = interview.notes or ""
            if prefix not in notes:
                joined = "；".join(str(s) for s in suggestions)
                interview.notes = f"{notes}\n{prefix} {joined}".strip()
                await InterviewRepository(self.session).update(interview)

        # 回写候选人匹配报告摘要
        if result.get("summary"):
            c = await CandidateRepository(self.session).get_by_id(candidate_id)
            if c:
                c.match_report = str(result["summary"])
                await CandidateRepository(self.session).update(c)

        return report
