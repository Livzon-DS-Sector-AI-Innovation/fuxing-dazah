"""有毒有害可燃探测器相关 API 端点。"""

from __future__ import annotations

import logging
from typing import Any, cast
from uuid import UUID

from fastapi import Depends, Query, Request
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.response import paginated_response, success_response
from app.modules.meter import repository as repo
from app.modules.meter import service
from app.modules.meter.api._helpers import _build_report_items
from app.modules.meter.api._router import router
from app.modules.meter.schemas import (
    BatchCreateResult,
    BatchDeleteRequest,
    DateStatsResponse,
    ExportReportRequest,
    GasDetectorBatchCreateRequest,
    GasDetectorCreate,
    GasDetectorFilter,
    GasDetectorFilterOptions,
    GasDetectorListResponse,
    GasDetectorResponse,
    GasDetectorUpdate,
    LedgerImportResult,
    _normalize_department,
)

logger = logging.getLogger(__name__)


@router.post("/gas-detectors/batch", summary="批量新增有毒有害可燃探测器（单次最多 200 条）")
async def batch_create_gas_detectors(
    body: dict[str, Any],
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    req = GasDetectorBatchCreateRequest(**body)
    result = await service.batch_create_gas_detectors(
        db, [item.model_dump() for item in req.items]
    )
    return success_response(
        BatchCreateResult(**result).model_dump(mode="json"),
        status_code=201,
    )



@router.post("/gas-detectors/import-ledger", summary="导入有毒有害探测器台账Excel（按产品编号更新）")
async def import_gas_detector_ledger(
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    """上传探测器台账 Excel，按产品编号更新现有探测器数据。

    仅处理探测器 sheet（含"可燃""有毒""探测器"等关键词的 sheet 或 Sheet 0）。
    支持 .et (WPS) 和 .xlsx 格式，文件限制 50MB。
    """
    from app.core.config import get_settings

    settings_obj = get_settings()
    form = await request.form(max_part_size=settings_obj.MAX_UPLOAD_SIZE_MB * 1024 * 1024)

    file_obj = form.get("file")
    file_obj = cast(Any, file_obj)
    if file_obj is None or not hasattr(file_obj, "read"):
        return JSONResponse(status_code=400, content={"code": 400, "message": "缺少 file 参数"})

    max_size = 50 * 1024 * 1024
    filename = (file_obj.filename if hasattr(file_obj, "filename") else None) or "unknown"

    try:
        file_data = await file_obj.read()
        if len(file_data) > max_size:
            return JSONResponse(
                status_code=400,
                content={"code": 400, "message": "文件大小超过 50MB 限制"},
            )

        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in ("et", "xlsx", "xls"):
            return JSONResponse(
                status_code=400,
                content={"code": 400, "message": "不支持的文件格式，请上传 .et 或 .xlsx 文件"},
            )

        result = await service.import_gas_detector_ledger(db, file_data, filename)
    except ValueError as e:
        return JSONResponse(
            status_code=400,
            content={"code": 400, "message": str(e)},
        )
    except Exception:
        logger.exception("import_gas_detector_ledger failed")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": "导入失败，请检查文件格式是否正确"},
        )

    return success_response(
        LedgerImportResult(**result).model_dump(mode="json"),
        status_code=201,
    )



# ═══════════════════════════════════════════
# 有毒有害可燃探测器
# ═══════════════════════════════════════════


@router.get("/gas-detectors", summary="有毒有害可燃探测器列表")
async def list_gas_detectors(
    filters: GasDetectorFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    records, total = await service.list_gas_detectors(db, filters)

    # 批量查询报告数量
    report_counts = await repo.count_reports_by_gas_detector_ids(db, [r.id for r in records])

    items: list[dict[str, Any]] = []
    for r in records:
        anomaly = r.anomaly_flags or {}
        items.append(
            GasDetectorListResponse(
                id=str(r.id),
                department=r.department,
                instrument_name=r.instrument_name,
                detection_model=r.detection_model,
                measurement_range=r.measurement_range,
                product_number=r.product_number,
                installation_type=r.installation_type,
                installation_location=r.installation_location,
                medium=r.medium,
                calibration_factor=r.calibration_factor,
                manufacturer_supplier=r.manufacturer_supplier,
                manufacturer=r.manufacturer,
                status=service.compute_status(r.status, r.next_calibration_date),
                calibration_date=r.calibration_date,
                next_calibration_date=r.next_calibration_date,
                detection_unit=r.detection_unit,
                calibration_result=r.calibration_result,
                has_anomaly=bool(anomaly),
                report_count=report_counts.get(r.id, 0),
                remark=r.remark,
                updated_at=r.updated_at,
            ).model_dump(mode="json")
        )

    return paginated_response(
        data=items,
        page=filters.page,
        page_size=filters.page_size,
        total=total,
    )



@router.get("/gas-detectors/filter-options", summary="获取有毒有害可燃探测器筛选选项")
async def get_gas_detector_filter_options(
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    options = await service.get_gas_detector_filter_options(db)
    if "department" in options:
        options["department"] = [d for d in (_normalize_department(x) for x in options["department"]) if d is not None]
    return success_response(GasDetectorFilterOptions(**options).model_dump(mode="json"))


@router.get("/gas-detectors/filter-options/search", summary="按字段搜索探测器筛选项（typeahead）")
async def search_gas_detector_filter_options(
    field: str = Query(..., description="字段名（白名单）"),
    q: str | None = Query(default=None, max_length=200, description="搜索关键字，空则返回前 limit 个"),
    limit: int = Query(default=50, ge=1, le=200, description="返回上限"),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    try:
        data = await service.search_gas_detector_filter_options(db, field=field, q=q, limit=limit)
    except ValueError as e:
        return JSONResponse(status_code=400, content={"code": 400, "message": str(e)})
    return success_response(data)


@router.get("/gas-detectors/date-stats", summary="有毒有害可燃探测器日期聚合统计")
async def get_gas_detector_date_stats(
    field: str = Query(default="calibration_date", pattern="^(calibration_date|next_calibration_date)$", description="统计的日期字段"),
    filters: GasDetectorFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    stats = await service.get_gas_detector_date_stats(db, filters, field)
    return success_response(DateStatsResponse(**stats).model_dump(mode="json"))



@router.get("/gas-detectors/export-excel", summary="导出有毒有害可燃探测器为 Excel")
async def export_gas_detectors_excel(
    filters: GasDetectorFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    import io as io_mod

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    # 导出全量数据，绕过 page_size 上限
    object.__setattr__(filters, 'page_size', 100000)
    records, _ = await service.list_gas_detectors(db, filters)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "有毒有害可燃探测器台账"

    headers = [
        "部门", "器具名称", "检测型号", "量程", "产品编号",
        "安装方式", "安装位置", "使用介质", "标定系数", "制造商/供应商",
        "检定时间", "检测单位", "下次检定时间", "检定结论", "制造单位",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, r in enumerate(records, 2):
        values = [
            _normalize_department(r.department), r.instrument_name, r.detection_model,
            r.measurement_range, r.product_number, r.installation_type,
            r.installation_location, r.medium, r.calibration_factor,
            r.manufacturer_supplier, r.calibration_date.isoformat() if r.calibration_date else "",
            r.detection_unit, r.next_calibration_date.isoformat() if r.next_calibration_date else "",
            r.calibration_result, r.manufacturer,
        ]
        for col_idx, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)  # type: ignore[union-attr]

    buf = io_mod.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=gas_detectors_export.xlsx"},
    )



@router.get("/gas-detectors/ids", summary="获取筛选条件下所有探测器 ID（用于跨页全选）")
async def get_gas_detector_ids(
    filters: GasDetectorFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    ids = await service.get_all_gas_detector_ids(db, filters)
    return success_response([str(i) for i in ids])



@router.get("/gas-detectors/{detector_id}", summary="有毒有害可燃探测器详情")
async def get_gas_detector(
    detector_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    obj = await service.get_gas_detector(db, detector_id)
    reports = _build_report_items(obj.reports if hasattr(obj, 'reports') else [])
    return success_response(
        GasDetectorResponse(
            id=str(obj.id),
            instrument_name=obj.instrument_name,
            detection_model=obj.detection_model,
            measurement_range=obj.measurement_range,
            product_number=obj.product_number,
            installation_type=obj.installation_type,
            installation_location=obj.installation_location,
            medium=obj.medium,
            calibration_factor=obj.calibration_factor,
            manufacturer_supplier=obj.manufacturer_supplier,
            calibration_date=obj.calibration_date,
            calibration_result=obj.calibration_result,
            detection_unit=obj.detection_unit,
            next_calibration_date=obj.next_calibration_date,
            manufacturer=obj.manufacturer,
            status=service.compute_status(obj.status, obj.next_calibration_date),
            department=obj.department,
            sheet_name=obj.sheet_name,
            anomaly_flags=obj.anomaly_flags,
            is_deleted=obj.is_deleted,
            created_at=obj.created_at,
            updated_at=obj.updated_at,
            reports=cast(Any, reports),
        ).model_dump(mode="json")
    )



@router.post("/gas-detectors", summary="新增有毒有害可燃探测器")
async def create_gas_detector(
    data: GasDetectorCreate,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    record = await service.create_gas_detector(db, data)
    return success_response(
        GasDetectorResponse(
            id=str(record.id),
            instrument_name=record.instrument_name,
            detection_model=record.detection_model,
            measurement_range=record.measurement_range,
            product_number=record.product_number,
            installation_type=record.installation_type,
            installation_location=record.installation_location,
            medium=record.medium,
            calibration_factor=record.calibration_factor,
            manufacturer_supplier=record.manufacturer_supplier,
            calibration_date=record.calibration_date,
            calibration_result=record.calibration_result,
            detection_unit=record.detection_unit,
            next_calibration_date=record.next_calibration_date,
            manufacturer=record.manufacturer,
            status=service.compute_status(record.status, record.next_calibration_date),
            department=record.department,
            sheet_name=record.sheet_name,
            anomaly_flags=record.anomaly_flags,
            is_deleted=record.is_deleted,
            created_at=record.created_at,
            updated_at=record.updated_at,
            reports=[],
        ).model_dump(mode="json"),
        status_code=201,
    )



@router.put("/gas-detectors/{detector_id}", summary="更新有毒有害可燃探测器")
async def update_gas_detector(
    detector_id: UUID,
    data: GasDetectorUpdate,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    record = await service.update_gas_detector(db, detector_id, data)
    reports = _build_report_items(record.reports if hasattr(record, 'reports') else [])
    return success_response(
        GasDetectorResponse(
            id=str(record.id),
            instrument_name=record.instrument_name,
            detection_model=record.detection_model,
            measurement_range=record.measurement_range,
            product_number=record.product_number,
            installation_type=record.installation_type,
            installation_location=record.installation_location,
            medium=record.medium,
            calibration_factor=record.calibration_factor,
            manufacturer_supplier=record.manufacturer_supplier,
            calibration_date=record.calibration_date,
            calibration_result=record.calibration_result,
            detection_unit=record.detection_unit,
            next_calibration_date=record.next_calibration_date,
            manufacturer=record.manufacturer,
            status=service.compute_status(record.status, record.next_calibration_date),
            department=record.department,
            sheet_name=record.sheet_name,
            anomaly_flags=record.anomaly_flags,
            is_deleted=record.is_deleted,
            created_at=record.created_at,
            updated_at=record.updated_at,
            reports=cast(Any, reports),
        ).model_dump(mode="json")
    )



@router.delete("/gas-detectors/{detector_id}", summary="删除有毒有害可燃探测器（软删除）")
async def delete_gas_detector(
    detector_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    await service.delete_gas_detector(db, detector_id)
    return success_response(message="删除成功")



@router.post("/gas-detectors/batch-delete", summary="批量删除有毒有害可燃探测器（软删除）")
async def batch_delete_gas_detectors(
    body: BatchDeleteRequest,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    ids = [UUID(i) for i in body.ids]
    deleted_count = await service.batch_delete_gas_detectors(db, ids)
    return success_response({"deleted_count": deleted_count}, message=f"成功删除 {deleted_count} 条记录")



@router.post("/gas-detectors/export-reports", summary="批量导出探测器最新报告 ZIP（单次最多 200 份）")
async def export_gas_detector_reports(
    body: ExportReportRequest,
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    ids = [UUID(i) for i in body.ids]
    zip_data, filename, count = await service.export_gas_detector_reports(db, ids)
    return StreamingResponse(
        iter([zip_data]),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Report-Count": str(count),
        },
    )



@router.get("/departments/gas-detectors", summary="获取有毒有害可燃探测器部门列表")
async def list_gas_detector_departments(
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    depts = await service.get_gas_detector_departments(db)
    return success_response([_normalize_department(d) for d in depts])
