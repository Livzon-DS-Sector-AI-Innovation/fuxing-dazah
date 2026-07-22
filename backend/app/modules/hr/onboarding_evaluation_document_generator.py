"""员工上岗评估表 Excel 文档生成器."""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from pydantic import BaseModel, Field


class OnboardingEvaluationInput(BaseModel):
    employee_name: str = Field(..., max_length=64, description="员工姓名")
    employee_number: str | None = Field(None, max_length=32, description="工作卡号")
    gender: str | None = Field(None, max_length=8, description="性别")
    department_position: str | None = Field(None, max_length=128, description="所在部门/岗位")
    hire_date: date | None = Field(None, description="入厂时间")
    training_period: str | None = Field(None, max_length=64, description="培训/考核期")
    regularization_date: date | None = Field(None, description="转正时间")
    assessment_contents: list[str] = Field(default_factory=list, description="上岗培训期内考核内容")
    comprehensive_comment: str | None = Field(None, max_length=1024, description="培训/考核期综合评语")
    is_qualified: bool | None = Field(None, description="是否同意上岗")
    assigned_position: str | None = Field(None, max_length=64, description="担任岗位")
    assessment_method: str | None = Field(None, max_length=32, description="考核方式")
    dept_manager_signature: str | None = Field(None, max_length=64, description="部门负责人签名")
    signature_date: date | None = Field(None, description="签名日期")
    remarks: str | None = Field(None, max_length=512, description="备注")
    dept_manager_agree: bool | None = Field(None, description="部门负责人是否同意")
    hr_manager_agree: bool | None = Field(None, description="人事行政部负责人是否同意")
    qa_manager_agree: bool | None = Field(None, description="质量管理负责人是否同意")
    dept_manager: str | None = Field(None, max_length=64, description="部门负责人")
    hr_manager: str | None = Field(None, max_length=64, description="人事行政部负责人")
    qa_manager: str | None = Field(None, max_length=64, description="质量管理负责人")
    approval_date: date | None = Field(None, description="审批日期")


def _cell_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_onboarding_evaluation(data: OnboardingEvaluationInput) -> BytesIO:
    """根据填写的评估信息生成员工上岗评估表 Excel 文档."""
    wb = Workbook()
    ws = wb.active
    ws.title = "员工上岗评估表"

    # 列宽设置（A-F，共6列）
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    # R1: 格式编号
    ws.merge_cells("A1:F1")
    ws["A1"] = "QR.SOP.PM.003/18（格式）                                        P9/12"
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # R2: 公司名称
    ws.merge_cells("A2:F2")
    ws["A2"] = "丽珠集团福州福兴医药有限公司"
    ws["A2"].font = Font(name="宋体", size=14, bold=True)
    ws["A2"].alignment = _center_align()
    ws.row_dimensions[2].height = 28

    # R3: 标题
    ws.merge_cells("A3:F3")
    ws["A3"] = "员工上岗评估表"
    ws["A3"].font = Font(name="宋体", size=16, bold=True)
    ws["A3"].alignment = _center_align()
    ws.row_dimensions[3].height = 32

    # R4: 姓名 / 性别 / 所在部门/岗位
    ws["A4"] = "姓名"
    ws["A4"].alignment = _center_align()
    ws["A4"].border = _cell_border()
    ws["A4"].font = Font(bold=True)
    ws["B4"] = data.employee_name or ""
    ws["B4"].alignment = _center_align()
    ws["B4"].border = _cell_border()
    ws["C4"] = "性别"
    ws["C4"].alignment = _center_align()
    ws["C4"].border = _cell_border()
    ws["C4"].font = Font(bold=True)
    ws["D4"] = data.gender or ""
    ws["D4"].alignment = _center_align()
    ws["D4"].border = _cell_border()
    ws["E4"] = "所在部门/岗位"
    ws["E4"].alignment = _center_align()
    ws["E4"].border = _cell_border()
    ws["E4"].font = Font(bold=True)
    ws["F4"] = data.department_position or ""
    ws["F4"].alignment = _center_align()
    ws["F4"].border = _cell_border()
    ws.row_dimensions[4].height = 24

    # R5: 工作卡号
    ws["A5"] = "工作卡号"
    ws["A5"].alignment = _center_align()
    ws["A5"].border = _cell_border()
    ws["A5"].font = Font(bold=True)
    ws.merge_cells("B5:F5")
    ws["B5"] = data.employee_number or ""
    ws["B5"].alignment = _center_align()
    ws["B5"].border = _cell_border()
    for col in ["C", "D", "E", "F"]:
        ws[f"{col}5"].border = _cell_border()
    ws.row_dimensions[5].height = 24

    # R6: 入厂时间 / 培训/考核期 / 转正时间
    ws["A6"] = "入厂时间"
    ws["A6"].alignment = _center_align()
    ws["A6"].border = _cell_border()
    ws["A6"].font = Font(bold=True)
    hire_date_str = data.hire_date.strftime("%Y-%m-%d") if data.hire_date else ""
    ws["B6"] = hire_date_str
    ws["B6"].alignment = _center_align()
    ws["B6"].border = _cell_border()
    ws["C6"] = "培训/考核期"
    ws["C6"].alignment = _center_align()
    ws["C6"].border = _cell_border()
    ws["C6"].font = Font(bold=True)
    ws["D6"] = data.training_period or ""
    ws["D6"].alignment = _center_align()
    ws["D6"].border = _cell_border()
    ws["E6"] = "转正时间"
    ws["E6"].alignment = _center_align()
    ws["E6"].border = _cell_border()
    ws["E6"].font = Font(bold=True)
    reg_date_str = data.regularization_date.strftime("%Y-%m-%d") if data.regularization_date else ""
    ws["F6"] = reg_date_str
    ws["F6"].alignment = _center_align()
    ws["F6"].border = _cell_border()
    ws.row_dimensions[6].height = 24

    # R7: 上岗培训期内考核内容、培训内容和结果
    ws.merge_cells("A7:F7")
    ws["A7"] = "上岗培训期内考核内容、培训内容和结果"
    ws["A7"].font = Font(bold=True)
    ws["A7"].alignment = _center_align()
    ws["A7"].border = _cell_border()
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}7"].border = _cell_border()
    ws.row_dimensions[7].height = 24

    # R8-R13: 考核内容填写区域（6行）
    for i in range(6):
        row = 8 + i
        ws.merge_cells(f"A{row}:F{row}")
        content = data.assessment_contents[i] if i < len(data.assessment_contents) else ""
        ws[f"A{row}"] = content
        ws[f"A{row}"].alignment = _left_align()
        ws[f"A{row}"].border = _cell_border()
        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].border = _cell_border()
        ws.row_dimensions[row].height = 24

    # R14: 培训/考核期综合评语
    ws.merge_cells("A14:F14")
    ws["A14"] = "培训/考核期综合评语："
    ws["A14"].font = Font(bold=True)
    ws["A14"].alignment = _left_align()
    ws["A14"].border = _cell_border()
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}14"].border = _cell_border()
    ws.row_dimensions[14].height = 24

    # R15: 评语填写区域
    ws.merge_cells("A15:F15")
    ws["A15"] = data.comprehensive_comment or ""
    ws["A15"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws["A15"].border = _cell_border()
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}15"].border = _cell_border()
    ws.row_dimensions[15].height = 48

    # R16: 同意上岗
    ws.merge_cells("A16:F16")
    position = data.assigned_position or "____"
    agree_str = "☑" if data.is_qualified is True else "□"
    ws["A16"] = f" {agree_str}经考核该员工培训期表现优秀/确认，同意该员工正式上岗，担任{position}岗位。"
    ws["A16"].alignment = _left_align()
    ws["A16"].border = _cell_border()
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}16"].border = _cell_border()
    ws.row_dimensions[16].height = 28

    # R17: 不同意上岗
    ws.merge_cells("A17:F17")
    disagree_str = "☑" if data.is_qualified is False else "□"
    ws["A17"] = f" {disagree_str}经考核该员工培训期内表现不符合此岗位要求，不准上岗。"
    ws["A17"].alignment = _left_align()
    ws["A17"].border = _cell_border()
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}17"].border = _cell_border()
    ws.row_dimensions[17].height = 28

    # R18: 考核方式
    ws.merge_cells("A18:F18")
    method_map = {
        "理论": "☑理论 □实操 □现场",
        "实操": "□理论 ☑实操 □现场",
        "现场": "□理论 □实操 ☑现场",
    }
    method_str = method_map.get(data.assessment_method, "□理论 □实操 □现场")
    ws["A18"] = f" 考核方式：{method_str}"
    ws["A18"].alignment = _left_align()
    ws["A18"].border = _cell_border()
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}18"].border = _cell_border()
    ws.row_dimensions[18].height = 28

    # R19: 部门负责人签名 / 日期
    ws.merge_cells("A19:F19")
    sig_date = data.signature_date.strftime("%Y年%m月%d日") if data.signature_date else ""
    ws["A19"] = f" 部门负责人签名：{data.dept_manager_signature or ''}                   日期：{sig_date}"
    ws["A19"].alignment = _left_align()
    ws["A19"].border = _cell_border()
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}19"].border = _cell_border()
    ws.row_dimensions[19].height = 28

    # R20: 备注
    ws.merge_cells("A20:F20")
    ws["A20"] = f" 备注：{data.remarks or '培训期延长或转岗，由部门主管决定。'}"
    ws["A20"].alignment = _left_align()
    ws["A20"].border = _cell_border()
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}20"].border = _cell_border()
    ws.row_dimensions[20].height = 28

    # R21: 上岗考核审批
    ws.merge_cells("A21:F21")
    ws["A21"] = "上岗考核审批"
    ws["A21"].font = Font(bold=True)
    ws["A21"].alignment = _center_align()
    ws["A21"].border = _cell_border()
    for col in ["B", "C", "D", "E", "F"]:
        ws[f"{col}21"].border = _cell_border()
    ws.row_dimensions[21].height = 28

    # R22-R24: 审批行
    approvals = [
        ("部门负责人", data.dept_manager, data.dept_manager_agree),
        ("人事行政部负责人", data.hr_manager, data.hr_manager_agree),
        ("质量管理负责人", data.qa_manager, data.qa_manager_agree),
    ]
    for i, (title, name, agree) in enumerate(approvals):
        row = 22 + i
        ws.merge_cells(f"A{row}:B{row}")
        agree_str = "☑同意  □不同意" if agree is True else "□同意  ☑不同意" if agree is False else "□同意  □不同意"
        ws[f"A{row}"] = agree_str
        ws[f"A{row}"].alignment = _center_align()
        ws[f"A{row}"].border = _cell_border()
        ws[f"B{row}"].border = _cell_border()

        ws[f"C{row}"] = title
        ws[f"C{row}"].alignment = _center_align()
        ws[f"C{row}"].border = _cell_border()
        ws[f"C{row}"].font = Font(bold=True)

        ws[f"D{row}"] = name or ""
        ws[f"D{row}"].alignment = _center_align()
        ws[f"D{row}"].border = _cell_border()

        ws[f"E{row}"] = "日期"
        ws[f"E{row}"].alignment = _center_align()
        ws[f"E{row}"].border = _cell_border()
        ws[f"E{row}"].font = Font(bold=True)

        app_date = data.approval_date.strftime("%Y-%m-%d") if data.approval_date else ""
        ws[f"F{row}"] = app_date
        ws[f"F{row}"].alignment = _center_align()
        ws[f"F{row}"].border = _cell_border()

        ws.row_dimensions[row].height = 28

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
