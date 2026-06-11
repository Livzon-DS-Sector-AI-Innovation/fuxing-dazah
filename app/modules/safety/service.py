"""Safety business workflows."""

import asyncio
import logging
import os
import uuid
from datetime import datetime
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.safety.models import (
    Accident,
    AIWorkflowConfig,
    APICallConfig,
    Contractor,
    ContractorWorkRecord,
    DailyRiskReport,
    EhsChange,
    HazardReport,
    OhHazardMonitor,
    OhHealthExam,
    SafetyCheck,
    SafetyKnowledgeArticle,
    SafetyTraining,
    SpecialOperationPermit,
    SpecialOperationPersonnel,
    SpecialOperationReport,
    TrainingRecord,
)
from app.modules.safety.repository import SafetyRepository
from app.modules.safety.schemas import (
    AccidentCreate,
    AccidentUpdate,
    AIWorkflowConfigCreate,
    AIWorkflowConfigUpdate,
    APICallConfigCreate,
    APICallConfigUpdate,
    ContractorCreate,
    ContractorUpdate,
    ContractorWorkRecordCreate,
    ContractorWorkRecordUpdate,
    DailyRiskReportCreate,
    DailyRiskReportUpdate,
    EhsChangeCreate,
    EhsChangeUpdate,
    HazardReportCreate,
    HazardReportUpdate,
    SafetyCheckCreate,
    SafetyCheckUpdate,
    SafetyKnowledgeArticleCreate,
    SafetyKnowledgeArticleUpdate,
    SafetyTrainingCreate,
    SafetyTrainingUpdate,
    SpecialOperationPermitCreate,
    SpecialOperationPermitUpdate,
    SpecialOperationPersonnelCreate,
    SpecialOperationPersonnelUpdate,
    SpecialOperationReportCreate,
    SpecialOperationReportUpdate,
    TrainingRecordCreate,
    TrainingRecordUpdate,
)
from app.platform.integrations.ai.client import AIOutputError, AIService
from app.platform.integrations.ai.document_parser import DocumentParser
from app.platform.integrations.ai.prompts import (
    HAZARD_WORKFLOW_STEP_CONFIG,
    SCRIPT_CONFIG,
    STANDALONE_WORKFLOW_CONFIG,
    build_prompt,
)
from app.modules.safety.feishu.notification import send_user_card
from app.core.config import get_settings

logger = logging.getLogger(__name__)

# ── AI 配置默认值（仅用于自动种子和 temperature fallback）──
_DEFAULT_TEXT_AI_CONFIG = {
    "api_base_url": "https://api.deepseek.com",
    "model_name": "deepseek-v4-flash",
    "temperature": 0.1,
    "timeout_seconds": 120,
}
_DEFAULT_VISION_AI_CONFIG = {
    "api_base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1",
    "model_name": "qwen-vl-max",
    "temperature": 0.1,
    "timeout_seconds": 120,
}


async def _ensure_ai_config_seeded(
    session: AsyncSession, config_type: str
) -> None:
    """自动种子：将 env 中的 AI 配置迁移到数据库。

    仅在数据库中尚无该类型活跃配置时执行，
    从环境变量读取 API Key，结合模块默认值创建配置。
    """
    from app.modules.safety.repository import SafetyRepository

    repo = SafetyRepository(session)
    existing = await repo.get_active_api_call_config(config_type)
    if existing:
        return  # 已有活跃配置，无需种子

    api_key = os.environ.get(
        "VISION_AI_API_KEY" if config_type == "vision" else "AI_API_KEY", ""
    ).strip()
    if not api_key:
        logger.warning("环境变量中未找到 %s API Key，跳过自动种子", config_type)
        return

    defaults = (
        _DEFAULT_VISION_AI_CONFIG
        if config_type == "vision"
        else _DEFAULT_TEXT_AI_CONFIG
    )
    logger.info("自动种子 AI 配置: config_type=%s model=%s", config_type, defaults["model_name"])
    await repo.create_api_call_config({
        "config_name": f"默认{'视觉' if config_type == 'vision' else '文本'}模型配置",
        "config_type": config_type,
        "api_base_url": os.environ.get(
            "VISION_AI_BASE_URL" if config_type == "vision" else "AI_BASE_URL",
            defaults["api_base_url"],
        ),
        "api_key": api_key,
        "model_name": os.environ.get(
            "VISION_AI_MODEL" if config_type == "vision" else "AI_MODEL",
            defaults["model_name"],
        ),
        "temperature": defaults["temperature"],
        "timeout_seconds": defaults["timeout_seconds"],
        "is_active": True,
    })
    await session.flush()


class SafetyService:
    """Safety module service"""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.repo = SafetyRepository(session)

    # ==================== SafetyCheck Operations ====================

    async def get_checks(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        check_type: str | None = None,
        department: str | None = None,
    ) -> tuple[list[SafetyCheck], int]:
        """获取安全检查列表"""
        return await self.repo.get_checks(skip, limit, status, check_type, department)

    async def get_check(self, check_id: uuid.UUID) -> SafetyCheck | None:
        """获取安全检查详情"""
        return await self.repo.get_check_by_id(check_id)

    async def create_check(self, data: SafetyCheckCreate) -> SafetyCheck:
        """创建安全检查"""
        check_data = data.model_dump()
        return await self.repo.create_check(check_data)

    async def update_check(
        self, check_id: uuid.UUID, data: SafetyCheckUpdate
    ) -> SafetyCheck | None:
        """更新安全检查"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_check(check_id, update_data)

    async def submit_check(self, check_id: uuid.UUID) -> SafetyCheck | None:
        """提交安全检查（草稿→已提交）"""
        check = await self.repo.get_check_by_id(check_id)
        if not check or check.status != "draft":
            return None
        return await self.repo.update_check(check_id, {"status": "submitted"})

    async def review_check(
        self, check_id: uuid.UUID, result: str
    ) -> SafetyCheck | None:
        """审核安全检查"""
        check = await self.repo.get_check_by_id(check_id)
        if not check or check.status not in ("submitted",):
            return None
        return await self.repo.update_check(
            check_id, {"status": "reviewed", "result": result}
        )

    async def close_check(self, check_id: uuid.UUID) -> SafetyCheck | None:
        """关闭安全检查"""
        check = await self.repo.get_check_by_id(check_id)
        if not check or check.status not in ("reviewed",):
            return None
        return await self.repo.update_check(check_id, {"status": "closed"})

    async def confirm_check(
        self, check_id: uuid.UUID, role: str
    ) -> SafetyCheck | None:
        """确认安全检查（检查人员 / 安全办）"""
        check = await self.repo.get_check_by_id(check_id)
        if not check:
            return None
        if role == "inspector":
            return await self.repo.update_check(
                check_id, {"inspector_confirmed": True}
            )
        elif role == "safety_officer":
            return await self.repo.update_check(
                check_id, {"safety_officer_confirmed": True}
            )
        return None

    async def delete_check(self, check_id: uuid.UUID) -> bool:
        """删除安全检查"""
        return await self.repo.delete_check(check_id)

    # ==================== HazardReport Operations ====================

    async def get_hazards(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        hazard_type: str | None = None,
        hazard_level: str | None = None,
        hazard_category: str | None = None,
        department: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list[HazardReport], int]:
        """获取隐患列表"""
        return await self.repo.get_hazards(
            skip, limit, status, hazard_type, hazard_level, hazard_category,
            department, keyword,
        )

    async def get_hazard(self, hazard_id: uuid.UUID) -> HazardReport | None:
        """获取隐患详情"""
        return await self.repo.get_hazard_by_id(hazard_id)

    async def create_hazard(self, data: HazardReportCreate) -> HazardReport:
        """创建隐患（hazard_no 留空时自动生成），创建后自动执行 AI 工作流。"""
        hazard_data = data.model_dump(exclude_none=True)
        if not hazard_data.get("hazard_no"):
            # 自动生成：HZ-年月日-序号
            today = datetime.now().strftime("%Y%m%d")
            existing = await self.repo.count_hazards_today(today)
            hazard_data["hazard_no"] = f"HZ-{today}-{existing + 1:03d}"
        # 填充默认值以兼容DB NOT NULL约束
        hazard_data.setdefault("hazard_type", "unsafe_condition")
        hazard_data.setdefault("hazard_level", "general")
        hazard_data.setdefault("description", "待AI填写")
        hazard_data.setdefault("discovered_at", datetime.now())
        # 初始化 AI 状态
        hazard_data.setdefault("ai_node_progress", "pending_script1")
        hazard_data.setdefault("overall_status", "draft")
        hazard_data.setdefault("ai_generated", False)

        item = await self.repo.create_hazard(hazard_data)

        # ── 自动执行 AI 工作流 ──
        if item.defect_photos:
            try:
                # Step 1: AI隐患识别（视觉模型）
                item = await self.run_hazard_ai_script(item.id, 1)
                if item and not item.ai_error_message:
                    # Step 2: AI整改建议（文本模型）
                    item = await self.run_hazard_ai_script(item.id, 2)
            except Exception as e:
                logger.error(f"自动执行 AI 工作流失败(hazard {item.id}): {e}")
                # 即使 AI 失败也返回记录，用户可在台账中手动重试

        return item

    async def update_hazard(
        self, hazard_id: uuid.UUID, data: HazardReportUpdate
    ) -> HazardReport | None:
        """更新隐患"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_hazard(hazard_id, update_data)

    async def upload_hazard_photo(
        self, hazard_id: uuid.UUID, file_name: str, file_path: str
    ) -> HazardReport | None:
        """保存隐患图片路径，追加到 defect_photos JSON 数组"""
        import json

        hazard = await self.repo.get_hazard_by_id(hazard_id)
        if not hazard:
            return None
        # 标准化为 / 分隔符：避免 Windows 反斜杠在 JSON 中被当作非法转义符
        safe_path = file_path.replace("\\", "/")
        try:
            photos = json.loads(hazard.defect_photos) if hazard.defect_photos else []
        except (json.JSONDecodeError, TypeError):
            # 兼容历史损坏数据：尝试替换反斜杠后解析，仍失败则视为空列表
            try:
                photos = json.loads(hazard.defect_photos.replace("\\", "/"))
            except Exception:
                photos = []
        if not isinstance(photos, list):
            photos = []
        photos.append(safe_path)
        return await self.repo.update_hazard(
            hazard_id, {"defect_photos": json.dumps(photos, ensure_ascii=False)}
        )

    async def upload_rectification_photo(
        self, hazard_id: uuid.UUID, file_path: str
    ) -> HazardReport | None:
        """保存整改后图片路径，追加到 rectification_photos JSON 数组"""
        import json

        hazard = await self.repo.get_hazard_by_id(hazard_id)
        if not hazard:
            return None
        safe_path = file_path.replace("\\", "/")
        try:
            photos = json.loads(hazard.rectification_photos) if hazard.rectification_photos else []
        except (json.JSONDecodeError, TypeError):
            try:
                photos = json.loads(hazard.rectification_photos.replace("\\", "/"))
            except Exception:
                photos = []
        if not isinstance(photos, list):
            photos = []
        photos.append(safe_path)
        return await self.repo.update_hazard(
            hazard_id, {"rectification_photos": json.dumps(photos, ensure_ascii=False)}
        )

    async def start_rectification(self, hazard_id: uuid.UUID) -> HazardReport | None:
        """开始整改"""
        hazard = await self.repo.get_hazard_by_id(hazard_id)
        if not hazard or hazard.rectification_status != "pending":
            return None
        return await self.repo.update_hazard(
            hazard_id, {"rectification_status": "in_progress"}
        )

    async def complete_rectification(
        self,
        hazard_id: uuid.UUID,
        actual_completion_date: datetime | None = None,
        rectification_photos: str | None = None,
        corrective_preventive_measures: str | None = None,
    ) -> HazardReport | None:
        """完成整改"""
        hazard = await self.repo.get_hazard_by_id(hazard_id)
        if not hazard or hazard.rectification_status != "in_progress":
            return None
        update_data: dict[str, Any] = {
            "rectification_status": "completed",
            "actual_completion_date": actual_completion_date or datetime.now(),
        }
        if rectification_photos:
            update_data["rectification_photos"] = rectification_photos
        if corrective_preventive_measures:
            update_data["corrective_preventive_measures"] = corrective_preventive_measures
        return await self.repo.update_hazard(hazard_id, update_data)

    async def extend_deadline(
        self, hazard_id: uuid.UUID, extended_deadline: datetime
    ) -> HazardReport | None:
        """延期整改"""
        hazard = await self.repo.get_hazard_by_id(hazard_id)
        if not hazard or hazard.rectification_status != "in_progress":
            return None
        return await self.repo.update_hazard(
            hazard_id, {"extended_deadline": extended_deadline}
        )

    async def reply_rectification(
        self,
        hazard_id: uuid.UUID,
        reply_content: str,
        rectification_photos: str | None,
        user_id: uuid.UUID,
        user_name: str,
    ) -> HazardReport | None:
        """整改回复：pending / in_progress → replied"""
        hazard = await self.repo.get_hazard_by_id(hazard_id)
        if not hazard or hazard.rectification_status not in ("pending", "in_progress"):
            return None
        update_data: dict[str, Any] = {
            "rectification_status": "replied",
            "rectification_reply": reply_content,
            "rectification_replied_at": datetime.now(),
            "rectification_replied_by": user_id,
            "rectification_replied_by_name": user_name,
        }
        if rectification_photos is not None:
            update_data["rectification_photos"] = rectification_photos
        updated = await self.repo.update_hazard(hazard_id, update_data)

        # 整改回复后，通知一级复核人
        if updated:
            asyncio.create_task(self._send_verify_notification(updated, 1))

        return updated

    async def verify_level(
        self,
        hazard_id: uuid.UUID,
        level: int,
        action: str,
        opinion: str | None,
        user_id: uuid.UUID,
        user_name: str,
    ) -> HazardReport | None:
        """三级复核：按级别审批或驳回。

        一般隐患：仅需一级（部门负责人）+ 三级（隐患发现人），跳过二级。
        较大/重大隐患：三级全流程。
        """
        hazard = await self.repo.get_hazard_by_id(hazard_id)
        if not hazard:
            return None

        # 判断是否一般隐患（可跳过二级复核）
        is_general = hazard.hazard_level == "general"

        # 检查当前状态是否允许该级别复核
        if level == 1 and hazard.rectification_status != "replied":
            return None
        if level == 2 and hazard.verify_level_1_status != "approved":
            return None
        if level == 3:
            # 一般隐患：一级通过后可直接三级复核
            if is_general:
                if hazard.verify_level_1_status != "approved":
                    return None
            else:
                if hazard.verify_level_2_status != "approved":
                    return None

        now = datetime.now()
        if action == "rejected":
            update_data: dict[str, Any] = {
                "rectification_status": "rejected",
                f"verify_level_{level}_status": "rejected",
                f"verify_level_{level}_by": user_id,
                f"verify_level_{level}_by_name": user_name,
                f"verify_level_{level}_at": now,
                f"verify_level_{level}_opinion": opinion,
            }
            return await self.repo.update_hazard(hazard_id, update_data)

        # action == "approved"
        if level == 1 and is_general:
            # 一般隐患：一级通过后自动跳过二级，状态直接到 level2_approved
            update_data: dict[str, Any] = {
                "rectification_status": "level2_approved",
                "verify_level_1_status": "approved",
                "verify_level_1_by": user_id,
                "verify_level_1_by_name": user_name,
                "verify_level_1_at": now,
                "verify_level_1_opinion": opinion,
                "verify_level_2_status": "approved",
                "verify_level_2_by_name": "（自动通过：一般隐患）",
                "verify_level_2_at": now,
            }
        elif level == 3:
            update_data = {
                "rectification_status": "closed",
                "status": "closed",
                "verify_level_3_status": "approved",
                "verify_level_3_by": user_id,
                "verify_level_3_by_name": user_name,
                "verify_level_3_at": now,
                "verify_level_3_opinion": opinion,
            }
        else:
            next_status = f"level{level}_approved"
            update_data = {
                "rectification_status": next_status,
                f"verify_level_{level}_status": "approved",
                f"verify_level_{level}_by": user_id,
                f"verify_level_{level}_by_name": user_name,
                f"verify_level_{level}_at": now,
                f"verify_level_{level}_opinion": opinion,
            }
        updated = await self.repo.update_hazard(hazard_id, update_data)

        # 审核通过后，通知下一级复核人
        if updated and action == "approved":
            if level == 1 and is_general:
                # 一般隐患：跳过二级，直接通知三级
                asyncio.create_task(self._send_verify_notification(updated, 3))
            elif level < 3:
                asyncio.create_task(self._send_verify_notification(updated, level + 1))

        return updated

    async def rework_rectification(
        self,
        hazard_id: uuid.UUID,
        reply_content: str,
        rectification_photos: str | None,
        user_id: uuid.UUID,
        user_name: str,
    ) -> HazardReport | None:
        """重新整改：rejected → replied，重置所有复核级别"""
        hazard = await self.repo.get_hazard_by_id(hazard_id)
        if not hazard or hazard.rectification_status != "rejected":
            return None
        update_data: dict[str, Any] = {
            "rectification_status": "replied",
            "rectification_reply": reply_content,
            "rectification_replied_at": datetime.now(),
            "rectification_replied_by": user_id,
            "rectification_replied_by_name": user_name,
            "verify_level_1_status": "pending",
            "verify_level_1_by": None,
            "verify_level_1_by_name": None,
            "verify_level_1_at": None,
            "verify_level_1_opinion": None,
            "verify_level_2_status": "pending",
            "verify_level_2_by": None,
            "verify_level_2_by_name": None,
            "verify_level_2_at": None,
            "verify_level_2_opinion": None,
            "verify_level_3_status": "pending",
            "verify_level_3_by": None,
            "verify_level_3_by_name": None,
            "verify_level_3_at": None,
            "verify_level_3_opinion": None,
        }
        if rectification_photos is not None:
            update_data["rectification_photos"] = rectification_photos
        updated = await self.repo.update_hazard(hazard_id, update_data)

        # 重新整改后，通知一级复核人
        if updated:
            asyncio.create_task(self._send_verify_notification(updated, 1))

        return updated

    async def delete_hazard(self, hazard_id: uuid.UUID) -> bool:
        """删除隐患"""
        return await self.repo.delete_hazard(hazard_id)

    # ── AI 工作流 ──

    async def _build_hazard_context(self, script_number: int, item: HazardReport) -> str:
        """构建隐患 AI 工作流上下文信息。"""
        parts = [f"隐患编号：{item.hazard_no}"]
        if item.department:
            parts.append(f"责任部门：{item.department}")
        if item.location:
            parts.append(f"地点/部位：{item.location}")
        if item.discovered_by_name:
            parts.append(f"发现人：{item.discovered_by_name}")
        if item.discovered_at:
            parts.append(f"发现时间：{item.discovered_at.isoformat()}")
        if item.defect_photos:
            parts.append(f"缺陷图片：{item.defect_photos}")

        # 步骤2：包含步骤1已确认的输出
        if script_number == 2:
            if item.hazard_type:
                parts.append(f"隐患分类（已确认）：{item.hazard_type}")
            if item.hazard_level:
                parts.append(f"隐患等级（已确认）：{item.hazard_level}")
            if item.hazard_category:
                parts.append(f"隐患类别（已确认）：{item.hazard_category}")
            if item.description:
                parts.append(f"隐患描述（已确认）：{item.description}")
            if item.location:
                parts.append(f"地点/部位（已确认）：{item.location}")
            if item.key_defect:
                parts.append(f"重点缺陷（已确认）：{item.key_defect}")
            if item.major_hazard_basis:
                parts.append(f"重大隐患判定依据（已确认）：{item.major_hazard_basis}")

        return "\n".join(parts)

    async def _generate_hazard_ai_output(
        self, script_number: int, item: HazardReport
    ) -> dict:
        """调用 AI 服务为隐患模块生成工作流输出。失败时抛出 AIOutputError。

        优先从数据库 ai_workflow_configs 表读取配置，fallback 到 prompts.py 的硬编码 HAZARD_WORKFLOW_STEP_CONFIG。
        """
        # 优先从数据库读取工作流配置
        workflow_config = await self.repo.get_ai_workflow_config_by_module("hazard")
        if (
            workflow_config
            and workflow_config.is_enabled
            and workflow_config.script_configs
        ):
            raw = workflow_config.script_configs
            if isinstance(raw, list):
                scripts = raw
            elif isinstance(raw, dict):
                scripts = raw.get("scripts", [])
            else:
                scripts = []
            db_script = next(
                (s for s in scripts if s.get("script_number") == script_number), None
            )
            if db_script and db_script.get("is_enabled", True):
                prompt_template = build_prompt(db_script)
                expected_keys = db_script.get("expected_keys", [])
                logger.debug("使用数据库工作流配置(hazard): 步骤%d - %s", script_number, db_script.get("name"))
            else:
                config = HAZARD_WORKFLOW_STEP_CONFIG[script_number]
                prompt_template = build_prompt(config)
                expected_keys = config["expected_keys"]
                logger.debug("数据库未找到步骤 %d，fallback 到硬编码(hazard)", script_number)
        else:
            config = HAZARD_WORKFLOW_STEP_CONFIG[script_number]
            prompt_template = build_prompt(config)
            expected_keys = config["expected_keys"]
            logger.debug("无数据库工作流配置(hazard)，使用硬编码步骤 %d", script_number)

        context_text = await self._build_hazard_context(script_number, item)

        if '{context}' in prompt_template:
            prompt = prompt_template.replace('{context}', context_text)
        else:
            prompt = f"## 上下文信息\n{context_text}\n\n{prompt_template}"

        # ── Step1 + 有缺陷图片 → 走视觉模型 ──
        if script_number == 1 and item.defect_photos:
            image_urls = self._parse_defect_photo_urls(item.defect_photos)
            if image_urls:
                logger.debug("Step1 使用视觉模型, 图片数: %d", len(image_urls))
                db_config = await self.repo.get_active_api_call_config(config_type="vision")
                temperature = db_config.temperature if db_config else _DEFAULT_VISION_AI_CONFIG["temperature"]
                vision_service = await self._get_vision_ai_service()
                try:
                    result = await vision_service.chat_vision_parsed(
                        text_prompt=prompt,
                        image_urls=image_urls,
                        expected_keys=expected_keys,
                        temperature=temperature,
                    )
                    return result
                finally:
                    await vision_service.close()

        # ── 文本模型 ──
        messages = [
            {"role": "system", "content": "你是一个专业的化工安全与隐患排查专家助手，服务于原料药生产企业。"},
            {"role": "user", "content": prompt},
        ]

        ai_service = await self._get_ai_service()
        db_config = await self.repo.get_active_api_call_config(config_type="text")
        temperature = db_config.temperature if db_config else _DEFAULT_TEXT_AI_CONFIG["temperature"]
        try:
            result = await ai_service.chat_parsed(
                messages=messages,
                expected_keys=expected_keys,
                temperature=temperature,
            )
            return result
        finally:
            await ai_service.close()

    def _parse_defect_photo_urls(self, defect_photos: str) -> list[str]:
        """从 defect_photos JSON 字段提取图片，本地路径转 base64 data URI。"""
        import base64
        import json as _json

        try:
            photos = _json.loads(defect_photos)
        except (_json.JSONDecodeError, TypeError):
            # Windows 反斜杠路径经 json.dumps 写入后，json.loads 会报 Invalid \escape
            # 尝试把反斜杠替换为正斜杠再解析
            try:
                photos = _json.loads(defect_photos.replace("\\", "/"))
            except (_json.JSONDecodeError, TypeError):
                return []
        if isinstance(photos, str):
            photos = [photos] if photos else []
        elif not isinstance(photos, list):
            return []

        urls: list[str] = []
        for p in photos:
            p_str = str(p)
            # 已经是 http/data URL，直接使用
            if p_str.startswith("http://") or p_str.startswith("https://") or p_str.startswith("data:"):
                urls.append(p_str)
                continue
            # 本地路径 → base64 data URI
            # 兼容 Windows 反斜杠路径：同时检查原始路径和正斜杠版本
            check_paths = [p_str]
            if "\\" in p_str:
                check_paths.append(p_str.replace("\\", "/"))
            elif "/" in p_str:
                check_paths.append(p_str.replace("/", "\\"))
            found_path = None
            for cp in check_paths:
                if cp and os.path.exists(cp):
                    found_path = cp
                    break
            if found_path:
                try:
                    ext = os.path.splitext(found_path)[1].lower()
                    mime = {
                        ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                        ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp",
                    }.get(ext, "image/png")
                    with open(found_path, "rb") as f:
                        b64 = base64.b64encode(f.read()).decode()
                    urls.append(f"data:{mime};base64,{b64}")
                    logger.debug("转换本地图片为 data URI: %s (%s)", found_path, mime)
                except Exception as exc:
                    logger.warning("无法读取图片 %s: %s", found_path, exc)
        return urls

    async def run_hazard_ai_script(
        self, hazard_id: uuid.UUID, script_number: int
    ) -> HazardReport | None:
        """执行隐患AI工作流脚本（1=AI隐患识别, 2=AI整改建议）。

        简化状态机：AI 自动执行，不再需要逐步骤人工审核。
        Step1 完成后自动进入 Step2，Step2 完成后进入 completed。
        """
        item = await self.repo.get_hazard_by_id(hazard_id)
        if not item:
            return None

        # ── 状态机校验 ──
        expected_node = f"pending_script{script_number}"
        if item.ai_node_progress != expected_node:
            # 允许从 completed 状态重跑（台账中重新执行）
            if item.ai_node_progress != "completed":
                return None

        # Step2 需要 Step1 已产生输出（hazard_type 已被填写）
        if script_number == 2:
            if not item.hazard_type or item.hazard_type == "unsafe_condition":
                # Step1 可能未执行或失败，先执行 Step1
                pass

        update_data: dict[str, Any] = {}

        try:
            generated = await self._generate_hazard_ai_output(script_number, item)
            self._map_hazard_ai_output(script_number, generated, update_data)
            if script_number == 1:
                update_data["ai_node_progress"] = "pending_script2"
            else:
                update_data["ai_node_progress"] = "completed"
                update_data["overall_status"] = "completed"
            update_data["ai_error_message"] = None
            update_data["ai_generated"] = True
        except AIOutputError as e:
            logger.error(f"Hazard script {script_number} AI 输出错误: {e}")
            update_data["ai_error_message"] = str(e)
        except Exception as e:
            logger.error(f"Hazard script {script_number} 执行异常: {e}")
            update_data["ai_error_message"] = f"AI 服务调用失败：{e}"

        return await self.repo.update_hazard(hazard_id, update_data)

    async def review_hazard_ai_script(
        self, hazard_id: uuid.UUID, script_number: int, action: str
    ) -> HazardReport | None:
        """审核隐患AI输出（action=approved|rejected）。

        新流程：在台账中进行统一审核，而非逐步骤审核。
        审核通过 → overall_status='reviewed', status→'open'（进入整改流程）
        驳回 → 重置 AI 状态，允许重新执行。
        """
        item = await self.repo.get_hazard_by_id(hazard_id)
        if not item:
            return None

        # 审核仅允许在 AI 完成后进行
        if item.overall_status != "completed":
            return None

        update_data: dict[str, Any] = {}

        if action == "approved":
            # 审核通过：标记为已审核，状态改为 open 进入整改流程
            update_data["script1_review_status"] = "approved"
            update_data["script2_review_status"] = "approved"
            update_data["overall_status"] = "reviewed"
            update_data["status"] = "open"
        elif action == "rejected":
            # 驳回：重置 AI 状态，允许重新执行
            update_data["script1_review_status"] = "pending"
            update_data["script2_review_status"] = "pending"
            update_data["ai_node_progress"] = "pending_script1"
            update_data["overall_status"] = "draft"
            update_data["ai_error_message"] = None
        else:
            return None

        updated = await self.repo.update_hazard(hazard_id, update_data)

        # 审核通过后，异步推送飞书通知
        if action == "approved" and updated:
            asyncio.create_task(self._send_hazard_notification(updated))

        return updated

    async def _send_hazard_notification(self, hazard: HazardReport) -> None:
        """异步发送隐患整改通知到飞书（测试阶段：固定发送给许康福）。"""
        try:
            # 测试阶段：固定通知人 open_id（许康福）
            target_open_id = "ou_5773c42e1fc7ce3e554b83242c87aa0b"

            settings = get_settings()
            detail_url = f"{settings.FRONTEND_URL}/safety/hazard/{hazard.id}"

            # 隐患等级标签
            level_labels = {"general": "一般隐患", "serious": "较大隐患", "major": "重大隐患"}
            level_text = level_labels.get(hazard.hazard_level, hazard.hazard_level or "未分级")

            # 整改期限
            deadline_text = hazard.deadline.strftime("%Y-%m-%d") if hazard.deadline else "未设置"

            content = (
                f"**隐患编号：** {hazard.hazard_no or '-'}\n"
                f"**隐患等级：** {level_text}\n"
                f"**隐患描述：** {hazard.description or '-'}\n"
                f"**地点/部位：** {hazard.location or '-'}\n"
                f"**责任部门：** {hazard.department or '-'}\n"
                f"**整改期限：** {deadline_text}"
            )

            elements = [
                {
                    "tag": "action",
                    "actions": [
                        {
                            "tag": "button",
                            "text": {"tag": "plain_text", "content": "📋 前往整改"},
                            "type": "primary",
                            "url": detail_url,
                        }
                    ],
                }
            ]

            success = await send_user_card(
                open_id=target_open_id,
                title="🔔 隐患整改通知",
                content=content,
                elements=elements,
            )
            if success:
                logger.info("隐患整改通知已发送: hazard_no=%s, open_id=%s", hazard.hazard_no, target_open_id)
            else:
                logger.warning("隐患整改通知发送失败: hazard_no=%s", hazard.hazard_no)
        except Exception as e:
            logger.warning("隐患整改通知异常: %s", e)

    async def _send_verify_notification(self, hazard: HazardReport, level: int) -> None:
        """异步发送复核通知到飞书（测试阶段：固定发送给许康福）。"""
        try:
            target_open_id = "ou_5773c42e1fc7ce3e554b83242c87aa0b"

            settings = get_settings()
            detail_url = f"{settings.FRONTEND_URL}/safety/hazard/{hazard.id}"

            level_labels = {1: "一级", 2: "二级", 3: "三级"}
            level_text = level_labels.get(level, f"{level}级")

            content = (
                f"**隐患编号：** {hazard.hazard_no or '-'}\n"
                f"**隐患描述：** {hazard.description or '-'}\n"
                f"**地点/部位：** {hazard.location or '-'}\n"
                f"**责任部门：** {hazard.department or '-'}\n"
                f"**整改回复：** {hazard.rectification_reply or '-'}\n"
            )

            elements = [
                {
                    "tag": "action",
                    "actions": [
                        {
                            "tag": "button",
                            "text": {"tag": "plain_text", "content": f"📋 前往{level_text}复核"},
                            "type": "primary",
                            "url": detail_url,
                        }
                    ],
                }
            ]

            success = await send_user_card(
                open_id=target_open_id,
                title=f"🔔 隐患{level_text}复核通知",
                content=content,
                elements=elements,
            )
            if success:
                logger.info("复核通知已发送: hazard_no=%s, level=%s", hazard.hazard_no, level)
            else:
                logger.warning("复核通知发送失败: hazard_no=%s, level=%s", hazard.hazard_no, level)
        except Exception as e:
            logger.warning("复核通知异常: %s", e)

    # ── AI 输出字段约束（DB Enum 限制）──
    _VALID_HAZARD_TYPES = {"unsafe_condition", "unsafe_action", "management_defect", "environmental"}
    _VALID_HAZARD_LEVELS = {"general", "serious", "major"}
    _VALID_HAZARD_CATEGORIES = {
        "equipment", "hazardous_storage", "emergency_mgmt", "instrument_electrical",
        "lightning_antistatic", "occupational_health", "violation_operation", "six_s",
        "label_signage", "process_mgmt", "contractor_defect", "documentation", "special_operation",
    }

    def _map_hazard_ai_output(
        self, script_number: int, output: dict, update_data: dict[str, Any]
    ) -> None:
        """将 AI JSON 输出映射到 HazardReport 字段（校验枚举值）。"""
        if script_number == 1:
            key_map = {
                "hazard_type": "hazard_type",
                "hazard_level": "hazard_level",
                "hazard_category": "hazard_category",
                "description": "description",
                "location": "location",
                "key_defect": "key_defect",
                "major_hazard_basis": "major_hazard_basis",
            }
        else:  # script_number == 2
            key_map = {
                "control_measures": "control_measures",
                "corrective_preventive_measures": "corrective_preventive_measures",
            }

        for json_key, db_field in key_map.items():
            if json_key in output and output[json_key]:
                val = output[json_key]
                # 校验枚举字段，非法值跳过（保留默认值）
                if db_field == "hazard_type" and val not in self._VALID_HAZARD_TYPES:
                    logger.debug("AI 输出非法 hazard_type=%s，跳过", val)
                    continue
                if db_field == "hazard_level" and val not in self._VALID_HAZARD_LEVELS:
                    logger.debug("AI 输出非法 hazard_level=%s，跳过", val)
                    continue
                if db_field == "hazard_category" and val not in self._VALID_HAZARD_CATEGORIES:
                    logger.debug("AI 输出非法 hazard_category=%s，跳过", val)
                    continue
                update_data[db_field] = val

    # ==================== Accident Operations ====================

    async def get_accidents(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        accident_type: str | None = None,
        accident_level: str | None = None,
        department: str | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        keyword: str | None = None,
    ) -> tuple[list[Accident], int]:
        """获取事故列表"""
        return await self.repo.get_accidents(
            skip, limit, status, accident_type, accident_level,
            department, date_from, date_to, keyword,
        )

    async def get_accident(self, accident_id: uuid.UUID) -> Accident | None:
        """获取事故详情"""
        return await self.repo.get_accident_by_id(accident_id)

    async def create_accident(self, data: AccidentCreate) -> Accident:
        """创建事故"""
        accident_data = data.model_dump()
        return await self.repo.create_accident(accident_data)

    async def update_accident(
        self, accident_id: uuid.UUID, data: AccidentUpdate
    ) -> Accident | None:
        """更新事故"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_accident(accident_id, update_data)

    async def investigate_accident(
        self,
        accident_id: uuid.UUID,
        investigator: uuid.UUID,
        investigator_name: str,
    ) -> Accident | None:
        """开始调查事故"""
        accident = await self.repo.get_accident_by_id(accident_id)
        if not accident or accident.status != "reported":
            return None
        return await self.repo.update_accident(
            accident_id,
            {
                "status": "investigating",
                "investigator": investigator,
                "investigator_name": investigator_name,
            },
        )

    async def resolve_accident(
        self,
        accident_id: uuid.UUID,
        direct_cause: str,
        root_cause: str,
        handling_measures: str,
        corrective_actions: str | None = None,
        investigation_findings: str | None = None,
        investigation_method: str | None = None,
        investigation_team: list | None = None,
    ) -> Accident | None:
        """完成调查事故"""
        accident = await self.repo.get_accident_by_id(accident_id)
        if not accident or accident.status != "investigating":
            return None
        update_data: dict[str, Any] = {
            "status": "investigated",
            "direct_cause": direct_cause,
            "root_cause": root_cause,
            "handling_measures": handling_measures,
            "corrective_actions": corrective_actions,
            "investigation_findings": investigation_findings,
            "investigation_method": investigation_method,
        }
        if investigation_team is not None:
            update_data["investigation_team"] = investigation_team
        return await self.repo.update_accident(accident_id, update_data)

    async def start_capa(
        self,
        accident_id: uuid.UUID,
        corrective_action_deadline: datetime,
        corrective_action_responsible: str,
    ) -> Accident | None:
        """启动 CAPA"""
        accident = await self.repo.get_accident_by_id(accident_id)
        if not accident or accident.status != "investigated":
            return None
        return await self.repo.update_accident(
            accident_id,
            {
                "status": "capa_in_progress",
                "corrective_action_deadline": corrective_action_deadline,
                "corrective_action_responsible": corrective_action_responsible,
                "corrective_action_status": "in_progress",
            },
        )

    async def verify_capa(
        self,
        accident_id: uuid.UUID,
        verified_by: uuid.UUID,
        verified_by_name: str,
    ) -> Accident | None:
        """验证 CAPA 并关闭事故"""
        accident = await self.repo.get_accident_by_id(accident_id)
        if not accident or accident.status != "capa_in_progress":
            return None
        return await self.repo.update_accident(
            accident_id,
            {
                "status": "closed",
                "corrective_action_status": "verified",
                "verified_by": verified_by,
                "verified_by_name": verified_by_name,
                "verified_at": datetime.now(),
            },
        )

    async def close_accident(self, accident_id: uuid.UUID) -> Accident | None:
        """直接关闭事故（无CAPA时）"""
        accident = await self.repo.get_accident_by_id(accident_id)
        if not accident or accident.status != "investigated":
            return None
        return await self.repo.update_accident(accident_id, {"status": "closed"})

    async def delete_accident(self, accident_id: uuid.UUID) -> bool:
        """删除事故"""
        return await self.repo.delete_accident(accident_id)

    # ==================== Contractor Operations ====================

    async def get_contractors(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        qualification_type: str | None = None,
        training_status: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list[Contractor], int]:
        """获取承包商列表"""
        return await self.repo.get_contractors(
            skip, limit, status, qualification_type, training_status, keyword,
        )

    async def get_contractor(self, contractor_id: uuid.UUID) -> Contractor | None:
        """获取承包商详情"""
        return await self.repo.get_contractor_by_id(contractor_id)

    async def create_contractor(self, data: "ContractorCreate") -> Contractor:
        """创建承包商"""
        contractor_data = data.model_dump()
        return await self.repo.create_contractor(contractor_data)

    async def update_contractor(
        self, contractor_id: uuid.UUID, data: "ContractorUpdate"
    ) -> Contractor | None:
        """更新承包商"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_contractor(contractor_id, update_data)

    async def blacklist_contractor(self, contractor_id: uuid.UUID) -> Contractor | None:
        """将承包商加入黑名单"""
        contractor = await self.repo.get_contractor_by_id(contractor_id)
        if not contractor:
            return None
        return await self.repo.update_contractor(
            contractor_id, {"status": "blacklisted", "blacklisted": True}
        )

    async def activate_contractor(self, contractor_id: uuid.UUID) -> Contractor | None:
        """激活承包商"""
        contractor = await self.repo.get_contractor_by_id(contractor_id)
        if not contractor:
            return None
        return await self.repo.update_contractor(
            contractor_id, {"status": "active", "blacklisted": False}
        )

    async def update_contractor_training(
        self, contractor_id: uuid.UUID, training_status: str
    ) -> Contractor | None:
        """更新承包商培训状态"""
        contractor = await self.repo.get_contractor_by_id(contractor_id)
        if not contractor:
            return None
        return await self.repo.update_contractor(
            contractor_id,
            {
                "training_status": training_status,
                "training_date": datetime.now(),
            },
        )

    async def delete_contractor(self, contractor_id: uuid.UUID) -> bool:
        """删除承包商"""
        return await self.repo.delete_contractor(contractor_id)

    # ── 施工记录 ──

    async def get_work_records(
        self, contractor_id: uuid.UUID
    ) -> list[ContractorWorkRecord]:
        """获取承包商的施工记录"""
        return await self.repo.get_work_records_by_contractor(contractor_id)

    async def create_work_record(
        self, contractor_id: uuid.UUID, data: "ContractorWorkRecordCreate"
    ) -> ContractorWorkRecord:
        """创建施工记录"""
        record_data = data.model_dump()
        record_data["contractor_id"] = str(contractor_id)
        return await self.repo.create_work_record(record_data)

    async def update_work_record(
        self, record_id: uuid.UUID, data: "ContractorWorkRecordUpdate"
    ) -> ContractorWorkRecord | None:
        """更新施工记录"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_work_record(record_id, update_data)

    async def evaluate_work_record(
        self, record_id: uuid.UUID, score: int, comments: str | None = None,
        evaluator: str | None = None,
    ) -> ContractorWorkRecord | None:
        """评价施工记录"""
        record = await self.repo.get_work_record_by_id(record_id)
        if not record:
            return None
        return await self.repo.update_work_record(
            record_id,
            {
                "status": "evaluated",
                "evaluation": {
                    "score": score,
                    "comments": comments,
                    "evaluator": evaluator,
                    "date": datetime.now().isoformat(),
                },
            },
        )

    async def delete_work_record(self, record_id: uuid.UUID) -> bool:
        """删除施工记录"""
        return await self.repo.delete_work_record(record_id)

    # ==================== SafetyTraining Operations ====================

    async def get_trainings(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        training_type: str | None = None,
        department: str | None = None,
    ) -> tuple[list[SafetyTraining], int]:
        """获取安全培训列表"""
        return await self.repo.get_trainings(skip, limit, status, training_type, department)

    async def get_training(self, training_id: uuid.UUID) -> SafetyTraining | None:
        """获取安全培训详情"""
        return await self.repo.get_training_by_id(training_id)

    async def create_training(self, data: SafetyTrainingCreate) -> SafetyTraining:
        """创建安全培训"""
        training_data = data.model_dump()
        return await self.repo.create_training(training_data)

    async def update_training(
        self, training_id: uuid.UUID, data: SafetyTrainingUpdate
    ) -> SafetyTraining | None:
        """更新安全培训"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_training(training_id, update_data)

    async def start_training(self, training_id: uuid.UUID) -> SafetyTraining | None:
        """开始培训（草稿→进行中）"""
        training = await self.repo.get_training_by_id(training_id)
        if not training or training.status != "draft":
            return None
        return await self.repo.update_training(training_id, {"status": "in_progress"})

    async def complete_training(self, training_id: uuid.UUID) -> SafetyTraining | None:
        """完成培训"""
        training = await self.repo.get_training_by_id(training_id)
        if not training or training.status != "in_progress":
            return None
        return await self.repo.update_training(training_id, {"status": "completed"})

    async def archive_training(self, training_id: uuid.UUID) -> SafetyTraining | None:
        """归档培训"""
        training = await self.repo.get_training_by_id(training_id)
        if not training or training.status != "completed":
            return None
        return await self.repo.update_training(training_id, {"status": "archived"})

    async def delete_training(self, training_id: uuid.UUID) -> bool:
        """删除安全培训"""
        return await self.repo.delete_training(training_id)

    # ==================== TrainingRecord Operations ====================

    async def get_training_records(self, training_id: uuid.UUID) -> list[TrainingRecord]:
        """获取培训记录列表"""
        return await self.repo.get_records_by_training(training_id)

    async def create_training_record(self, data: TrainingRecordCreate) -> TrainingRecord:
        """创建培训记录"""
        record_data = data.model_dump()
        return await self.repo.create_training_record(record_data)

    async def update_training_record(
        self, record_id: uuid.UUID, data: TrainingRecordUpdate
    ) -> TrainingRecord | None:
        """更新培训记录"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_training_record(record_id, update_data)

    async def batch_create_records(
        self, training_id: uuid.UUID, records: list[TrainingRecordCreate]
    ) -> list[TrainingRecord]:
        """批量创建培训签到记录"""
        result = []
        for record in records:
            record_data = record.model_dump()
            record_data["training_id"] = training_id
            item = await self.repo.create_training_record(record_data)
            result.append(item)
        return result

    async def delete_training_record(self, record_id: uuid.UUID) -> bool:
        """删除培训记录"""
        return await self.repo.delete_training_record(record_id)

    # ── 培训证书 ──

    async def get_training_certificates(
        self,
        skip: int = 0,
        limit: int = 20,
        certificate_status: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list[TrainingRecord], int]:
        """获取培训证书列表"""
        return await self.repo.get_training_certificates(
            skip, limit, certificate_status, keyword,
        )

    async def get_expiring_certificates(self) -> list[TrainingRecord]:
        """获取即将过期的证书（30天内）"""
        return await self.repo.get_expiring_certificates()

    # ==================== HazardIdentification Operations ====================

    async def get_hazard_identifications(
        self,
        skip: int = 0,
        limit: int = 20,
        department: str | None = None,
        overall_status: str | None = None,
        ai_node_progress: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list, int]:
        """获取危险源辨识列表"""
        return await self.repo.get_hazard_identifications(
            skip, limit, department, overall_status, ai_node_progress, keyword
        )

    async def get_hazard_identification(self, hid: uuid.UUID):
        """获取危险源辨识详情"""
        return await self.repo.get_hazard_identification_by_id(hid)

    async def create_hazard_identification(self, data) -> Any:
        """创建危险源辨识记录"""

        create_data = data.model_dump()
        create_data["ai_node_progress"] = "pending_input"
        create_data["overall_status"] = "draft"
        return await self.repo.create_hazard_identification(create_data)

    async def update_hazard_identification(self, hid: uuid.UUID, data) -> Any | None:
        """更新危险源辨识"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_hazard_identification(hid, update_data)

    async def delete_hazard_identification(self, hid: uuid.UUID) -> bool:
        """删除危险源辨识"""
        return await self.repo.delete_hazard_identification(hid)

    async def submit_hazard_identification(self, hid: uuid.UUID) -> Any | None:
        """提交基础信息 → 进入脚本1（待AI解析附件）"""
        item = await self.repo.get_hazard_identification_by_id(hid)
        if not item or item.overall_status not in ("draft",):
            return None
        return await self.repo.update_hazard_identification(
            hid, {"ai_node_progress": "pending_script1", "overall_status": "in_progress"}
        )

    # ── 工作流状态机 ──

    SCRIPT_NODE_MAP = {
        1: ("pending_script1", "pending_script2", "script1_review_status"),
        2: ("pending_script2", "pending_script3", "script2_review_status"),
        3: ("pending_script3", "pending_script4", "script3_review_status"),
        4: ("pending_script4", "pending_script5", "script4_review_status"),
        5: ("pending_script5", "pending_script6", "script5_review_status"),
        6: ("pending_script6", "pending_script7", "script6_review_status"),
        7: ("pending_script7", "completed", "script7_review_status"),
    }

    # ── AI 集成 ──

    async def _get_ai_service(self) -> AIService:
        """获取文本模型 AIService（安全模块数据库配置）"""
        config = await self.repo.get_active_api_call_config(config_type="text")
        if config:
            logger.debug("使用数据库 API 配置: %s (%s)", config.config_name, config.model_name)
            return AIService(
                api_key=config.api_key,
                base_url=config.api_base_url,
                model=config.model_name,
                timeout=config.timeout_seconds,
            )

        # 自动种子：从环境变量迁移到数据库
        await _ensure_ai_config_seeded(self.session, "text")
        config = await self.repo.get_active_api_call_config(config_type="text")
        if config:
            return AIService(
                api_key=config.api_key,
                base_url=config.api_base_url,
                model=config.model_name,
                timeout=config.timeout_seconds,
            )

        raise AIOutputError("安全模块未配置文本 AI 模型，请在 API 配置页面进行配置")

    async def _get_vision_ai_service(self) -> AIService:
        """获取视觉模型 AIService（安全模块数据库配置）"""
        config = await self.repo.get_active_api_call_config(config_type="vision")
        if config:
            logger.debug(
                "使用数据库视觉API配置: %s (%s)", config.config_name, config.model_name
            )
            return AIService(
                api_key=config.api_key,
                base_url=config.api_base_url,
                model=config.model_name,
                timeout=config.timeout_seconds,
            )

        # 自动种子：从环境变量迁移到数据库
        await _ensure_ai_config_seeded(self.session, "vision")
        config = await self.repo.get_active_api_call_config(config_type="vision")
        if config:
            return AIService(
                api_key=config.api_key,
                base_url=config.api_base_url,
                model=config.model_name,
                timeout=config.timeout_seconds,
            )

        raise AIOutputError("安全模块未配置视觉 AI 模型，请在 API 配置页面进行配置")

    def _build_context(self, script_number: int, item: Any) -> str:
        """从当前记录构建供 AI 使用的上下文字符串"""
        parts: list[str] = [
            f"部门：{item.department or '未知'}",
            f"岗位：{item.position or '未知'}",
            f"生产步骤：{item.production_step or '未知'}",
        ]

        # Script 1 基础字段
        if script_number >= 1:
            if item.specific_activity:
                parts.append(f"具体作业活动：{item.specific_activity}")
            if item.equipment_facilities:
                parts.append(f"设备设施：{item.equipment_facilities}")
            if item.raw_auxiliary_materials:
                parts.append(f"原辅料：{item.raw_auxiliary_materials}")
            if item.operation_frequency:
                parts.append(f"作业频次：{item.operation_frequency}")
            if item.operator_count is not None:
                parts.append(f"操作人数：{item.operator_count}")

        # Script 2 输出
        if script_number >= 3:
            if item.hazard_type:
                parts.append(f"危险类型：{item.hazard_type}")
            if item.possible_accident:
                parts.append(f"可能导致事故：{item.possible_accident}")
            if item.unsafe_behavior:
                parts.append(f"不规范作业行为表现：{item.unsafe_behavior}")

        # Script 3 输出
        if script_number >= 4:
            if item.l_inherent is not None:
                parts.append(f"可能性 L（固有）：{item.l_inherent}")
            if item.e_inherent is not None:
                parts.append(f"暴露频率 E（固有）：{item.e_inherent}")
            if item.c_inherent is not None:
                parts.append(f"严重性 C（固有）：{item.c_inherent}")
            if item.d_inherent is not None:
                parts.append(f"风险值 D（固有）：{item.d_inherent}")
            if item.inherent_risk_label:
                parts.append(f"固有风险等级：{item.inherent_risk_label}")

        # Script 4 输出
        if script_number >= 5:
            if item.existing_engineering_controls:
                parts.append(f"现有工程控制措施：{item.existing_engineering_controls}")
            if item.existing_management_controls:
                parts.append(f"现有管理控制措施：{item.existing_management_controls}")
            if item.existing_ppe:
                parts.append(f"现有个人防护措施：{item.existing_ppe}")
            if item.existing_emergency_measures:
                parts.append(f"现有应急措施：{item.existing_emergency_measures}")

        # Script 5 输出
        if script_number >= 6:
            if item.l_residual is not None:
                parts.append(f"可能性 L（残余）：{item.l_residual}")
            if item.e_residual is not None:
                parts.append(f"暴露频率 E（残余）：{item.e_residual}")
            if item.c_residual is not None:
                parts.append(f"严重性 C（残余）：{item.c_residual}")
            if item.d_residual is not None:
                parts.append(f"风险值 D（残余）：{item.d_residual}")
            if item.residual_risk_label:
                parts.append(f"残余风险等级：{item.residual_risk_label}")
            if item.control_level:
                parts.append(f"管控等级：{item.control_level}")

        # Script 6 输出
        if script_number >= 7:
            if item.recommendation_content:
                parts.append(f"建议措施内容：{item.recommendation_content}")
            if item.recommendation_type:
                parts.append(f"建议措施类型：{item.recommendation_type}")

        return "\n".join(parts)

    async def _generate_ai_output(
        self, script_number: int, item: Any
    ) -> dict:
        """调用 AI 服务生成工作流输出。失败时抛出 AIOutputError。

        优先从数据库 ai_workflow_configs 表读取对应模块的工作流配置，
        fallback 到 prompts.py 的硬编码 WORKFLOW_STEP_CONFIG。
        """
        # 优先从数据库读取工作流配置
        workflow_config = await self.repo.get_ai_workflow_config_by_module("hazard-identification")
        if (
            workflow_config
            and workflow_config.is_enabled
            and workflow_config.script_configs
        ):
            raw = workflow_config.script_configs
            if isinstance(raw, list):
                scripts = raw
            elif isinstance(raw, dict):
                scripts = raw.get("scripts", [])
            else:
                scripts = []
            db_script = next(
                (s for s in scripts if s.get("script_number") == script_number), None
            )
            if db_script and db_script.get("is_enabled", True):
                prompt_template = build_prompt(db_script)
                expected_keys = db_script.get("expected_keys", [])
                logger.debug("使用数据库工作流配置: 步骤%d - %s", script_number, db_script.get("name"))
            else:
                # DB 中有 workflow 但没有对应步骤 → fallback
                config = SCRIPT_CONFIG[script_number]
                prompt_template = build_prompt(config)
                expected_keys = config["expected_keys"]
                logger.debug(
                    "数据库工作流配置中未找到步骤 %d，fallback 到硬编码", script_number
                )
        else:
            # DB 无配置 → fallback 到硬编码
            config = SCRIPT_CONFIG[script_number]
            prompt_template = build_prompt(config)
            expected_keys = config["expected_keys"]
            logger.debug("无数据库工作流配置，使用硬编码步骤 %d", script_number)

        context_text = self._build_context(script_number, item)

        # Script 1 特殊处理：解析附件
        if script_number == 1:
            attachment_text = ""
            if item.attachment_path:
                try:
                    attachment_text = DocumentParser.extract_text(
                        item.attachment_path, max_chars=30000
                    )
                except Exception as e:
                    logger.warning(f"附件解析失败: {e}")
            if attachment_text:
                context_text += f"\n\n### 附件文档内容\n{attachment_text}"
            else:
                context_text += "\n\n### 附件文档内容\n（未上传附件或附件无法解析）"

        # 使用 replace 而非 format()，避免 AI 输出示例中的花括号被误解析
        if '{context}' in prompt_template:
            prompt = prompt_template.replace('{context}', context_text)
        else:
            prompt = f"## 上下文信息\n{context_text}\n\n{prompt_template}"
        messages = [
            {"role": "system", "content": "你是一个专业的危险源辨识与风险评价专家助手，服务于原料药生产企业。"},
            {"role": "user", "content": prompt},
        ]

        ai_service = await self._get_ai_service()
        # 优先使用数据库配置的 temperature
        db_config = await self.repo.get_active_api_call_config(config_type="text")
        temperature = db_config.temperature if db_config else _DEFAULT_TEXT_AI_CONFIG["temperature"]
        try:
            result = await ai_service.chat_parsed(
                messages=messages,
                expected_keys=expected_keys,
                temperature=temperature,
            )
            return result
        finally:
            await ai_service.close()

    async def run_script(
        self, hid: uuid.UUID, script_number: int, ai_output: dict | None = None
    ) -> Any | None:
        """执行AI脚本（状态机推进）"""

        item = await self.repo.get_hazard_identification_by_id(hid)
        if not item:
            return None

        if script_number not in self.SCRIPT_NODE_MAP:
            return None

        current_node, next_node, review_field = self.SCRIPT_NODE_MAP[script_number]

        # Check current state allows this script
        if item.ai_node_progress != current_node:
            return None

        # Check previous script's review status is approved (for scripts 2-7)
        if script_number > 1:
            prev_review = getattr(item, f"script{script_number - 1}_review_status")
            if prev_review != "approved":
                return None

        update_data: dict[str, Any] = {}

        # ── AI 生成或使用传入数据 ──
        if ai_output is not None:
            # 人工覆盖 / 前端 demo 数据
            self._map_ai_output(script_number, ai_output, update_data)
            self._calculate_risk_levels(script_number, update_data)
            update_data["ai_node_progress"] = next_node
            update_data["ai_error_message"] = None
        else:
            # 调用真实 AI
            try:
                generated = await self._generate_ai_output(script_number, item)
                self._map_ai_output(script_number, generated, update_data)
                self._calculate_risk_levels(script_number, update_data)
                update_data["ai_node_progress"] = next_node
                update_data["ai_error_message"] = None
            except AIOutputError as e:
                logger.error(f"Script {script_number} AI 输出错误: {e}")
                update_data[f"script{script_number}_review_status"] = "rejected"
                update_data["ai_error_message"] = str(e)
            except Exception as e:
                logger.error(f"Script {script_number} 执行异常: {e}")
                update_data[f"script{script_number}_review_status"] = "rejected"
                update_data["ai_error_message"] = f"AI 服务调用失败：{e}"

        result = await self.repo.update_hazard_identification(hid, update_data)
        return result

    async def review_script(
        self, hid: uuid.UUID, script_number: int, action: str
    ) -> Any | None:
        """审核确认或驳回脚本输出"""
        item = await self.repo.get_hazard_identification_by_id(hid)
        if not item:
            return None

        if script_number not in self.SCRIPT_NODE_MAP:
            return None

        current_node, next_node, review_field = self.SCRIPT_NODE_MAP[script_number]

        # Current node must match
        expected_current = current_node if action == "approved" else current_node

        update_data: dict[str, Any] = {
            f"script{script_number}_review_status": action,
        }

        if action == "approved":
            update_data["ai_node_progress"] = next_node

        # 当完成脚本7审核 → overall_status = completed
        if action == "approved" and script_number == 7:
            update_data["overall_status"] = "completed"
        elif action == "rejected":
            # 驳回：回退到之前节点，允许重新生成
            update_data["ai_node_progress"] = current_node

        result = await self.repo.update_hazard_identification(hid, update_data)
        return result

    @staticmethod
    def _safe_float(value: Any) -> float | None:
        """将 AI 输出值转为 float；若为'待人工确认'等非数值字符串则返回 None"""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            stripped = value.strip()
            if stripped == "待人工确认":
                return None
            try:
                return float(stripped)
            except ValueError:
                return None
        return None

    def _map_ai_output(
        self, script_number: int, ai_output: dict, update_data: dict[str, Any]
    ) -> None:
        """将 AI 输出映射到模型字段。

        - 脚本1：仅输出 3 个字段（标准文件规定）
        - 脚本3/5/7：AI 同时输出 L/E/C/D 值和风险等级；若为"待人工确认"则存 None
        - 脚本6：needs_recommendation 为字符串三态（是/否/待人工确认）
        """
        if script_number == 1:
            for f in ("specific_activity", "equipment_facilities", "raw_auxiliary_materials"):
                if f in ai_output:
                    update_data[f] = ai_output[f]

        elif script_number == 2:
            for f in ("hazard_type", "possible_accident", "unsafe_behavior"):
                if f in ai_output:
                    update_data[f] = ai_output[f]

        elif script_number == 3:
            for f in ("l_inherent", "e_inherent", "c_inherent"):
                if f in ai_output:
                    update_data[f] = self._safe_float(ai_output[f])
            # AI 直接输出 D 值和风险等级
            if "d_inherent" in ai_output:
                update_data["d_inherent"] = self._safe_float(ai_output["d_inherent"])
            if "inherent_risk_level" in ai_output:
                update_data["inherent_risk_level"] = ai_output["inherent_risk_level"]

        elif script_number == 4:
            for f in ("existing_engineering_controls", "existing_management_controls",
                      "existing_ppe", "existing_emergency_measures"):
                if f in ai_output:
                    update_data[f] = ai_output[f]

        elif script_number == 5:
            for f in ("l_residual", "e_residual", "c_residual"):
                if f in ai_output:
                    update_data[f] = self._safe_float(ai_output[f])
            if "d_residual" in ai_output:
                update_data["d_residual"] = self._safe_float(ai_output["d_residual"])
            if "residual_risk_level" in ai_output:
                update_data["residual_risk_level"] = ai_output["residual_risk_level"]

        elif script_number == 6:
            for f in ("needs_recommendation", "recommendation_type",
                      "recommendation_content", "recommendation_priority"):
                if f in ai_output:
                    update_data[f] = ai_output[f]

        elif script_number == 7:
            for f in ("l_post", "e_post", "c_post"):
                if f in ai_output:
                    update_data[f] = self._safe_float(ai_output[f])
            if "d_post" in ai_output:
                update_data["d_post"] = self._safe_float(ai_output["d_post"])
            if "post_risk_level" in ai_output:
                update_data["post_risk_level"] = ai_output["post_risk_level"]

    def _calculate_risk_levels(
        self, script_number: int, update_data: dict[str, Any]
    ) -> None:
        """补全风险等级和管控信息。

        AI 在脚本3/5/7中直接输出 D 值和风险等级；此处仅在后端可计算且 AI
        未提供对应字段时做兜底计算，同时补充 label、control_level 等展示字段。
        """
        from app.modules.safety.schemas import RISK_LEVELS, get_risk_level

        if script_number == 3:
            l = update_data.get("l_inherent")
            e = update_data.get("e_inherent")
            c = update_data.get("c_inherent")
            if all(v is not None for v in (l, e, c)):
                # D 值：优先用 AI 输出，否则后端计算
                if update_data.get("d_inherent") is None:
                    update_data["d_inherent"] = l * e * c
                # 风险等级 key：优先用 AI 输出
                if update_data.get("inherent_risk_level") is None:
                    level = get_risk_level(update_data["d_inherent"])
                    update_data["inherent_risk_level"] = level["key"]
                # 补充 label / control_level / responsible_person（后端计算）
                for rl in RISK_LEVELS:
                    if rl["key"] == update_data.get("inherent_risk_level"):
                        update_data["inherent_risk_label"] = rl["label"]
                        update_data["control_level"] = rl["control_level"]
                        update_data["responsible_person"] = rl["responsible_person"]
                        break

        elif script_number == 5:
            l = update_data.get("l_residual")
            e = update_data.get("e_residual")
            c = update_data.get("c_residual")
            if all(v is not None for v in (l, e, c)):
                if update_data.get("d_residual") is None:
                    update_data["d_residual"] = l * e * c
                if update_data.get("residual_risk_level") is None:
                    level = get_risk_level(update_data["d_residual"])
                    update_data["residual_risk_level"] = level["key"]
                for rl in RISK_LEVELS:
                    if rl["key"] == update_data.get("residual_risk_level"):
                        update_data["residual_risk_label"] = rl["label"]
                        break

        elif script_number == 7:
            l = update_data.get("l_post")
            e = update_data.get("e_post")
            c = update_data.get("c_post")
            if all(v is not None for v in (l, e, c)):
                if update_data.get("d_post") is None:
                    update_data["d_post"] = l * e * c
                if update_data.get("post_risk_level") is None:
                    level = get_risk_level(update_data["post_risk_level"])
                    update_data["post_risk_level"] = level["key"]
                for rl in RISK_LEVELS:
                    if rl["key"] == update_data.get("post_risk_level"):
                        update_data["post_risk_label"] = rl["label"]
                        break

    # ==================== 附件上传 ====================

    async def upload_attachment(
        self, hid: uuid.UUID, file_name: str, file_path: str
    ) -> Any | None:
        """保存附件路径到记录"""
        return await self.repo.update_hazard_identification(
            hid,
            {
                "attachment_path": file_path,
                "attachment_original_name": file_name,
            },
        )

    # ── AI 导出 ──

    async def parse_hazard_export_query(self, natural_query: str) -> dict:
        """使用 AI 将自然语言筛选条件解析为结构化参数"""
        wf_config = await self._get_workflow_config("hazard-identification-export")
        if wf_config:
            prompt = build_prompt(wf_config) + "\n\n用户查询：" + natural_query
        else:
            prompt = natural_query

        try:
            ai = await self._get_ai_service()
            messages = [
                {"role": "system", "content": "你是一个数据库查询助手。只返回 JSON。"},
                {"role": "user", "content": prompt},
            ]
            response_text = await ai.chat(messages, response_format="json_object")
            import json

            result = json.loads(response_text)
            # 清除 None 值
            return {k: v for k, v in result.items() if v is not None}
        except Exception as e:
            logger.warning("AI 自然语言解析失败(hazard-identification): %s", e)
            return {
                "explanation": f"AI 解析失败，将使用原始查询: {natural_query}",
                "keyword": natural_query,
            }

    async def export_hazard_ledger_pdf(
        self,
        department: str | None = None,
        position: str | None = None,
        risk_level: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        keyword: str | None = None,
    ) -> bytes:
        """导出危险源辨识台账为 PDF 文件（A4 横向，标准格式）"""
        import io
        from datetime import datetime as dt_module

        # 获取数据
        items, _ = await self.repo.get_hazard_identifications(
            skip=0,
            limit=10000,
            department=department,
            overall_status="completed",
            position=position,
            risk_level=risk_level,
            date_from=date_from,
            date_to=date_to,
            keyword=keyword,
        )

        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        # A4 横向
        page_w, page_h = landscape(A4)
        margin = 15 * mm
        buf = io.BytesIO()

        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=margin,
            rightMargin=margin,
            topMargin=margin,
            bottomMargin=margin,
            title="危险源辨识台账",
        )

        # 字体注册
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        # 尝试注册中文字体
        _font_name = "Helvetica"
        _font_name_bold = "Helvetica-Bold"
        _chinese_fonts = [
            ("C:/Windows/Fonts/simsun.ttc", "SimSun"),
            ("C:/Windows/Fonts/simhei.ttf", "SimHei"),
            ("C:/Windows/Fonts/msyh.ttc", "MicrosoftYaHei"),
        ]
        for font_path, font_alias in _chinese_fonts:
            try:
                pdfmetrics.registerFont(TTFont(font_alias, font_path))
                if font_alias == "SimSun":
                    _font_name = "SimSun"
                if font_alias == "SimHei":
                    _font_name_bold = "SimHei"
            except Exception:
                pass
        # 回退：如果没有中文字体，使用 Helvetica（英文/数字可用）
        logger.debug("PDF fonts: body=%s, bold=%s", _font_name, _font_name_bold)

        styles = getSampleStyleSheet()
        body_style = ParagraphStyle(
            "BodyCN",
            parent=styles["Normal"],
            fontName=_font_name,
            fontSize=8,
            leading=12,
        )
        header_style = ParagraphStyle(
            "HeaderCN",
            parent=styles["Normal"],
            fontName=_font_name_bold,
            fontSize=8,
            leading=12,
            textColor=colors.white,
        )
        title_style = ParagraphStyle(
            "TitleCN",
            parent=styles["Title"],
            fontName=_font_name_bold,
            fontSize=16,
            leading=22,
            alignment=TA_CENTER,
            spaceAfter=4,
        )
        subtitle_style = ParagraphStyle(
            "SubtitleCN",
            parent=styles["Normal"],
            fontName=_font_name,
            fontSize=9,
            leading=14,
            alignment=TA_CENTER,
            textColor=colors.grey,
        )

        # ── 构建内容 ──
        elements: list = []

        # 标题
        elements.append(Paragraph("危险源辨识台账", title_style))

        # 副标题：筛选条件 + 时间
        filter_parts: list[str] = []
        if department:
            filter_parts.append(f"部门：{department}")
        if position:
            filter_parts.append(f"岗位：{position}")
        if risk_level:
            level_map = {
                "level_1": "一级/重大风险", "level_2": "二级/较大风险",
                "level_3": "三级/一般风险", "level_4": "四级/低风险",
            }
            filter_parts.append(f"风险等级：{level_map.get(risk_level, risk_level)}")
        if date_from:
            filter_parts.append(f"起：{date_from}")
        if date_to:
            filter_parts.append(f"止：{date_to}")
        if keyword:
            filter_parts.append(f"关键词：{keyword}")
        filter_text = "；".join(filter_parts) if filter_parts else "全部记录"
        export_time = dt_module.now().strftime("%Y-%m-%d %H:%M")
        elements.append(Paragraph(
            f"筛选条件：{filter_text}　|　导出时间：{export_time}　|　共 {len(items)} 条",
            subtitle_style,
        ))
        elements.append(Spacer(1, 6 * mm))

        # ── 数据表 ──
        level_label_map = {
            "level_1": "重大", "level_2": "较大",
            "level_3": "一般", "level_4": "低",
        }
        headers = [
            "序号", "编号", "部门", "岗位", "作业活动",
            "危险类型", "固有风险", "残余风险",
            "管控层级", "责任人", "控制措施摘要",
        ]
        col_widths = [
            25, 70, 50, 50, 80,
            55, 55, 55,
            45, 50, 160,
        ]

        table_data = [headers]
        for idx, item in enumerate(items, 1):
            inherent_label = item.inherent_risk_label or level_label_map.get(
                item.inherent_risk_level or "", ""
            )
            inherent_d = f"{inherent_label}(D={int(item.d_inherent)})" if item.d_inherent and inherent_label else inherent_label or "-"
            residual_label = item.residual_risk_label or level_label_map.get(
                item.residual_risk_level or "", ""
            )
            residual_d = f"{residual_label}(D={int(item.d_residual)})" if item.d_residual and residual_label else residual_label or "-"

            # 控制措施摘要
            controls_parts = []
            if item.existing_engineering_controls:
                controls_parts.append(f"工程：{item.existing_engineering_controls[:60]}")
            if item.existing_management_controls:
                controls_parts.append(f"管理：{item.existing_management_controls[:60]}")
            if item.existing_ppe:
                controls_parts.append(f"PPE：{item.existing_ppe[:40]}")
            controls_summary = "；".join(controls_parts) if controls_parts else "-"

            row = [
                str(idx),
                item.hazard_id_no or "",
                item.department or "",
                item.position or "",
                item.specific_activity or item.production_step or "",
                item.hazard_type or "",
                inherent_d,
                residual_d,
                item.control_level or "",
                item.responsible_person or "",
                controls_summary,
            ]
            table_data.append(row)

        table = Table(table_data, colWidths=[w * mm / 4 for w in col_widths], repeatRows=1)
        table.setStyle(TableStyle([
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#5645D4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), _font_name_bold),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # Body
            ("FONTNAME", (0, 1), (-1, -1), _font_name),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (1, 1), (1, -1), "CENTER"),
            # Grid
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
            ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#3D2DA6")),
            # Row striping
            *[
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F7F6FB"))
                for i in range(2, len(table_data) + 1, 2)
            ],
        ]))
        elements.append(table)
        elements.append(Spacer(1, 10 * mm))

        # ── 签发栏 ──
        sign_style = ParagraphStyle(
            "SignCN",
            parent=body_style,
            fontSize=10,
            leading=16,
        )
        sign_table = Table(
            [[
                Paragraph("编制人：______________", sign_style),
                Paragraph("审核人：______________", sign_style),
                Paragraph("批准人：______________", sign_style),
            ]],
            colWidths=[page_w / 3 - 20, page_w / 3 - 20, page_w / 3 - 20],
        )
        sign_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, -1), _font_name),
            ("LEFTPADDING", (0, 0), (-1, -1), 30),
            ("RIGHTPADDING", (0, 0), (-1, -1), 30),
        ]))
        sign_data = [
            [Paragraph("日期：______________", sign_style),
             Paragraph("日期：______________", sign_style),
             Paragraph("日期：______________", sign_style)],
        ]
        sign_table2 = Table(
            sign_data,
            colWidths=[page_w / 3 - 20, page_w / 3 - 20, page_w / 3 - 20],
        )
        sign_table2.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, -1), _font_name),
            ("LEFTPADDING", (0, 0), (-1, -1), 30),
            ("RIGHTPADDING", (0, 0), (-1, -1), 30),
        ]))
        elements.append(sign_table)
        elements.append(Spacer(1, 4 * mm))
        elements.append(sign_table2)

        # ── 生成 PDF ──
        def add_page_number(canvas, doc_obj):
            canvas.saveState()
            canvas.setFont(_font_name, 8)
            canvas.drawCentredString(
                page_w / 2, 10 * mm,
                f"第 {canvas.getPageNumber()} 页",
            )
            canvas.restoreState()

        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        buf.seek(0)
        return buf.getvalue()


# ==================== 操规修订 Service ====================


class RegulationService:
    """安全操规修订业务服务

    两大核心能力：
    1. 安全操规修订 — 人工修订 / AI修订
    2. 危险源辨识修订 — 工艺变更自动触发
    """

    def __init__(self, session: AsyncSession):
        self.session = session
        self.repo = SafetyRepository(session)

    # ==================== 安全操作规程 CRUD ====================

    async def get_regulations(
        self,
        skip: int = 0,
        limit: int = 20,
        position: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list, int]:
        """获取操规列表"""
        return await self.repo.get_regulations(skip, limit, position, keyword)

    async def get_regulation(self, regulation_id: uuid.UUID):
        """获取操规详情"""
        return await self.repo.get_regulation_by_id(regulation_id)

    async def create_regulation(self, data) -> Any:
        """创建安全操作规程"""

        create_data = data.model_dump() if not isinstance(data, dict) else data
        return await self.repo.create_regulation(create_data)

    async def update_regulation(self, regulation_id: uuid.UUID, data) -> Any | None:
        """更新安全操作规程"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_regulation(regulation_id, update_data)

    async def delete_regulation(self, regulation_id: uuid.UUID) -> bool:
        """删除安全操作规程"""
        return await self.repo.delete_regulation(regulation_id)

    # ==================== 修订记录 CRUD ====================

    async def get_revisions(
        self,
        skip: int = 0,
        limit: int = 20,
        regulation_id: uuid.UUID | None = None,
        revision_type: str | None = None,
        review_opinion: str | None = None,
        revision_scope: str | None = None,
    ) -> tuple[list, int]:
        """获取修订记录列表"""
        return await self.repo.get_revisions(
            skip, limit, regulation_id, revision_type, review_opinion, revision_scope
        )

    async def get_revision(self, revision_id: uuid.UUID):
        """获取修订记录详情"""
        return await self.repo.get_revision_by_id(revision_id)

    async def create_revision(self, data) -> Any:
        """创建修订记录

        自动从安全操作规程表获取当前文档链接填入旧文档链接。
        """
        reg = await self.repo.get_regulation_by_id(data.regulation_id)
        if not reg:
            return None

        revision_data = {
            "revision_no": data.revision_no,
            "regulation_id": data.regulation_id,
            "regulation_name": reg.regulation_name,
            "old_document_path": reg.document_path,
            "revision_type": data.revision_type.value if hasattr(data.revision_type, 'value') else data.revision_type,
            "revision_opinion": data.revision_opinion,
            "reviser": data.reviser,
            "reviser_name": data.reviser_name,
            "revision_time": datetime.now(),
            "notes": data.notes,
        }
        return await self.repo.create_revision(revision_data)

    async def update_revision(self, revision_id: uuid.UUID, data) -> Any | None:
        """更新修订记录"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_revision(revision_id, update_data)

    async def delete_revision(self, revision_id: uuid.UUID) -> bool:
        """删除修订记录"""
        return await self.repo.delete_revision(revision_id)

    # ==================== 人工修订流程 ====================

    async def manual_revision_complete(
        self,
        revision_id: uuid.UUID,
        new_document_path: str,
        new_document_name: str | None = None,
    ) -> Any | None:
        """完成人工修订：

        1. 将新文档路径填入修订记录
        2. 更新操规表的文档链接和更新日期
        3. 审核意见直接填"已审核"（人工修订无需额外审核）
        4. 触发修订范围识别
        """
        revision = await self.repo.get_revision_by_id(revision_id)
        if not revision or revision.revision_type != "manual":
            return None

        # 更新修订记录：新文档链接 + 审核通过
        await self.repo.update_revision(
            revision_id,
            {
                "new_document_path": new_document_path,
                "review_opinion": "approved",
            },
        )

        # 同步更新操规表：最新文档链接 + 更新时间
        await self.repo.update_regulation(
            revision.regulation_id,
            {
                "document_path": new_document_path,
                "document_original_name": new_document_name,
            },
        )

        # 刷新修订记录以获取最新数据
        await self.session.flush()
        return await self.repo.get_revision_by_id(revision_id)

    # ==================== AI 修订流程 ====================

    async def ai_revision_generate(
        self,
        revision_id: uuid.UUID,
    ) -> dict | None:
        """AI 根据修订意见生成修订版本（不持久化，返回给用户确认）

        返回 {"generated_content": str} 供前端展示对比。
        """
        revision = await self.repo.get_revision_by_id(revision_id)
        if not revision or revision.revision_type != "ai":
            return None

        reg = await self.repo.get_regulation_by_id(revision.regulation_id)
        if not reg:
            return None

        # 读取当前操规文档内容
        current_content = ""
        if reg.document_path:
            try:
                current_content = self._read_document(reg.document_path)
            except Exception as e:
                logger.warning(f"读取文档失败: {e}")
                current_content = "（无法读取当前文档）"

        # 调用 AI 生成修订版本
        generated = await self._ai_generate_revision(
            regulation_name=reg.regulation_name,
            current_content=current_content,
            revision_opinion=revision.revision_opinion or "",
        )

        return {"generated_content": generated}

    async def ai_revision_confirm(
        self,
        revision_id: uuid.UUID,
        generated_content: str,
        document_name: str | None = None,
    ) -> Any | None:
        """用户确认 AI 修订内容后：

        1. 保存新文档到 uploads/
        2. 更新修订记录的 new_document_path
        3. 同步更新操规表
        4. 审核意见填"已审核"
        """
        import os

        revision = await self.repo.get_revision_by_id(revision_id)
        if not revision or revision.revision_type != "ai":
            return None

        # 保存生成的文档
        upload_dir = os.path.join("uploads", "safety", "regulations")
        os.makedirs(upload_dir, exist_ok=True)

        safe_name = f"revision_{revision_id}_{int(datetime.now().timestamp())}.md"
        file_path = os.path.join(upload_dir, safe_name)

        with open(file_path, "w", encoding="utf-8") as f:
            f.write(generated_content)

        doc_name = document_name or f"{revision.regulation_name}_修订版_{int(datetime.now().timestamp())}.md"

        # 更新修订记录
        await self.repo.update_revision(
            revision_id,
            {
                "new_document_path": file_path,
                "review_opinion": "approved",
            },
        )

        # 同步更新操规表
        await self.repo.update_regulation(
            revision.regulation_id,
            {
                "document_path": file_path,
                "document_original_name": doc_name,
            },
        )

        await self.session.flush()
        return await self.repo.get_revision_by_id(revision_id)

    # ==================== 修订范围识别（AI） ====================

    async def identify_revision_scope(self, revision_id: uuid.UUID) -> Any | None:
        """AI 识别修订范围（工艺/安全要求）

        在修订完成后（新文档链接已填充）调用。
        分析修订意见内容，识别属于工艺变更还是安全要求变更。
        """
        revision = await self.repo.get_revision_by_id(revision_id)
        if not revision:
            return None

        if not revision.revision_opinion:
            # 无修订意见，默认仅安全要求
            await self.repo.update_revision(revision_id, {"revision_scope": "safety_requirement"})
            await self.session.flush()
            return await self.repo.get_revision_by_id(revision_id)

        # 调用 AI 识别修订范围
        scope_result = await self._ai_identify_scope(
            revision_opinion=revision.revision_opinion,
            revision_type=revision.revision_type,
            regulation_name=revision.regulation_name,
        )

        await self.repo.update_revision(revision_id, {"revision_scope": scope_result})
        await self.session.flush()

        updated = await self.repo.get_revision_by_id(revision_id)

        return updated

    # ==================== AI 辅助方法 ====================

    async def _get_ai_client(self) -> AIService:
        """获取文本模型 AI 服务客户端（安全模块数据库配置）"""
        config = await self.repo.get_active_api_call_config(config_type="text")
        if config:
            logger.debug("使用数据库 API 配置: %s (%s)", config.config_name, config.model_name)
            return AIService(
                api_key=config.api_key,
                base_url=config.api_base_url,
                model=config.model_name,
                timeout=config.timeout_seconds,
            )

        # 自动种子：从环境变量迁移到数据库
        await _ensure_ai_config_seeded(self.session, "text")
        config = await self.repo.get_active_api_call_config(config_type="text")
        if config:
            return AIService(
                api_key=config.api_key,
                base_url=config.api_base_url,
                model=config.model_name,
                timeout=config.timeout_seconds,
            )

        raise AIOutputError("安全模块未配置文本 AI 模型，请在 API 配置页面进行配置")

    async def _ai_identify_scope(
        self,
        revision_opinion: str,
        revision_type: str,
        regulation_name: str,
    ) -> str:
        """AI 识别修订范围

        Returns:
            逗号分隔的范围字符串，如 "process,safety_requirement"
        """
        prompt = f"""你是一个专业的安全生产管理专家。请分析以下操规修订意见，判断修订范围属于"工艺"还是"安全要求"。

操规名称：{regulation_name}
修订类型：{"人工修订" if revision_type == "manual" else "AI修订"}
修订意见：
{revision_opinion}

判断标准：
- 工艺（process）：涉及工艺参数调整、操作步骤变更、工艺条件修改、设备参数调整等
- 安全要求（safety_requirement）：涉及安全措施、防护要求、警示标识、联锁装置、应急措施等

请返回 JSON 格式：
{{"scope": "process" 或 "safety_requirement" 或 "process,safety_requirement"（两者都有时逗号分隔）, "reasoning": "识别依据说明"}}"""

        try:
            ai = await self._get_ai_client()
            result = await ai.chat_parsed(
                messages=[
                    {"role": "system", "content": "你是一个专业的安全生产管理专家，擅长识别操规修订的影响范围。"},
                    {"role": "user", "content": prompt},
                ],
                expected_keys=["scope", "reasoning"],
            )
            await ai.close()
            return result.get("scope", "safety_requirement")
        except AIOutputError:
            logger.warning("AI 识别修订范围失败，默认标记为安全要求")
            return "safety_requirement"
        except Exception as e:
            logger.error(f"AI 识别修订范围异常: {e}")
            return "safety_requirement"

    async def _ai_generate_revision(
        self,
        regulation_name: str,
        current_content: str,
        revision_opinion: str,
    ) -> str:
        """AI 根据修订意见生成新版本的操规文档"""
        prompt = f"""请根据以下修订意见，对安全操作规程进行修订，生成完整的修订后文档。

操规名称：{regulation_name}

当前操规内容：
{current_content if current_content else "（无当前内容，请根据操规名称和修订意见生成完整文档）"}

修订意见：
{revision_opinion}

要求：
1. 生成完整的修订后文档，而非仅修改部分
2. 保持文档的结构和格式
3. 用注释标注修改过的部分，格式为：【修订】原内容 → 新内容
4. 在文档末尾添加修订说明

请直接输出修订后的完整文档内容。"""

        try:
            ai = await self._get_ai_client()
            messages = [
                {"role": "system", "content": "你是一个专业的安全操作规程编写专家，服务于原料药生产企业。"},
                {"role": "user", "content": prompt},
            ]
            # 自由文本生成，使用 chat 方法
            result = await ai.chat(
                messages=messages,
                response_format="text",
                max_tokens=16384,
            )
            await ai.close()
            return result if result else ""
        except Exception as e:
            logger.error(f"AI 生成修订版本失败: {e}")
            raise AIOutputError(f"AI 生成修订版本失败: {e}")

    async def _ai_diff_analysis(
        self,
        old_content: str,
        new_content: str,
        regulation_name: str,
    ) -> dict:
        """AI 识别新旧文档差异（人工修订时使用）"""
        prompt = f"""请对比以下安全操作规程的新旧版本，识别具体差异。

操规名称：{regulation_name}

【旧版本】
{old_content}

【新版本】
{new_content}

请输出 JSON 格式：
{{"has_changes": true/false, "changes": [{{"section": "章节/条款号", "old_text": "旧内容摘要", "new_text": "新内容摘要", "change_type": "新增/修改/删除"}}], "summary": "差异摘要说明"}}"""

        try:
            ai = await self._get_ai_client()
            result = await ai.chat_parsed(
                messages=[
                    {"role": "system", "content": "你是一个专业的文档对比分析专家。"},
                    {"role": "user", "content": prompt},
                ],
                expected_keys=["has_changes", "changes", "summary"],
            )
            await ai.close()
            return result
        except Exception as e:
            logger.error(f"AI 差异分析失败: {e}")
            raise AIOutputError(f"AI 差异分析失败: {e}")

    @staticmethod
    def _read_document(path: str, max_chars: int = 50000) -> str:
        """读取文档内容"""
        import os

        if not os.path.exists(path):
            raise FileNotFoundError(f"文档不存在: {path}")

        with open(path, encoding="utf-8") as f:
            content = f.read(max_chars)
        return content

    # ==================== 文档上传处理 ====================

    async def upload_regulation_document(
        self, regulation_id: uuid.UUID, file_name: str, file_path: str
    ) -> Any | None:
        """上传操规文档并更新操规记录"""
        return await self.repo.update_regulation(
            regulation_id,
            {
                "document_path": file_path,
                "document_original_name": file_name,
            },
        )


# ==================== AI 配置 Service ====================


class ConfigService:
    """AI 工作流配置 & API 调用配置业务服务"""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.repo = SafetyRepository(session)

    # ==================== AI 工作流配置 CRUD ====================

    async def get_ai_workflow_configs(
        self,
        skip: int = 0,
        limit: int = 100,
        module_code: str | None = None,
        is_enabled: bool | None = None,
    ) -> tuple[list[AIWorkflowConfig], int]:
        """获取 AI 工作流配置列表"""
        return await self.repo.get_ai_workflow_configs(
            skip, limit, module_code, is_enabled
        )

    async def get_ai_workflow_config(self, config_id: uuid.UUID) -> AIWorkflowConfig | None:
        """获取 AI 工作流配置详情"""
        return await self.repo.get_ai_workflow_config_by_id(config_id)

    async def get_ai_workflow_config_by_module(
        self, module_code: str
    ) -> AIWorkflowConfig | None:
        """按模块代码获取 AI 工作流配置"""
        return await self.repo.get_ai_workflow_config_by_module(module_code)

    async def create_ai_workflow_config(self, data: AIWorkflowConfigCreate) -> AIWorkflowConfig:
        """创建 AI 工作流配置"""
        create_data = data.model_dump() if not isinstance(data, dict) else data
        return await self.repo.create_ai_workflow_config(create_data)

    async def update_ai_workflow_config(
        self, config_id: uuid.UUID, data: AIWorkflowConfigUpdate
    ) -> AIWorkflowConfig | None:
        """更新 AI 工作流配置"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_ai_workflow_config(config_id, update_data)

    async def delete_ai_workflow_config(self, config_id: uuid.UUID) -> bool:
        """删除 AI 工作流配置"""
        return await self.repo.delete_ai_workflow_config(config_id)

    # ==================== API 调用配置 CRUD ====================

    async def get_api_call_configs(
        self,
        skip: int = 0,
        limit: int = 100,
        is_active: bool | None = None,
    ) -> tuple[list[APICallConfig], int]:
        """获取 API 调用配置列表"""
        return await self.repo.get_api_call_configs(skip, limit, is_active)

    async def get_api_call_config(self, config_id: uuid.UUID) -> APICallConfig | None:
        """获取 API 调用配置详情"""
        return await self.repo.get_api_call_config_by_id(config_id)

    async def get_active_api_call_config(
        self, config_type: str = "text"
    ) -> APICallConfig | None:
        """获取当前激活的 API 调用配置"""
        return await self.repo.get_active_api_call_config(config_type)

    async def create_api_call_config(self, data: APICallConfigCreate) -> APICallConfig:
        """创建 API 调用配置"""
        create_data = data.model_dump() if not isinstance(data, dict) else data
        # 如果新配置标记为激活，停用同类型其他配置（允许 text/vision 各有一个激活）
        if create_data.get("is_active"):
            cfg_type = create_data.get("config_type", "text")
            await self.repo.deactivate_all_api_call_configs(cfg_type)
        return await self.repo.create_api_call_config(create_data)

    async def update_api_call_config(
        self, config_id: uuid.UUID, data: APICallConfigUpdate
    ) -> APICallConfig | None:
        """更新 API 调用配置"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        # 如果更新为激活，停用同类型其他配置
        if update_data.get("is_active"):
            cfg_type = update_data.get("config_type", "text")
            await self.repo.deactivate_all_api_call_configs(cfg_type)
        return await self.repo.update_api_call_config(config_id, update_data)

    async def activate_api_call_config(self, config_id: uuid.UUID) -> APICallConfig | None:
        """激活指定的 API 调用配置（停用同类型其他配置）"""
        config = await self.repo.get_api_call_config_by_id(config_id)
        if not config:
            return None
        await self.repo.deactivate_all_api_call_configs(config.config_type)
        return await self.repo.update_api_call_config(config_id, {"is_active": True})

    async def delete_api_call_config(self, config_id: uuid.UUID) -> bool:
        """删除 API 调用配置"""
        return await self.repo.delete_api_call_config(config_id)


# ==================== 特殊作业管理 Service ====================


class SpecialOperationService:
    """特殊作业管理业务服务

    两大核心能力：
    1. 特殊作业人员资质管理
    2. 特殊作业票管理（含工作流状态机）
    """

    def __init__(self, session: AsyncSession):
        self.session = session
        self.repo = SafetyRepository(session)

    # ==================== 人员资质 CRUD ====================

    async def get_personnel(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        certificate_type: str | None = None,
        department: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list[SpecialOperationPersonnel], int]:
        """获取特殊作业人员资质列表"""
        return await self.repo.get_special_operation_personnel(
            skip, limit, status, certificate_type, department, keyword
        )

    async def get_personnel_by_id(
        self, personnel_id: uuid.UUID
    ) -> SpecialOperationPersonnel | None:
        """获取人员资质详情"""
        return await self.repo.get_special_operation_personnel_by_id(personnel_id)

    async def create_personnel(
        self, data: SpecialOperationPersonnelCreate
    ) -> SpecialOperationPersonnel:
        """创建人员资质"""
        create_data = data.model_dump()
        return await self.repo.create_special_operation_personnel(create_data)

    async def update_personnel(
        self, personnel_id: uuid.UUID, data: SpecialOperationPersonnelUpdate
    ) -> SpecialOperationPersonnel | None:
        """更新人员资质"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_special_operation_personnel(
            personnel_id, update_data
        )

    async def delete_personnel(self, personnel_id: uuid.UUID) -> bool:
        """删除人员资质"""
        return await self.repo.delete_special_operation_personnel(personnel_id)

    # ==================== 作业票 CRUD ====================

    async def get_permits(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        operation_type: str | None = None,
        operation_level: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list[SpecialOperationPermit], int]:
        """获取特殊作业票列表"""
        return await self.repo.get_special_operation_permits(
            skip, limit, status, operation_type, operation_level, keyword
        )

    async def get_permit(
        self, permit_id: uuid.UUID
    ) -> SpecialOperationPermit | None:
        """获取作业票详情"""
        return await self.repo.get_special_operation_permit_by_id(permit_id)

    async def create_permit(
        self, data: SpecialOperationPermitCreate
    ) -> SpecialOperationPermit:
        """创建作业票"""
        create_data = data.model_dump()
        return await self.repo.create_special_operation_permit(create_data)

    async def update_permit(
        self, permit_id: uuid.UUID, data: SpecialOperationPermitUpdate
    ) -> SpecialOperationPermit | None:
        """更新作业票"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_special_operation_permit(permit_id, update_data)

    async def delete_permit(self, permit_id: uuid.UUID) -> bool:
        """删除作业票"""
        return await self.repo.delete_special_operation_permit(permit_id)

    # ==================== 作业票工作流 ====================

    async def submit_permit(self, permit_id: uuid.UUID) -> SpecialOperationPermit | None:
        """提交作业票（草稿→已提交）"""
        permit = await self.repo.get_special_operation_permit_by_id(permit_id)
        if not permit or permit.status != "draft":
            return None
        return await self.repo.update_special_operation_permit(
            permit_id, {"status": "submitted"}
        )

    async def approve_permit(
        self, permit_id: uuid.UUID
    ) -> SpecialOperationPermit | None:
        """审批作业票（已提交→已审批）"""
        permit = await self.repo.get_special_operation_permit_by_id(permit_id)
        if not permit or permit.status != "submitted":
            return None
        return await self.repo.update_special_operation_permit(
            permit_id, {"status": "approved"}
        )

    async def reject_permit(
        self, permit_id: uuid.UUID, reason: str
    ) -> SpecialOperationPermit | None:
        """驳回作业票（已提交→已驳回）"""
        permit = await self.repo.get_special_operation_permit_by_id(permit_id)
        if not permit or permit.status != "submitted":
            return None
        return await self.repo.update_special_operation_permit(
            permit_id, {"status": "rejected", "rejection_reason": reason}
        )

    async def start_permit(
        self, permit_id: uuid.UUID
    ) -> SpecialOperationPermit | None:
        """开始作业（已审批→作业中）"""
        permit = await self.repo.get_special_operation_permit_by_id(permit_id)
        if not permit or permit.status != "approved":
            return None
        return await self.repo.update_special_operation_permit(
            permit_id,
            {"status": "in_progress", "actual_start_time": datetime.now()},
        )

    async def complete_permit(
        self, permit_id: uuid.UUID, method: str
    ) -> SpecialOperationPermit | None:
        """完工（作业中→已完工）"""
        permit = await self.repo.get_special_operation_permit_by_id(permit_id)
        if not permit or permit.status != "in_progress":
            return None
        return await self.repo.update_special_operation_permit(
            permit_id,
            {
                "status": "completed",
                "actual_end_time": datetime.now(),
                "completion_method": method,
            },
        )

    async def archive_permit(
        self, permit_id: uuid.UUID
    ) -> SpecialOperationPermit | None:
        """归档作业票（已完工→已归档）"""
        permit = await self.repo.get_special_operation_permit_by_id(permit_id)
        if not permit or permit.status != "completed":
            return None
        return await self.repo.update_special_operation_permit(
            permit_id, {"status": "archived"}
        )


# ==================== 安全知识库 Service ====================


class KnowledgeService:
    """安全知识库业务服务"""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.repo = SafetyRepository(session)

    async def get_articles(
        self,
        skip: int = 0,
        limit: int = 20,
        category: str | None = None,
        status: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list[SafetyKnowledgeArticle], int]:
        """获取知识库文章列表"""
        return await self.repo.get_knowledge_articles(
            skip, limit, category, status, keyword
        )

    async def get_article(self, article_id: uuid.UUID) -> SafetyKnowledgeArticle | None:
        """获取文章详情（浏览计数+1）"""
        article = await self.repo.get_knowledge_article_by_id(article_id)
        if article:
            await self.repo.update_knowledge_article(
                article_id, {"view_count": article.view_count + 1}
            )
        return article

    async def create_article(
        self, data: SafetyKnowledgeArticleCreate
    ) -> SafetyKnowledgeArticle:
        """创建知识库文章"""
        article_data = data.model_dump()
        return await self.repo.create_knowledge_article(article_data)

    async def update_article(
        self, article_id: uuid.UUID, data: SafetyKnowledgeArticleUpdate
    ) -> SafetyKnowledgeArticle | None:
        """更新知识库文章"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_knowledge_article(article_id, update_data)

    async def delete_article(self, article_id: uuid.UUID) -> bool:
        """删除知识库文章"""
        return await self.repo.delete_knowledge_article(article_id)

    async def publish_article(self, article_id: uuid.UUID) -> SafetyKnowledgeArticle | None:
        """发布文章（草稿→已发布）"""
        article = await self.repo.get_knowledge_article_by_id(article_id)
        if not article or article.status != "draft":
            return None
        return await self.repo.update_knowledge_article(
            article_id, {"status": "published"}
        )

    async def archive_article(self, article_id: uuid.UUID) -> SafetyKnowledgeArticle | None:
        """归档文章（已发布→已归档）"""
        article = await self.repo.get_knowledge_article_by_id(article_id)
        if not article or article.status != "published":
            return None
        return await self.repo.update_knowledge_article(
            article_id, {"status": "archived"}
        )


# ==================== 风险作业报备 Services ====================


class SpecialOperationReportService:
    """八大特殊作业报备业务服务"""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.repo = SafetyRepository(session)

    # ── CRUD ──

    async def get_reports(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        operation_type: str | None = None,
        department: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list[SpecialOperationReport], int]:
        """获取特殊作业报备列表"""
        return await self.repo.get_special_operation_reports(
            skip, limit, status, operation_type, department, keyword
        )

    async def get_report(self, report_id: uuid.UUID) -> SpecialOperationReport | None:
        """获取报备详情"""
        return await self.repo.get_special_operation_report_by_id(report_id)

    async def create_report(
        self, data: SpecialOperationReportCreate
    ) -> SpecialOperationReport:
        """创建报备"""
        return await self.repo.create_special_operation_report(data.model_dump())

    async def update_report(
        self, report_id: uuid.UUID, data: SpecialOperationReportUpdate
    ) -> SpecialOperationReport | None:
        """更新报备"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_special_operation_report(report_id, update_data)

    async def delete_report(self, report_id: uuid.UUID) -> bool:
        """删除报备"""
        return await self.repo.delete_special_operation_report(report_id)

    # ── 工作流 ──

    async def submit_report(
        self, report_id: uuid.UUID
    ) -> SpecialOperationReport | None:
        """提交报备（草稿→已提交），并自动判定是否为关键作业"""
        report = await self.repo.get_special_operation_report_by_id(report_id)
        if not report or report.status != "draft":
            return None

        # AI 自动判定关键作业
        is_critical, critical_reason = await self._identify_critical(report)

        return await self.repo.update_special_operation_report(
            report_id, {
                "status": "submitted",
                "is_critical": is_critical,
                "is_critical_reason": critical_reason,
            }
        )

    async def approve_report(
        self, report_id: uuid.UUID
    ) -> SpecialOperationReport | None:
        """审批报备（已提交→已审批）"""
        report = await self.repo.get_special_operation_report_by_id(report_id)
        if not report or report.status != "submitted":
            return None
        return await self.repo.update_special_operation_report(
            report_id,
            {"status": "approved", "approved_at": datetime.now()},
        )

    async def reject_report(
        self, report_id: uuid.UUID, reason: str
    ) -> SpecialOperationReport | None:
        """驳回报备（已提交→已驳回）"""
        report = await self.repo.get_special_operation_report_by_id(report_id)
        if not report or report.status != "submitted":
            return None
        return await self.repo.update_special_operation_report(
            report_id,
            {"status": "rejected", "rejection_reason": reason},
        )

    async def set_critical_manual(
        self, report_id: uuid.UUID, is_critical: bool, reason: str | None, updated_by: str | None
    ) -> SpecialOperationReport | None:
        """手动修改关键作业标记"""
        report = await self.repo.get_special_operation_report_by_id(report_id)
        if not report:
            return None
        return await self.repo.update_special_operation_report(
            report_id, {
                "is_critical": is_critical,
                "is_critical_reason": reason,
                "is_critical_updated_by": updated_by,
            }
        )

    # ── AI 判定 ──

    async def _get_workflow_config(self, module_code: str) -> dict:
        """获取独立工作流配置（DB 优先，fallback 到硬编码 STANDALONE_WORKFLOW_CONFIG）"""
        workflow = await self.repo.get_ai_workflow_config_by_module(module_code)
        if workflow and workflow.is_enabled and workflow.script_configs:
            scripts = workflow.script_configs
            if isinstance(scripts, list) and len(scripts) > 0:
                first = scripts[0]
                if first.get("is_enabled", True):
                    logger.debug(
                        "使用 DB 工作流配置: %s — %s", module_code, first.get("name")
                    )
                    return first
        # Fallback 到硬编码
        config = STANDALONE_WORKFLOW_CONFIG.get(module_code, {})
        if config:
            logger.debug("使用硬编码工作流配置: %s — %s", module_code, config.get("name"))
        return config

    async def _get_ai_service(self) -> "AIService":
        """获取文本模型 AI 服务客户端（安全模块数据库配置）"""
        config = await self.repo.get_active_api_call_config(config_type="text")
        if config:
            logger.debug("使用数据库 API 配置: %s (%s)", config.config_name, config.model_name)
            return AIService(
                api_key=config.api_key,
                base_url=config.api_base_url,
                model=config.model_name,
                timeout=config.timeout_seconds,
            )

        # 自动种子：从环境变量迁移到数据库
        await _ensure_ai_config_seeded(self.session, "text")
        config = await self.repo.get_active_api_call_config(config_type="text")
        if config:
            return AIService(
                api_key=config.api_key,
                base_url=config.api_base_url,
                model=config.model_name,
                timeout=config.timeout_seconds,
            )

        raise AIOutputError("安全模块未配置文本 AI 模型，请在 API 配置页面进行配置")

    async def _identify_critical(
        self, report: "SpecialOperationReport"
    ) -> tuple[bool, str | None]:
        """判定报备是否为关键作业（AI 优先，失败时基于规则 fallback）"""
        try:
            ai = await self._get_ai_service()
            return await self._ai_identify_critical(ai, report)
        except Exception as e:
            logger.warning("AI 关键作业判定失败，使用规则 fallback: %s", e)
            return self._rule_based_identify_critical(report)

    async def _ai_identify_critical(
        self, ai: "AIService", report: "SpecialOperationReport"
    ) -> tuple[bool, str | None]:
        """使用 AI 判定关键作业（提示词由工作流配置提供）"""
        OP_TYPE_LABELS = {
            "hot_work": "动火作业", "confined_space": "受限空间",
            "height_work": "高处作业", "temporary_electricity": "临时用电",
            "blind_plate": "盲板抽堵", "excavation": "动土作业",
            "lifting": "起重吊装", "road_breaking": "断路作业",
        }
        op_label = OP_TYPE_LABELS.get(report.operation_type, report.operation_type)

        context = (
            f"作业类型：{op_label}\n"
            f"作业级别：{report.operation_level or '未指定'}\n"
            f"作业地点：{report.location or '未指定'}\n"
            f"作业部门：{report.department or '未指定'}\n"
            f"作业内容：{report.work_description or '未指定'}\n"
            f"风险等级：{report.risk_level or '未指定'}\n"
            f"安全措施：{report.safety_measures or '未指定'}\n"
            f"风险评估：{report.risk_assessment or '未指定'}\n"
            f"应急消防器材：{report.emergency_equipment or '未指定'}"
        )

        # 从工作流配置构建提示词（DB 优先，fallback 到硬编码）
        wf_config = await self._get_workflow_config("special-ops-critical")
        if wf_config:
            prompt = build_prompt(wf_config) + "\n\n## 本次判定输入\n" + context
        else:
            prompt = context  # 极端情况：无任何配置可用

        messages = [
            {"role": "system", "content": "你是一名化工安全专家，严格按照 GB 30871-2022 标准判定。只返回 JSON。"},
            {"role": "user", "content": prompt},
        ]

        response_text = await ai.chat(messages, response_format="json_object")
        import json
        result = json.loads(response_text)
        return result.get("is_critical", False), result.get("reason")

    def _rule_based_identify_critical(
        self, report: "SpecialOperationReport"
    ) -> tuple[bool, str | None]:
        """基于规则的 fallback 关键作业判定"""
        high_risk_types = {"hot_work", "confined_space", "height_work", "lifting"}
        critical_reasons: list[str] = []

        # 特级/一级作业
        if report.operation_level in ("special", "grade1"):
            critical_reasons.append(f"作业级别为{report.operation_level}")

        # 高风险等级
        if report.risk_level in ("level_1", "level_2"):
            critical_reasons.append(f"风险等级为{report.risk_level}")

        # 高风险作业类型 + 特级/一级
        if report.operation_type in high_risk_types and report.operation_level in ("special", "grade1"):
            if not critical_reasons:
                critical_reasons.append(
                    f"{report.operation_type} 作业类型属于高风险作业"
                )

        if critical_reasons:
            return True, "；".join(critical_reasons)
        return False, None

    # ── 台账查询 ──

    async def get_ledger(
        self,
        skip: int = 0,
        limit: int = 20,
        status_list: list[str] | None = None,
        operation_type: str | None = None,
        operation_level: str | None = None,
        risk_level: str | None = None,
        department: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        keyword: str | None = None,
        is_critical: bool | None = None,
    ) -> tuple[list[SpecialOperationReport], int]:
        """获取特殊作业台账列表"""
        return await self.repo.get_special_operation_ledger(
            skip=skip,
            limit=limit,
            status_list=status_list,
            operation_type=operation_type,
            operation_level=operation_level,
            risk_level=risk_level,
            department=department,
            date_from=date_from,
            date_to=date_to,
            keyword=keyword,
            is_critical=is_critical,
        )

    async def get_ledger_stats(
        self, status_list: list[str] | None = None
    ) -> list[dict]:
        """获取台账统计"""
        return await self.repo.get_special_operation_ledger_stats(status_list)

    # ── AI 导出 ──

    async def parse_natural_query(
        self, natural_query: str
    ) -> dict:
        """使用 AI 将自然语言筛选条件解析为结构化参数（提示词由工作流配置提供）"""
        OP_TYPE_LABELS = {
            "动火作业": "hot_work", "受限空间": "confined_space",
            "高处作业": "height_work", "临时用电": "temporary_electricity",
            "盲板抽堵": "blind_plate", "动土作业": "excavation",
            "起重吊装": "lifting", "断路作业": "road_breaking",
        }

        # 从工作流配置构建提示词（DB 优先，fallback 到硬编码）
        wf_config = await self._get_workflow_config("special-ops-export")
        if wf_config:
            prompt = build_prompt(wf_config) + "\n\n用户查询：" + natural_query
        else:
            prompt = natural_query

        try:
            ai = await self._get_ai_service()
            messages = [
                {"role": "system", "content": "你是一个数据库查询助手。只返回 JSON。"},
                {"role": "user", "content": prompt},
            ]
            response_text = await ai.chat(messages, response_format="json_object")
            import json
            result = json.loads(response_text)
            # 验证 operation_type 值
            if result.get("operation_type"):
                op_type = result["operation_type"]
                if op_type in OP_TYPE_LABELS:
                    result["operation_type"] = OP_TYPE_LABELS[op_type]
            # 清除 None 值
            return {k: v for k, v in result.items() if v is not None}
        except Exception as e:
            logger.warning("AI 自然语言解析失败: %s", e)
            return {"explanation": f"AI 解析失败，将使用原始查询: {natural_query}", "keyword": natural_query}

    async def export_ledger_excel(
        self,
        operation_type: str | None = None,
        operation_level: str | None = None,
        risk_level: str | None = None,
        department: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        keyword: str | None = None,
        is_critical: bool | None = None,
    ) -> bytes:
        """导出台账为 Excel 文件"""
        import io

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        items, _ = await self.get_ledger(
            skip=0, limit=10000,  # 导出全部（上限1万条）
            operation_type=operation_type, operation_level=operation_level,
            risk_level=risk_level, department=department,
            date_from=date_from, date_to=date_to,
            keyword=keyword, is_critical=is_critical,
        )

        OP_TYPE_LABELS = {
            "hot_work": "动火作业", "confined_space": "受限空间",
            "height_work": "高处作业", "temporary_electricity": "临时用电",
            "blind_plate": "盲板抽堵", "excavation": "动土作业",
            "lifting": "起重吊装", "road_breaking": "断路作业",
        }
        OP_LEVEL_LABELS = {"special": "特级", "grade1": "一级", "grade2": "二级", "not_applicable": "不涉及"}
        STATUS_LABELS = {"draft": "草稿", "submitted": "审批中", "approved": "已审批", "rejected": "已驳回"}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "特殊作业台账"

        # 标题样式
        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="5645D4", end_color="5645D4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_font = Font(name="微软雅黑", size=10)
        cell_alignment = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        critical_fill = PatternFill(start_color="FDE0EC", end_color="FDE0EC", fill_type="solid")

        # 表头
        headers = [
            "序号", "报备编号", "作业类型", "作业级别", "作业地点",
            "作业内容", "作业部门", "计划开始", "计划结束",
            "报备人", "审批人", "审批时间",
            "状态", "是否关键作业", "关键作业判定理由", "备注",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据行
        for row_idx, item in enumerate(items, 2):
            values = [
                row_idx - 1,
                item.report_no or "",
                OP_TYPE_LABELS.get(item.operation_type, item.operation_type or ""),
                OP_LEVEL_LABELS.get(item.operation_level, item.operation_level or ""),
                item.location or "",
                item.work_description or "",
                item.department or "",
                item.planned_start_time.strftime("%Y-%m-%d %H:%M") if item.planned_start_time else "",
                item.planned_end_time.strftime("%Y-%m-%d %H:%M") if item.planned_end_time else "",
                item.applicant_name or "",
                item.approver_name or "",
                item.approved_at.strftime("%Y-%m-%d %H:%M") if item.approved_at else "",
                STATUS_LABELS.get(item.status, item.status or ""),
                "是" if item.is_critical else "否",
                item.is_critical_reason or "",
                item.notes or "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                # 关键作业行高亮
                if item.is_critical:
                    cell.fill = critical_fill

        # 列宽
        col_widths = [6, 14, 10, 8, 14, 24, 12, 16, 16, 8, 8, 16, 8, 10, 28, 20]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class DailyRiskReportService:
    """每日风险作业报备业务服务"""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.repo = SafetyRepository(session)

    # ── CRUD ──

    async def get_reports(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        department: str | None = None,
        report_date: datetime | None = None,
        keyword: str | None = None,
    ) -> tuple[list[DailyRiskReport], int]:
        """获取每日风险作业报备列表"""
        return await self.repo.get_daily_risk_reports(
            skip, limit, status, department, report_date, keyword
        )

    async def get_report(self, report_id: uuid.UUID) -> DailyRiskReport | None:
        """获取报备详情"""
        return await self.repo.get_daily_risk_report_by_id(report_id)

    async def create_report(
        self, data: DailyRiskReportCreate
    ) -> DailyRiskReport:
        """创建报备"""
        return await self.repo.create_daily_risk_report(data.model_dump())

    async def update_report(
        self, report_id: uuid.UUID, data: DailyRiskReportUpdate
    ) -> DailyRiskReport | None:
        """更新报备"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_daily_risk_report(report_id, update_data)

    async def delete_report(self, report_id: uuid.UUID) -> bool:
        """删除报备"""
        return await self.repo.delete_daily_risk_report(report_id)

    # ── 工作流 ──

    async def submit_report(
        self, report_id: uuid.UUID
    ) -> DailyRiskReport | None:
        """提交报备（草稿→已提交）"""
        report = await self.repo.get_daily_risk_report_by_id(report_id)
        if not report or report.status != "draft":
            return None
        return await self.repo.update_daily_risk_report(
            report_id, {"status": "submitted"}
        )

    async def approve_report(
        self, report_id: uuid.UUID
    ) -> DailyRiskReport | None:
        """审批报备（已提交→已审批）"""
        report = await self.repo.get_daily_risk_report_by_id(report_id)
        if not report or report.status != "submitted":
            return None
        return await self.repo.update_daily_risk_report(
            report_id,
            {"status": "approved", "approved_at": datetime.now()},
        )

    async def reject_report(
        self, report_id: uuid.UUID, reason: str
    ) -> DailyRiskReport | None:
        """驳回报备（已提交→已驳回）"""
        report = await self.repo.get_daily_risk_report_by_id(report_id)
        if not report or report.status != "submitted":
            return None
        return await self.repo.update_daily_risk_report(
            report_id,
            {"status": "rejected", "rejection_reason": reason},
        )


# ==================== EHS变更管理 (MOC) Service ====================


class EhsChangeService:
    """EHS变更管理业务服务（基于 T/CCSAS 007-2020）"""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.repo = SafetyRepository(session)

    # ── CRUD ──

    async def get_ehs_changes(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        change_type: str | None = None,
        change_grade: str | None = None,
        change_duration: str | None = None,
        department: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list[EhsChange], int]:
        """获取EHS变更列表"""
        return await self.repo.get_ehs_changes(
            skip, limit, status, change_type, change_grade, change_duration, department, keyword
        )

    async def get_ehs_change(self, change_id: uuid.UUID) -> EhsChange | None:
        """获取EHS变更详情"""
        return await self.repo.get_ehs_change_by_id(change_id)

    async def create_ehs_change(self, data: EhsChangeCreate) -> EhsChange:
        """创建EHS变更"""
        create_data = data.model_dump()
        return await self.repo.create_ehs_change(create_data)

    async def update_ehs_change(
        self, change_id: uuid.UUID, data: EhsChangeUpdate
    ) -> EhsChange | None:
        """更新EHS变更"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_ehs_change(change_id, update_data)

    async def delete_ehs_change(self, change_id: uuid.UUID) -> bool:
        """删除EHS变更（软删除）"""
        return await self.repo.delete_ehs_change(change_id)

    # ── 工作流状态机 ──

    async def submit_change(self, change_id: uuid.UUID) -> EhsChange | None:
        """提交变更（草稿→审核中；紧急变更自动批准）"""
        change = await self.repo.get_ehs_change_by_id(change_id)
        if not change or change.status != "draft":
            return None

        # 紧急变更：自动批准，保留审批链追溯
        if change.change_duration == "emergency":
            approval_chain = list(change.approval_chain or [])
            approval_chain.append({
                "level": 1,
                "approver_role": "系统自动批准（紧急变更）",
                "approver": "系统",
                "decision": "approved",
                "comments": "紧急变更，自动批准。需在48小时内补办审批手续。",
                "decided_at": datetime.now().isoformat(),
            })
            return await self.repo.update_ehs_change(
                change_id,
                {"status": "approved", "approval_chain": approval_chain},
            )

        return await self.repo.update_ehs_change(
            change_id, {"status": "under_review"}
        )

    async def approve_change(
        self, change_id: uuid.UUID, decision: str, comments: str | None = None
    ) -> EhsChange | None:
        """审批变更（审核中→已批准/已驳回）"""
        change = await self.repo.get_ehs_change_by_id(change_id)
        if not change or change.status != "under_review":
            return None

        if decision == "approved":
            return await self.repo.update_ehs_change(
                change_id, {"status": "approved"}
            )
        elif decision == "rejected":
            return await self.repo.update_ehs_change(
                change_id, {"status": "rejected"}
            )
        return None

    async def reject_change(
        self, change_id: uuid.UUID, comments: str | None = None
    ) -> EhsChange | None:
        """驳回变更（审核中→已驳回）"""
        change = await self.repo.get_ehs_change_by_id(change_id)
        if not change or change.status != "under_review":
            return None
        return await self.repo.update_ehs_change(
            change_id, {"status": "rejected"}
        )

    async def start_implementation(self, change_id: uuid.UUID) -> EhsChange | None:
        """开始实施（已批准→实施中）"""
        change = await self.repo.get_ehs_change_by_id(change_id)
        if not change or change.status != "approved":
            return None
        return await self.repo.update_ehs_change(
            change_id,
            {"status": "in_progress", "actual_start": datetime.now()},
        )

    async def commission_change(self, change_id: uuid.UUID) -> EhsChange | None:
        """投用（实施中→已投用）"""
        change = await self.repo.get_ehs_change_by_id(change_id)
        if not change or change.status != "in_progress":
            return None
        return await self.repo.update_ehs_change(
            change_id,
            {"status": "commissioned", "actual_completion": datetime.now()},
        )

    async def close_change(
        self,
        change_id: uuid.UUID,
        closed_by: str | None = None,
        temp_expiry_date: str | None = None,
        restored_date: str | None = None,
    ) -> EhsChange | None:
        """关闭变更（已投用→已关闭）"""
        change = await self.repo.get_ehs_change_by_id(change_id)
        if not change or change.status != "commissioned":
            return None

        closure_data = {
            "closed_by": closed_by,
            "closed_date": datetime.now().isoformat(),
        }
        if temp_expiry_date:
            closure_data["temp_expiry_date"] = temp_expiry_date
        if restored_date:
            closure_data["restored_date"] = restored_date

        return await self.repo.update_ehs_change(
            change_id,
            {"status": "closed", "closure": closure_data},
        )

    async def cancel_change(self, change_id: uuid.UUID) -> EhsChange | None:
        """取消变更（草稿→已关闭）"""
        change = await self.repo.get_ehs_change_by_id(change_id)
        if not change or change.status != "draft":
            return None
        closure_data = {
            "closed_by": None,
            "closed_date": datetime.now().isoformat(),
        }
        return await self.repo.update_ehs_change(
            change_id,
            {"status": "closed", "closure": closure_data},
        )

    # ── JSON 子记录操作 ──

    async def add_risk_assessment(
        self, change_id: uuid.UUID, item: dict
    ) -> EhsChange | None:
        """追加风险评估记录"""
        change = await self.repo.get_ehs_change_by_id(change_id)
        if not change:
            return None
        assessments = list(change.risk_assessments or [])
        assessments.append(item)
        return await self.repo.update_ehs_change(
            change_id, {"risk_assessments": assessments}
        )

    async def update_action_item(
        self, change_id: uuid.UUID, index: int, status: str
    ) -> EhsChange | None:
        """更新行动项状态"""
        change = await self.repo.get_ehs_change_by_id(change_id)
        if not change:
            return None
        items = list(change.action_items or [])
        if index < 0 or index >= len(items):
            return None
        items[index] = {**items[index], "status": status}
        if status == "completed":
            items[index]["completed_at"] = datetime.now().isoformat()
        return await self.repo.update_ehs_change(
            change_id, {"action_items": items}
        )

    async def update_pssr_checklist(
        self, change_id: uuid.UUID, items: list[dict]
    ) -> EhsChange | None:
        """更新PSSR检查清单"""
        change = await self.repo.get_ehs_change_by_id(change_id)
        if not change:
            return None
        return await self.repo.update_ehs_change(
            change_id, {"pssr_checklist": items}
        )

    async def submit_verification(
        self, change_id: uuid.UUID, data: dict
    ) -> EhsChange | None:
        """提交变更验证数据"""
        change = await self.repo.get_ehs_change_by_id(change_id)
        if not change:
            return None
        return await self.repo.update_ehs_change(
            change_id, {"verification": data}
        )


# ==================== 职业危害因素监测 Service ====================


class OhHazardMonitorService:
    """职业危害因素监测服务"""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.repo = SafetyRepository(session)

    # ── CRUD ──

    async def get_monitors(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        detection_type: str | None = None,
        workplace: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list[OhHazardMonitor], int]:
        """获取监测列表"""
        return await self.repo.get_hazard_monitors(
            skip, limit, status, detection_type, workplace, keyword
        )

    async def get_monitor(self, monitor_id: uuid.UUID) -> OhHazardMonitor | None:
        """获取监测详情"""
        return await self.repo.get_hazard_monitor_by_id(monitor_id)

    async def create_monitor(self, data: Any) -> OhHazardMonitor:
        """创建监测记录"""
        create_data = data.model_dump()
        return await self.repo.create_hazard_monitor(create_data)

    async def update_monitor(self, monitor_id: uuid.UUID, data: Any) -> OhHazardMonitor | None:
        """更新监测记录"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_hazard_monitor(monitor_id, update_data)

    async def delete_monitor(self, monitor_id: uuid.UUID) -> bool:
        """删除监测记录（软删除）"""
        return await self.repo.delete_hazard_monitor(monitor_id)

    # ── 工作流 ──

    async def start_monitoring(self, monitor_id: uuid.UUID) -> OhHazardMonitor | None:
        """开始监测（草稿→检测中）"""
        monitor = await self.repo.get_hazard_monitor_by_id(monitor_id)
        if not monitor or monitor.status != "draft":
            return None
        return await self.repo.update_hazard_monitor(
            monitor_id, {"status": "in_progress"}
        )

    async def complete_monitoring(self, monitor_id: uuid.UUID) -> OhHazardMonitor | None:
        """完成监测（检测中→已完成），自动计算OEL合规状态并生成异常记录"""
        monitor = await self.repo.get_hazard_monitor_by_id(monitor_id)
        if not monitor or monitor.status != "in_progress":
            return None

        # 自动计算合规状态
        results = list(monitor.detection_results or [])
        abnormality_records = list(monitor.abnormality_records or [])
        has_exceeding = False

        for i, item in enumerate(results):
            value = item.get("detection_value")
            limit_val = item.get("oel_limit")
            if value is not None and limit_val is not None and limit_val > 0:
                ratio = value / limit_val
                if ratio > 1.0:
                    results[i]["compliance_status"] = "exceeding"
                    has_exceeding = True
                    # 自动创建异常记录
                    abnormality_records.append({
                        "abnormality_desc": (
                            f"{item.get('factor_name', '未知因素')} 检测值 {value} {item.get('unit', '')} "
                            f"超过OEL限值 {limit_val} {item.get('unit', '')}"
                        ),
                        "corrective_action": "",
                        "responsible_person": "",
                        "deadline": "",
                        "status": "open",
                        "completed_at": "",
                        "remarks": f"标准参考: {item.get('standard_ref', '')}",
                    })
                elif ratio >= 0.8:
                    results[i]["compliance_status"] = "marginal"
                else:
                    results[i]["compliance_status"] = "compliant"
            else:
                results[i]["compliance_status"] = "compliant"

        return await self.repo.update_hazard_monitor(
            monitor_id,
            {
                "status": "completed",
                "detection_results": results,
                "abnormality_records": abnormality_records,
            },
        )

    async def verify_monitoring(
        self, monitor_id: uuid.UUID, verified_by: str | None, comments: str | None
    ) -> OhHazardMonitor | None:
        """验证监测（已完成→已验证）"""
        monitor = await self.repo.get_hazard_monitor_by_id(monitor_id)
        if not monitor or monitor.status != "completed":
            return None
        update_data: dict[str, Any] = {"status": "verified"}
        if verified_by:
            update_data["verifier_name"] = verified_by
        if comments:
            update_data["notes"] = (
                f"{(monitor.notes or '')}\n验证意见: {comments}".strip()
            )
        return await self.repo.update_hazard_monitor(monitor_id, update_data)

    # ── JSON 子记录操作 ──

    async def add_detection_result(
        self, monitor_id: uuid.UUID, item: dict
    ) -> OhHazardMonitor | None:
        """追加检测结果"""
        monitor = await self.repo.get_hazard_monitor_by_id(monitor_id)
        if not monitor:
            return None
        results = list(monitor.detection_results or [])
        results.append(item)
        return await self.repo.update_hazard_monitor(
            monitor_id, {"detection_results": results}
        )

    async def update_detection_result(
        self, monitor_id: uuid.UUID, index: int, data: dict
    ) -> OhHazardMonitor | None:
        """更新检测结果"""
        monitor = await self.repo.get_hazard_monitor_by_id(monitor_id)
        if not monitor:
            return None
        results = list(monitor.detection_results or [])
        if index < 0 or index >= len(results):
            return None
        results[index] = {**results[index], **data}
        return await self.repo.update_hazard_monitor(
            monitor_id, {"detection_results": results}
        )

    async def remove_detection_result(
        self, monitor_id: uuid.UUID, index: int
    ) -> OhHazardMonitor | None:
        """删除检测结果"""
        monitor = await self.repo.get_hazard_monitor_by_id(monitor_id)
        if not monitor:
            return None
        results = list(monitor.detection_results or [])
        if index < 0 or index >= len(results):
            return None
        results.pop(index)
        return await self.repo.update_hazard_monitor(
            monitor_id, {"detection_results": results}
        )

    async def add_abnormality_record(
        self, monitor_id: uuid.UUID, item: dict
    ) -> OhHazardMonitor | None:
        """追加异常处置记录"""
        monitor = await self.repo.get_hazard_monitor_by_id(monitor_id)
        if not monitor:
            return None
        records = list(monitor.abnormality_records or [])
        records.append(item)
        return await self.repo.update_hazard_monitor(
            monitor_id, {"abnormality_records": records}
        )

    async def update_abnormality_record_status(
        self, monitor_id: uuid.UUID, index: int, status: str
    ) -> OhHazardMonitor | None:
        """更新异常处置状态"""
        monitor = await self.repo.get_hazard_monitor_by_id(monitor_id)
        if not monitor:
            return None
        records = list(monitor.abnormality_records or [])
        if index < 0 or index >= len(records):
            return None
        records[index] = {**records[index], "status": status}
        if status == "closed":
            records[index]["completed_at"] = datetime.now().isoformat()
        return await self.repo.update_hazard_monitor(
            monitor_id, {"abnormality_records": records}
        )


# ==================== 职业健康体检 Service ====================


class OhHealthExamService:
    """职业健康体检服务"""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.repo = SafetyRepository(session)

    # ── CRUD ──

    async def get_exams(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        exam_type: str | None = None,
        department: str | None = None,
        keyword: str | None = None,
    ) -> tuple[list[OhHealthExam], int]:
        """获取体检列表"""
        return await self.repo.get_health_exams(
            skip, limit, status, exam_type, department, keyword
        )

    async def get_exam(self, exam_id: uuid.UUID) -> OhHealthExam | None:
        """获取体检详情"""
        return await self.repo.get_health_exam_by_id(exam_id)

    async def create_exam(self, data: Any) -> OhHealthExam:
        """创建体检记录"""
        create_data = data.model_dump()
        return await self.repo.create_health_exam(create_data)

    async def update_exam(self, exam_id: uuid.UUID, data: Any) -> OhHealthExam | None:
        """更新体检记录"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_health_exam(exam_id, update_data)

    async def delete_exam(self, exam_id: uuid.UUID) -> bool:
        """删除体检记录（软删除）"""
        return await self.repo.delete_health_exam(exam_id)

    # ── 工作流 ──

    async def start_exam(self, exam_id: uuid.UUID) -> OhHealthExam | None:
        """开始体检（已安排→体检中）"""
        exam = await self.repo.get_health_exam_by_id(exam_id)
        if not exam or exam.status != "scheduled":
            return None
        return await self.repo.update_health_exam(
            exam_id, {"status": "in_progress"}
        )

    async def complete_exam(self, exam_id: uuid.UUID) -> OhHealthExam | None:
        """完成体检（体检中→已完成）"""
        exam = await self.repo.get_health_exam_by_id(exam_id)
        if not exam or exam.status != "in_progress":
            return None
        return await self.repo.update_health_exam(
            exam_id, {"status": "completed"}
        )

    async def archive_exam(self, exam_id: uuid.UUID) -> OhHealthExam | None:
        """归档体检（已完成→已归档）"""
        exam = await self.repo.get_health_exam_by_id(exam_id)
        if not exam or exam.status != "completed":
            return None
        return await self.repo.update_health_exam(
            exam_id, {"status": "archived"}
        )

    # ── JSON 子记录操作 ──

    async def add_exam_item(
        self, exam_id: uuid.UUID, item: dict
    ) -> OhHealthExam | None:
        """追加体检项目"""
        exam = await self.repo.get_health_exam_by_id(exam_id)
        if not exam:
            return None
        items = list(exam.exam_items or [])
        items.append(item)
        return await self.repo.update_health_exam(
            exam_id, {"exam_items": items}
        )

    async def update_exam_item(
        self, exam_id: uuid.UUID, index: int, data: dict
    ) -> OhHealthExam | None:
        """更新体检项目"""
        exam = await self.repo.get_health_exam_by_id(exam_id)
        if not exam:
            return None
        items = list(exam.exam_items or [])
        if index < 0 or index >= len(items):
            return None
        items[index] = {**items[index], **data}
        return await self.repo.update_health_exam(
            exam_id, {"exam_items": items}
        )

    async def remove_exam_item(
        self, exam_id: uuid.UUID, index: int
    ) -> OhHealthExam | None:
        """删除体检项目"""
        exam = await self.repo.get_health_exam_by_id(exam_id)
        if not exam:
            return None
        items = list(exam.exam_items or [])
        if index < 0 or index >= len(items):
            return None
        items.pop(index)
        return await self.repo.update_health_exam(
            exam_id, {"exam_items": items}
        )

    async def set_conclusion(
        self, exam_id: uuid.UUID, conclusion: str, remarks: str | None = None
    ) -> OhHealthExam | None:
        """设置体检结论（若为异常结论则自动创建异常记录）"""
        exam = await self.repo.get_health_exam_by_id(exam_id)
        if not exam:
            return None

        update_data: dict[str, Any] = {"overall_conclusion": conclusion}

        # 异常结论自动创建处置记录
        alarming_conclusions = {"suspected_od", "od_diagnosed", "contraindicated"}
        if conclusion in alarming_conclusions:
            conclusion_labels = {
                "suspected_od": "疑似职业病",
                "od_diagnosed": "职业病确诊",
                "contraindicated": "职业禁忌证",
            }
            label = conclusion_labels.get(conclusion, conclusion)
            records = list(exam.abnormality_records or [])
            records.append({
                "abnormality_desc": (
                    f"员工 {exam.employee_name} 体检结论为「{label}」"
                    + (f"，备注: {remarks}" if remarks else "")
                ),
                "corrective_action": "",
                "responsible_person": "",
                "deadline": "",
                "status": "open",
                "completed_at": "",
                "remarks": remarks or "",
            })
            update_data["abnormality_records"] = records

        return await self.repo.update_health_exam(exam_id, update_data)

    async def add_abnormality_record(
        self, exam_id: uuid.UUID, item: dict
    ) -> OhHealthExam | None:
        """追加异常处置记录"""
        exam = await self.repo.get_health_exam_by_id(exam_id)
        if not exam:
            return None
        records = list(exam.abnormality_records or [])
        records.append(item)
        return await self.repo.update_health_exam(
            exam_id, {"abnormality_records": records}
        )

    async def update_abnormality_record_status(
        self, exam_id: uuid.UUID, index: int, status: str
    ) -> OhHealthExam | None:
        """更新异常处置状态"""
        exam = await self.repo.get_health_exam_by_id(exam_id)
        if not exam:
            return None
        records = list(exam.abnormality_records or [])
        if index < 0 or index >= len(records):
            return None
        records[index] = {**records[index], "status": status}
        if status == "closed":
            records[index]["completed_at"] = datetime.now().isoformat()
        return await self.repo.update_health_exam(
            exam_id, {"abnormality_records": records}
        )
