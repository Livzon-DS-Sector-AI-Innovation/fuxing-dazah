"""Document text extraction for uploaded attachments.

Supports PDF, DOCX, XLSX, TXT, and Markdown files.
"""

from pathlib import Path


class DocumentParser:
    """Extract text content from common document formats."""

    SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xls", ".txt", ".md"}

    @staticmethod
    def extract_text(file_path: str, max_chars: int = 50000) -> str:
        """Extract text from a file. Truncates to *max_chars* to avoid token limits."""
        ext = Path(file_path).suffix.lower()
        if ext == ".pdf":
            text = DocumentParser._extract_pdf(file_path)
        elif ext == ".docx":
            text = DocumentParser._extract_docx(file_path)
        elif ext in (".xlsx", ".xls"):
            text = DocumentParser._extract_xlsx(file_path)
        elif ext in (".txt", ".md"):
            text = DocumentParser._extract_txt(file_path)
        else:
            raise ValueError(f"Unsupported file format: {ext}")

        if len(text) > max_chars:
            text = text[:max_chars] + "\n\n...（文档过长，已截断）"
        return text.strip()

    @staticmethod
    def _extract_pdf(path: str) -> str:
        from pypdf import PdfReader

        reader = PdfReader(path)
        parts: list[str] = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                parts.append(t)
        return "\n".join(parts)

    @staticmethod
    def _extract_docx(path: str) -> str:
        from docx import Document

        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text)

    @staticmethod
    def _extract_xlsx(path: str) -> str:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## Sheet: {sheet_name}")
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                row_str = "\t".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                if row_str.strip():
                    rows.append(row_str)
            parts.append("\n".join(rows))
        wb.close()
        return "\n\n".join(parts)

    @staticmethod
    def _extract_txt(path: str) -> str:
        with open(path, encoding="utf-8", errors="replace") as f:
            return f.read()
