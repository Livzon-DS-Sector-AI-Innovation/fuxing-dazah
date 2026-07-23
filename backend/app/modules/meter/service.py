"""Meter business workflows live here."""

from __future__ import annotations

import logging
import uuid
from collections.abc import AsyncGenerator
from datetime import date, time
from typing import Any
from uuid import UUID

from fastapi import UploadFile
from sqlalchemy import insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import DuplicateException, NotFoundException
from app.core.storage import get_object, is_enabled, upload_object
from app.modules.meter import ai_service as ai_svc
from app.modules.meter import repository as repo
from app.modules.meter.models import (
    CalibrationReport,
    Department,
    GasDetectorRecord,
    InstrumentRecord,
)
from app.modules.meter.schemas import (
    DepartmentCreate,
    DepartmentUpdate,
    GasDetectorCreate,
    GasDetectorFilter,
    GasDetectorUpdate,
    InstrumentCreate,
    InstrumentFilter,
    InstrumentUpdate,
    _normalize_department,
)

logger = logging.getLogger(__name__)

MODULE_CODE = "meter"


def compute_status(record_status: str | None, next_calibration_date: date | None) -> str | None:
    """计算显示状态。

    - 在用/超期 + 下次检定已过期 → 显示"超期"
    - 手动设为超期 + 下次检定未过期（或无日期） → 自动恢复为"在用"
    - 停用永不变为超期，原样返回
    """
    if record_status in ("在用", "超期") and next_calibration_date is not None and next_calibration_date < date.today():
        return "超期"
    if record_status == "超期" and (next_calibration_date is None or next_calibration_date >= date.today()):
        return "在用"
    return record_status


def _auto_calc_next_calibration_date(item: dict[str, Any]) -> None:
    """自动计算下次检定日期 = 检定日期 + 检定周期(月) - 1 天。

    仅在「已提供检定日期和检定周期、但未提供下次检定日期」时生效，
    不会覆盖用户已手动填写的下次检定日期。
    """
    from datetime import timedelta

    from dateutil.relativedelta import relativedelta  # type: ignore[import-untyped]

    cal_date = item.get("calibration_date")
    cycle = item.get("calibration_cycle_months")
    next_date = item.get("next_calibration_date")

    if cal_date is not None and cycle is not None and next_date is None:
        if isinstance(cal_date, date):
            item["next_calibration_date"] = cal_date + relativedelta(months=cycle) - timedelta(days=1)


# ═══════════════════════════════════════════
# 标准计量器具
# ═══════════════════════════════════════════


async def create_instrument(
    db: AsyncSession, data: InstrumentCreate
) -> InstrumentRecord:
    if await repo.exists_instrument_by_asset_number(db, data.asset_number):
        raise DuplicateException("资产编号", data.asset_number)
    values = data.model_dump()
    _auto_calc_next_calibration_date(values)
    max_order = await repo.get_max_instrument_sort_order(db)
    values["sort_order"] = max_order + 1
    return await repo.create_instrument(db, values)


async def batch_create_instruments(
    db: AsyncSession, items: list[dict[str, Any]]
) -> dict[str, Any]:
    """批量新增标准计量器具。"""
    results: list[dict[str, Any]] = []
    created = skipped = 0

    max_order = await repo.get_max_instrument_sort_order(db)

    for i, item in enumerate(items):
        asset_number = item.get("asset_number")
        instrument_name = item.get("instrument_name", "").strip()

        if not instrument_name:
            skipped += 1
            results.append({
                "index": i, "asset_number": asset_number,
                "status": "skipped", "id": None, "message": "器具名称为空",
            })
            continue

        if asset_number:
            try:
                exists = await repo.exists_instrument_by_asset_number(db, asset_number)
            except Exception:
                skipped += 1
                results.append({
                    "index": i, "asset_number": asset_number,
                    "status": "skipped", "id": None, "message": "查询资产编号失败",
                })
                continue
            if exists:
                skipped += 1
                results.append({
                    "index": i, "asset_number": asset_number,
                    "status": "skipped", "id": None, "message": f"资产编号 {asset_number} 已存在",
                })
                continue

        try:
            _auto_calc_next_calibration_date(item)
            max_order += 1
            item["sort_order"] = max_order
            record = await repo.create_instrument(db, item)
            created += 1
            results.append({
                "index": i, "asset_number": asset_number,
                "status": "created", "id": str(record.id), "message": None,
            })
        except Exception as e:
            skipped += 1
            results.append({
                "index": i, "asset_number": asset_number,
                "status": "skipped", "id": None, "message": str(e),
            })

    await db.commit()
    return {"total": len(items), "created": created, "skipped": skipped, "results": results}


async def batch_create_gas_detectors(
    db: AsyncSession, items: list[dict[str, Any]]
) -> dict[str, Any]:
    """批量新增有毒有害可燃探测器。"""
    results: list[dict[str, Any]] = []
    created = skipped = 0

    max_order = await repo.get_max_gas_detector_sort_order(db)

    for i, item in enumerate(items):
        instrument_name = item.get("instrument_name", "").strip()

        if not instrument_name:
            skipped += 1
            results.append({
                "index": i, "asset_number": None,
                "status": "skipped", "id": None, "message": "器具名称为空",
            })
            continue

        try:
            max_order += 1
            item["sort_order"] = max_order
            record = await repo.create_gas_detector(db, item)
            created += 1
            results.append({
                "index": i, "asset_number": None,
                "status": "created", "id": str(record.id), "message": None,
            })
        except Exception as e:
            skipped += 1
            results.append({
                "index": i, "asset_number": None,
                "status": "skipped", "id": None, "message": str(e),
            })

    await db.commit()
    return {"total": len(items), "created": created, "skipped": skipped, "results": results}


async def get_instrument(
    db: AsyncSession, instrument_id: UUID
) -> InstrumentRecord:
    obj = await repo.get_instrument_by_id(db, instrument_id)
    if obj is None:
        raise NotFoundException("标准计量器具", str(instrument_id))
    return obj


async def list_instruments(
    db: AsyncSession, filters: InstrumentFilter
) -> tuple[list[InstrumentRecord], int]:
    return await repo.list_instruments(
        db,
        department=filters.department,
        asset_number=filters.asset_number,
        instrument_name=filters.instrument_name,
        model_spec=filters.model_spec,
        measurement_range=filters.measurement_range,
        accuracy_grade=filters.accuracy_grade,
        serial_number=filters.serial_number,
        location=filters.location,
        manufacturer=filters.manufacturer,
        status=filters.status,
        calibration_unit=filters.calibration_unit,
        calibration_result=filters.calibration_result,
        color_marking=filters.color_marking,
        next_calibration_before=filters.next_calibration_before,
        next_calibration_after=filters.next_calibration_after,
        calibration_date_before=filters.calibration_date_before,
        calibration_date_after=filters.calibration_date_after,
        keyword=filters.keyword,
        page=filters.page,
        page_size=filters.page_size,
    )


async def update_instrument(
    db: AsyncSession, instrument_id: UUID, data: InstrumentUpdate
) -> InstrumentRecord:
    obj = await repo.get_instrument_by_id(db, instrument_id, include_reports=False)
    if obj is None:
        raise NotFoundException("标准计量器具", str(instrument_id))

    updates = data.model_dump(exclude_unset=True)
    if not updates:
        # 无变更时也需 re-fetch reports，否则 API 层访问 .reports 会触发 MissingGreenlet
        obj = await repo.get_instrument_by_id(db, instrument_id, include_reports=True)
        if obj is None:
            raise NotFoundException("标准计量器具", str(instrument_id))
        return obj

    # 如果修改了 asset_number，检查唯一性
    if "asset_number" in updates and updates["asset_number"] != obj.asset_number:
        if await repo.exists_instrument_by_asset_number(
            db, updates["asset_number"], exclude_id=instrument_id
        ):
            raise DuplicateException("资产编号", updates["asset_number"])

    # 自动计算下次检定日期（如果只改了检定日期/周期而未提供 next）
    _auto_calc_next_calibration_date(updates)

    updated = await repo.update_instrument(db, instrument_id, updates)
    if updated is None:
        raise NotFoundException("标准计量器具", str(instrument_id))
    return updated


async def delete_instrument(db: AsyncSession, instrument_id: UUID) -> None:
    deleted = await repo.soft_delete_instrument(db, instrument_id)
    if not deleted:
        raise NotFoundException("标准计量器具", str(instrument_id))


async def batch_delete_instruments(db: AsyncSession, ids: list[UUID]) -> int:
    """批量软删除标准计量器具，返回实际删除数。"""
    return await repo.batch_soft_delete_instruments(db, ids)


async def get_all_instrument_ids(
    db: AsyncSession, filters: InstrumentFilter
) -> list[UUID]:
    """获取当前筛选条件下的所有记录 ID。"""
    return await repo.get_all_instrument_ids(
        db,
        department=filters.department,
        asset_number=filters.asset_number,
        instrument_name=filters.instrument_name,
        model_spec=filters.model_spec,
        measurement_range=filters.measurement_range,
        accuracy_grade=filters.accuracy_grade,
        serial_number=filters.serial_number,
        location=filters.location,
        manufacturer=filters.manufacturer,
        status=filters.status,
        calibration_unit=filters.calibration_unit,
        calibration_result=filters.calibration_result,
        color_marking=filters.color_marking,
        next_calibration_before=filters.next_calibration_before,
        next_calibration_after=filters.next_calibration_after,
        calibration_date_before=filters.calibration_date_before,
        calibration_date_after=filters.calibration_date_after,
        keyword=filters.keyword,
    )


async def get_instrument_departments(db: AsyncSession) -> list[str]:
    return await repo.get_instrument_departments(db)


# ═══════════════════════════════════════════
# 有毒有害可燃探测器
# ═══════════════════════════════════════════


async def create_gas_detector(
    db: AsyncSession, data: GasDetectorCreate
) -> GasDetectorRecord:
    if data.product_number and await repo.exists_gas_detector_by_product_number(
        db, data.product_number
    ):
        raise DuplicateException("产品编号", data.product_number)
    values = data.model_dump()
    max_order = await repo.get_max_gas_detector_sort_order(db)
    values["sort_order"] = max_order + 1
    return await repo.create_gas_detector(db, values)


async def get_gas_detector(
    db: AsyncSession, detector_id: UUID
) -> GasDetectorRecord:
    obj = await repo.get_gas_detector_by_id(db, detector_id)
    if obj is None:
        raise NotFoundException("有毒有害可燃探测器", str(detector_id))
    return obj


async def list_gas_detectors(
    db: AsyncSession, filters: GasDetectorFilter
) -> tuple[list[GasDetectorRecord], int]:
    return await repo.list_gas_detectors(
        db,
        department=filters.department,
        instrument_name=filters.instrument_name,
        detection_model=filters.detection_model,
        product_number=filters.product_number,
        measurement_range=filters.measurement_range,
        installation_type=filters.installation_type,
        installation_location=filters.installation_location,
        medium=filters.medium,
        detection_unit=filters.detection_unit,
        calibration_result=filters.calibration_result,
        calibration_factor=filters.calibration_factor,
        manufacturer_supplier=filters.manufacturer_supplier,
        manufacturer=filters.manufacturer,
        status=filters.status,
        next_calibration_before=filters.next_calibration_before,
        next_calibration_after=filters.next_calibration_after,
        calibration_date_before=filters.calibration_date_before,
        calibration_date_after=filters.calibration_date_after,
        keyword=filters.keyword,
        page=filters.page,
        page_size=filters.page_size,
    )


async def update_gas_detector(
    db: AsyncSession, detector_id: UUID, data: GasDetectorUpdate
) -> GasDetectorRecord:
    obj = await repo.get_gas_detector_by_id(db, detector_id, include_reports=False)
    if obj is None:
        raise NotFoundException("有毒有害可燃探测器", str(detector_id))

    updates = data.model_dump(exclude_unset=True)
    if not updates:
        # 无变更时也需 re-fetch reports，否则 API 层访问 .reports 会触发 MissingGreenlet
        obj = await repo.get_gas_detector_by_id(db, detector_id, include_reports=True)
        if obj is None:
            raise NotFoundException("有毒有害可燃探测器", str(detector_id))
        return obj

    if "product_number" in updates and updates["product_number"] != obj.product_number:
        if updates["product_number"] and await repo.exists_gas_detector_by_product_number(
            db, updates["product_number"], exclude_id=detector_id
        ):
            raise DuplicateException("产品编号", updates["product_number"])

    updated = await repo.update_gas_detector(db, detector_id, updates)
    if updated is None:
        raise NotFoundException("有毒有害可燃探测器", str(detector_id))
    return updated


async def delete_gas_detector(db: AsyncSession, detector_id: UUID) -> None:
    deleted = await repo.soft_delete_gas_detector(db, detector_id)
    if not deleted:
        raise NotFoundException("有毒有害可燃探测器", str(detector_id))


async def batch_delete_gas_detectors(db: AsyncSession, ids: list[UUID]) -> int:
    """批量软删除有毒有害可燃探测器，返回实际删除数。"""
    return await repo.batch_soft_delete_gas_detectors(db, ids)


async def get_all_gas_detector_ids(
    db: AsyncSession, filters: GasDetectorFilter
) -> list[UUID]:
    """获取当前筛选条件下的所有记录 ID。"""
    return await repo.get_all_gas_detector_ids(
        db,
        department=filters.department,
        instrument_name=filters.instrument_name,
        detection_model=filters.detection_model,
        product_number=filters.product_number,
        measurement_range=filters.measurement_range,
        installation_type=filters.installation_type,
        installation_location=filters.installation_location,
        medium=filters.medium,
        detection_unit=filters.detection_unit,
        calibration_result=filters.calibration_result,
        calibration_factor=filters.calibration_factor,
        manufacturer_supplier=filters.manufacturer_supplier,
        manufacturer=filters.manufacturer,
        status=filters.status,
        next_calibration_before=filters.next_calibration_before,
        next_calibration_after=filters.next_calibration_after,
        calibration_date_before=filters.calibration_date_before,
        calibration_date_after=filters.calibration_date_after,
        keyword=filters.keyword,
    )


async def get_gas_detector_departments(db: AsyncSession) -> list[str]:
    return await repo.get_gas_detector_departments(db)


async def get_instrument_filter_options(db: AsyncSession) -> dict[str, list[str]]:
    """获取标准计量器具所有筛选列的 distinct 值。"""
    options = await repo.get_instrument_filter_options(db)
    # "超期" 是动态计算的状态，不在数据库中存储，需要手动加入筛选选项
    if "超期" not in options.get("status", []):
        options.setdefault("status", []).insert(0, "超期")
    return options


async def get_gas_detector_filter_options(db: AsyncSession) -> dict[str, list[str]]:
    """获取有毒有害可燃探测器所有筛选列的 distinct 值。"""
    options = await repo.get_gas_detector_filter_options(db)
    # "超期" 是动态计算的状态，不在数据库中存储，需要手动加入筛选选项
    if "超期" not in options.get("status", []):
        options.setdefault("status", []).insert(0, "超期")
    return options


# ═══════════════════════════════════════════
# 检测报告
# ═══════════════════════════════════════════


def _build_report_path(record_id: UUID, filename: str) -> str:
    """构建 MinIO 对象路径：reports/{record_id}/{uuid}.{ext}"""
    ext = filename.rsplit(".", 1)[-1] if "." in filename else "bin"
    return f"reports/{record_id}/{uuid.uuid4().hex}.{ext}"


async def upload_report(
    db: AsyncSession,
    *,
    file: UploadFile,
    instrument_id: UUID | None = None,
    gas_detector_id: UUID | None = None,
    report_date: date | None = None,
    remark: str | None = None,
) -> CalibrationReport:
    """上传检测报告文件到 MinIO 并创建元数据记录。"""
    if not is_enabled():
        raise RuntimeError("MinIO 未启用，无法上传文件")

    # 校验：必须且只能关联一种仪表
    if (instrument_id is None) == (gas_detector_id is None):
        raise ValueError("必须且只能指定 instrument_id 或 gas_detector_id 中的一个")

    # 校验目标仪表存在
    if instrument_id:
        target_inst = await repo.get_instrument_by_id(db, instrument_id, include_reports=False)
        if target_inst is None:
            raise NotFoundException("标准计量器具", str(instrument_id))
    else:
        assert gas_detector_id is not None
        target_det = await repo.get_gas_detector_by_id(db, gas_detector_id, include_reports=False)
        if target_det is None:
            raise NotFoundException("有毒有害可燃探测器", str(gas_detector_id))

    # 读取文件内容
    file_data = await file.read()
    file_size = len(file_data)
    content_type = file.content_type or "application/octet-stream"

    # 上传到 MinIO
    object_path = _build_report_path(instrument_id or gas_detector_id, file.filename)  # type: ignore[arg-type]
    upload_object(MODULE_CODE, object_path, file_data, file_size, content_type)

    # 创建元数据记录
    report = await repo.create_report(
        db,
        {
            "instrument_id": instrument_id,
            "gas_detector_id": gas_detector_id,
            "file_name": file.filename,
            "file_path": object_path,
            "file_size": file_size,
            "content_type": content_type,
            "report_date": report_date,
            "remark": remark,
        },
    )

    return report


async def get_report(db: AsyncSession, report_id: UUID) -> CalibrationReport:
    report = await repo.get_report_by_id(db, report_id)
    if report is None:
        raise NotFoundException("检测报告", str(report_id))
    return report


async def download_report_data(report: CalibrationReport) -> tuple[bytes, str] | None:
    """从 MinIO 下载报告文件的实际内容。"""
    if not is_enabled():
        return None
    return get_object(MODULE_CODE, report.file_path)


async def delete_report(db: AsyncSession, report_id: UUID) -> None:
    deleted = await repo.soft_delete_report(db, report_id)
    if not deleted:
        raise NotFoundException("检测报告", str(report_id))


async def list_instrument_reports(db: AsyncSession, instrument_id: UUID) -> list[CalibrationReport]:
    """获取某个标准计量器具的所有报告。"""
    return await repo.list_reports_by_instrument(db, instrument_id)


async def list_gas_detector_reports(db: AsyncSession, gas_detector_id: UUID) -> list[CalibrationReport]:
    """获取某个探测器的所有报告。"""
    return await repo.list_reports_by_gas_detector(db, gas_detector_id)


# ═══════════════════════════════════════════
# 文件匹配
# ═══════════════════════════════════════════


def _parse_filename(filename: str) -> tuple[str, str]:
    """从文件名中解析器具名称和编号。右边最后一个 _ 为分界。"""
    stem = filename.rsplit(".", 1)[0]  # 去掉扩展名
    if "_" not in stem:
        return stem, ""
    idx = stem.rfind("_")
    return stem[:idx], stem[idx + 1:]


async def match_filenames(
    db: AsyncSession, filenames: list[str]
) -> list[dict[str, Any]]:
    """批量匹配文件名到仪表记录。"""
    results: list[dict[str, Any]] = []
    for fn in filenames:
        name, code = _parse_filename(fn)
        matched_type = None
        matched_id = None
        matched_name = None
        matched_dept = None

        if code:
            # 先匹配标准计量器具
            inst = await repo.find_instrument_by_name_and_serial(db, name, code)
            if inst:
                matched_type = "instrument"
                matched_id = str(inst.id)
                matched_name = f"{inst.instrument_name} [{inst.asset_number}]"
                matched_dept = inst.department
            else:
                # 再匹配探测器
                det = await repo.find_gas_detector_by_name_and_product(db, name, code)
                if det:
                    matched_type = "gas_detector"
                    matched_id = str(det.id)
                    matched_name = f"{det.instrument_name} [{det.product_number}]"
                    matched_dept = det.department

        results.append({
            "filename": fn,
            "matched_type": matched_type,
            "matched_id": matched_id,
            "matched_name": matched_name,
            "matched_department": matched_dept,
        })
    return results


# ═══════════════════════════════════════════
# 批量上传
# ═══════════════════════════════════════════


async def batch_upload_reports(
    db: AsyncSession,
    files: list[tuple[str, bytes, str]],  # (filename, data, content_type)
    items: list[dict[str, Any]],
    report_date: date | None = None,
    remark: str | None = None,
) -> dict[str, Any]:
    """批量上传报告文件，按 items 中的 instrument_id / gas_detector_id 关联。"""
    if not is_enabled():
        raise RuntimeError("MinIO 未启用，无法上传文件")

    success = 0
    failed = 0
    errors: list[str] = []
    report_ids: list[str] = []
    file_map = {f[0]: f for f in files}

    for item in items:
        fn = item["filename"]
        instrument_id = item.get("instrument_id")
        gas_detector_id = item.get("gas_detector_id")

        if fn not in file_map:
            failed += 1
            errors.append(f"{fn}: 文件未找到")
            continue
        if not instrument_id and not gas_detector_id:
            failed += 1
            errors.append(f"{fn}: 未关联仪表")
            continue

        file_data, content_type = file_map[fn][1], file_map[fn][2]

        try:
            if instrument_id:
                target = await repo.get_instrument_by_id(db, UUID(instrument_id), include_reports=False)
                if not target:
                    raise NotFoundException("标准计量器具", instrument_id)
                target_id = UUID(instrument_id)
            else:
                target = await repo.get_gas_detector_by_id(db, UUID(gas_detector_id), include_reports=False)  # type: ignore[assignment]
                if not target:
                    raise NotFoundException("有毒有害可燃探测器", str(gas_detector_id))
                target_id = UUID(gas_detector_id)

            object_path = _build_report_path(target_id, fn)
            upload_object(MODULE_CODE, object_path, file_data, len(file_data), content_type)

            report = await repo.create_report(
                db,
                {
                    "instrument_id": UUID(instrument_id) if instrument_id else None,
                    "gas_detector_id": UUID(gas_detector_id) if gas_detector_id else None,
                    "file_name": fn,
                    "file_path": object_path,
                    "file_size": len(file_data),
                    "content_type": content_type,
                    "report_date": report_date,
                    "remark": remark,
                },
            )
            report_ids.append(str(report.id))
            success += 1
        except Exception as e:
            failed += 1
            errors.append(f"{fn}: {str(e)}")

    return {"success": success, "failed": failed, "errors": errors, "report_ids": report_ids}


# ═══════════════════════════════════════════
# 批量导出报告
# ═══════════════════════════════════════════


async def export_instrument_reports(
    db: AsyncSession, ids: list[UUID]
) -> tuple[bytes, str, int]:
    """导出指定仪表的最新报告为 ZIP。返回 (zip_bytes, filename, count)。"""
    import io as io_mod
    import zipfile

    if not is_enabled():
        raise RuntimeError("MinIO 未启用，无法导出报告")

    zip_buf = io_mod.BytesIO()
    count = 0
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for inst_id in ids:
            reports = await repo.list_reports_by_instrument(db, inst_id)
            if not reports:
                continue
            latest = reports[0]  # 按 report_date desc 排列，第一条是最新的
            data = get_object(MODULE_CODE, latest.file_path)
            if data is None:
                continue
            file_data, _ = data
            # 获取仪表名称和资产编号
            inst = await repo.get_instrument_by_id(db, inst_id, include_reports=False)
            if inst is None:
                continue
            safe_name = f"{inst.instrument_name}_{inst.asset_number or inst.id}"
            safe_name = safe_name.replace("/", "_").replace("\\", "_")
            ext = latest.file_name.rsplit(".", 1)[-1] if "." in latest.file_name else "pdf"
            zf.writestr(f"{safe_name}.{ext}", file_data)
            count += 1

    zip_buf.seek(0)
    return zip_buf.getvalue(), "instruments_reports.zip", count


async def export_gas_detector_reports(
    db: AsyncSession, ids: list[UUID]
) -> tuple[bytes, str, int]:
    """导出指定探测器的最新报告为 ZIP。"""
    import io as io_mod
    import zipfile

    if not is_enabled():
        raise RuntimeError("MinIO 未启用，无法导出报告")

    zip_buf = io_mod.BytesIO()
    count = 0
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for det_id in ids:
            reports = await repo.list_reports_by_gas_detector(db, det_id)
            if not reports:
                continue
            latest = reports[0]
            data = get_object(MODULE_CODE, latest.file_path)
            if data is None:
                continue
            file_data, _ = data
            det = await repo.get_gas_detector_by_id(db, det_id, include_reports=False)
            if det is None:
                continue
            safe_name = f"{det.instrument_name}_{det.product_number or det.id}"
            safe_name = safe_name.replace("/", "_").replace("\\", "_")
            ext = latest.file_name.rsplit(".", 1)[-1] if "." in latest.file_name else "pdf"
            zf.writestr(f"{safe_name}.{ext}", file_data)
            count += 1

    zip_buf.seek(0)
    return zip_buf.getvalue(), "gas_detectors_reports.zip", count


# ═══════════════════════════════════════════
# 检定到期提醒
# ═══════════════════════════════════════════


async def get_calibration_alerts(
    db: AsyncSession, *, days_before: int = 30, department: str | None = None, source: str | None = None
) -> list[dict[str, Any]]:
    """查询仪表中未来 N 天内到期的记录，合并返回。source 可选 instrument/gas_detector。"""
    instruments: list[InstrumentRecord] = []
    detectors: list[GasDetectorRecord] = []

    if source is None or source == "instrument":
        instruments = await repo.list_instruments_due_for_calibration(db, days_before=days_before)
    if source is None or source == "gas_detector":
        detectors = await repo.list_gas_detectors_due_for_calibration(db, days_before=days_before)

    today = date.today()
    alerts: list[dict[str, Any]] = []

    for obj in instruments:
        days = (obj.next_calibration_date - today).days if obj.next_calibration_date else None
        dept = _normalize_department(obj.department)
        if department and (not dept or department not in dept):
            continue
        alerts.append({
            "source": "instrument",
            "id": str(obj.id),
            "serial_number": obj.serial_number,
            "instrument_name": obj.instrument_name,
            "location": obj.location,
            "department": dept,
            "next_calibration_date": obj.next_calibration_date,
            "days_until_due": days,
        })

    for det in detectors:
        days = (det.next_calibration_date - today).days if det.next_calibration_date else None
        dept = _normalize_department(det.department)
        if department and (not dept or department not in dept):
            continue
        alerts.append({
            "source": "gas_detector",
            "id": str(det.id),
            "serial_number": det.product_number,
            "instrument_name": det.instrument_name,
            "location": det.installation_location,
            "department": dept,
            "next_calibration_date": det.next_calibration_date,
            "days_until_due": days,
        })

    return alerts


# ═══════════════════════════════════════════
# 批量 AI 日期提取
# ═══════════════════════════════════════════


async def batch_extract_dates(
    db: AsyncSession, report_ids: list[UUID]
) -> dict[str, Any]:
    """批量提取报告中的校准日期并更新对应仪表。"""

    results: list[dict[str, Any]] = []
    success = 0
    failed = 0

    # 从环境变量读取 AI 配置
    config = ai_svc.get_meter_ai_config()
    if not config:
        return {
            "total": len(report_ids),
            "success": 0,
            "failed": len(report_ids),
            "results": [
                {"report_id": str(rid), "file_name": "", "success": False,
                 "error": "请先配置环境变量 METER_AI_BASE_URL / METER_AI_API_KEY / METER_AI_MODEL"}
                for rid in report_ids
            ],
        }

    for rid in report_ids:
        report = await repo.get_report_by_id(db, rid)
        if report is None:
            failed += 1
            results.append({
                "report_id": str(rid), "file_name": "", "success": False,
                "error": "报告不存在",
            })
            continue

        # 下载报告文件
        data = await download_report_data(report)
        if data is None:
            failed += 1
            results.append({
                "report_id": str(rid), "file_name": report.file_name, "success": False,
                "error": "文件不存在或 MinIO 未启用",
            })
            continue

        pdf_data, _ = data

        # 获取关联仪表的检定周期
        calibration_cycle = None
        if report.instrument_id:
            inst = await repo.get_instrument_by_id(db, report.instrument_id, include_reports=False)
            if inst:
                calibration_cycle = inst.calibration_cycle_months
        elif report.gas_detector_id:
            det = await repo.get_gas_detector_by_id(db, report.gas_detector_id, include_reports=False)
            if det:
                calibration_cycle = getattr(det, 'calibration_cycle_months', None)

        # AI 识别
        try:
            ai_result = await ai_svc.extract_and_update_date(
                pdf_data, config["api_url"], config["api_key"], config["model"], calibration_cycle
            )
        except Exception as e:
            logger.exception(f"batch_extract_dates AI error for report {rid}")
            failed += 1
            results.append({
                "report_id": str(rid), "file_name": report.file_name, "success": False,
                "error": f"AI 提取失败: {e}",
            })
            continue

        if not ai_result["success"]:
            failed += 1
            results.append({
                "report_id": str(rid), "file_name": report.file_name, "success": False,
                "error": ai_result.get("error", "AI 未识别到日期"),
            })
            continue

        # 更新仪表日期
        try:
            updates: dict[str, Any] = {
                "calibration_date": date.fromisoformat(ai_result["calibration_date"])
            }
            if ai_result.get("next_calibration_date"):
                updates["next_calibration_date"] = date.fromisoformat(ai_result["next_calibration_date"])
            if report.instrument_id:
                await repo.update_instrument(db, report.instrument_id, updates)
            elif report.gas_detector_id:
                await repo.update_gas_detector(db, report.gas_detector_id, updates)
        except Exception as e:
            logger.exception(f"batch_extract_dates update error for report {rid}")
            failed += 1
            results.append({
                "report_id": str(rid), "file_name": report.file_name, "success": False,
                "error": f"更新仪表失败: {e}",
            })
            continue

        success += 1
        results.append({
            "report_id": str(rid),
            "file_name": report.file_name,
            "success": True,
            "calibration_date": ai_result["calibration_date"],
            "next_calibration_date": ai_result.get("next_calibration_date"),
        })

    await db.commit()
    return {
        "total": len(report_ids),
        "success": success,
        "failed": failed,
        "results": results,
    }


async def batch_extract_dates_stream(
    db: AsyncSession, report_ids: list[UUID],
) -> AsyncGenerator[str, None]:
    """逐份 AI 识别报告日期并推送 SSE 进度事件。

    与 batch_extract_dates 不同，此函数每处理完一份报告就 commit 一次，
    确保客户端断开后已处理的数据不丢失。
    """
    yield _sse_event("start", {"total": len(report_ids)})

    # 从环境变量读取 AI 配置
    config = ai_svc.get_meter_ai_config()
    if not config:
        yield _sse_event("error", {
            "message": "请先配置环境变量 METER_AI_BASE_URL / METER_AI_API_KEY / METER_AI_MODEL",
        })
        yield _sse_event("complete", {"total": len(report_ids), "success": 0, "failed": len(report_ids), "interrupted": False})
        return

    success = 0
    failed = 0
    interrupted = False

    for idx, rid in enumerate(report_ids):
        current = idx + 1
        total = len(report_ids)

        # 获取报告
        report = await repo.get_report_by_id(db, rid)
        if report is None:
            failed += 1
            yield _sse_event("progress", {"current": current, "total": total, "report_id": str(rid), "file_name": ""})
            yield _sse_event("result", {
                "report_id": str(rid), "file_name": "", "status": "failed",
                "error": "报告不存在",
            })
            continue

        # 推送进度事件
        yield _sse_event("progress", {
            "current": current, "total": total,
            "report_id": str(rid), "file_name": report.file_name,
        })

        # 下载文件
        data = await download_report_data(report)
        if data is None:
            failed += 1
            yield _sse_event("result", {
                "report_id": str(rid), "file_name": report.file_name, "status": "failed",
                "error": "文件不存在或 MinIO 未启用",
            })
            continue

        pdf_data, _ = data

        # 获取关联仪表的检定周期
        calibration_cycle = None
        if report.instrument_id:
            inst = await repo.get_instrument_by_id(db, report.instrument_id, include_reports=False)
            if inst:
                calibration_cycle = inst.calibration_cycle_months
        elif report.gas_detector_id:
            det = await repo.get_gas_detector_by_id(db, report.gas_detector_id, include_reports=False)
            if det:
                calibration_cycle = getattr(det, 'calibration_cycle_months', None)

        # AI 识别
        try:
            ai_result = await ai_svc.extract_and_update_date(
                pdf_data, config["api_url"], config["api_key"], config["model"], calibration_cycle,
            )
        except Exception as e:
            logger.exception(f"batch_extract_dates_stream AI error for report {rid}")
            failed += 1
            yield _sse_event("result", {
                "report_id": str(rid), "file_name": report.file_name, "status": "failed",
                "error": f"AI 提取失败: {e}",
            })
            continue

        if not ai_result["success"]:
            failed += 1
            yield _sse_event("result", {
                "report_id": str(rid), "file_name": report.file_name, "status": "failed",
                "error": ai_result.get("error", "AI 未识别到日期"),
            })
            continue

        # 更新仪表日期
        try:
            updates: dict[str, Any] = {
                "calibration_date": date.fromisoformat(ai_result["calibration_date"])
            }
            if ai_result.get("next_calibration_date"):
                updates["next_calibration_date"] = date.fromisoformat(ai_result["next_calibration_date"])
            if report.instrument_id:
                await repo.update_instrument(db, report.instrument_id, updates)
            elif report.gas_detector_id:
                await repo.update_gas_detector(db, report.gas_detector_id, updates)
        except Exception as e:
            logger.exception(f"batch_extract_dates_stream update error for report {rid}")
            failed += 1
            yield _sse_event("result", {
                "report_id": str(rid), "file_name": report.file_name, "status": "failed",
                "error": f"更新仪表失败: {e}",
            })
            continue

        # 逐份 commit，保证中断后已处理的数据不丢失
        await db.commit()

        success += 1
        yield _sse_event("result", {
            "report_id": str(rid), "file_name": report.file_name, "status": "success",
            "calibration_date": ai_result["calibration_date"],
            "next_calibration_date": ai_result.get("next_calibration_date"),
        })

    yield _sse_event("complete", {
        "total": len(report_ids), "success": success, "failed": failed,
        "interrupted": interrupted,
    })


def _sse_event(event: str, data: dict[str, Any]) -> str:
    """构建 SSE 事件字符串。"""
    import json
    payload = json.dumps(data, ensure_ascii=False)
    return f"event: {event}\ndata: {payload}\n\n"


async def get_meter_overview(db: AsyncSession, source: str) -> dict[str, int]:
    """获取仪表总览统计数据。"""
    if source == "instrument":
        return await repo.get_instrument_overview(db)
    elif source == "gas_detector":
        return await repo.get_gas_detector_overview(db)
    else:
        raise ValueError(f"不支持的数据源: {source}")


# ═══════════════════════════════════════════
# 日期聚合统计
# ═══════════════════════════════════════════


def _build_date_stats_tree(rows: list[dict[str, int]]) -> list[dict[str, Any]]:
    """将 repo 返回的扁平行 [{year, month, day, count}] 组装为嵌套 year→month→day 结构。"""
    years_map: dict[int, dict[str, Any]] = {}
    for row in rows:
        y = row["year"]
        m = row["month"]
        d = row["day"]
        c = row["count"]

        if y not in years_map:
            years_map[y] = {"year": y, "count": 0, "months": {}}
        years_map[y]["count"] += c

        months_map: dict[int, dict[str, Any]] = years_map[y]["months"]
        if m not in months_map:
            months_map[m] = {"month": m, "count": 0, "days": {}}
        months_map[m]["count"] += c

        days_map: dict[int, dict[str, Any]] = months_map[m]["days"]
        days_map[d] = {"day": d, "count": c}

    # 转为列表并按年份降序、月降序、日降序排列
    result: list[dict[str, Any]] = []
    for y in sorted(years_map.keys(), reverse=True):
        y_data = years_map[y]
        months_list: list[dict[str, Any]] = []
        for m in sorted(y_data["months"].keys(), reverse=True):
            m_data = y_data["months"][m]
            days_list = [
                {"day": d, "count": m_data["days"][d]["count"]}
                for d in sorted(m_data["days"].keys(), reverse=True)
            ]
            months_list.append({"month": m, "count": m_data["count"], "days": days_list})
        result.append({"year": y, "count": y_data["count"], "months": months_list})
    return result


async def get_instrument_date_stats(
    db: AsyncSession, filters: InstrumentFilter, field: str
) -> dict[str, Any]:
    """获取标准计量器具的日期聚合统计。"""
    rows = await repo.get_instrument_date_stats(
        db,
        field=field,
        department=filters.department,
        asset_number=filters.asset_number,
        instrument_name=filters.instrument_name,
        model_spec=filters.model_spec,
        measurement_range=filters.measurement_range,
        accuracy_grade=filters.accuracy_grade,
        serial_number=filters.serial_number,
        location=filters.location,
        manufacturer=filters.manufacturer,
        status=filters.status,
        calibration_unit=filters.calibration_unit,
        calibration_result=filters.calibration_result,
        color_marking=filters.color_marking,
        keyword=filters.keyword,
    )
    return {"field": field, "years": _build_date_stats_tree(rows)}


async def get_gas_detector_date_stats(
    db: AsyncSession, filters: GasDetectorFilter, field: str
) -> dict[str, Any]:
    """获取有毒有害可燃探测器的日期聚合统计。"""
    rows = await repo.get_gas_detector_date_stats(
        db,
        field=field,
        department=filters.department,
        instrument_name=filters.instrument_name,
        detection_model=filters.detection_model,
        product_number=filters.product_number,
        measurement_range=filters.measurement_range,
        installation_type=filters.installation_type,
        installation_location=filters.installation_location,
        medium=filters.medium,
        detection_unit=filters.detection_unit,
        calibration_result=filters.calibration_result,
        calibration_factor=filters.calibration_factor,
        manufacturer_supplier=filters.manufacturer_supplier,
        manufacturer=filters.manufacturer,
        status=filters.status,
        keyword=filters.keyword,
    )
    return {"field": field, "years": _build_date_stats_tree(rows)}


# ═══════════════════════════════════════════
# 部门管理
# ═══════════════════════════════════════════


async def get_personnel_candidates(db: AsyncSession) -> list[dict[str, Any]]:
    """从平台 identity.users 查询所有用户，作为负责人候选人列表。"""
    from sqlalchemy import select as sa_select

    from app.platform.identity.models import User

    stmt = sa_select(User).where(
        User.is_deleted == False,  # noqa: E712
    ).order_by(User.name)
    result = await db.execute(stmt)
    users = result.scalars().all()

    return [
        {
            "name": u.name,
            "feishu_open_id": u.feishu_open_id or "",
            "department": u.department,
        }
        for u in users
    ]


async def create_department(
    db: AsyncSession, data: DepartmentCreate
) -> Department:
    name = data.name.strip()
    if not name:
        raise ValueError("部门名称不能为空")
    existing = await repo.get_department_by_source_and_name(db, data.source, name)
    if existing:
        raise DuplicateException("部门名称", name)
    return await repo.create_department(db, {
        "source": data.source,
        "name": name,
        "heads": data.heads or [],
    })


async def list_departments(
    db: AsyncSession, source: str | None = None
) -> list[dict[str, Any]]:
    depts = await repo.list_departments(db, source=source)
    results: list[dict[str, Any]] = []
    for d in depts:
        counts = await repo.count_records_by_department(db, d.name)
        record_count = counts["instrument_count"] if d.source == "instrument" else counts["gas_detector_count"]
        results.append({
            "id": str(d.id),
            "source": d.source,
            "name": d.name,
            "heads": d.heads or [],
            "auto_notify_enabled": d.auto_notify_enabled,
            "record_count": record_count,
            "created_at": d.created_at,
            "updated_at": d.updated_at,
        })
    return results


async def update_department(
    db: AsyncSession, dept_id: UUID, data: DepartmentUpdate
) -> Department:
    dept = await repo.get_department_by_id(db, dept_id)
    if dept is None:
        raise NotFoundException("部门", str(dept_id))

    new_name = data.name.strip()
    if not new_name:
        raise ValueError("部门名称不能为空")
    if new_name != dept.name:
        # 检查新名称是否与其他记录冲突
        conflict = await repo.get_department_by_source_and_name(
            db, dept.source, new_name, exclude_id=dept_id
        )
        if conflict:
            raise DuplicateException("部门名称", new_name)

        # 联动更新对应表中所有匹配记录
        await repo.rename_department_in_records(db, dept.name, new_name, dept.source)

    # 构建更新字段
    update_fields: dict[str, Any] = {"name": new_name}
    if data.heads is not None:
        update_fields["heads"] = data.heads
    if data.auto_notify_enabled is not None:
        update_fields["auto_notify_enabled"] = data.auto_notify_enabled

    updated = await repo.update_department(db, dept_id, update_fields)
    if updated is None:
        raise NotFoundException("部门", str(dept_id))
    return updated


async def delete_department(db: AsyncSession, dept_id: UUID) -> None:
    dept = await repo.get_department_by_id(db, dept_id)
    if dept is None:
        raise NotFoundException("部门", str(dept_id))

    # 检查是否还有本来源的记录使用该部门名
    counts = await repo.count_records_by_department(db, dept.name)
    total = counts["instrument_count"] if dept.source == "instrument" else counts["gas_detector_count"]
    if total > 0:
        raise DuplicateException(
            "部门", f"{dept.name}（仍有 {total} 条记录使用，无法删除）"
        )

    deleted = await repo.soft_delete_department(db, dept_id)
    if not deleted:
        raise NotFoundException("部门", str(dept_id))


# ═══════════════════════════════════════════
# 检定到期飞书通知
# ═══════════════════════════════════════════

_CALIBRATION_RANGE_LABELS: dict[str, str] = {
    "due_today": "⚠️ 今天到期",
    "due_7d": "📅 未来 7 天到期",
    "due_30d": "📅 未来 30 天到期",
    "due_90d": "📅 未来 90 天到期",
}


def _build_reminder_card(
    department_name: str,
    groups: dict[str, list[dict[str, Any]]],
) -> str:
    """构建飞书卡片 markdown 正文。空窗口自动隐藏。"""
    lines = [
        f"**📋 仪表到期提醒 — {department_name}**",
        "",
    ]

    for key in ("due_today", "due_7d", "due_30d", "due_90d"):
        items = groups.get(key, [])
        if not items:
            continue  # 空窗口不显示
        label = _CALIBRATION_RANGE_LABELS[key]
        lines.append(f"**{label}**：{len(items)} 台")
        lines.append("")
        for item in items:
            lines.append(
                f"器具名称：{item.get('instrument_name', '-')}\n"
                f"器具编号：{item.get('serial_number') or '-'}\n"
                f"位置信息：{item.get('location') or '-'}"
            )
            lines.append("")
        lines.append("")

    return "\n".join(lines)


async def send_calibration_reminders(db: AsyncSession) -> dict[str, Any]:
    """扫描所有开启提醒的部门，发送检定到期飞书通知。

    在每个部门的通知中，按 4 个时间节点分组展示标准器具和探测器的到期记录。
    通知发送失败不影响其他部门的处理。
    """
    from app.core.config import get_settings
    from app.platform.integrations.feishu.notification import send_user_card

    settings = get_settings()
    if not settings.METER_CALIBRATION_AUTO_NOTIFY_ENABLED:
        logger.info("全局自动提醒开关已关闭，跳过发送")
        return {"sent": 0, "skipped": 0, "errors": 0}

    depts = await repo.get_notifiable_departments(db)
    if not depts:
        logger.info("没有需要发送提醒的部门")
        return {"sent": 0, "skipped": 0, "errors": 0}

    # 预加载 identity.users 的 name → feishu_user_id 映射（user_id 跨应用有效）
    from sqlalchemy import select as sa_select

    from app.platform.identity.models import User

    name_to_user_id: dict[str, str] = {}
    users_result = await db.execute(
        sa_select(User.name, User.feishu_user_id).where(
            User.is_deleted == False,  # noqa: E712
            User.feishu_user_id.isnot(None),
            User.feishu_user_id != "",
        )
    )
    for row in users_result.all():
        name_to_user_id[row[0]] = row[1]

    today = date.today()
    sent = 0
    skipped = 0
    errors = 0

    for dept in depts:
        heads_list: list[dict[str, str]] = dept.heads or []  # type: ignore[assignment]
        if not heads_list:
            skipped += 1
            continue

        # 查询标准器具 + 探测器到期记录
        inst_groups = await repo.list_instruments_due_grouped(db, dept.name)
        det_groups = await repo.list_gas_detectors_due_grouped(db, dept.name)

        # 合并两个数据源，按 4 节点分组
        merged: dict[str, list[dict[str, Any]]] = {
            "due_today": [],
            "due_7d": [],
            "due_30d": [],
            "due_90d": [],
        }

        for key in merged:
            for inst in inst_groups.get(key, []):
                # 跳过检定单位为"计量室"或已停用的器具
                if inst.calibration_unit == "计量室":
                    continue
                if inst.status == "停用":
                    continue
                days = (inst.next_calibration_date - today).days if inst.next_calibration_date else None
                merged[key].append({
                    "source": "instrument",
                    "serial_number": inst.serial_number,
                    "instrument_name": inst.instrument_name,
                    "location": inst.location,
                    "next_calibration_date_str": inst.next_calibration_date.isoformat() if inst.next_calibration_date else None,
                    "days_until_due": days,
                })
            for det in det_groups.get(key, []):
                # 跳过检测单位为"计量室"或已停用的探测器
                if det.detection_unit == "计量室":
                    continue
                if det.status == "停用":
                    continue
                days = (det.next_calibration_date - today).days if det.next_calibration_date else None
                merged[key].append({
                    "source": "gas_detector",
                    "serial_number": det.product_number,
                    "instrument_name": det.instrument_name,
                    "location": det.installation_location,
                    "next_calibration_date_str": det.next_calibration_date.isoformat() if det.next_calibration_date else None,
                    "days_until_due": days,
                })

        total_items = sum(len(v) for v in merged.values())
        if total_items == 0:
            skipped += 1
            continue

        # 构建并发送卡片（给每个负责人各发一份）
        content = _build_reminder_card(dept.name, merged)
        title = f"📋 仪表到期提醒 - {dept.name}"

        all_ok = True
        for head in heads_list:
            head_name = head.get("name", "未知")
            # 优先使用 user_id（应用无关），其次回退到 open_id
            feishu_id = name_to_user_id.get(head_name, "")
            receive_id_type: str = "user_id"
            if not feishu_id:
                feishu_id = head.get("feishu_open_id", "").strip()
                receive_id_type = "open_id"
            if not feishu_id:
                continue
            ok = await send_user_card(
                open_id=feishu_id,
                title=title,
                content=content,
                receive_id_type=receive_id_type,
            )
            if not ok:
                all_ok = False
                logger.error(
                    "检定到期提醒发送失败: 部门=%s, 负责人=%s",
                    dept.name, head_name,
                )

        if all_ok:
            sent += 1
            logger.info(
                "检定到期提醒已发送: 部门=%s, 负责人数=%d, 共 %d 条",
                dept.name, len(heads_list), total_items,
            )
        else:
            errors += 1

    logger.info(
        "检定到期提醒发送完成: sent=%d, skipped=%d, errors=%d",
        sent, skipped, errors,
    )
    return {"sent": sent, "skipped": skipped, "errors": errors}


# ═══════════════════════════════════════════
# Excel 台账导入
# ═══════════════════════════════════════════

# 列头匹配：标准化后的列名 → DB 字段名
INSTRUMENT_COLUMN_MAP: dict[str, str] = {
    "资产编号": "asset_number",
    "器具名称": "instrument_name",
    "型号规格": "model_spec",
    "测量范围": "measurement_range",
    "精度等级": "accuracy_grade",
    "器具编号": "serial_number",
    "检定周期(月)": "calibration_cycle_months",
    "使用地点": "location",
    "器具制造商": "manufacturer",
    "器具状态": "status",
    "彩色标志": "color_marking",
    "检定日期": "calibration_date",
    "检定单位": "calibration_unit",
    "检定结论": "calibration_result",
    "下次检定日期": "next_calibration_date",
}

GAS_DETECTOR_COLUMN_MAP: dict[str, str] = {
    "器具名称": "instrument_name",
    "规格型号": "detection_model",
    "量程": "measurement_range",
    "产品编号": "product_number",
    "安装形式": "installation_type",
    "安装位置": "installation_location",
    "使用介质": "medium",
    "标定系数": "calibration_factor",
    "传感器出厂日期": "manufacturer_supplier",
    "检定时间": "calibration_date",
    "检测单位": "detection_unit",
    "下次检定时间": "next_calibration_date",
    "检定结论": "calibration_result",
    "生产厂家": "manufacturer",
    "器具状态": "status",
    "部门": "department",
}


def _normalize_header(raw: str) -> str:
    """标准化列头：去换行、空格、全角括号统一为半角。"""
    result = raw.replace("\n", "").replace("\r", "").replace(" ", "").replace("　", "")
    result = result.replace("（", "(").replace("）", ")")
    # 去除末尾的冒号和多余符号
    result = result.rstrip("：:")
    return result


def _parse_department(cell_value: str) -> str | None:
    """从 Row 2 单元格中提取部门名。例如 '部门：质量控制部' → '质量控制部'。"""
    if not cell_value:
        return None
    text = str(cell_value).strip()
    # 尝试匹配 "部门：XXX" 或 "部门:XXX"
    for sep in ("部门：", "部门:", "部门 "):
        if text.startswith(sep):
            return text[len(sep):].strip()
    # 如果包含冒号，取冒号后的部分
    if "：" in text:
        return text.split("：", 1)[1].strip()
    if ":" in text:
        return text.split(":", 1)[1].strip()
    return text if text else None


def _excel_serial_to_date(
    cell_value: float, datemode: int = 0
) -> date | None:
    """将 Excel 序列号转换为 Python date。"""
    import xlrd  # type: ignore[import-untyped]
    try:
        dt = xlrd.xldate_as_datetime(cell_value, datemode)
        return dt.date()  # type: ignore[no-any-return]
    except Exception:
        return None


def _parse_workbook_xlrd(file_content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """用 xlrd 解析 .et/.xls 文件。

    返回 (sheets_data, errors)：
    - sheets_data: [{"name": sheet_name, "headers": [...], "rows": [[...], ...], "dept": str|None}, ...]
    """
    import xlrd

    errors: list[str] = []
    sheets_data: list[dict[str, Any]] = []

    try:
        wb = xlrd.open_workbook(file_contents=file_content, encoding_override="gbk")
    except Exception as e:
        errors.append(f"无法打开文件: {e}")
        return sheets_data, errors

    for sheet in wb.sheets():
        if sheet.nrows < 4:
            # 至少需要 header row + 1 行数据
            continue

        sheet_name = sheet.name.strip()

        # Row 2: department
        dept_raw = str(sheet.cell_value(2, 0)).strip() if sheet.ncols > 0 else ""
        dept = _parse_department(dept_raw)

        # Row 3: headers
        headers = [
            _normalize_header(str(sheet.cell_value(3, c)))
            for c in range(sheet.ncols)
        ]

        # Row 4+: data
        rows = []
        for row_idx in range(4, sheet.nrows):
            row_data = [sheet.cell_value(row_idx, c) for c in range(sheet.ncols)]
            # 跳过全空行
            if all(v == "" or v is None for v in row_data):
                continue
            rows.append(row_data)

        if rows:
            sheets_data.append({
                "name": sheet_name,
                "headers": headers,
                "rows": rows,
                "dept": dept,
                "datemode": wb.datemode,
            })

    return sheets_data, errors


def _parse_workbook_xlsx(file_content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """用 openpyxl 解析 .xlsx 文件。"""
    import io as io_mod

    import openpyxl  # type: ignore[import-untyped]

    errors: list[str] = []
    sheets_data: list[dict[str, Any]] = []

    try:
        wb = openpyxl.load_workbook(io_mod.BytesIO(file_content), read_only=True, data_only=True)
    except Exception as e:
        errors.append(f"无法打开文件: {e}")
        return sheets_data, errors

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # 收集所有行
        all_rows: list[list[Any]] = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            all_rows.append(list(row))

        if len(all_rows) < 4:
            continue

        name_clean = sheet_name.strip()

        # Row 2 (index 1): department
        dept_raw = str(all_rows[1][0]).strip() if all_rows[1] and len(all_rows[1]) > 0 else ""
        dept = _parse_department(dept_raw)

        # Row 3 (index 2): headers
        header_row = all_rows[2]
        headers = [_normalize_header(str(c)) if c is not None else "" for c in header_row]

        # Row 4+ (index 3+): data
        rows = []
        for row_data in all_rows[3:]:
            if all(v is None or str(v).strip() == "" for v in row_data):
                continue
            rows.append([
                v.value if hasattr(v, "value") else v
                for v in row_data
            ])

        if rows:
            sheets_data.append({
                "name": name_clean,
                "headers": headers,
                "rows": rows,
                "dept": dept,
                "datemode": 0,  # openpyxl uses 1900 date system by default
            })

    wb.close()
    return sheets_data, errors


def _map_and_convert_rows(
    sheets_data: list[dict[str, Any]],
    column_map: dict[str, str],
    datemode: int = 0,
    use_sheet_name_as_dept: bool = False,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """将原始行数据映射到 DB 字段，并转换日期。

    返回 (mapped_rows, warnings)。
    不拦截任何行，全部导入；warnings 仅提示哪些字段为空，供用户后续补全。
    """
    warnings: list[dict[str, Any]] = []
    all_rows: list[dict[str, Any]] = []

    for sheet in sheets_data:
        headers = sheet["headers"]
        rows = sheet["rows"]
        sheet_name = sheet["name"]
        # 标准器具：直接用 sheet 名称作为部门；探测器：优先 Row 2 部门名，回退 sheet 名称
        if use_sheet_name_as_dept:
            dept = sheet_name.strip()
        else:
            dept = (sheet.get("dept") or sheet_name).strip()

        # 构建列索引映射：header_index → db_field
        col_mapping: dict[int, str] = {}
        # 反向映射：db_field → 中文名（用于错误提示）
        db_field_cn: dict[str, str] = {v: k for k, v in column_map.items()}
        for idx, header in enumerate(headers):
            if not header:
                continue
            if header in column_map:
                col_mapping[idx] = column_map[header]

        if not col_mapping:
            continue

        for row_idx, row_data in enumerate(rows):
            excel_row = row_idx + 5  # Excel 行号（Row 1-4 是标题等）
            record: dict[str, Any] = {}
            missing: list[str] = []

            # 部门
            if dept:
                record["department"] = dept
            record["sheet_name"] = sheet_name

            # 器具名称 — 不阻断，空也导入
            name_idx = None
            for idx, field in col_mapping.items():
                if field == "instrument_name":
                    name_idx = idx
                    break
            if name_idx is not None and name_idx < len(row_data):
                name_val = row_data[name_idx]
                record["instrument_name"] = str(name_val).strip() if name_val else ""
            else:
                record["instrument_name"] = ""
            if not record["instrument_name"]:
                missing.append("器具名称")

            # 映射其他字段，记录缺失的选填字段
            for col_idx, db_field in col_mapping.items():
                if db_field == "instrument_name":
                    continue
                if col_idx >= len(row_data):
                    missing.append(db_field_cn.get(db_field, db_field))
                    continue
                value = row_data[col_idx]

                # 跳过空值 — 记录为缺失
                if value is None or str(value).strip() == "":
                    missing.append(db_field_cn.get(db_field, db_field))
                    continue

                # 日期字段处理
                if db_field in ("calibration_date", "next_calibration_date", "report_date"):
                    if isinstance(value, (int, float)) and value > 1:
                        dt = _excel_serial_to_date(float(value), datemode)
                        if dt:
                            record[db_field] = dt
                        else:
                            missing.append(db_field_cn.get(db_field, db_field))
                    elif isinstance(value, str):
                        from datetime import datetime
                        parsed = False
                        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
                            try:
                                record[db_field] = datetime.strptime(value.strip(), fmt).date()
                                parsed = True
                                break
                            except ValueError:
                                continue
                        if not parsed:
                            missing.append(db_field_cn.get(db_field, db_field))
                elif db_field == "calibration_cycle_months":
                    try:
                        record[db_field] = int(float(str(value)))
                    except (ValueError, TypeError):
                        missing.append(db_field_cn.get(db_field, db_field))
                else:
                    # Excel 会把整数读成 float（如 4699.0），str() 后变成 "4699.0"
                    val = value
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                    record[db_field] = str(val).strip()

            # 记录缺失字段警告
            if missing:
                warnings.append({
                    "sheet": sheet_name, "row": excel_row, "type": "warning",
                    "message": f"缺少字段: {', '.join(missing)}",
                    "missing_fields": missing,
                })

            # 自动计算下次检定日期
            _auto_calc_next_calibration_date(record)

            all_rows.append(record)

    return all_rows, warnings


async def _soft_delete_all_instruments(db: AsyncSession) -> int:
    """软删除所有标准计量器具。返回删除数量。"""
    from sqlalchemy import update as sa_update

    stmt = (
        sa_update(InstrumentRecord)
        .where(InstrumentRecord.is_deleted == False)  # noqa: E712
        .values(is_deleted=True)
    )
    result = await db.execute(stmt)
    return result.rowcount or 0  # type: ignore[attr-defined]


async def _soft_delete_all_gas_detectors(db: AsyncSession) -> int:
    """软删除所有有毒有害探测器。返回删除数量。"""
    from sqlalchemy import update as sa_update

    stmt = (
        sa_update(GasDetectorRecord)
        .where(GasDetectorRecord.is_deleted == False)  # noqa: E712
        .values(is_deleted=True)
    )
    result = await db.execute(stmt)
    return result.rowcount or 0  # type: ignore[attr-defined]


async def import_instrument_ledger(
    db: AsyncSession, file_content: bytes, filename: str
) -> dict[str, Any]:
    """导入标准计量器具台账 Excel。

    流程：解析文件 → 软删除全部旧记录 → 批量插入新记录。
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    # 1. 解析文件
    if ext in ("et", "xls"):
        sheets_data, parse_errors = _parse_workbook_xlrd(file_content)
        datemode = sheets_data[0]["datemode"] if sheets_data else 0
    elif ext == "xlsx":
        sheets_data, parse_errors = _parse_workbook_xlsx(file_content)
        datemode = 0
    else:
        raise ValueError(f"不支持的文件格式: .{ext}，请上传 .et 或 .xlsx 文件")

    if parse_errors:
        logger.warning(f"Parse errors during file import: {parse_errors}")

    # 2. 过滤：跳过探测器 sheet
    instrument_sheets = []
    detector_keywords = ["可燃", "有毒", "探测器", "气体检测"]
    for sheet in sheets_data:
        is_detector = any(kw in sheet["name"] for kw in detector_keywords)
        if not is_detector:
            instrument_sheets.append(sheet)

    if not instrument_sheets:
        raise ValueError("未找到标准计量器具数据 sheet（已跳过探测器 sheet）")

    # 3. 映射和转换
    mapped_rows, map_warnings = _map_and_convert_rows(
        instrument_sheets, INSTRUMENT_COLUMN_MAP, datemode, use_sheet_name_as_dept=True
    )

    if not mapped_rows:
        raise ValueError("文件中未找到有效的计量器具数据")

    # 按 Excel 解析顺序赋予全局 sort_order
    for idx, row in enumerate(mapped_rows):
        row["sort_order"] = idx

    # 4. 软删除全部 + 批量插入（在事务中）
    deleted_count = await _soft_delete_all_instruments(db)
    await db.flush()

    # 确定所有去重后行的字段并集，填充缺失字段为 None
    all_keys: set[str] = set()
    for row in mapped_rows:
        all_keys.update(row.keys())
    # anomaly_flags 有 server_default，不能设 None；统一设 {}
    all_keys.discard("anomaly_flags")
    for row in mapped_rows:
        for key in all_keys:
            if key not in row:
                row[key] = None
        # anomaly_flags 始终为空 dict，让 server_default 也能正常工作
        if "anomaly_flags" not in row or row.get("anomaly_flags") is None:
            row["anomaly_flags"] = {}

    batch_size = 500
    for i in range(0, len(mapped_rows), batch_size):
        batch = mapped_rows[i:i + batch_size]
        stmt = insert(InstrumentRecord).values(batch)
        await db.execute(stmt)

    # 6. 同步部门到 departments 表
    dept_names: set[str] = {r["department"] for r in mapped_rows if r.get("department")}
    synced = await repo.sync_departments(db, "instrument", dept_names)
    logger.info(f"Sync departments (instrument): {synced} new from {len(dept_names)} unique")

    # 7. 构建 sheet 详情
    sheet_details: list[dict[str, Any]] = []
    for sheet in instrument_sheets:
        dept = sheet.get("dept")
        # 统计该 sheet 的去重后行数
        sheet_rows = sum(
            1 for r in mapped_rows
            if r.get("sheet_name") == sheet["name"]
        )
        if sheet_rows > 0:
            sheet_details.append({
                "sheet_name": sheet["name"],
                "department": dept,
                "rows": sheet_rows,
            })

    await db.commit()

    return {
        "deleted_count": deleted_count,
        "imported_count": len(mapped_rows),
        "sheet_count": len(sheet_details),
        "sheet_details": sheet_details,
        "warnings": list(map_warnings)[:200],
    }


async def import_gas_detector_ledger(
    db: AsyncSession, file_content: bytes, filename: str
) -> dict[str, Any]:
    """导入有毒有害探测器台账 Excel。

    只处理 Sheet 0（探测器 sheet）。
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    # 1. 解析文件
    if ext in ("et", "xls"):
        sheets_data, parse_errors = _parse_workbook_xlrd(file_content)
        datemode = sheets_data[0]["datemode"] if sheets_data else 0
    elif ext == "xlsx":
        sheets_data, parse_errors = _parse_workbook_xlsx(file_content)
        datemode = 0
    else:
        raise ValueError(f"不支持的文件格式: .{ext}，请上传 .et 或 .xlsx 文件")

    if parse_errors:
        logger.warning(f"Parse errors during detector import: {parse_errors}")

    # 2. 找到探测器 sheet
    detector_keywords = ["可燃", "有毒", "探测器", "气体检测"]
    detector_sheet = None
    for sheet in sheets_data:
        if any(kw in sheet["name"] for kw in detector_keywords):
            detector_sheet = sheet
            break

    if detector_sheet is None:
        # 如果没有匹配的 sheet 名，尝试使用第一个 sheet 作为探测器 sheet
        if sheets_data:
            detector_sheet = sheets_data[0]
        else:
            raise ValueError("文件中未找到任何数据 sheet")

    # 3. 映射和转换
    mapped_rows, map_warnings = _map_and_convert_rows(
        [detector_sheet], GAS_DETECTOR_COLUMN_MAP, datemode
    )

    if not mapped_rows:
        raise ValueError("文件中未找到有效的探测器数据")

    # 按 Excel 解析顺序赋予全局 sort_order
    for idx, row in enumerate(mapped_rows):
        row["sort_order"] = idx

    # 4. 软删除全部 + 批量插入
    deleted_count = await _soft_delete_all_gas_detectors(db)
    await db.flush()

    # 确定所有去重后行的字段并集，填充缺失字段为 None
    all_keys_gd: set[str] = set()
    for row in mapped_rows:
        all_keys_gd.update(row.keys())
    all_keys_gd.discard("anomaly_flags")
    for row in mapped_rows:
        for key in all_keys_gd:
            if key not in row:
                row[key] = None
        if "anomaly_flags" not in row or row.get("anomaly_flags") is None:
            row["anomaly_flags"] = {}

    batch_size = 500
    for i in range(0, len(mapped_rows), batch_size):
        batch = mapped_rows[i:i + batch_size]
        stmt = insert(GasDetectorRecord).values(batch)
        await db.execute(stmt)

    # 6. 同步部门到 departments 表
    dept_names: set[str] = {r["department"] for r in mapped_rows if r.get("department")}
    synced = await repo.sync_departments(db, "gas_detector", dept_names)
    logger.info(f"Sync departments (gas_detector): {synced} new from {len(dept_names)} unique")

    # 7. sheet 详情
    sheet_details: list[dict[str, Any]] = [{
        "sheet_name": detector_sheet["name"],
        "department": detector_sheet.get("dept"),
        "rows": len(mapped_rows),
    }]

    await db.commit()

    return {
        "deleted_count": deleted_count,
        "imported_count": len(mapped_rows),
        "sheet_count": 1,
        "sheet_details": sheet_details,
        "warnings": list(map_warnings)[:200],
    }


# ═══════════════════════════════════════════
# 全局设置
# ═══════════════════════════════════════════


async def get_meter_settings(db: AsyncSession) -> dict[str, str]:
    """获取全局 meter 设置。"""
    cfg = await repo.get_or_create_meter_settings(db)
    return {"notify_time": cfg.notify_time.strftime("%H:%M")}


async def update_meter_settings(
    db: AsyncSession, notify_time_str: str,
) -> dict[str, str]:
    """校验并更新提醒时间。"""
    try:
        h, m = map(int, notify_time_str.split(":"))
        if not (0 <= h < 24 and 0 <= m < 60):
            raise ValueError("时间值超出范围")
    except (ValueError, TypeError):
        raise ValueError(f"无效的时间格式: {notify_time_str}，期望 HH:MM") from None

    cfg = await repo.update_meter_settings(db, time(h, m))
    return {"notify_time": cfg.notify_time.strftime("%H:%M")}

