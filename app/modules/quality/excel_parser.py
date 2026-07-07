"""液相计算表 Excel 解析器。

支持多种产品/标准的计算表模板，通过识别关键标签定位数据区域。
当前已适配：盐酸万古霉素 USP 标准（EX-HA-5246-001），其他模板可扩展。
"""

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


# ─── 解析结果数据结构 ───


@dataclass
class QualityStandard:
    """单项质量标准"""

    name: str  # 项目名称，如"万古霉素B"
    limit: float | None = None  # 合格限度值
    oot_haf: float | None = None  # OOT 阈值（HAF 产品线）
    oot_haa: float | None = None  # OOT 阈值（HAA 产品线）
    operator: str = "≤"  # 比较运算符：≤、≥、<、>


@dataclass
class ImpurityPeakArea:
    """单个杂质峰面积（一式两份）"""

    name: str  # 杂质名称，如"RS1"
    first: float  # 第一份峰面积
    second: float  # 第二份峰面积


@dataclass
class ImpurityResult:
    """单个杂质计算结果（百分比）"""

    name: str
    first_percent: float  # 第一份%
    second_percent: float  # 第二份%
    limit: float | None = None  # 合格标准
    oot_haf: float | None = None
    oot_haa: float | None = None


@dataclass
class CalculatedResult:
    """计算结果（百分比值，四舍五入后）"""

    name: str
    first_percent: float
    second_percent: float
    rounded_first: float  # 四舍五入后的值（报告用）
    rounded_second: float
    limit: float | None = None
    oot_haf: float | None = None
    oot_haa: float | None = None


@dataclass
class LcReportData:
    """液相计算表完整解析结果"""

    # 基本信息
    product_name: str = ""  # 产品名称
    batch_number: str = ""  # 批号
    form_id: str = ""  # 表格编号，如 EX-HA-5246-001
    standard_type: str = ""  # 标准类型：USP、EP、CP 等

    # 供试液 A 原始峰面积
    total_peak_area_a_first: float = 0  # 供试液A 总峰面积 第一份
    total_peak_area_a_second: float = 0  # 供试液A 总峰面积 第二份
    main_peak_area_a_first: float = 0  # 供试液A 主峰面积 第一份
    main_peak_area_a_second: float = 0  # 供试液A 主峰面积 第二份
    total_impurity_area_first: float = 0  # 杂质总峰面积 At 第一份
    total_impurity_area_second: float = 0  # 杂质总峰面积 At 第二份
    any_unknown_impurity_first: float = 0  # 任何未知杂质 Ax 第一份
    any_unknown_impurity_second: float = 0  # 任何未知杂质 Ax 第二份

    # 供试液 B 原始峰面积
    main_peak_area_b_first: float = 0  # Ab 第一份
    main_peak_area_b_second: float = 0  # Ab 第二份

    # 各杂质峰面积
    impurity_peaks: list[ImpurityPeakArea] = field(default_factory=list)

    # 计算结果
    vancomycin_b: CalculatedResult | None = None  # 万古霉素B%
    total_impurities: CalculatedResult | None = None  # 总杂质%
    impurity_results: list[ImpurityResult] = field(default_factory=list)

    # 质量标准
    standards: list[QualityStandard] = field(default_factory=list)

    # 原始解析信息
    raw_rows: int = 0
    raw_cols: int = 0


# ─── 解析器 ───


class LcExcelParser:
    """液相计算表解析器基类。

    子类针对不同模板实现 _do_parse()。
    """

    @classmethod
    def parse(cls, file_bytes: bytes, filename: str = "") -> LcReportData:
        """解析 Excel 文件，自动识别模板类型。"""
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active if wb.active else wb[wb.sheetnames[0]]

        # 读取前几行识别模板
        rows = list(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True))

        # 根据标题行判断模板类型
        title = str(rows[0][0]) if rows and rows[0] and rows[0][0] else ""

        if "USP" in title:
            parser = _UsrVancomycinParser()
        elif "EP" in title or "Ph.Eur" in title or "欧洲" in title:
            parser = _UsrVancomycinParser()  # TODO: 后续扩展 EP 模板
        elif "CP" in title or "中国药典" in title:
            parser = _UsrVancomycinParser()  # TODO: 后续扩展 CP 模板
        else:
            # 默认尝试 USP 格式
            parser = _UsrVancomycinParser()

        return parser._do_parse(ws, filename)


class _UsrVancomycinParser:
    """盐酸万古霉素 USP 标准计算表解析器（EX-HA-5246-001）。"""

    def _do_parse(self, ws, filename: str) -> LcReportData:
        data = LcReportData()
        data.raw_rows = ws.max_row
        data.raw_cols = ws.max_column

        # 将所有非空单元格读入内存
        cells: dict[tuple[int, int], str] = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            for cell in row:
                if cell.value is not None:
                    cells[(cell.row, cell.column)] = str(cell.value).strip()

        # ── R1: 标题行 ──
        title_raw = str(ws.cell(1, 1).value or "")
        data.product_name = self._extract_product_name(title_raw)
        data.form_id = self._extract_form_id(title_raw)
        data.standard_type = "USP" if "USP" in title_raw else ""

        # ── R2: 批号 + 标准头 ──
        data.batch_number = str(ws.cell(2, 2).value or "").strip()

        # ── 解析质量标准（R2 的 O列之后 + R5-R18 的标准区域）──
        self._parse_standards(ws, data)

        # ── 解析供试液A 峰面积数据（R4-R19）──
        self._parse_peak_areas(ws, data)

        # ── 解析供试液B（R20）──
        data.main_peak_area_b_first = self._safe_float(ws.cell(20, 6).value)
        data.main_peak_area_b_second = self._safe_float(ws.cell(20, 10).value)

        # ── 解析计算结果 ──
        self._parse_results(ws, data)

        # ── 解析杂质计算明细 ──
        self._parse_impurity_details(ws, data)

        return data

    # ─── 辅助方法 ───

    def _extract_product_name(self, title: str) -> str:
        """从标题提取产品名。"""
        # "盐酸万古霉素USP含量与有关物质计算表" → "盐酸万古霉素"
        for suffix in ["USP", "EP", "CP", "含量与有关物质计算表"]:
            idx = title.find(suffix)
            if idx > 0:
                return title[:idx].strip()
        return title.strip()

    def _extract_form_id(self, title: str) -> str:
        """从标题提取表格编号。"""
        m = re.search(r"EX-[A-Z]+-\d+-\d+", title)
        return m.group(0) if m else ""

    def _safe_float(self, val: Any) -> float:
        """安全转换为 float。"""
        if val is None:
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    def _pick_value(self, values: list[Any]) -> float:
        """从列表中取最后一个有效数值（Excel 公式拆分后，结果在最后）。"""
        for v in reversed(values):
            f = self._safe_float(v)
            if f != 0.0:
                return f
        return 0.0

    def _parse_standards(self, ws, data: LcReportData) -> None:
        """解析质量标准区域（O列及之后）。"""
        # O列（col 15）开始是合格标准，Q列（col 17）是 OOT(HAF)，R列（col 18）是 OOT(HAA)
        standards_map: dict[str, QualityStandard] = {}

        for row_idx in range(2, ws.max_row + 1):
            name = str(ws.cell(row_idx, 15).value or "").strip()  # O列
            if not name or name == "合格标准" or name.startswith("注意"):
                continue

            std = QualityStandard(
                name=name,
                limit=self._safe_float(ws.cell(row_idx, 16).value),  # P列
                oot_haf=self._safe_float(ws.cell(row_idx, 17).value),  # Q列
                oot_haa=self._safe_float(ws.cell(row_idx, 18).value),  # R列
            )
            # 万古霉素B 是 ≥，其他都是 ≤
            if "万古霉素" in name or "Vancomycin" in name.lower():
                std.operator = "≥"
            standards_map[name] = std

        data.standards = list(standards_map.values())

    def _parse_peak_areas(self, ws, data: LcReportData) -> None:
        """解析供试液A 各峰面积（R4-R19）。"""
        # C列=项目名，F列=第一份值，J列=第二份值
        col_name = 2  # B列
        col_first = 6  # F列
        col_second = 10  # J列

        peak_map: dict[str, tuple[float, float]] = {}

        for row_idx in range(4, 20):
            name = str(ws.cell(row_idx, col_name).value or "").strip()
            if not name:
                continue
            first = self._safe_float(ws.cell(row_idx, col_first).value)
            second = self._safe_float(ws.cell(row_idx, col_second).value)
            peak_map[name] = (first, second)

        # 提取特殊字段
        data.total_peak_area_a_first, data.total_peak_area_a_second = peak_map.pop("总峰面积", (0, 0))
        data.main_peak_area_a_first, data.main_peak_area_a_second = peak_map.pop("主峰面积", (0, 0))
        data.total_impurity_area_first, data.total_impurity_area_second = peak_map.pop(
            "杂质总峰面积（At）", peak_map.pop("杂质总峰面积", (0, 0))
        )
        data.any_unknown_impurity_first, data.any_unknown_impurity_second = peak_map.pop(
            "任何未知杂质（Ax）", peak_map.pop("任何未知杂质", (0, 0))
        )

        # 其余为各杂质峰面积
        for name, (first, second) in peak_map.items():
            # 从名称中提取杂质代号
            imp_name = self._extract_impurity_code(name)
            data.impurity_peaks.append(ImpurityPeakArea(name=imp_name, first=first, second=second))

    def _extract_impurity_code(self, label: str) -> str:
        """从标签提取杂质代号，如 '杂质RS1（ARS1）' → '杂质RS1'。"""
        if "（" in label:
            return label[: label.index("（")].strip()
        return label.strip()

    def _parse_results(self, ws, data: LcReportData) -> None:
        """解析万古霉素B和总杂质的计算结果。

        列映射（从实际 .xlsx 调试确认）：
        - Col 11 (K) = "＝"（分隔符，文本）
        - Col 12 (L) = 精确值（如 0.9563）
        - Col 13 (M) = 报告取整值（如 0.956）
        - 第一份值：标签行 Col 12/13
        - 第二份值：标签行+2 Col 12（可能无 Col 13）
        """
        COL_RAW = 12  # L列：精确值
        COL_RND = 13  # M列：报告取整值

        # 万古霉素B（R21-R24）
        vb_first_raw = self._safe_float(ws.cell(21, COL_RAW).value)
        vb_second_raw = self._safe_float(ws.cell(23, COL_RAW).value)
        vb_first_rnd = self._safe_float(ws.cell(21, COL_RND).value)
        vb_second_rnd = self._safe_float(ws.cell(23, COL_RND).value)

        vb_std = self._find_standard(data, "万古霉素B")
        data.vancomycin_b = CalculatedResult(
            name="万古霉素B",
            first_percent=vb_first_raw,  # 保持小数形式（0.9563 = 95.63%）
            second_percent=vb_second_raw,
            rounded_first=vb_first_rnd,
            rounded_second=vb_second_rnd,
            limit=vb_std.limit if vb_std else None,
            oot_haf=vb_std.oot_haf if vb_std else None,
            oot_haa=vb_std.oot_haa if vb_std else None,
        )

        # 总杂质（R77，K列=11 是 "＝"，L列=12 是值）
        total_first_raw = self._safe_float(ws.cell(77, COL_RAW).value)
        total_second = 0.0  # 总杂质只有一份
        ts = self._find_standard(data, "总杂质")
        data.total_impurities = CalculatedResult(
            name="总杂质",
            first_percent=total_first_raw,  # 保持小数形式（0.044 = 4.4%）
            second_percent=total_second,
            rounded_first=total_first_raw,
            rounded_second=total_second,
            limit=ts.limit if ts else None,
            oot_haf=ts.oot_haf if ts else None,
            oot_haa=ts.oot_haa if ts else None,
        )

    def _parse_impurity_details(self, ws, data: LcReportData) -> None:
        """解析各杂质百分比计算结果（R25-R76）。

        实际格式（从 .xlsx 调试确认）：
        - R[N]:   杂质XX％ 标签 + 第一份公式元素 + L列=精确值 + M列=报告取整值
        - R[N+1]: 分母行（25×Ab+At 的具体值），跳过
        - R[N+2]: 第二份＝ + 公式元素 + L列=精确值（可能无M列取整值）
        - R[N+3]: 分母行，跳过

        列映射：Col 11(K)="＝", Col 12(L)=精确值, Col 13(M)=报告取整值
        """
        COL_RAW = 12  # L列：精确值
        COL_RND = 13  # M列：报告取整值

        row = 25
        while row <= 76:
            label = str(ws.cell(row, 1).value or "").strip()
            if not label:
                row += 1
                continue

            # 提取杂质名
            imp_name = label.rstrip("％").rstrip("%").strip()
            if not imp_name or imp_name.startswith("25×") or imp_name == "总杂质％":
                row += 1
                continue

            # 第一份结果：标签行本身就有值（Col 12=精确值, Col 13=取整值）
            first_raw = self._safe_float(ws.cell(row, COL_RAW).value)
            first_rnd = self._safe_float(ws.cell(row, COL_RND).value)
            # 第二份结果：标签行+2
            second_raw = self._safe_float(ws.cell(row + 2, COL_RAW).value)
            second_rnd = self._safe_float(ws.cell(row + 2, COL_RND).value)

            # 保持小数形式，与限度值单位一致（0.00203 = 0.203%）
            first_pct = first_raw
            second_pct = second_raw

            std = self._find_standard(data, imp_name)
            data.impurity_results.append(
                ImpurityResult(
                    name=imp_name,
                    first_percent=first_pct,
                    second_percent=second_pct,
                    limit=std.limit if std else None,
                    oot_haf=std.oot_haf if std else None,
                    oot_haa=std.oot_haa if std else None,
                )
            )
            row += 4  # 跳到下一个杂质组

    def _find_standard(self, data: LcReportData, name: str) -> QualityStandard | None:
        """按名称匹配质量标准（支持模糊匹配）。"""
        # 精确匹配
        for s in data.standards:
            if s.name == name:
                return s
        # 包含匹配：如 "杂质RS1" 匹配 "RS1"，"杂质A" 匹配 "杂质A"
        for s in data.standards:
            if s.name in name or name in s.name:
                return s
        # 去掉"杂质"前缀再试
        clean = name.replace("杂质", "").strip()
        if clean != name:
            for s in data.standards:
                if s.name == clean or s.name in clean or clean in s.name:
                    return s
        return None


# ─── 便捷函数 ───


def parse_lc_excel(file_bytes: bytes, filename: str = "") -> LcReportData:
    """解析液相计算表 Excel 文件。"""
    return LcExcelParser.parse(file_bytes, filename)
