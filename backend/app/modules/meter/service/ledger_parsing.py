"""Excel 台账解析：列映射、表头/部门/日期解析与行映射。"""

from __future__ import annotations

from datetime import date, datetime
from typing import Any, cast

from app.modules.meter.service.common import _auto_calc_next_calibration_date

# 列头匹配：标准化后的列名 → DB 字段名
INSTRUMENT_COLUMN_MAP: dict[str, str] = {
    "资产编号": "asset_number",
    "器具名称": "instrument_name",
    "型号规格": "model_spec",
    "测量范围": "measurement_range",
    "精度等级": "accuracy_grade",
    "器具编号": "serial_number",
    "检定周期(月)": "calibration_cycle_months",
    "使用地点": "location",
    "器具制造商": "manufacturer",
    "器具状态": "status",
    "彩色标志": "color_marking",
    "检定日期": "calibration_date",
    "检定单位": "calibration_unit",
    "检定结论": "calibration_result",
    "下次检定日期": "next_calibration_date",
}

GAS_DETECTOR_COLUMN_MAP: dict[str, str] = {
    "器具名称": "instrument_name",
    "规格型号": "detection_model",
    "量程": "measurement_range",
    "产品编号": "product_number",
    "安装形式": "installation_type",
    "安装位置": "installation_location",
    "使用介质": "medium",
    "标定系数": "calibration_factor",
    "传感器出厂日期": "manufacturer_supplier",
    "检定时间": "calibration_date",
    "检测单位": "detection_unit",
    "下次检定时间": "next_calibration_date",
    "检定结论": "calibration_result",
    "生产厂家": "manufacturer",
    "器具状态": "status",
    "部门": "department",
}


def _normalize_header(raw: str) -> str:
    """标准化列头：去换行、空格、全角括号统一为半角。"""
    result = raw.replace("\n", "").replace("\r", "").replace(" ", "").replace("　", "")
    result = result.replace("（", "(").replace("）", ")")
    # 去除末尾的冒号和多余符号
    result = result.rstrip("：:")
    return result


def _parse_department(cell_value: str) -> str | None:
    """从 Row 2 单元格中提取部门名。例如 '部门：质量控制部' → '质量控制部'。"""
    if not cell_value:
        return None
    text = str(cell_value).strip()
    # 尝试匹配 "部门：XXX" 或 "部门:XXX"
    for sep in ("部门：", "部门:", "部门 "):
        if text.startswith(sep):
            return text[len(sep):].strip()
    # 如果包含冒号，取冒号后的部分
    if "：" in text:
        return text.split("：", 1)[1].strip()
    if ":" in text:
        return text.split(":", 1)[1].strip()
    return text if text else None


def _excel_serial_to_date(
    cell_value: float, datemode: int = 0
) -> date | None:
    """将 Excel 序列号转换为 Python date。"""
    import xlrd  # type: ignore[import-untyped]
    try:
        dt = xlrd.xldate_as_datetime(cell_value, cast(Any, datemode))
        return dt.date()  # type: ignore[no-any-return]
    except Exception:
        return None


def _parse_workbook_xlrd(file_content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """用 xlrd 解析 .et/.xls 文件。

    返回 (sheets_data, errors)：
    - sheets_data: [{"name": sheet_name, "headers": [...], "rows": [[...], ...], "dept": str|None}, ...]
    """
    import xlrd

    errors: list[str] = []
    sheets_data: list[dict[str, Any]] = []

    try:
        wb = xlrd.open_workbook(file_contents=file_content, encoding_override="gbk")
    except Exception as e:
        errors.append(f"无法打开文件: {e}")
        return sheets_data, errors

    for sheet in wb.sheets():
        if sheet.nrows < 4:
            # 至少需要 header row + 1 行数据
            continue

        sheet_name = sheet.name.strip()

        # Row 2: department
        dept_raw = str(sheet.cell_value(2, 0)).strip() if sheet.ncols > 0 else ""
        dept = _parse_department(dept_raw)

        # Row 3: headers
        headers = [
            _normalize_header(str(sheet.cell_value(3, c)))
            for c in range(sheet.ncols)
        ]

        # Row 4+: data
        rows = []
        for row_idx in range(4, sheet.nrows):
            row_data = [sheet.cell_value(row_idx, c) for c in range(sheet.ncols)]
            # 跳过全空行
            if all(v == "" or v is None for v in row_data):
                continue
            rows.append(row_data)

        if rows:
            sheets_data.append({
                "name": sheet_name,
                "headers": headers,
                "rows": rows,
                "dept": dept,
                "datemode": wb.datemode,
            })

    return sheets_data, errors


def _parse_workbook_xlsx(file_content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """用 openpyxl 解析 .xlsx 文件。"""
    import io as io_mod

    import openpyxl

    errors: list[str] = []
    sheets_data: list[dict[str, Any]] = []

    try:
        wb = openpyxl.load_workbook(io_mod.BytesIO(file_content), read_only=True, data_only=True)
    except Exception as e:
        errors.append(f"无法打开文件: {e}")
        return sheets_data, errors

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # 收集所有行
        all_rows: list[list[Any]] = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            all_rows.append(list(row))

        if len(all_rows) < 4:
            continue

        name_clean = sheet_name.strip()

        # 表头行探测：前 4 行中非空单元格最多的行。
        # 兼容「标题/部门/表头/数据」与「标题/标题/部门/表头/数据」两种版式，
        # 避免与 xlrd 解析器及旧导入脚本出现整行错位。
        best_idx, best_score = 0, -1
        for i in range(min(4, len(all_rows))):
            score = sum(
                1 for v in all_rows[i]
                if v is not None and str(v).strip() != ""
            )
            if score > best_score:
                best_idx, best_score = i, score
        header_idx = best_idx

        # 部门行 = 表头行的上一行
        dept_raw = ""
        dept_idx = header_idx - 1
        if dept_idx >= 0 and all_rows[dept_idx]:
            cell = all_rows[dept_idx][0]
            dept_raw = str(cell).strip() if cell is not None else ""
        dept = _parse_department(dept_raw)

        # 表头行
        header_row = all_rows[header_idx]
        headers = [_normalize_header(str(c)) if c is not None else "" for c in header_row]

        # 表头之后的行 = 数据
        rows = []
        for row_data in all_rows[header_idx + 1:]:
            if all(v is None or str(v).strip() == "" for v in row_data):
                continue
            rows.append([
                v.value if hasattr(v, "value") else v
                for v in row_data
            ])

        if rows:
            sheets_data.append({
                "name": name_clean,
                "headers": headers,
                "rows": rows,
                "dept": dept,
                "datemode": 0,  # openpyxl uses 1900 date system by default
            })

    wb.close()
    return sheets_data, errors


def _map_and_convert_rows(
    sheets_data: list[dict[str, Any]],
    column_map: dict[str, str],
    datemode: int = 0,
    use_sheet_name_as_dept: bool = False,
    required_fields: tuple[str, ...] = (),
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], set[str]]:
    """将原始行数据映射到 DB 字段，并转换日期。

    返回 (mapped_rows, warnings, skipped_keys)。
    required_fields 中任一字段为空/空字符串的行会被跳过（记入 warnings）；
    不传则全部导入，warnings 仅提示哪些字段为空，供用户后续补全。
    skipped_keys：因必填字段为空被跳过的行中出现的资产/产品编号，
    供导入方在「文件未出现即软删除」清理时保留这些旧记录，避免误删。
    """
    warnings: list[dict[str, Any]] = []
    all_rows: list[dict[str, Any]] = []
    skipped_keys: set[str] = set()

    for sheet in sheets_data:
        headers = sheet["headers"]
        rows = sheet["rows"]
        sheet_name = sheet["name"]
        # 标准器具：直接用 sheet 名称作为部门；探测器：优先 Row 2 部门名，回退 sheet 名称
        if use_sheet_name_as_dept:
            dept = sheet_name.strip()
        else:
            dept = (sheet.get("dept") or sheet_name).strip()

        # 构建列索引映射：header_index → db_field
        col_mapping: dict[int, str] = {}
        # 反向映射：db_field → 中文名（用于错误提示）
        db_field_cn: dict[str, str] = {v: k for k, v in column_map.items()}
        for idx, header in enumerate(headers):
            if not header:
                continue
            if header in column_map:
                col_mapping[idx] = column_map[header]

        if not col_mapping:
            warnings.append({
                "sheet": sheet_name, "type": "warning",
                "department": dept,
                "message": "表头列名无法匹配，已跳过该 sheet 的数据",
            })
            continue

        for row_idx, row_data in enumerate(rows):
            excel_row = row_idx + 5  # Excel 行号（Row 1-4 是标题等）
            record: dict[str, Any] = {}
            missing: list[str] = []

            # 部门
            if dept:
                record["department"] = dept
            record["sheet_name"] = sheet_name

            # 器具名称 — 不阻断，空也导入
            name_idx = None
            for idx, field in col_mapping.items():
                if field == "instrument_name":
                    name_idx = idx
                    break
            if name_idx is not None and name_idx < len(row_data):
                name_val = row_data[name_idx]
                record["instrument_name"] = str(name_val).strip() if name_val else ""
            else:
                record["instrument_name"] = ""
            if not record["instrument_name"]:
                missing.append("器具名称")

            # 映射其他字段，记录缺失的选填字段
            for col_idx, db_field in col_mapping.items():
                if db_field == "instrument_name":
                    continue
                if col_idx >= len(row_data):
                    missing.append(db_field_cn.get(db_field, db_field))
                    continue
                value = row_data[col_idx]

                # 跳过空值 — 记录为缺失
                if value is None or str(value).strip() == "":
                    missing.append(db_field_cn.get(db_field, db_field))
                    continue

                # 日期字段处理
                if db_field in ("calibration_date", "next_calibration_date", "report_date"):
                    if isinstance(value, (int, float)) and value > 1:
                        dt = _excel_serial_to_date(float(value), datemode)
                        if dt:
                            record[db_field] = dt
                        else:
                            missing.append(db_field_cn.get(db_field, db_field))
                    elif isinstance(value, str):
                        parsed = False
                        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
                            try:
                                record[db_field] = datetime.strptime(value.strip(), fmt).date()
                                parsed = True
                                break
                            except ValueError:
                                continue
                        if not parsed:
                            missing.append(db_field_cn.get(db_field, db_field))
                    elif isinstance(value, datetime):
                        # Excel 日期格式单元格：openpyxl 读出 datetime 对象
                        record[db_field] = value.date()
                    elif isinstance(value, date):
                        # 纯日期单元格（xlrd .et 文件可能直接给出 date）
                        record[db_field] = value
                elif db_field == "calibration_cycle_months":
                    try:
                        record[db_field] = int(float(str(value)))
                    except (ValueError, TypeError):
                        missing.append(db_field_cn.get(db_field, db_field))
                else:
                    # Excel 会把整数读成 float（如 4699.0），str() 后变成 "4699.0"
                    val = value
                    if isinstance(val, float) and val == int(val):
                        val = int(val)
                    record[db_field] = str(val).strip()

            # 记录缺失字段警告
            if missing:
                warnings.append({
                    "sheet": sheet_name, "row": excel_row, "type": "warning",
                    "message": f"缺少字段: {', '.join(missing)}",
                    "missing_fields": missing,
                })

            # 必填字段过滤：任一为空/空字符串则跳过该行
            if required_fields:
                empty_required = [f for f in required_fields if not record.get(f)]
                if empty_required:
                    for key_field in ("asset_number", "product_number"):
                        if record.get(key_field):
                            skipped_keys.add(str(record[key_field]))
                    warnings.append({
                        "sheet": sheet_name, "row": excel_row, "type": "warning",
                        "message": f"必填字段为空（{', '.join(empty_required)}），已跳过该行",
                        "missing_fields": empty_required,
                    })
                    continue

            # 自动计算下次检定日期
            _auto_calc_next_calibration_date(record)

            all_rows.append(record)

    return all_rows, warnings, skipped_keys
