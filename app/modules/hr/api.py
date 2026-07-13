from datetime import date
from io import BytesIO
from urllib.parse import quote
from uuid import UUID

from fastapi import Depends, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.response import paginated_response, success_response
from app.modules.hr.analysis_api import router as analysis_router
from app.modules.hr.document_generator import generate_onboarding_training_record
from app.modules.hr.evaluation_document_generator import generate_training_evaluation
from app.modules.hr.notification_document_generator import (
    generate_training_notification,
)
from app.modules.hr.onboarding_evaluation_document_generator import (
    generate_onboarding_evaluation,
)
from app.modules.hr.prejob_document_generator import generate_prejob_training_plan
from app.modules.hr.schemas import (
    AnnualTrainingPlanCreate,
    AnnualTrainingPlanItemBatchUpdate,
    AnnualTrainingPlanItemResponse,
    AnnualTrainingPlanResponse,
    AnnualTrainingPlanUpdate,
    DepartmentCreate,
    DepartmentResponse,
    DepartmentUpdate,
    DepartureRecordCreate,
    DepartureRecordResponse,
    DepartureRecordUpdate,
    EmployeeCreate,
    EmployeeResponse,
    EmployeeUpdate,
    OffboardingRecordCreate,
    OffboardingRecordResponse,
    OffboardingRecordUpdate,
    OnboardingEvaluationInput,
    OnboardingRecordResponse,
    TeamCreate,
    TeamResponse,
    TeamUpdate,
    TrainingEvaluationInput,
    TrainingLedgerCreate,
    TrainingLedgerPageCreate,
    TrainingLedgerPageResponse,
    TrainingLedgerResponse,
    TrainingLedgerUpdate,
    TrainingNotificationInput,
    TrainingSignInSheetInput,
    TrainerResponse,
    TrainerListResponse,
    SopCatalogResponse,
    SopCatalogListResponse,
)
from app.modules.hr.service import (
    AnnualTrainingPlanItemService,
    AnnualTrainingPlanService,
    DepartmentService,
    DepartureRecordService,
    EmployeeService,
    OffboardingRecordService,
    OnboardingRecordService,
    TeamService,
    TrainingLedgerPageService,
    TrainingLedgerService,
)
from app.modules.hr.signin_document_generator import generate_training_sign_in_sheet
from app.shared.module_api import create_module_router
from app.shared.module_registry import MODULES_BY_CODE
from app.shared.schemas import PageParams

router = create_module_router(MODULES_BY_CODE["hr"])


def get_employee_service(session: AsyncSession = Depends(get_db)) -> EmployeeService:
    return EmployeeService(session)


def get_department_service(
    session: AsyncSession = Depends(get_db),
) -> DepartmentService:
    return DepartmentService(session)


def get_offboarding_service(
    session: AsyncSession = Depends(get_db),
) -> OffboardingRecordService:
    return OffboardingRecordService(session)


def get_team_service(
    session: AsyncSession = Depends(get_db),
) -> TeamService:
    return TeamService(session)


def get_onboarding_service(
    session: AsyncSession = Depends(get_db),
) -> OnboardingRecordService:
    return OnboardingRecordService(session)


def get_departure_service(
    session: AsyncSession = Depends(get_db),
) -> DepartureRecordService:
    return DepartureRecordService(session)


def get_training_ledger_service(
    session: AsyncSession = Depends(get_db),
) -> TrainingLedgerService:
    return TrainingLedgerService(session)


def get_training_ledger_page_service(
    session: AsyncSession = Depends(get_db),
) -> TrainingLedgerPageService:
    return TrainingLedgerPageService(session)


def get_annual_training_plan_service(
    session: AsyncSession = Depends(get_db),
) -> AnnualTrainingPlanService:
    return AnnualTrainingPlanService(session)


def get_annual_training_plan_item_service(
    session: AsyncSession = Depends(get_db),
) -> AnnualTrainingPlanItemService:
    return AnnualTrainingPlanItemService(session)


# ─── Employee Routes ───

@router.get("/employees", summary="员工列表")
async def list_employees(
    department: str | None = Query(None, description="部门筛选"),
    status: str | None = Query(None, description="状态筛选"),
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    page_params: PageParams = Depends(),
    service: EmployeeService = Depends(get_employee_service),
):
    employees, total = await service.list_employees(
        department=department,
        status=status,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        EmployeeResponse.model_validate(e).model_dump(mode="json")
        for e in employees
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/employees", summary="创建员工")
async def create_employee(
    payload: EmployeeCreate,
    service: EmployeeService = Depends(get_employee_service),
):
    employee = await service.create_employee(payload)
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
        message="员工创建成功",
        status_code=201,
    )


@router.post("/employees/upload", summary="上传人员名单")
async def upload_employees(
    file: UploadFile,
    service: EmployeeService = Depends(get_employee_service),
):
    """上传 Excel 人员名单，按工号自动新增或更新。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 格式")
    try:
        content = await file.read()
        result = await service.upload_employees(content)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return success_response(data=result, message=f"新增 {result['created']}，更新 {result['updated']}")


@router.get("/roster", summary="下载花名册")
async def download_roster(
    department: str | None = Query(None),
    session: AsyncSession = Depends(get_db),
):
    """按部门下载员工花名册 Excel。"""
    from app.modules.hr.roster_generator import generate_roster_sync
    from app.modules.hr.models import Employee
    r = await session.execute(
        select(Employee).where(Employee.is_deleted == False, Employee.status != "离职").order_by(Employee.department, Employee.employee_number)
    )
    employees = [(e.name, e.department, e.gender or "", e.education or "", e.hire_date, e.status) for e in r.scalars().all()]
    if department:
        employees = [e for e in employees if e[1] == department]
    buffer = generate_roster_sync(employees, department)
    filename = f"花名册_{department or '全部'}.docx"
    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{quote(filename)}"}
    )


@router.get("/employees/by-number/{employee_number}", summary="根据工号查询员工")
async def get_employee_by_number(
    employee_number: str,
    service: EmployeeService = Depends(get_employee_service),
):
    employee = await service.get_employee_by_number(employee_number)
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
    )


@router.get("/employees/{employee_id}", summary="员工详情")
async def get_employee(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
):
    employee = await service.get_employee(employee_id)
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
    )


@router.put("/employees/{employee_id}", summary="更新员工")
async def update_employee(
    employee_id: UUID,
    payload: EmployeeUpdate,
    service: EmployeeService = Depends(get_employee_service),
):
    employee = await service.update_employee(employee_id, payload)
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
        message="员工更新成功",
    )


@router.delete("/employees/{employee_id}", summary="删除员工")
async def delete_employee(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
):
    await service.delete_employee(employee_id)
    return success_response(message="员工删除成功")


@router.get(
    "/employees/{employee_number}/onboarding-training-record",
    summary="导出员工入职培训记录",
)
async def export_onboarding_training_record(
    employee_number: str,
    service: EmployeeService = Depends(get_employee_service),
):
    """根据员工工号自动生成并下载入职培训记录 Word 文档。"""
    employee = await service.get_employee_by_number(employee_number)
    try:
        buffer: BytesIO = generate_onboarding_training_record(employee)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    filename = f"onboarding_training_record_{employee.employee_number}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class TrainingItem(BaseModel):
    sop_number: str = ""
    file_name: str = ""
    content: str = ""
    method: str = ""
    trainer: str = ""


class TrainingExportRequest(BaseModel):
    training_items: list[TrainingItem] = []


@router.post(
    "/employees/{employee_number}/onboarding-training-record",
    summary="导出员工入职培训记录（带培训项）",
)
async def export_onboarding_training_record_with_items(
    employee_number: str,
    body: TrainingExportRequest,
    service: EmployeeService = Depends(get_employee_service),
):
    """根据员工工号和前端选中的培训项，生成入职培训记录 Word 文档。"""
    employee = await service.get_employee_by_number(employee_number)
    items = [it.model_dump() for it in body.training_items]
    try:
        buffer: BytesIO = generate_onboarding_training_record(employee, training_items=items)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    filename = f"onboarding_training_record_{employee.employee_number}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/employees/{employee_id}/prejob-training-plan",
    summary="导出员工岗前培训计划",
)
async def export_prejob_training_plan(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
):
    """根据员工数据自动生成并下载岗前培训计划 Excel 文档。"""
    employee = await service.get_employee(employee_id)
    try:
        buffer: BytesIO = generate_prejob_training_plan(employee)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    filename = f"prejob_training_plan_{employee.employee_number}.xlsx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/employees/{employee_id}/onboarding-evaluation",
    summary="导出员工上岗评估表",
)
async def export_onboarding_evaluation_by_employee(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
):
    """根据员工档案预填基本信息并导出上岗评估表 Excel 文档。"""
    employee = await service.get_employee(employee_id)

    payload = OnboardingEvaluationInput(
        employee_name=employee.name or "",
        employee_number=employee.employee_number or None,
        gender=employee.gender or None,
        department_position=f"{employee.department or ''}/{employee.position or ''}",
        hire_date=employee.hire_date,
    )
    buffer: BytesIO = generate_onboarding_evaluation(payload)

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(employee.hire_date).replace("-", "") if employee.hire_date else "nodate"
    filename = f"onboarding_evaluation_{employee.employee_number}_{safe_date}.xlsx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )




@router.post("/training-sign-in-sheet", summary="生成培训签到表")
async def export_training_sign_in_sheet(
    payload: TrainingSignInSheetInput,
):
    """根据填写的培训信息自动生成培训签到表 Word 文档。

    超过一页（20人）时自动分页，所有页面合并在一个 docx 中。
    """
    safe_date = str(payload.training_date).replace("-", "")
    try:
        buffer: BytesIO = generate_training_sign_in_sheet(payload)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_filename = f"training_sign_in_sheet_{safe_date}.docx"
    from urllib.parse import quote

    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=utf-8''{quote('培训签到表_' + safe_date + '.docx')}"
        },
    )


@router.post("/training-notification", summary="生成培训通知")
async def export_training_notification(
    payload: TrainingNotificationInput,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    """根据填写的培训信息生成培训通知 Word 文档。培训台账需通过「添加到培训台账」按钮手动录入。"""
    try:
        buffer: BytesIO = generate_training_notification(payload)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(payload.training_date).replace("-", "")
    filename = f"training_notification_{safe_date}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/training-evaluation", summary="生成培训效果评估表")
async def export_training_evaluation(
    payload: TrainingEvaluationInput,
):
    """根据填写的培训信息自动生成培训效果评估表 Word 文档。"""
    buffer: BytesIO = generate_training_evaluation(payload)

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(payload.training_date).replace("-", "") if payload.training_date else "nodate"
    from urllib.parse import quote
    safe_filename = f"training_evaluation_{safe_date}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=utf-8''{quote('培训效果评估表_' + safe_date + '.docx')}"
        },
    )


@router.post("/onboarding-evaluation", summary="生成员工上岗评估表")
async def export_onboarding_evaluation(
    payload: OnboardingEvaluationInput,
):
    """根据填写的评估信息自动生成员工上岗评估表 Excel 文档。"""
    buffer: BytesIO = generate_onboarding_evaluation(payload)

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(payload.evaluation_date).replace("-", "") if payload.evaluation_date else "nodate"
    filename = f"onboarding_evaluation_{safe_date}.xlsx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Department Routes ───

@router.get("/departments", summary="部门列表")
async def list_departments(
    keyword: str | None = Query(None, description="部门名称或编码关键词"),
    page_params: PageParams = Depends(),
    service: DepartmentService = Depends(get_department_service),
):
    departments, total = await service.list_departments(
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        DepartmentResponse.model_validate(d).model_dump(mode="json")
        for d in departments
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/departments", summary="创建部门")
async def create_department(
    payload: DepartmentCreate,
    service: DepartmentService = Depends(get_department_service),
):
    department = await service.create_department(payload)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
        message="部门创建成功",
        status_code=201,
    )


@router.get("/departments/{department_id}", summary="部门详情")
async def get_department(
    department_id: UUID,
    service: DepartmentService = Depends(get_department_service),
):
    department = await service.get_department(department_id)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
    )


@router.put("/departments/{department_id}", summary="更新部门")
async def update_department(
    department_id: UUID,
    payload: DepartmentUpdate,
    service: DepartmentService = Depends(get_department_service),
):
    department = await service.update_department(department_id, payload)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
        message="部门更新成功",
    )


@router.delete("/departments/{department_id}", summary="删除部门")
async def delete_department(
    department_id: UUID,
    service: DepartmentService = Depends(get_department_service),
):
    await service.delete_department(department_id)
    return success_response(message="部门删除成功")


# ─── Position Routes ───

@router.get("/positions", summary="职位列表（按部门筛选）")
async def list_positions(
    department: str | None = Query(None, description="部门名称筛选"),
    session: AsyncSession = Depends(get_db),
):
    """返回职位列表，可按部门筛选。用于入职管理和员工档案的职位下拉选项。"""
    from app.modules.hr.models import HrPosition

    q = select(HrPosition).where(HrPosition.is_deleted == False)
    if department:
        q = q.where(HrPosition.department == department)
    q = q.order_by(HrPosition.sort_order)

    result = await session.execute(q)
    rows = result.scalars().all()
    data = [
        {"id": str(r.id), "department": r.department, "name": r.name}
        for r in rows
    ]
    return success_response(data=data)


class PositionCreate(BaseModel):
    department: str
    name: str


@router.post("/positions", summary="新建职位")
async def create_position(
    payload: PositionCreate,
    session: AsyncSession = Depends(get_db),
):
    """手动新建一个职位，写入 hr.positions 表。"""
    from app.modules.hr.models import HrPosition

    pos = HrPosition(department=payload.department, name=payload.name)
    session.add(pos)
    await session.flush()
    return success_response(
        data={"id": str(pos.id), "department": pos.department, "name": pos.name},
        message="职位创建成功",
        status_code=201,
    )


@router.delete("/positions/by-name/{position_name}", summary="按名称删除职位")
async def delete_position_by_name(
    position_name: str,
    department: str = Query(..., description="部门名称"),
    session: AsyncSession = Depends(get_db),
):
    """删除指定部门和名称的职位，同时清除关联的 SOP 目录条目。"""
    from app.modules.hr.models import HrPosition, SopCatalog

    pos = (await session.execute(
        select(HrPosition).where(
            HrPosition.department == department,
            HrPosition.name == position_name,
            HrPosition.is_deleted == False,
        )
    )).scalar_one_or_none()
    if not pos:
        raise HTTPException(404, "职位不存在")
    pos.is_deleted = True

    # 同时清除关联SOP条目
    await session.execute(
        select(SopCatalog).where(
            SopCatalog.department == department,
            SopCatalog.position_name == position_name,
            SopCatalog.is_deleted == False,
        )
    )
    sops = (await session.execute(
        select(SopCatalog).where(
            SopCatalog.department == department,
            SopCatalog.position_name == position_name,
            SopCatalog.is_deleted == False,
        )
    )).scalars().all()
    for sop in sops:
        sop.is_deleted = True

    await session.flush()
    return success_response(message="删除成功")


@router.get("/positions/departments", summary="职位表中所有部门")
async def list_position_departments(
    session: AsyncSession = Depends(get_db),
):
    """返回职位表中所有不重复的部门名称。"""
    from app.modules.hr.models import HrPosition

    result = await session.execute(
        select(HrPosition.department)
        .where(HrPosition.is_deleted == False)
        .distinct()
        .order_by(HrPosition.department)
    )
    return success_response(data=[row[0] for row in result.all()])


# ─── Position Training Routes ───

@router.get("/position-trainings", summary="岗位培训内容列表")
async def list_position_trainings(
    department: str | None = Query(None, description="部门筛选"),
    position_name: str | None = Query(None, description="岗位名称筛选"),
    session: AsyncSession = Depends(get_db),
):
    """按部门和岗位查询关联的培训内容（SOP/文件）。"""
    from app.modules.hr.models import PositionTraining

    q = select(PositionTraining).where(PositionTraining.is_deleted == False)
    if department:
        q = q.where(PositionTraining.department == department)
    if position_name:
        q = q.where(PositionTraining.position_name == position_name)
    q = q.order_by(PositionTraining.department, PositionTraining.position_name, PositionTraining.sort_order)

    result = await session.execute(q)
    rows = result.scalars().all()
    data = [
        {
            "id": str(r.id),
            "position_name": r.position_name,
            "department": r.department,
            "variety": r.variety,
            "training_category": r.training_category,
            "trainer": r.trainer,
            "training_method": r.training_method,
            "sop_number": r.sop_number,
            "file_name": r.file_name,
        }
        for r in rows
    ]
    return success_response(data=data)


class PositionTrainingCreate(BaseModel):
    department: str
    position_name: str
    training_category: str
    sop_number: str | None = None
    file_name: str


@router.post("/position-trainings", summary="新建岗位培训内容")
async def create_position_training(
    payload: PositionTrainingCreate,
    session: AsyncSession = Depends(get_db),
):
    """手动新建一条岗位培训关联记录。"""
    from app.modules.hr.models import PositionTraining, SopCatalog

    pt = PositionTraining(
        department=payload.department,
        position_name=payload.position_name,
        training_category=payload.training_category,
        sop_number=payload.sop_number,
        file_name=payload.file_name,
    )
    session.add(pt)

    # 同步写入 SOP 目录
    sc = SopCatalog(
        department=payload.department,
        category=payload.training_category,
        sop_number=payload.sop_number,
        file_name=payload.file_name,
        position_name=payload.position_name,
    )
    session.add(sc)

    await session.flush()
    return success_response(data={"id": str(pt.id)}, message="创建成功", status_code=201)


@router.get("/position-trainings/departments", summary="岗位培训内容中的部门列表")
async def list_pt_departments(
    session: AsyncSession = Depends(get_db),
):
    from app.modules.hr.models import PositionTraining

    result = await session.execute(
        select(PositionTraining.department)
        .where(PositionTraining.is_deleted == False)
        .distinct()
        .order_by(PositionTraining.department)
    )
    return success_response(data=[row[0] for row in result.all()])


# ─── Team Routes ───

@router.get("/teams", summary="班组列表")
async def list_teams(
    department_id: UUID | None = Query(None, description="部门筛选"),
    keyword: str | None = Query(None, description="班组名称或编码关键词"),
    page_params: PageParams = Depends(),
    service: TeamService = Depends(get_team_service),
):
    teams, total = await service.list_teams(
        department_id=department_id,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        TeamResponse.model_validate(t).model_dump(mode="json")
        for t in teams
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/teams", summary="创建班组")
async def create_team(
    payload: TeamCreate,
    service: TeamService = Depends(get_team_service),
):
    team = await service.create_team(payload)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
        message="班组创建成功",
        status_code=201,
    )


@router.get("/teams/{team_id}", summary="班组详情")
async def get_team(
    team_id: UUID,
    service: TeamService = Depends(get_team_service),
):
    team = await service.get_team(team_id)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
    )


@router.put("/teams/{team_id}", summary="更新班组")
async def update_team(
    team_id: UUID,
    payload: TeamUpdate,
    service: TeamService = Depends(get_team_service),
):
    team = await service.update_team(team_id, payload)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
        message="班组更新成功",
    )


@router.delete("/teams/{team_id}", summary="删除班组")
async def delete_team(
    team_id: UUID,
    service: TeamService = Depends(get_team_service),
):
    await service.delete_team(team_id)
    return success_response(message="班组删除成功")


# ─── OffboardingRecord Routes ───

@router.get("/offboarding-records", summary="离职记录列表")
async def list_offboarding_records(
    employee_id: UUID | None = Query(None, description="员工ID筛选"),
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    page_params: PageParams = Depends(),
    service: OffboardingRecordService = Depends(get_offboarding_service),
):
    records, total = await service.list_records(
        employee_id=employee_id,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        OffboardingRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/offboarding-records", summary="创建离职记录")
async def create_offboarding_record(
    payload: OffboardingRecordCreate,
    service: OffboardingRecordService = Depends(get_offboarding_service),
):
    record = await service.create_record(payload)
    # 手动构建响应，避免触发未加载的 relationship
    data = {
        "id": str(record.id),
        "employee_id": str(record.employee_id),
        "offboarding_date": (
            record.offboarding_date.isoformat()
            if record.offboarding_date else None
        ),
        "offboarding_type": record.offboarding_type,
        "reason": record.reason,
        "handover_status": record.handover_status,
        "notes": record.notes,
        "created_at": record.created_at.isoformat() if record.created_at else None,
        "updated_at": record.updated_at.isoformat() if record.updated_at else None,
    }
    return success_response(
        data=data,
        message="离职记录创建成功，员工状态已更新为离职",
        status_code=201,
    )


@router.get("/offboarding-records/{record_id}", summary="离职记录详情")
async def get_offboarding_record(
    record_id: UUID,
    service: OffboardingRecordService = Depends(get_offboarding_service),
):
    record = await service.get_record(record_id)
    return success_response(
        data=OffboardingRecordResponse.model_validate(record).model_dump(mode="json"),
    )


@router.put("/offboarding-records/{record_id}", summary="更新离职记录")
async def update_offboarding_record(
    record_id: UUID,
    payload: OffboardingRecordUpdate,
    service: OffboardingRecordService = Depends(get_offboarding_service),
):
    record = await service.update_record(record_id, payload)
    return success_response(
        data=OffboardingRecordResponse.model_validate(record).model_dump(mode="json"),
        message="离职记录更新成功",
    )


@router.delete("/offboarding-records/{record_id}", summary="删除离职记录")
async def delete_offboarding_record(
    record_id: UUID,
    service: OffboardingRecordService = Depends(get_offboarding_service),
):
    await service.delete_record(record_id)
    return success_response(message="离职记录删除成功")


# ─── OnboardingRecord Routes ───

@router.get("/onboarding-records", summary="老厂入职台账列表")
async def list_onboarding_records(
    department: str | None = Query(None, description="部门筛选"),
    position: str | None = Query(None, description="岗位筛选"),
    is_employed: str | None = Query(None, description="是否在职筛选"),
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    sort_by: str = Query("hire_date", description="排序字段"),
    sort_order: str = Query("desc", description="排序方向"),
    page_params: PageParams = Depends(),
    service: OnboardingRecordService = Depends(get_onboarding_service),
):
    records, total = await service.list_records(
        department=department,
        position=position,
        is_employed=is_employed,
        keyword=keyword,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        OnboardingRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.get("/onboarding-records/{record_id}", summary="入职记录详情")
async def get_onboarding_record(
    record_id: UUID,
    service: OnboardingRecordService = Depends(get_onboarding_service),
):
    record = await service.get_record(record_id)
    return success_response(
        data=OnboardingRecordResponse.model_validate(record).model_dump(mode="json"),
    )


@router.delete("/onboarding-records/{record_id}", summary="删除入职台账记录")
async def delete_onboarding_record(
    record_id: UUID,
    service: OnboardingRecordService = Depends(get_onboarding_service),
):
    await service.delete_record(record_id)
    return success_response(message="删除成功")


# ─── DepartureRecord Routes ───

@router.get("/departure-records", summary="老厂离职台账列表")
async def list_departure_records(
    department: str | None = Query(None, description="部门筛选"),
    offboarding_type: str | None = Query(None, description="离职类型筛选"),
    keyword: str | None = Query(None, description="姓名/部门/职位关键词"),
    sort_by: str = Query("offboarding_date", description="排序字段"),
    sort_order: str = Query("desc", description="排序方向"),
    page_params: PageParams = Depends(),
    service: DepartureRecordService = Depends(get_departure_service),
):
    records, total = await service.list_records(
        department=department,
        offboarding_type=offboarding_type,
        keyword=keyword,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        DepartureRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/departure-records", summary="创建离职台账记录")
async def create_departure_record(
    payload: DepartureRecordCreate,
    service: DepartureRecordService = Depends(get_departure_service),
):
    record = await service.create_record(payload)
    return success_response(
        data=DepartureRecordResponse.model_validate(record).model_dump(mode="json"),
        message="离职台账记录创建成功",
        status_code=201,
    )


@router.get("/departure-records/{record_id}", summary="离职台账记录详情")
async def get_departure_record(
    record_id: UUID,
    service: DepartureRecordService = Depends(get_departure_service),
):
    record = await service.get_record(record_id)
    return success_response(
        data=DepartureRecordResponse.model_validate(record).model_dump(mode="json"),
    )


@router.put("/departure-records/{record_id}", summary="更新离职台账记录")
async def update_departure_record(
    record_id: UUID,
    payload: DepartureRecordUpdate,
    service: DepartureRecordService = Depends(get_departure_service),
):
    record = await service.update_record(record_id, payload)
    return success_response(
        data=DepartureRecordResponse.model_validate(record).model_dump(mode="json"),
        message="离职台账记录更新成功",
    )


@router.delete("/departure-records/{record_id}", summary="删除离职台账记录")
async def delete_departure_record(
    record_id: UUID,
    service: DepartureRecordService = Depends(get_departure_service),
):
    await service.delete_record(record_id)
    return success_response(message="离职台账记录删除成功")


# ─── TrainingLedger Routes ───

@router.get("/training-ledgers", summary="培训台账列表")
async def list_training_ledgers(
    employee_number: str | None = Query(None, description="工号筛选"),
    date_from: date | None = Query(None, description="培训日期起"),
    date_to: date | None = Query(None, description="培训日期止"),
    page_params: PageParams = Depends(),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    records, total = await service.list_records(
        employee_number=employee_number,
        date_from=date_from,
        date_to=date_to,
        page=page_params.page,
        page_size=page_params.page_size,
        sort_by="training_date",
        sort_order="asc",
    )
    data = [
        TrainingLedgerResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/training-ledgers", summary="创建培训台账记录")
async def create_training_ledger(
    payload: TrainingLedgerCreate,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    record = await service.create_record(payload)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
        message="培训台账记录创建成功",
        status_code=201,
    )


# ─── TrainingLedgerPage Routes (must be before /{record_id}) ───

@router.get("/training-ledgers/pages", summary="已创建的培训台账页面列表")
async def list_training_ledger_pages(
    service: TrainingLedgerPageService = Depends(get_training_ledger_page_service),
):
    pages_with_dept = await service.list_pages_with_department()
    data = [
        {
            "id": str(page.id),
            "employee_number": page.employee_number,
            "employee_name": page.employee_name,
            "department": dept or "未知部门",
            "created_at": page.created_at.isoformat() if page.created_at else None,
            "updated_at": page.updated_at.isoformat() if page.updated_at else None,
        }
        for page, dept in pages_with_dept
    ]
    return success_response(data=data)


@router.post("/training-ledgers/pages", summary="创建培训台账页面")
async def create_training_ledger_page(
    payload: TrainingLedgerPageCreate,
    service: TrainingLedgerPageService = Depends(get_training_ledger_page_service),
):
    page = await service.create_page(payload)
    return success_response(
        data=TrainingLedgerPageResponse(
            id=page.id,
            employee_number=page.employee_number,
            employee_name=page.employee_name,
            department=None,
            created_at=page.created_at,
            updated_at=page.updated_at,
        ).model_dump(mode="json"),
        message="培训台账页面创建成功",
        status_code=201,
    )


def _generate_training_ledger_excel(employee: dict, records: list[dict]) -> BytesIO:
    """Generate training ledger Excel based on employee training ledger format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "员工培训台账"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=16)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    ws.merge_cells("A1:G1")
    ws["A1"] = "丽珠集团福州福兴医药有限公司"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = "员工培训台账"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center_align
    ws.row_dimensions[2].height = 24

    ws["A3"] = "姓名"
    ws["A3"].font = bold_font
    ws["A3"].alignment = center_align
    ws["A3"].border = thin_border
    ws["B3"] = employee.get("name", "")
    ws["B3"].border = thin_border
    ws["C3"] = "性别"
    ws["C3"].font = bold_font
    ws["C3"].alignment = center_align
    ws["C3"].border = thin_border
    ws["D3"] = employee.get("gender", "")
    ws["D3"].border = thin_border
    ws["E3"] = "工作卡号"
    ws["E3"].font = bold_font
    ws["E3"].alignment = center_align
    ws["E3"].border = thin_border
    ws.merge_cells("F3:G3")
    ws["F3"] = employee.get("employee_number", "")
    ws["F3"].border = thin_border
    ws["G3"].border = thin_border

    ws["A4"] = "部门"
    ws["A4"].font = bold_font
    ws["A4"].alignment = center_align
    ws["A4"].border = thin_border
    ws["B4"] = employee.get("department", "")
    ws["B4"].border = thin_border
    ws["C4"] = "岗位/职务"
    ws["C4"].font = bold_font
    ws["C4"].alignment = center_align
    ws["C4"].border = thin_border
    ws["D4"] = employee.get("position", "")
    ws["D4"].border = thin_border
    ws["E4"] = "入厂时间"
    ws["E4"].font = bold_font
    ws["E4"].alignment = center_align
    ws["E4"].border = thin_border
    ws.merge_cells("F4:G4")
    ws["F4"] = employee.get("factory_entry_date") or employee.get("hire_date", "")
    ws["F4"].border = thin_border
    ws["G4"].border = thin_border

    ws["A5"] = "岗位变动"
    ws["A5"].font = bold_font
    ws["A5"].alignment = center_align
    ws["A5"].border = thin_border
    ws.merge_cells("B5:G5")
    ws["B5"] = employee.get("transfer_history", "无")
    ws["B5"].border = thin_border
    for c in range(3, 8):
        ws.cell(row=5, column=c).border = thin_border

    ws["A6"] = "记录"
    ws["A6"].font = bold_font
    ws["A6"].alignment = center_align
    ws["A6"].border = thin_border
    ws.merge_cells("B6:G6")
    ws["B6"] = ""
    ws["B6"].border = thin_border
    for c in range(3, 8):
        ws.cell(row=6, column=c).border = thin_border

    headers = ["年月日", "培训课程", "培训方式", "课时", "培训单位/培训师", "考核成绩", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[7].height = 24

    for idx, record in enumerate(records, 8):
        values = [
            record.get("training_date", ""),
            record.get("training_subject", ""),
            record.get("training_method", ""),
            record.get("duration_hours", ""),
            record.get("trainer", ""),
            record.get("assessment_result", ""),
            record.get("remarks", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align if col in (1, 3, 4, 6, 7) else left_align

    while len(records) < 12:
        row = 8 + len(records)
        for col in range(1, 8):
            ws.cell(row=row, column=col, value="").border = thin_border
        records.append({})

    footer_row = 8 + len(records)
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    ws.cell(row=footer_row, column=1, value="备注：笔试考核设置为满分100分，考试合格线为80分。")
    ws.cell(row=footer_row, column=1).alignment = left_align
    ws.cell(row=footer_row, column=1).border = thin_border
    for c in range(2, 8):
        ws.cell(row=footer_row, column=c).border = thin_border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/training-ledgers/export", summary="导出培训台账Excel")
async def export_training_ledger(
    employee_number: str = Query(..., description="员工工号"),
    ledger_service: TrainingLedgerService = Depends(get_training_ledger_service),
    employee_service: EmployeeService = Depends(get_employee_service),
):
    """根据员工数据生成并导出培训台账 Excel 文件。"""
    employee = await employee_service.get_employee_by_number(employee_number)
    if not employee:
        raise HTTPException(status_code=404, detail="未找到该员工")

    records, _ = await ledger_service.list_records(
        employee_number=employee_number,
        page=1,
        page_size=1000,
        sort_by="training_date",
        sort_order="asc",
    )

    employee_dict = EmployeeResponse.model_validate(employee).model_dump(mode="json")
    record_dicts = [
        TrainingLedgerResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]

    buffer = _generate_training_ledger_excel(employee_dict, record_dicts)
    buffer.seek(0)

    safe_name = employee.name or "unknown"
    filename = f"{safe_name}培训台账.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"
        },
    )


@router.get("/training-ledgers/{record_id}", summary="培训台账记录详情")
async def get_training_ledger(
    record_id: UUID,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    record = await service.get_record(record_id)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
    )


@router.put("/training-ledgers/{record_id}", summary="更新培训台账记录")
async def update_training_ledger(
    record_id: UUID,
    payload: TrainingLedgerUpdate,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    record = await service.update_record(record_id, payload)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
        message="培训台账记录更新成功",
    )


@router.delete("/training-ledgers/{record_id}", summary="删除培训台账记录")
async def delete_training_ledger(
    record_id: UUID,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    await service.delete_record(record_id)
    return success_response(message="培训台账记录删除成功")


# ─── AnnualTrainingPlan Routes ───

@router.post("/annual-training-plans/upload", summary="上传年度培训计划")
async def upload_annual_training_plan(
    file: UploadFile,
    service: EmployeeService = Depends(get_employee_service),
):
    """上传 Excel 年度培训计划，按年度+部门自动分类为计划项。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 格式")
    try:
        content = await file.read()
        result = await service.upload_annual_plan(content)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return success_response(data=result, message=f"新增 {result['created']} 条计划项")


@router.get("/annual-training-plans", summary="年度培训计划列表")
async def list_annual_training_plans(
    year: int | None = Query(None, description="年度筛选"),
    department: str | None = Query(None, description="部门筛选"),
    page_params: PageParams = Depends(),
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    plans, total = await service.list_plans(
        year=year,
        department=department,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        AnnualTrainingPlanResponse.model_validate(p).model_dump(mode="json")
        for p in plans
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/annual-training-plans", summary="创建年度培训计划")
async def create_annual_training_plan(
    payload: AnnualTrainingPlanCreate,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    plan = await service.create_plan(payload)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
        message="年度培训计划创建成功",
        status_code=201,
    )


@router.get("/annual-training-plans/{plan_id}", summary="年度培训计划详情")
async def get_annual_training_plan(
    plan_id: UUID,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    plan = await service.get_plan(plan_id)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
    )


@router.put("/annual-training-plans/{plan_id}", summary="更新年度培训计划")
async def update_annual_training_plan(
    plan_id: UUID,
    payload: AnnualTrainingPlanUpdate,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    plan = await service.update_plan(plan_id, payload)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
        message="年度培训计划更新成功",
    )


@router.delete("/annual-training-plans/{plan_id}", summary="删除年度培训计划")
async def delete_annual_training_plan(
    plan_id: UUID,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    await service.delete_plan(plan_id)
    return success_response(message="年度培训计划删除成功")


@router.get("/annual-training-plans/{plan_id}/items", summary="年度计划明细列表")
async def list_annual_training_plan_items(
    plan_id: UUID,
    service: AnnualTrainingPlanItemService = Depends(get_annual_training_plan_item_service),
):
    items = await service.list_items(plan_id)
    data = [
        AnnualTrainingPlanItemResponse.model_validate(i).model_dump(mode="json")
        for i in items
    ]
    return success_response(data=data)


@router.delete("/annual-training-plans/{plan_id}/items/{item_id}", summary="删除年度计划明细")
async def delete_annual_training_plan_item(
    plan_id: UUID,
    item_id: UUID,
    session: AsyncSession = Depends(get_db),
):
    """软删除一条年度计划明细。"""
    from app.modules.hr.models import AnnualTrainingPlanItem
    item = (await session.execute(
        select(AnnualTrainingPlanItem).where(
            AnnualTrainingPlanItem.id == item_id,
            AnnualTrainingPlanItem.plan_id == plan_id,
            AnnualTrainingPlanItem.is_deleted == False,
        )
    )).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "明细不存在")
    # 同时物理删除关联的培训台账记录
    if item.content_and_textbook:
        from app.modules.hr.models import TrainingLedger
        ledgers = (await session.execute(
            select(TrainingLedger).where(
                TrainingLedger.training_subject == item.content_and_textbook,
            )
        )).scalars().all()
        for ledger in ledgers:
            session.delete(ledger)
    session.delete(item)
    await session.flush()
    return success_response(message="删除成功")


@router.put("/annual-training-plans/{plan_id}/items/batch", summary="批量更新年度计划明细")
async def batch_update_annual_training_plan_items(
    plan_id: UUID,
    payload: AnnualTrainingPlanItemBatchUpdate,
    service: AnnualTrainingPlanItemService = Depends(get_annual_training_plan_item_service),
):
    items = await service.batch_update_items(plan_id, payload)
    data = [
        AnnualTrainingPlanItemResponse.model_validate(i).model_dump(mode="json")
        for i in items
    ]
    return success_response(
        data=data,
        message="年度计划明细更新成功",
    )


def _generate_annual_plan_excel(plan: dict, items: list[dict]) -> BytesIO:
    """Generate annual training plan Excel based on 7.7 template format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "年度培训计划"

    # Styles
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=16)

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 10

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{plan['year']} 年培训计划"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Department row
    ws.merge_cells("A2:I2")
    ws["A2"] = f"部门：{plan['department']}"
    ws["A2"].font = bold_font
    ws["A2"].alignment = left_align
    ws.row_dimensions[2].height = 22

    # Header row
    headers = ["序号", "培训季度及课时", "培训内容及使用教材", "培训对象",
               "授课单位及授课人", "考核方式", "培训跟踪", "确认人/日期", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[3].height = 28

    # Data rows
    for idx, item in enumerate(items, 1):
        row = 3 + idx
        quarter = item.get("month") or ""
        hours = item.get("duration_hours")
        quarter_hours = f"{quarter}\n{hours}课时" if hours else quarter

        values = [
            idx,
            quarter_hours,
            item.get("content_and_textbook") or "",
            item.get("target_audience") or "",
            item.get("position_and_count") or "",
            item.get("training_method") or "",
            item.get("tracking_status") or "",
            f"{item.get('confirmer') or ''}{' / ' + str(item.get('confirm_date')) if item.get('confirm_date') else ''}",
            item.get("remarks") or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align if col in (1, 2, 6, 7) else left_align
        ws.row_dimensions[row].height = 36

    # Pad to at least 12 rows
    while len(items) < 12:
        row = 3 + len(items) + 1
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = thin_border
        ws.row_dimensions[row].height = 36
        items.append({})

    # Footer row
    footer_row = 4 + len(items) + 1
    ws.merge_cells(f"A{footer_row}:E{footer_row}")
    ws.cell(row=footer_row, column=1, value="制表人/日期：")
    ws.cell(row=footer_row, column=1).alignment = left_align
    ws.cell(row=footer_row, column=1).border = thin_border
    for c in range(2, 6):
        ws.cell(row=footer_row, column=c).border = thin_border

    ws.merge_cells(f"F{footer_row}:I{footer_row}")
    ws.cell(row=footer_row, column=6, value="部门负责人/日期：")
    ws.cell(row=footer_row, column=6).alignment = left_align
    ws.cell(row=footer_row, column=6).border = thin_border
    for c in range(7, 10):
        ws.cell(row=footer_row, column=c).border = thin_border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.post("/annual-training-plans/complete-by-content", summary="按培训内容标记完成")
async def complete_plan_by_content(
    payload: dict = None,
    session: AsyncSession = Depends(get_db),
):
    """根据培训内容标记所有匹配的年度计划明细为已完成。"""
    from app.modules.hr.models import AnnualTrainingPlanItem
    content = payload.get("content", "") if payload else ""
    if not content:
        raise HTTPException(400, "缺少培训内容")
    items = (await session.execute(
        select(AnnualTrainingPlanItem).where(
            AnnualTrainingPlanItem.content_and_textbook == content,
            AnnualTrainingPlanItem.is_deleted == False,
        )
    )).scalars().all()
    for item in items:
        item.tracking_status = "完成"
    await session.flush()
    return success_response(data={"updated": len(items)}, message=f"已标记 {len(items)} 条为完成")


@router.get("/annual-training-plans/{plan_id}/export", summary="导出年度培训计划Excel")
async def export_annual_training_plan(
    plan_id: UUID,
    plan_service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
    item_service: AnnualTrainingPlanItemService = Depends(get_annual_training_plan_item_service),
):
    """根据年度计划数据生成并导出Excel文件（7.7年度培训计划格式）。"""
    plan = await plan_service.get_plan(plan_id)
    items = await item_service.list_items(plan_id)

    plan_dict = AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json")
    item_dicts = [
        AnnualTrainingPlanItemResponse.model_validate(i).model_dump(mode="json")
        for i in items
    ]

    buffer = _generate_annual_plan_excel(plan_dict, item_dicts)
    buffer.seek(0)

    safe_dept = plan.department.replace(" ", "_")
    filename = f"{plan.year}年度培训计划_{safe_dept}.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"
        },
    )

# ─── Trainer Routes ───


@router.post("/trainers/upload", summary="上传内训师台账")
async def upload_trainers(
    file: UploadFile,
    service: EmployeeService = Depends(get_employee_service),
):
    """上传 Excel 内训师名单，按姓名+部门自动新增或更新。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 格式")
    try:
        content = await file.read()
        result = await service.upload_trainers(content)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return success_response(data=result, message=f"新增 {result['created']}，更新 {result['updated']}")


@router.get("/trainers", summary="内训师台账列表", response_model=TrainerListResponse)
async def list_trainers(
    department: str | None = Query(None),
    keyword: str | None = Query(None),
    is_level1: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    session: AsyncSession = Depends(get_db),
):
    from app.modules.hr.models import HrTrainer
    from app.core.response import paginated_response

    query = select(HrTrainer).where(HrTrainer.is_deleted == False)
    count_q = select(func.count()).select_from(HrTrainer).where(HrTrainer.is_deleted == False)
    if department:
        query = query.where(HrTrainer.department == department)
        count_q = count_q.where(HrTrainer.department == department)
    if is_level1:
        query = query.where(HrTrainer.is_level1 == is_level1)
        count_q = count_q.where(HrTrainer.is_level1 == is_level1)
    if keyword:
        query = query.where(
            or_(HrTrainer.name.ilike(f"%{keyword}%"), HrTrainer.trainable_departments.ilike(f"%{keyword}%"))
        )
        count_q = count_q.where(
            or_(HrTrainer.name.ilike(f"%{keyword}%"), HrTrainer.trainable_departments.ilike(f"%{keyword}%"))
        )

    total = (await session.execute(count_q)).scalar() or 0
    rows = (await session.execute(query.order_by(HrTrainer.department, HrTrainer.name).offset((page - 1) * page_size).limit(page_size))).scalars().all()

    return paginated_response(
        data=[TrainerResponse.model_validate(r).model_dump(mode="json") for r in rows],
        page=page, page_size=page_size, total=total,
    )


class TrainerCreate(BaseModel):
    name: str
    department: str | None = None
    trainable_departments: str | None = None
    qualification_scope: str | None = None
    certification_date: date | None = None
    confirmation_date: date | None = None
    confirmation_reminder: date | None = None
    is_level1: str | None = None
    admin: str | None = None
    remarks: str | None = None


@router.post("/trainers", summary="新增内训师")
async def create_trainer(payload: TrainerCreate, session: AsyncSession = Depends(get_db)):
    from app.modules.hr.models import HrTrainer
    t = HrTrainer(**payload.model_dump())
    session.add(t)
    await session.flush()
    return success_response(data=TrainerResponse.model_validate(t).model_dump(mode="json"), message="创建成功", status_code=201)


@router.put("/trainers/{trainer_id}", summary="更新内训师")
async def update_trainer(trainer_id: UUID, payload: TrainerCreate, session: AsyncSession = Depends(get_db)):
    from app.modules.hr.models import HrTrainer
    t = (await session.execute(select(HrTrainer).where(HrTrainer.id == trainer_id, HrTrainer.is_deleted == False))).scalar_one_or_none()
    if not t: raise HTTPException(404, "内训师不存在")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(t, k, v)
    await session.flush()
    return success_response(data=TrainerResponse.model_validate(t).model_dump(mode="json"), message="更新成功")


@router.delete("/trainers/{trainer_id}", summary="删除内训师")
async def delete_trainer(
    trainer_id: UUID,
    session: AsyncSession = Depends(get_db),
):
    from app.modules.hr.models import HrTrainer

    result = await session.execute(
        select(HrTrainer).where(HrTrainer.id == trainer_id, HrTrainer.is_deleted == False)
    )
    trainer = result.scalar_one_or_none()
    if not trainer:
        raise HTTPException(404, "内训师不存在")
    trainer.is_deleted = True
    trainer.deleted_at = func.now()
    await session.flush()
    return success_response(message="删除成功")


@router.delete("/trainers", summary="清空内训师台账")
async def clear_trainers(
    session: AsyncSession = Depends(get_db),
):
    from app.modules.hr.models import HrTrainer
    from sqlalchemy import update as sql_update

    await session.execute(
        sql_update(HrTrainer)
        .where(HrTrainer.is_deleted == False)
        .values(is_deleted=True, deleted_at=func.now())
    )
    await session.flush()
    return success_response(message="清空成功")


# ─── SOP Catalog Routes ───

@router.delete("/sop-catalog/{item_id}", summary="删除SOP目录条目")
async def delete_sop_catalog_item(
    item_id: UUID,
    session: AsyncSession = Depends(get_db),
):
    from app.modules.hr.models import SopCatalog
    item = (await session.execute(
        select(SopCatalog).where(SopCatalog.id == item_id, SopCatalog.is_deleted == False)
    )).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "SOP条目不存在")
    item.is_deleted = True
    await session.flush()
    return success_response(message="删除成功")


@router.post("/sop-catalog/upload", summary="上传SOP目录")
async def upload_sop_catalog(
    file: UploadFile,
    service: EmployeeService = Depends(get_employee_service),
):
    """上传 Excel SOP 目录，按 SOP编号 自动新增或更新。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 格式")
    try:
        content = await file.read()
        result = await service.upload_sop_catalog(content)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return success_response(data=result, message=f"新增 {result['created']}，更新 {result['updated']}")


@router.get("/sop-catalog", summary="SOP目录列表", response_model=SopCatalogListResponse)
async def list_sop_catalog(
    department: str | None = Query(None),
    category: str | None = Query(None),
    keyword: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    session: AsyncSession = Depends(get_db),
):
    from app.modules.hr.models import SopCatalog
    from app.core.response import paginated_response

    query = select(SopCatalog).where(SopCatalog.is_deleted == False)
    count_q = select(func.count()).select_from(SopCatalog).where(SopCatalog.is_deleted == False)
    if department:
        query = query.where(SopCatalog.department == department)
        count_q = count_q.where(SopCatalog.department == department)
    if category:
        query = query.where(SopCatalog.category == category)
        count_q = count_q.where(SopCatalog.category == category)
    if keyword:
        query = query.where(SopCatalog.file_name.ilike(f"%{keyword}%"))
        count_q = count_q.where(SopCatalog.file_name.ilike(f"%{keyword}%"))

    total = (await session.execute(count_q)).scalar() or 0
    rows = (await session.execute(query.order_by(SopCatalog.category, SopCatalog.file_name).offset((page - 1) * page_size).limit(page_size))).scalars().all()

    return paginated_response(
        data=[SopCatalogResponse.model_validate(r).model_dump(mode="json") for r in rows],
        page=page, page_size=page_size, total=total,
    )


@router.get("/sop-catalog/departments", summary="SOP目录部门列表")
async def list_sop_catalog_departments(
    session: AsyncSession = Depends(get_db),
):
    """返回 SOP 目录中所有不重复的部门名称。"""
    from app.modules.hr.models import SopCatalog

    result = await session.execute(
        select(SopCatalog.department)
        .where(SopCatalog.is_deleted == False, SopCatalog.department.isnot(None))
        .distinct()
        .order_by(SopCatalog.department)
    )
    departments = [row[0] for row in result.all()]
    return success_response(data=departments)


@router.get("/sop-catalog/categories", summary="SOP目录分类列表")
async def list_sop_catalog_categories(
    department: str | None = Query(None, description="按部门筛选"),
    session: AsyncSession = Depends(get_db),
):
    """返回 SOP 目录中所有不重复的分类名称，可按部门筛选。"""
    from app.modules.hr.models import SopCatalog

    stmt = (
        select(SopCatalog.category)
        .where(SopCatalog.is_deleted == False, SopCatalog.category.isnot(None))
        .distinct()
        .order_by(SopCatalog.category)
    )
    if department:
        stmt = stmt.where(SopCatalog.department == department)
    result = await session.execute(stmt)
    categories = [row[0] for row in result.all()]
    return success_response(data=categories)


# ─── Dept Training Personnel ───

@router.get("/dept-training-personnel", summary="部门培训人员列表")
async def list_dept_training_personnel(
    department: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    session: AsyncSession = Depends(get_db),
):
    tbl = "hr.dept_training_personnel"
    cols = "id, display_dept, department, admins, dept_head, primary_trainer, created_at"
    query = f"SELECT {cols} FROM {tbl} WHERE is_deleted = false"
    count_q = f"SELECT count(*) FROM {tbl} WHERE is_deleted = false"
    if department:
        query += f" AND (display_dept = :dept OR department = :dept)"
        count_q += f" AND (display_dept = :dept OR department = :dept)"
    query += " ORDER BY display_dept LIMIT :limit OFFSET :offset"

    total = (await session.execute(text(count_q), {"dept": department} if department else {})).scalar() or 0
    rows = (await session.execute(text(query), {"dept": department, "limit": page_size, "offset": (page-1)*page_size} if department else {"limit": page_size, "offset": (page-1)*page_size})).fetchall()

    from app.core.response import paginated_response
    data = [{"id": r[0], "display_dept": r[1], "department": r[2], "admins": r[3], "dept_head": r[4], "primary_trainer": r[5], "created_at": str(r[6]) if r[6] else None} for r in rows]
    return paginated_response(data=data, page=page, page_size=page_size, total=total)


# ─── Training Evaluation (补录) ───

@router.post("/training-evaluations", summary="保存培训评估补录数据")
async def save_training_evaluation(
    payload: TrainingEvaluationInput,
    session: AsyncSession = Depends(get_db),
):
    """保存评估数据，返回完整 Word 文档（含统计字段）。"""
    from app.modules.hr.evaluation_document_generator import generate_training_evaluation
    buffer = generate_training_evaluation(payload)

    def _iter():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(payload.training_date).replace("-", "") if payload.training_date else "nodate"
    from urllib.parse import quote
    safe_fn = f"evaluation_{safe_date}.docx"
    return StreamingResponse(
        _iter(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename=\"{safe_fn}\"; filename*=utf-8''{quote('培训效果评估表_' + safe_date + '.docx')}"},
    )


@router.get("/training-evaluations/list", summary="培训评估列表")
async def list_training_evaluations(
    keyword: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_db),
):
    tbl = "hr.training_evaluations"
    cols = "id, training_content, training_date, trainer, expected_count, actual_count, excellent_count, qualified_count, unqualified_count, training_method, trainer_name, assessment_method, created_at"
    where = "WHERE is_deleted = false"
    if keyword:
        where += f" AND training_content ILIKE '%{keyword}%'"
    count = (await session.execute(text(f"SELECT count(*) FROM {tbl} {where}"))).scalar() or 0
    rows = (await session.execute(text(f"SELECT {cols} FROM {tbl} {where} ORDER BY created_at DESC LIMIT :l OFFSET :o"), {"l": page_size, "o": (page-1)*page_size})).fetchall()
    data = [{"id": r[0], "training_content": r[1], "training_date": str(r[2]) if r[2] else None, "trainer": r[3],
             "expected_count": r[4], "actual_count": r[5], "excellent_count": r[6], "qualified_count": r[7],
             "unqualified_count": r[8], "training_method": r[9], "trainer_name": r[10],
             "assessment_method": r[11], "created_at": str(r[12]) if r[12] else None} for r in rows]
    from app.core.response import paginated_response
    return paginated_response(data=data, page=page, page_size=page_size, total=count)


@router.post("/training-evaluations/upsert", summary="同步评估补录的记录（培训内容+部门+应到人数）")
async def upsert_training_evaluation(
    training_content: str = Form(...),
    department: str = Form(...),
    expected_count: int = Form(0),
    training_method: str = Form(""),
    trainer_name: str = Form(""),
    assessment_method: str = Form(""),
    session: AsyncSession = Depends(get_db),
):
    """从培训通知面板同步：按培训内容+部门 upsert。"""
    from app.modules.hr.models import TrainingLedger
    existing = (await session.execute(
        text("SELECT id FROM hr.training_evaluations WHERE training_content = :c AND department = :d AND is_deleted = false"),
        {"c": training_content, "d": department}
    )).fetchone()
    if existing:
        await session.execute(
            text("UPDATE hr.training_evaluations SET expected_count = :n, training_method = :tm, trainer_name = :tn, assessment_method = :am, updated_at = now() WHERE id = :id"),
            {"n": expected_count, "id": existing[0], "tm": training_method, "tn": trainer_name, "am": assessment_method}
        )
        msg = "updated"
    else:
        await session.execute(
            text("INSERT INTO hr.training_evaluations (id, training_content, department, expected_count, training_method, trainer_name, assessment_method) VALUES (gen_random_uuid(), :c, :d, :n, :tm, :tn, :am)"),
            {"c": training_content, "d": department, "n": expected_count, "tm": training_method, "tn": trainer_name, "am": assessment_method}
        )
        msg = "created"
    await session.commit()
    return success_response(data={"status": msg}, message="同步成功")


@router.delete("/training-evaluations/{eval_id}", summary="删除评估补录记录")
async def delete_training_evaluation(
    eval_id: UUID,
    session: AsyncSession = Depends(get_db),
):
    await session.execute(text("DELETE FROM hr.training_evaluations WHERE id = :id"), {"id": eval_id})
    await session.commit()
    return success_response(message="删除成功")


@router.get("/training-evaluations/pending", summary="待评估的培训列表")
async def list_pending_evaluations(
    keyword: str | None = Query(None),
    session: AsyncSession = Depends(get_db),
):
    """Return trainings from the annual plan that haven't been evaluated yet."""
    rows = (await session.execute(text("""
        SELECT DISTINCT ON (p.content_and_textbook, pl.department)
               p.id, p.content_and_textbook, p.target_audience, p.training_method,
               p.training_hours, p.remarks, pl.department, pl.year,
               p.trainee_count
        FROM hr.annual_training_plan_items p
        JOIN hr.annual_training_plans pl ON pl.id = p.plan_id
        WHERE p.is_deleted = false
        AND NOT EXISTS (
            SELECT 1 FROM hr.training_evaluations e
            WHERE e.is_deleted = false
            AND e.training_content = p.content_and_textbook
            AND e.department = pl.department
        )
        ORDER BY p.content_and_textbook, pl.department, pl.year DESC
        LIMIT 200
    """))).fetchall()

    data = [{"id": str(r[0]), "content": r[1], "audience": r[2], "method": r[3],
             "hours": r[4], "remarks": r[5], "department": r[6], "year": r[7],
             "expected_count": r[8]} for r in rows]
    return success_response(data=data)


router.include_router(analysis_router)
