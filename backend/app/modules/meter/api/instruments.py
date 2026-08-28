"""标准计量器具相关 API 端点（含总览与台账导入）。"""

from __future__ import annotations

import logging
from typing import Any, cast
from uuid import UUID

from fastapi import Depends, Query, Request
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.response import paginated_response, success_response
from app.modules.meter import repository as repo
from app.modules.meter import service
from app.modules.meter.api._helpers import _build_report_items
from app.modules.meter.api._router import router
from app.modules.meter.schemas import (
    BatchCreateRequest,
    BatchCreateResult,
    BatchDeleteRequest,
    DateStatsResponse,
    ExportReportRequest,
    InstrumentCreate,
    InstrumentFilter,
    InstrumentFilterOptions,
    InstrumentListResponse,
    InstrumentResponse,
    InstrumentUpdate,
    LedgerImportResult,
    MeterOverviewResponse,
    _normalize_department,
)

logger = logging.getLogger(__name__)


@router.get("/overview", summary="仪表总览统计")
async def get_meter_overview(
    source: str = Query(default="instrument", pattern="^(instrument|gas_detector)$", description="数据源"),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    stats = await service.get_meter_overview(db, source)
    return success_response(MeterOverviewResponse(**stats).model_dump(mode="json"))



# ═══════════════════════════════════════════
# 标准计量器具
# ═══════════════════════════════════════════


@router.get("/instruments", summary="标准计量器具列表")
async def list_instruments(
    filters: InstrumentFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    records, total = await service.list_instruments(db, filters)

    # 批量查询报告数量
    report_counts = await repo.count_reports_by_instrument_ids(db, [r.id for r in records])

    items: list[dict[str, Any]] = []
    for r in records:
        anomaly = r.anomaly_flags or {}
        items.append(
            InstrumentListResponse(
                id=str(r.id),
                department=r.department,
                asset_number=r.asset_number,
                instrument_name=r.instrument_name,
                model_spec=r.model_spec,
                measurement_range=r.measurement_range,
                accuracy_grade=r.accuracy_grade,
                serial_number=r.serial_number,
                calibration_cycle_months=r.calibration_cycle_months,
                color_marking=r.color_marking,
                location=r.location,
                manufacturer=r.manufacturer,
                status=service.compute_status(r.status, r.next_calibration_date),
                calibration_date=r.calibration_date,
                calibration_unit=r.calibration_unit,
                calibration_result=r.calibration_result,
                next_calibration_date=r.next_calibration_date,
                has_anomaly=bool(anomaly),
                report_count=report_counts.get(r.id, 0),
                remark=r.remark,
                updated_at=r.updated_at,
            ).model_dump(mode="json")
        )

    return paginated_response(
        data=items,
        page=filters.page,
        page_size=filters.page_size,
        total=total,
    )



@router.get("/instruments/filter-options", summary="获取标准计量器具筛选选项")
async def get_instrument_filter_options(
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    options = await service.get_instrument_filter_options(db)
    if "department" in options:
        options["department"] = [d for d in (_normalize_department(x) for x in options["department"]) if d is not None]
    return success_response(InstrumentFilterOptions(**options).model_dump(mode="json"))



@router.get("/instruments/date-stats", summary="标准计量器具日期聚合统计")
async def get_instrument_date_stats(
    field: str = Query(default="calibration_date", pattern="^(calibration_date|next_calibration_date)$", description="统计的日期字段"),
    filters: InstrumentFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    stats = await service.get_instrument_date_stats(db, filters, field)
    return success_response(DateStatsResponse(**stats).model_dump(mode="json"))



@router.get("/instruments/export", summary="导出标准计量器具为 CSV")
async def export_instruments(
    filters: InstrumentFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    import csv
    import io

    # 导出全量数据（最多 20000 条），绕过 page_size 上限
    object.__setattr__(filters, 'page_size', 20000)
    records, _ = await service.list_instruments(db, filters)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "资产编号", "器具名称", "型号规格", "测量范围", "精度等级",
        "出厂编号", "检定周期(月)", "使用地点", "制造商", "状态",
        "检定日期", "检定单位", "检定结论", "下次检定日期", "部门",
    ])
    for r in records:
        writer.writerow([
            r.asset_number, r.instrument_name, r.model_spec, r.measurement_range,
            r.accuracy_grade, r.serial_number, r.calibration_cycle_months,
            r.location, r.manufacturer, service.compute_status(r.status, r.next_calibration_date),
            r.calibration_date.isoformat() if r.calibration_date else "",
            r.calibration_unit, r.calibration_result,
            r.next_calibration_date.isoformat() if r.next_calibration_date else "",
            _normalize_department(r.department),
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=instruments_export.csv"},
    )



@router.get("/instruments/export-excel", summary="导出标准计量器具为 Excel")
async def export_instruments_excel(
    filters: InstrumentFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    import io as io_mod

    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Font,
        PatternFill,
    )

    # 导出全量数据，绕过 page_size 上限
    object.__setattr__(filters, 'page_size', 100000)
    records, _ = await service.list_instruments(db, filters)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "标准计量器具台账"

    headers = [
        "资产编号", "器具名称", "型号规格", "测量范围", "精度等级",
        "出厂编号", "检定周期(月)", "使用地点", "制造商", "状态",
        "检定日期", "检定单位", "检定结论", "下次检定日期", "部门",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, r in enumerate(records, 2):
        values = [
            r.asset_number, r.instrument_name, r.model_spec, r.measurement_range,
            r.accuracy_grade, r.serial_number, r.calibration_cycle_months,
            r.location, r.manufacturer, service.compute_status(r.status, r.next_calibration_date),
            r.calibration_date.isoformat() if r.calibration_date else "",
            r.calibration_unit, r.calibration_result,
            r.next_calibration_date.isoformat() if r.next_calibration_date else "",
            _normalize_department(r.department),
        ]
        for col_idx, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)  # type: ignore[union-attr]

    buf = io_mod.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=instruments_export.xlsx"},
    )



@router.get("/instruments/ids", summary="获取筛选条件下所有标准计量器具 ID（用于跨页全选）")
async def get_instrument_ids(
    filters: InstrumentFilter = Depends(),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    ids = await service.get_all_instrument_ids(db, filters)
    return success_response([str(i) for i in ids])



@router.get("/instruments/{instrument_id}", summary="标准计量器具详情")
async def get_instrument(
    instrument_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    obj = await service.get_instrument(db, instrument_id)
    reports = _build_report_items(obj.reports if hasattr(obj, 'reports') else [])
    return success_response(
        InstrumentResponse(
            id=str(obj.id),
            asset_number=obj.asset_number,
            instrument_name=obj.instrument_name,
            model_spec=obj.model_spec,
            measurement_range=obj.measurement_range,
            accuracy_grade=obj.accuracy_grade,
            serial_number=obj.serial_number,
            calibration_cycle_months=obj.calibration_cycle_months,
            location=obj.location,
            manufacturer=obj.manufacturer,
            status=service.compute_status(obj.status, obj.next_calibration_date),
            color_marking=obj.color_marking,
            calibration_date=obj.calibration_date,
            calibration_unit=obj.calibration_unit,
            calibration_result=obj.calibration_result,
            next_calibration_date=obj.next_calibration_date,
            department=obj.department,
            sheet_name=obj.sheet_name,
            anomaly_flags=obj.anomaly_flags,
            is_deleted=obj.is_deleted,
            created_at=obj.created_at,
            updated_at=obj.updated_at,
            reports=cast(Any, reports),
        ).model_dump(mode="json")
    )



@router.post("/instruments", summary="新增标准计量器具")
async def create_instrument(
    data: InstrumentCreate,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    record = await service.create_instrument(db, data)
    return success_response(
        InstrumentResponse(
            id=str(record.id),
            asset_number=record.asset_number,
            instrument_name=record.instrument_name,
            model_spec=record.model_spec,
            measurement_range=record.measurement_range,
            accuracy_grade=record.accuracy_grade,
            serial_number=record.serial_number,
            calibration_cycle_months=record.calibration_cycle_months,
            location=record.location,
            manufacturer=record.manufacturer,
            status=service.compute_status(record.status, record.next_calibration_date),
            color_marking=record.color_marking,
            calibration_date=record.calibration_date,
            calibration_unit=record.calibration_unit,
            calibration_result=record.calibration_result,
            next_calibration_date=record.next_calibration_date,
            department=record.department,
            sheet_name=record.sheet_name,
            anomaly_flags=record.anomaly_flags,
            is_deleted=record.is_deleted,
            created_at=record.created_at,
            updated_at=record.updated_at,
            reports=[],
        ).model_dump(mode="json"),
        status_code=201,
    )



@router.post("/instruments/batch", summary="批量新增标准计量器具（单次最多 200 条）")
async def batch_create_instruments(
    body: dict[str, Any],
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    req = BatchCreateRequest(**body)
    result = await service.batch_create_instruments(
        db, [item.model_dump() for item in req.items]
    )
    return success_response(
        BatchCreateResult(**result).model_dump(mode="json"),
        status_code=201,
    )



# ═══════════════════════════════════════════
# Excel 台账导入
# ═══════════════════════════════════════════


@router.post("/instruments/import-ledger", summary="导入标准计量器具台账Excel（按资产编号更新）")
async def import_instrument_ledger(
    request: Request,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    """上传计量器具台账 Excel，按资产编号 upsert 标准计量器具数据。

    - 资产编号命中旧记录 → 更新字段（保留 id，报告关联不断链）
    - 新编号 → 插入；文件中未出现的旧记录 → 软删除
    支持 .et (WPS) 和 .xlsx 格式，文件限制 50MB。
    处理所有 sheet（跳过探测器 sheet）。
    """
    from app.core.config import get_settings

    settings_obj = get_settings()
    form = await request.form(max_part_size=settings_obj.MAX_UPLOAD_SIZE_MB * 1024 * 1024)

    file = form.get("file")
    file = cast(Any, file)
    if file is None or not hasattr(file, "read"):
        return JSONResponse(status_code=400, content={"code": 400, "message": "缺少 file 参数"})

    max_size = 50 * 1024 * 1024
    filename = (file.filename if hasattr(file, "filename") else None) or "unknown"

    try:
        file_data = await file.read()
        if len(file_data) > max_size:
            return JSONResponse(
                status_code=400,
                content={"code": 400, "message": "文件大小超过 50MB 限制"},
            )

        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in ("et", "xlsx", "xls"):
            return JSONResponse(
                status_code=400,
                content={"code": 400, "message": "不支持的文件格式，请上传 .et 或 .xlsx 文件"},
            )

        result = await service.import_instrument_ledger(db, file_data, filename)
    except ValueError as e:
        return JSONResponse(
            status_code=400,
            content={"code": 400, "message": str(e)},
        )
    except Exception as e:
        logger.exception("import_instrument_ledger failed")
        return JSONResponse(
            status_code=500,
            content={"code": 500, "message": f"导入失败: {e}"},
        )

    return success_response(
        LedgerImportResult(**result).model_dump(mode="json"),
        status_code=201,
    )



@router.put("/instruments/{instrument_id}", summary="更新标准计量器具")
async def update_instrument(
    instrument_id: UUID,
    data: InstrumentUpdate,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    record = await service.update_instrument(db, instrument_id, data)
    reports = _build_report_items(record.reports if hasattr(record, 'reports') else [])
    return success_response(
        InstrumentResponse(
            id=str(record.id),
            asset_number=record.asset_number,
            instrument_name=record.instrument_name,
            model_spec=record.model_spec,
            measurement_range=record.measurement_range,
            accuracy_grade=record.accuracy_grade,
            serial_number=record.serial_number,
            calibration_cycle_months=record.calibration_cycle_months,
            location=record.location,
            manufacturer=record.manufacturer,
            status=service.compute_status(record.status, record.next_calibration_date),
            color_marking=record.color_marking,
            calibration_date=record.calibration_date,
            calibration_unit=record.calibration_unit,
            calibration_result=record.calibration_result,
            next_calibration_date=record.next_calibration_date,
            department=record.department,
            sheet_name=record.sheet_name,
            anomaly_flags=record.anomaly_flags,
            is_deleted=record.is_deleted,
            created_at=record.created_at,
            updated_at=record.updated_at,
            reports=cast(Any, reports),
        ).model_dump(mode="json")
    )



@router.delete("/instruments/{instrument_id}", summary="删除标准计量器具（软删除）")
async def delete_instrument(
    instrument_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    await service.delete_instrument(db, instrument_id)
    return success_response(message="删除成功")



@router.post("/instruments/batch-delete", summary="批量删除标准计量器具（软删除）")
async def batch_delete_instruments(
    body: BatchDeleteRequest,
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    ids = [UUID(i) for i in body.ids]
    deleted_count = await service.batch_delete_instruments(db, ids)
    return success_response({"deleted_count": deleted_count}, message=f"成功删除 {deleted_count} 条记录")



@router.post("/instruments/export-reports", summary="批量导出标准计量器具最新报告 ZIP（单次最多 200 份）")
async def export_instrument_reports(
    body: ExportReportRequest,
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    ids = [UUID(i) for i in body.ids]
    zip_data, filename, count = await service.export_instrument_reports(db, ids)
    return StreamingResponse(
        iter([zip_data]),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Report-Count": str(count),
        },
    )



@router.get("/departments/instruments", summary="获取标准计量器具部门列表")
async def list_instrument_departments(
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    depts = await service.get_instrument_departments(db)
    return success_response([_normalize_department(d) for d in depts])
