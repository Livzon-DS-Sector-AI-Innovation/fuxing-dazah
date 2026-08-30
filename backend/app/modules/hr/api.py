import json
from datetime import date
from io import BytesIO
from typing import Any
from urllib.parse import quote
from uuid import UUID

from fastapi import Body, Depends, Form, HTTPException, Query, Response, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel
from sqlalchemy import and_, func, or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.exceptions import ForbiddenException
from app.core.response import paginated_response, success_response
from app.modules.hr.analysis_api import router as analysis_router
from app.modules.hr.candidate_routes import router as candidate_router
from app.modules.hr.deps import (
    HrAccessContext,
    check_sensitive_permission,
    get_hr_scope,
    require_hr_basic,
)
from app.modules.hr.document_generator import generate_onboarding_training_record
from app.modules.hr.evaluation_document_generator import generate_training_evaluation
from app.modules.hr.interview_routes import router as interview_router
from app.modules.hr.job_requirement_routes import router as job_requirement_router
from app.modules.hr.notification_document_generator import (
    generate_training_notification,
)
from app.modules.hr.onboarding_evaluation_document_generator import (
    generate_onboarding_evaluation,
)
from app.modules.hr.prejob_document_generator import generate_prejob_training_plan
from app.modules.hr.repository import PerformanceEvaluationRepository
from app.modules.hr.schemas import (
    AnnualTrainingPlanCreate,
    AnnualTrainingPlanItemBatchUpdate,
    AnnualTrainingPlanItemResponse,
    AnnualTrainingPlanResponse,
    AnnualTrainingPlanUpdate,
    CategoryScoreBatchInput,
    DepartmentCreate,
    DepartmentResponse,
    DepartmentUpdate,
    DepartureRecordCreate,
    DepartureRecordResponse,
    DepartureRecordUpdate,
    DeptTrainingPersonnelResponse,
    EmployeeCreate,
    EmployeeResponse,
    EmployeeUpdate,
    OffboardingRecordCreate,
    OffboardingRecordResponse,
    OffboardingRecordUpdate,
    OnboardingEvaluationInput,
    OnboardingRecordResponse,
    PerformanceCategoryCreate,
    PerformanceCategoryUpdate,
    PerformanceEvaluationCreate,
    PerformanceEvaluationUpdate,
    PerformanceListParams,
    SopCatalogListResponse,
    SopCatalogResponse,
    TeamCreate,
    TeamResponse,
    TeamUpdate,
    TrainerListResponse,
    TrainerResponse,
    TrainingEvaluationInput,
    TrainingLedgerCreate,
    TrainingLedgerPageCreate,
    TrainingLedgerPageResponse,
    TrainingLedgerResponse,
    TrainingLedgerUpdate,
    TrainingNotificationInput,
    TrainingSignInSheetInput,
    TransferCreate,
    _mask_id_card,
    mask_sensitive_fields,
)
from app.modules.hr.service import (
    AnnualTrainingPlanItemService,
    AnnualTrainingPlanService,
    DepartmentService,
    DepartureRecordService,
    EmployeeService,
    OffboardingRecordService,
    OnboardingRecordService,
    PerformanceEvaluationService,
    TeamService,
    TrainingLedgerPageService,
    TrainingLedgerService,
)
from app.modules.hr.signin_document_generator import generate_training_sign_in_sheet
from app.modules.hr.sop_training_routes import router as sop_training_router
from app.modules.hr.system_settings_routes import router as system_settings_router
from app.modules.hr.title_review.routes import router as title_review_router
from app.modules.hr.user_department_access_routes import router as uda_router
from app.shared.module_api import create_module_router
from app.shared.module_registry import MODULES_BY_CODE
from app.shared.schemas import PageParams

router = create_module_router(
    MODULES_BY_CODE["hr"],
    dependencies=[Depends(require_hr_basic)],
)
router.include_router(system_settings_router)
router.include_router(uda_router)
router.include_router(candidate_router)
router.include_router(interview_router)
router.include_router(job_requirement_router)
router.include_router(analysis_router)
router.include_router(sop_training_router)
router.include_router(title_review_router)


def get_employee_service(session: AsyncSession = Depends(get_db)) -> EmployeeService:
    return EmployeeService(session)


def get_department_service(
    session: AsyncSession = Depends(get_db),
) -> DepartmentService:
    return DepartmentService(session)


def get_offboarding_service(
    session: AsyncSession = Depends(get_db),
) -> OffboardingRecordService:
    return OffboardingRecordService(session)


def get_team_service(
    session: AsyncSession = Depends(get_db),
) -> TeamService:
    return TeamService(session)


def get_onboarding_service(
    session: AsyncSession = Depends(get_db),
) -> OnboardingRecordService:
    return OnboardingRecordService(session)


def get_departure_service(
    session: AsyncSession = Depends(get_db),
) -> DepartureRecordService:
    return DepartureRecordService(session)


def get_training_ledger_service(
    session: AsyncSession = Depends(get_db),
) -> TrainingLedgerService:
    return TrainingLedgerService(session)


def get_training_ledger_page_service(
    session: AsyncSession = Depends(get_db),
) -> TrainingLedgerPageService:
    return TrainingLedgerPageService(session)


def get_annual_training_plan_service(
    session: AsyncSession = Depends(get_db),
) -> AnnualTrainingPlanService:
    return AnnualTrainingPlanService(session)


def get_annual_training_plan_item_service(
    session: AsyncSession = Depends(get_db),
) -> AnnualTrainingPlanItemService:
    return AnnualTrainingPlanItemService(session)


async def _ensure_ledger_access(record, hr_scope: HrAccessContext, session: AsyncSession) -> None:
    """按台账工号反查员工并校验数据范围，越界抛 403。"""
    if hr_scope.is_unrestricted:
        return
    from app.modules.hr.repository import EmployeeRepository
    emp = await EmployeeRepository(session).get_by_employee_number(record.employee_number)
    if emp is None:
        raise ForbiddenException("数据范围限制：无法确认该记录所属员工")
    hr_scope.ensure_can_access_employee(emp)


def _ensure_qa_assessment_access(hr_scope: HrAccessContext, department: str | None) -> None:
    """考核场次数据范围校验：受限用户仅可访问授权部门场次（self_only 无对应场次）。"""
    if hr_scope.is_unrestricted:
        return
    if not hr_scope.scoped_departments:
        raise ForbiddenException("数据范围限制：仅可访问本人相关数据")
    if department not in hr_scope.scoped_departments:
        raise ForbiddenException("数据范围限制：仅可访问本部门考核场次")


# ─── Employee Routes ───

@router.get("/employees", summary="员工列表")
async def list_employees(
    department: str | None = Query(None, description="部门筛选"),
    status: str | None = Query("在职", description="状态筛选，默认仅显示在职"),
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    include_uncategorized: bool = Query(False, description="是否纳入「未分类」人员（按实际部门归属）"),
    page_params: PageParams = Depends(),
    service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
    has_sensitive: bool = Depends(check_sensitive_permission),
    session: AsyncSession = Depends(get_db),
):
    # 数据范围限制：非全部权限时强制收敛，忽略前端传入的 department 参数
    scoped_departments: set[str] | None = None
    if not hr_scope.is_unrestricted:
        if hr_scope.scoped_departments:
            if len(hr_scope.scoped_departments) == 1:
                department = next(iter(hr_scope.scoped_departments))
            else:
                # 多部门：SQL 层过滤（先过滤后分页，页内容不再稀疏/为空）
                scoped_departments = set(hr_scope.scoped_departments)
        else:
            keyword = hr_scope.employee_number
    employees, total = await service.list_employees(
        department=department,
        status=status,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
        include_uncategorized=include_uncategorized,
        departments=scoped_departments,
    )
    if not hr_scope.is_unrestricted and not hr_scope.scoped_departments:
        # self_only：keyword 为前缀匹配，再按工号精确收敛到本人
        employees = [e for e in employees if e.employee_number == hr_scope.employee_number]
        total = len(employees)
    data = [
        mask_sensitive_fields(
            EmployeeResponse.model_validate(e).model_dump(mode="json"),
            has_sensitive,
        )
        for e in employees
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/employees", summary="创建员工")
async def create_employee(
    payload: EmployeeCreate,
    service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
    has_sensitive: bool = Depends(check_sensitive_permission),
):
    hr_scope.ensure_dept_writable([payload.department])
    employee = await service.create_employee(payload)
    return success_response(
        data=mask_sensitive_fields(
            EmployeeResponse.model_validate(employee).model_dump(mode="json"), has_sensitive
        ),
        message="员工创建成功",
        status_code=201,
    )


@router.post("/employees/upload", summary="上传人员名单")
async def upload_employees(
    file: UploadFile,
    service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """上传 Excel 人员名单，按工号自动新增或更新（受限用户仅可操作授权部门行）。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 格式")
    try:
        content = await file.read()
        allowed_departments = None if hr_scope.is_unrestricted else set(hr_scope.scoped_departments)
        result = await service.upload_employees(content, allowed_departments=allowed_departments)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return success_response(data=result, message=f"新增 {result['created']}，更新 {result['updated']}")


@router.get("/roster", summary="下载花名册")
async def download_roster(
    department: str | None = Query(None),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """按部门下载员工花名册 Excel。"""
    from app.modules.hr.models import Employee
    from app.modules.hr.roster_generator import generate_roster_sync
    department = hr_scope.resolve_export_department(department)
    r = await session.execute(
        select(Employee).where(Employee.is_deleted == False, Employee.status != "离职").order_by(Employee.department, Employee.employee_number)
    )
    employees = [(e.name, e.department, e.gender or "", e.education or "", e.hire_date, e.status) for e in r.scalars().all()]
    if department:
        employees = [e for e in employees if e[1] == department]
    elif not hr_scope.is_unrestricted and hr_scope.scoped_departments:
        employees = [e for e in employees if e[1] in hr_scope.scoped_departments]
    buffer = generate_roster_sync(employees, department)
    filename = f"花名册_{department or '全部'}.docx"
    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{quote(filename)}"}
    )


@router.get("/training-registration", summary="下载个人培训登记表")
async def download_training_registration(
    department: str | None = Query(None),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """按部门下载个人培训登记表，整个部门合并为一个 docx 文件。"""
    from app.modules.hr.models import Employee
    from app.modules.hr.training_registration_generator import (
        generate_training_registration_sync,
    )

    department = hr_scope.resolve_export_department(department)
    stmt = select(Employee).where(
        Employee.is_deleted == False,
        Employee.status != "离职",
    ).order_by(Employee.department, Employee.employee_number)
    if department:
        stmt = stmt.where(Employee.department == department)
    elif not hr_scope.is_unrestricted and hr_scope.scoped_departments:
        stmt = stmt.where(Employee.department.in_(hr_scope.scoped_departments))

    r = await session.execute(stmt)
    records = []
    for e in r.scalars().all():
        records.append({
            "姓名": e.name or "",
            "性别": e.gender or "",
            "体现部门": e.department or e.actual_department or "",
            "体现岗位": e.position or e.job_category or "未设置",
            "学历": e.education or "",
            "毕业院校": e.school or "",
            "专业": e.major or "",
            "证书": "",
            "毕业时间": str(e.graduation_date) if e.graduation_date else "",
            "入职日期": str(e.hire_date) if e.hire_date else "",
        })

    buffer = generate_training_registration_sync(records)
    filename = f"个人培训登记表_{department or '全部'}.docx"
    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{quote(filename)}"}
    )


@router.get("/employees/training-candidates", summary="待培训人员列表")
async def training_candidates(
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """返回入职台账中在职的员工列表，同时关联员工表获取异动等完整信息。"""
    from app.modules.hr.models import Employee, OnboardingRecord, TransferRecord
    stmt = select(OnboardingRecord).where(
        OnboardingRecord.is_deleted == False,
        OnboardingRecord.is_employed == "是",
    )
    # 数据范围受限时强制收敛到本部门/本人
    # （OnboardingRecord 无 actual_department 列，不能用 apply_list_scope 的未分类分支）
    if not hr_scope.is_unrestricted:
        if hr_scope.data_scope == "self_only":
            stmt = stmt.where(OnboardingRecord.employee_number == hr_scope.employee_number)
        elif hr_scope.scoped_departments:
            stmt = stmt.where(OnboardingRecord.department.in_(hr_scope.scoped_departments))
        else:
            raise ForbiddenException("数据范围限制：无法确定您的部门，请联系管理员")
    if keyword:
        stmt = stmt.where(
            OnboardingRecord.name.ilike(f"{keyword}%")
            | OnboardingRecord.employee_number.ilike(f"{keyword}%")
        )
    stmt = stmt.order_by(OnboardingRecord.hire_date.desc().nulls_last())
    r = await session.execute(stmt)
    records = r.scalars().all()
    # 批量查询关联的 Employee 记录和异动记录
    emp_numbers = [rec.employee_number for rec in records if rec.employee_number]
    emp_map = {}
    transfer_map: dict[str, list] = {}
    if emp_numbers:
        emp_rows = (await session.execute(
            select(Employee).where(Employee.employee_number.in_(emp_numbers), Employee.is_deleted == False)
        )).scalars().all()
        emp_map = {e.employee_number: e for e in emp_rows}
        emp_ids = [e.id for e in emp_rows]
        if emp_ids:
            emp_id_map = {e.id: e.employee_number for e in emp_rows}
            t_rows = (await session.execute(
                select(TransferRecord).where(
                    TransferRecord.employee_id.in_(emp_ids),
                    TransferRecord.is_deleted == False,
                ).order_by(TransferRecord.effective_date.desc())
            )).scalars().all()
            for t in t_rows:
                en = emp_id_map.get(t.employee_id)
                if en:
                    transfer_map.setdefault(en, []).append({
                            "id": str(t.id), "transfer_type": t.transfer_type,
                            "from_department": t.from_department, "to_department": t.to_department,
                            "from_position": t.from_position, "to_position": t.to_position,
                            "effective_date": str(t.effective_date) if t.effective_date else None,
                            "reason": t.reason,
                        })
    return success_response(data=[
        {
            "id": str(rec.id),
            "employee_number": rec.employee_number,
            "name": rec.name,
            "department": rec.department,
            "position": rec.position,
            "hire_date": str(rec.hire_date) if rec.hire_date else None,
            "education": rec.education,
            "school": rec.school,
            "graduation_date": str(rec.graduation_date) if rec.graduation_date else None,
            "source": rec.source or "新入职",
            "transfers": transfer_map.get(rec.employee_number, []),
            "employee_id": str(emp_map[rec.employee_number].id) if rec.employee_number in emp_map else None,
        }
        for rec in records
    ])


@router.get("/employees/by-number/{employee_number}", summary="根据工号查询员工")
async def get_employee_by_number(
    employee_number: str,
    service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
    has_sensitive: bool = Depends(check_sensitive_permission),
):
    employee = await service.get_employee_by_number(employee_number)
    hr_scope.ensure_can_access_employee(employee)
    return success_response(
        data=mask_sensitive_fields(EmployeeResponse.model_validate(employee).model_dump(mode="json"), has_sensitive),
    )


def _build_export_columns() -> list[tuple[str, str]]:
    """按导入模板 _UPLOAD_COLUMN_MAP 的列头顺序构建导出列定义（去重取首个中文名）。

    每项为 (表头, 数据库字段名)。
    """
    from app.modules.hr.service import EmployeeService

    col_map = EmployeeService._UPLOAD_COLUMN_MAP
    seen: set[str] = set()
    columns: list[tuple[str, str]] = []

    # 去重：同一 DB 字段有多个中文别名时，取第一个
    for header, field in col_map.items():
        if field in seen:
            continue
        seen.add(field)
        columns.append((header, field))

    return columns


def _fmt_date(val) -> str:
    """日期 → YYYY-MM-DD 字符串，None → ""。"""
    if val is None:
        return ""
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return str(val)


def _fmt_list(val) -> str:
    """list / JSON 数组 → 逗号拼接字符串。"""
    if val is None:
        return ""
    if isinstance(val, list):
        return ",".join(str(v) for v in val)
    return str(val)


async def _ensure_record_in_scope(hr_scope: HrAccessContext, record: Any) -> None:
    """离职记录数据范围校验：记录按部门归属，越界抛 403。"""
    if hr_scope.is_unrestricted:
        return
    hr_scope.ensure_dept_writable([getattr(record, "department", None)])


def _xlsx_response(wb, filename: str) -> StreamingResponse:
    """openpyxl 工作簿 → xlsx 下载响应（统一保存/媒体类型/文件名头）。"""
    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/employees/export", summary="导出员工档案（Excel）")
async def export_employees(
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """导出员工档案为 Excel 文件（按数据范围限制），列头与导入模板一致（身份证已脱敏）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    from app.modules.hr.models import Employee as EmpModel

    stmt = select(EmpModel).where(EmpModel.is_deleted == False)  # noqa: E712
    stmt = hr_scope.apply_list_scope(stmt, EmpModel)
    employees = (await session.execute(stmt)).scalars().all()

    export_cols = _build_export_columns()

    wb = Workbook()
    ws = wb.active
    ws.title = "员工档案"

    # 写表头
    for col_idx, (header, _) in enumerate(export_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 写数据行
    for row_idx, emp in enumerate(employees, 2):
        for col_idx, (_, field_name) in enumerate(export_cols, 1):
            val = getattr(emp, field_name, None)

            # 特殊格式化
            if field_name == "id_card":
                val = _mask_id_card(val)
            elif field_name == "id_card_address":
                val = _mask_id_card(val)
            elif field_name in ("qualifications", "remarks"):
                val = _fmt_list(val)
            elif field_name == "_birth_date":
                # 出生年月：用 birth_year/birth_month/birth_day 拼装
                year = getattr(emp, "birth_year", None)
                month = getattr(emp, "birth_month", None)
                day = getattr(emp, "birth_day", None)
                if year and month and day:
                    val = f"{year}-{month:02d}-{day:02d}"
                elif year and month:
                    val = f"{year}-{month:02d}"
                elif year:
                    val = str(year)
                else:
                    val = ""
            elif isinstance(val, date):
                val = _fmt_date(val)
            elif isinstance(val, list):
                val = _fmt_list(val)

            cell_val = str(val) if val is not None else ""
            ws.cell(row=row_idx, column=col_idx, value=cell_val)

    filename = f"employee_export_{date.today().isoformat()}.xlsx"
    return _xlsx_response(wb, filename)


@router.get("/dashboard/stats", summary="人事仪表盘统计数据")
async def get_dashboard_stats(
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """返回仪表盘所有统计数据：KPI汇总、学历分布、部门人数、入职统计、离职统计。"""
    from calendar import month_abbr

    from sqlalchemy import and_ as sa_and
    from sqlalchemy import or_

    from app.modules.hr.models import Employee as EmpModel

    def _scoped():
        """数据范围过滤（未分类按实际部门归属），unrestricted 返回 []。

        self_only：按工号过滤到本人（与 apply_list_scope 语义一致）。
        部门范围解析失败时 fail-closed 抛 403，而不是放开成全公司数据。
        """
        if hr_scope.is_unrestricted:
            return []
        if hr_scope.data_scope == "self_only":
            return [EmpModel.employee_number == hr_scope.employee_number]
        if not hr_scope.scoped_departments:
            raise ForbiddenException("数据范围限制：无法确定您的授权部门，请联系管理员")
        return [or_(
            EmpModel.department.in_(hr_scope.scoped_departments),
            sa_and(EmpModel.department == "未分类", EmpModel.actual_department.in_(hr_scope.scoped_departments)),
        )]

    current_year = date.today().year
    current_month = date.today().month

    # 当前月份起止
    month_start = date(current_year, current_month, 1)
    if current_month == 12:
        month_end = date(current_year + 1, 1, 1)
    else:
        month_end = date(current_year, current_month + 1, 1)

    # ── KPI 汇总 ──
    total_employees = (await session.execute(
        select(func.count(EmpModel.id)).where(
            *_scoped(),
            EmpModel.is_deleted == False,  # noqa: E712
            EmpModel.status == "在职",
        )
    )).scalar() or 0

    new_hires_this_month = (await session.execute(
        select(func.count(EmpModel.id)).where(
            *_scoped(),
            EmpModel.is_deleted == False,  # noqa: E712
            EmpModel.hire_date >= month_start,
            EmpModel.hire_date < month_end,
        )
    )).scalar() or 0

    departures_this_month = (await session.execute(
        select(func.count(EmpModel.id)).where(
            *_scoped(),
            EmpModel.is_deleted == False,  # noqa: E712
            EmpModel.status == "离职",
            EmpModel.departure_date >= month_start,
            EmpModel.departure_date < month_end,
        )
    )).scalar() or 0

    # ── 学历分布 ──
    edu_rows = (await session.execute(
        select(EmpModel.education, func.count(EmpModel.id))
        .where(*_scoped(), EmpModel.is_deleted == False, EmpModel.status == "在职")  # noqa: E712
        .group_by(EmpModel.education)
    )).all()
    education_distribution = [
        {"name": r[0] if r[0] else "未知", "value": r[1]}
        for r in edu_rows
    ]

    # ── 部门人数分布 ──
    dept_rows = (await session.execute(
        select(EmpModel.department, func.count(EmpModel.id))
        .where(*_scoped(), EmpModel.is_deleted == False, EmpModel.status == "在职")  # noqa: E712
        .group_by(EmpModel.department)
        .order_by(func.count(EmpModel.id).desc())
    )).all()
    department_distribution = [
        {"name": r[0], "value": r[1]} for r in dept_rows
    ]

    # ── 本年度入职各部门月度统计 ──
    year_start = date(current_year, 1, 1)
    next_year_start = date(current_year + 1, 1, 1)
    hires_by_dept_rows = (await session.execute(
        select(
            func.to_char(EmpModel.hire_date, "YYYY-MM").label("month"),
            EmpModel.department,
            func.count(EmpModel.id),
        )
        .where(
            *_scoped(),
            EmpModel.is_deleted == False,  # noqa: E712
            EmpModel.hire_date >= year_start,
            EmpModel.hire_date < next_year_start,
        )
        .group_by("month", EmpModel.department)
        .order_by("month")
    )).all()
    monthly_hires_by_dept = [
        {"month": r.month, "department": r.department, "value": r[2]}
        for r in hires_by_dept_rows
    ]

    # ── 本年度离职月度统计 ──
    dep_rows = (await session.execute(
        select(
            func.to_char(EmpModel.departure_date, "YYYY-MM").label("month"),
            func.count(EmpModel.id),
        )
        .where(
            *_scoped(),
            EmpModel.is_deleted == False,  # noqa: E712
            EmpModel.status == "离职",
            EmpModel.departure_date >= year_start,
            EmpModel.departure_date < next_year_start,
        )
        .group_by("month")
        .order_by("month")
    )).all()
    monthly_departures = [
        {"month": r.month, "value": r[1]} for r in dep_rows
    ]

    # ── 离职月度对比（当年 vs 去年） ──
    dep_cur_rows = (await session.execute(
        select(
            func.to_char(EmpModel.departure_date, "MM").label("month_num"),
            func.count(EmpModel.id),
        )
        .where(
            *_scoped(),
            EmpModel.is_deleted == False,  # noqa: E712
            EmpModel.status == "离职",
            EmpModel.departure_date >= year_start,
            EmpModel.departure_date < next_year_start,
        )
        .group_by("month_num")
        .order_by("month_num")
    )).all()

    last_year_start = date(current_year - 1, 1, 1)
    last_year_end = date(current_year, 1, 1)
    dep_last_rows = (await session.execute(
        select(
            func.to_char(EmpModel.departure_date, "MM").label("month_num"),
            func.count(EmpModel.id),
        )
        .where(
            *_scoped(),
            EmpModel.is_deleted == False,  # noqa: E712
            EmpModel.status == "离职",
            EmpModel.departure_date >= last_year_start,
            EmpModel.departure_date < last_year_end,
        )
        .group_by("month_num")
        .order_by("month_num")
    )).all()

    cur_map = {int(r.month_num): r[1] for r in dep_cur_rows}
    last_map = {int(r.month_num): r[1] for r in dep_last_rows}

    monthly_departure_comparison = []
    for m in range(1, 13):
        monthly_departure_comparison.append({
            "month": f"{m:02d}月",
            "label": month_abbr[m],
            "last_year": last_map.get(m, 0),
            "current_year": cur_map.get(m, 0),
        })

    # ── 通用分布查询辅助 ──
    async def _query_distribution(field_name: str) -> list[dict]:
        col = getattr(EmpModel, field_name)
        rows = (await session.execute(
            select(col, func.count(EmpModel.id))
            .where(*_scoped(), EmpModel.is_deleted == False, EmpModel.status == "在职")  # noqa: E712
            .group_by(col)
            .order_by(func.count(EmpModel.id).desc())
        )).all()
        return [{"name": r[0] if r[0] else "未知", "value": r[1]} for r in rows]

    # ── 年龄区间分布（从出生年份动态计算）──
    current_year = date.today().year
    dynamic_age = current_year - EmpModel.birth_year  # 近似年龄，用于区间统计
    age_ranges = [(0, 25), (25, 30), (30, 35), (35, 40), (40, 45), (45, 50), (50, 55), (55, 999)]
    age_distribution = []
    for lo, hi in age_ranges:
        label = f"{lo}-{hi}岁" if hi < 999 else f"{lo}岁以上"
        cnt = (await session.execute(
            select(func.count(EmpModel.id)).where(
                *_scoped(),
                EmpModel.is_deleted == False,  # noqa: E712
                EmpModel.status == "在职",
                EmpModel.birth_year > 0,
                dynamic_age >= lo,
                dynamic_age < hi,
            )
        )).scalar() or 0
        age_distribution.append({"name": label, "value": cnt})

    # ── 入职时长（司龄）分布 ──
    tenure_ranges = [
        (0, 1, "不足1年"), (1, 3, "1~3年"), (3, 5, "3~5年"),
        (5, 10, "5~10年"), (10, 15, "10~15年"), (15, 20, "15~20年"),
        (20, 999, "20年以上"),
    ]
    tenure_distribution = []
    for lo, hi, label in tenure_ranges:
        lo_date = date(current_year - hi + 1, 1, 1)
        hi_date = date(current_year - lo + 1, 1, 1)
        cnt = (await session.execute(
            select(func.count(EmpModel.id)).where(
                *_scoped(),
                EmpModel.is_deleted == False,  # noqa: E712
                EmpModel.status == "在职",
                EmpModel.hire_date > lo_date,
                EmpModel.hire_date <= hi_date,
            )
        )).scalar() or 0
        tenure_distribution.append({"name": label, "value": cnt})

    gender_distribution = await _query_distribution("gender")
    level_distribution = await _query_distribution("level")
    position_distribution = await _query_distribution("position")
    job_category_distribution = await _query_distribution("job_category")
    political_status_distribution = await _query_distribution("political_status")
    marital_status_distribution = await _query_distribution("marital_status")
    contract_type_distribution = await _query_distribution("contract_type")
    household_type_distribution = await _query_distribution("household_type")

    # ── 每月入职人数（总计，不分部门）──
    monthly_hires_rows = (await session.execute(
        select(
            func.to_char(EmpModel.hire_date, "YYYY-MM").label("month"),
            func.count(EmpModel.id),
        )
        .where(
            *_scoped(),
            EmpModel.is_deleted == False,  # noqa: E712
            EmpModel.hire_date >= year_start,
            EmpModel.hire_date < next_year_start,
        )
        .group_by("month")
        .order_by("month")
    )).all()
    monthly_hires = [
        {"month": r.month, "value": r[1]} for r in monthly_hires_rows
    ]

    # ── 2026年离职各部门统计 ──
    dep_by_dept_rows = (await session.execute(
        select(
            EmpModel.department,
            func.count(EmpModel.id),
        )
        .where(
            *_scoped(),
            EmpModel.is_deleted == False,  # noqa: E712
            EmpModel.status == "离职",
            EmpModel.departure_date >= year_start,
            EmpModel.departure_date < next_year_start,
        )
        .group_by(EmpModel.department)
        .order_by(func.count(EmpModel.id).desc())
    )).all()
    departures_by_dept = [
        {"name": r[0], "value": r[1]} for r in dep_by_dept_rows
    ]

    # ── 在职员工平均年龄（从出生年份动态计算）──
    age_avg_result = (await session.execute(
        select(func.avg(current_year - EmpModel.birth_year)).where(
            *_scoped(),
            EmpModel.is_deleted == False,  # noqa: E712
            EmpModel.status == "在职",
            EmpModel.birth_year > 0,
        )
    )).scalar()
    age_avg = round(float(age_avg_result), 1) if age_avg_result else 0.0

    return success_response(data={
        "summary": {
            "total_employees": total_employees,
            "new_hires_this_month": new_hires_this_month,
            "departures_this_month": departures_this_month,
            "age_avg": age_avg,
        },
        "education_distribution": education_distribution,
        "department_distribution": department_distribution,
        "monthly_hires_by_dept": monthly_hires_by_dept,
        "monthly_hires": monthly_hires,
        "monthly_departures": monthly_departures,
        "monthly_departure_comparison": monthly_departure_comparison,
        "departures_by_dept": departures_by_dept,
        "distributions": {
            "gender": gender_distribution,
            "age": age_distribution,
            "level": level_distribution,
            "position": position_distribution,
            "job_category": job_category_distribution,
            "political_status": political_status_distribution,
            "marital_status": marital_status_distribution,
            "contract_type": contract_type_distribution,
            "household_type": household_type_distribution,
            "tenure": tenure_distribution,
        },
    })


@router.get("/employees/{employee_id}", summary="员工详情")
async def get_employee(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
    has_sensitive: bool = Depends(check_sensitive_permission),
):
    employee = await service.get_employee(employee_id)
    hr_scope.ensure_can_access_employee(employee)
    return success_response(
        data=mask_sensitive_fields(EmployeeResponse.model_validate(employee).model_dump(mode="json"), has_sensitive),
    )


@router.put("/employees/{employee_id}", summary="更新员工")
async def update_employee(
    employee_id: UUID,
    payload: EmployeeUpdate,
    service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
    has_sensitive: bool = Depends(check_sensitive_permission),
):
    employee = await service.get_employee(employee_id)
    hr_scope.ensure_can_access_employee(employee)
    # 目标部门也必须在数据范围内（防止把员工移出/移入越权部门）
    if "department" in payload.model_fields_set and payload.department:
        hr_scope.ensure_dept_writable([payload.department])
    employee = await service.update_employee(employee_id, payload)
    return success_response(
        data=mask_sensitive_fields(
            EmployeeResponse.model_validate(employee).model_dump(mode="json"), has_sensitive
        ),
        message="员工更新成功",
    )


@router.delete("/employees/{employee_id}", summary="删除员工")
async def delete_employee(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    employee = await service.get_employee(employee_id)
    hr_scope.ensure_can_access_employee(employee)
    await service.delete_employee(employee_id)
    return success_response(message="员工删除成功")


@router.get(
    "/employees/{employee_number}/onboarding-training-record",
    summary="导出员工入职培训记录",
)
async def export_onboarding_training_record(
    employee_number: str,
    service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """根据员工工号自动生成并下载入职培训记录 Word 文档。"""
    emp = await service.repo.get_by_employee_number(employee_number)
    if emp:
        hr_scope.ensure_can_access_employee(emp)
    employee = await service.get_employee_by_number(employee_number)
    try:
        buffer: BytesIO = generate_onboarding_training_record(employee)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    filename = f"onboarding_training_record_{employee.employee_number}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class TrainingItem(BaseModel):
    sop_number: str = ""
    file_name: str = ""
    content: str = ""
    method: str = ""
    trainer: str = ""
    plan_date: str = ""


class TrainingExportRequest(BaseModel):
    training_items: list[TrainingItem] = []


@router.post(
    "/employees/{employee_number}/onboarding-training-record",
    summary="导出员工入职培训记录（带培训项）",
)
async def export_onboarding_training_record_with_items(
    employee_number: str,
    body: TrainingExportRequest,
    service: EmployeeService = Depends(get_employee_service),
):
    """根据员工工号和前端选中的培训项，生成入职培训记录 Word 文档。"""
    employee = await service.get_employee_by_number(employee_number)
    items = [it.model_dump() for it in body.training_items]
    try:
        buffer: BytesIO = generate_onboarding_training_record(employee, training_items=items)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    filename = f"onboarding_training_record_{employee.employee_number}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/employees/{employee_number}/training-record",
    summary="导出新员工培训记录",
)
async def export_training_record(
    employee_number: str,
    body: TrainingExportRequest,
    service: EmployeeService = Depends(get_employee_service),
):
    """根据员工工号和培训项，生成新员工培训记录 Word 文档。"""
    from app.modules.hr.training_record_generator import generate_training_record

    employee = await service.get_employee_by_number(employee_number)
    items = [it.model_dump() for it in body.training_items]
    try:
        buffer: BytesIO = generate_training_record(employee, training_items=items)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    filename = f"training_record_{employee.employee_number}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/employees/{employee_number}/work-permit",
    summary="导出员工上岗证",
)
async def export_work_permit(
    employee_number: str,
    body: TrainingExportRequest,
    service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    emp = await service.repo.get_by_employee_number(employee_number)
    if emp:
        hr_scope.ensure_can_access_employee(emp)
    """根据员工工号和培训项，生成上岗证 Word 文档。"""
    from app.modules.hr.work_permit_generator import generate_work_permit

    employee = await service.get_employee_by_number(employee_number)
    items = [it.model_dump() for it in body.training_items]
    try:
        buffer: BytesIO = generate_work_permit(employee, training_items=items)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    filename = f"work_permit_{employee.employee_number}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class PrejobTrainingItems(BaseModel):
    training_items: list[dict] = []


@router.post(
    "/employees/{employee_id}/prejob-training-plan",
    summary="导出员工岗前培训计划",
)
async def export_prejob_training_plan(
    employee_id: UUID,
    payload: PrejobTrainingItems | None = None,
    service: EmployeeService = Depends(get_employee_service),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    employee = await service.get_employee(employee_id)
    hr_scope.ensure_can_access_employee(employee)
    """根据入职台账数据生成并下载岗前培训计划 Word 文档。

    优先使用入职台账（hr.onboarding_records）数据，
    找不到时回退到员工表（hr.employees）。
    支持 POST 传入前端填写的培训项（含计划完成日期、培训师、培训方式）。
    """
    from app.modules.hr.models import OnboardingRecord, PositionTraining

    # 优先查入职台账
    record = (await session.execute(
        select(OnboardingRecord).where(
            OnboardingRecord.id == employee_id,
            OnboardingRecord.is_deleted == False,
        )
    )).scalar_one_or_none()

    if record:
        person = record
        emp_no = record.employee_number
        dept = record.department
        pos = record.position
    else:
        person = await service.get_employee(employee_id)
        emp_no = person.employee_number
        dept = person.department
        pos = person.position

    # 前端传了培训项则优先用，否则查数据库
    if payload and payload.training_items:
        training_items = [
            {
                "training_category": item.get("content", "") or item.get("file_name", ""),
                "trainer": item.get("trainer", ""),
                "training_method": item.get("method", ""),
                "plan_date": item.get("plan_date", ""),
            }
            for item in payload.training_items
        ]
    else:
        # 查询岗位培训内容（精确匹配优先，无结果时模糊匹配）
        training_rows = (await session.execute(
            select(PositionTraining).where(
                PositionTraining.department == dept,
                PositionTraining.position_name == pos,
                PositionTraining.is_deleted == False,
            ).order_by(PositionTraining.sort_order)
        )).scalars().all()

        if not training_rows:
            training_rows = (await session.execute(
                select(PositionTraining).where(
                    PositionTraining.department == dept,
                    PositionTraining.position_name.ilike(f"%{pos}%"),
                    PositionTraining.is_deleted == False,
                ).order_by(PositionTraining.sort_order)
            )).scalars().all()

        training_items = [
            {
                "training_category": t.training_category or "",
                "trainer": t.trainer or "",
                "training_method": t.training_method or "",
                "plan_date": "",
            }
            for t in training_rows
        ]

    try:
        buffer: BytesIO = generate_prejob_training_plan(person, training_items)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    filename = f"prejob_training_plan_{emp_no}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/employees/{employee_id}/onboarding-evaluation",
    summary="导出员工上岗评估表",
)
async def export_onboarding_evaluation_by_employee(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """根据员工档案预填基本信息并导出上岗评估表 Excel 文档。"""
    employee = await service.get_employee(employee_id)
    hr_scope.ensure_can_access_employee(employee)

    payload = OnboardingEvaluationInput(
        employee_name=employee.name or "",
        employee_number=employee.employee_number or None,
        gender=employee.gender or None,
        department_position=f"{employee.department or ''}/{employee.position or ''}",
        hire_date=employee.hire_date,
    )
    buffer: BytesIO = generate_onboarding_evaluation(payload)

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(employee.hire_date).replace("-", "") if employee.hire_date else "nodate"
    filename = f"onboarding_evaluation_{employee.employee_number}_{safe_date}.xlsx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )




@router.get("/employee-varieties", summary="获取部门品种列表（去重）")
async def list_employee_varieties(
    department: str = Query(...),
    session: AsyncSession = Depends(get_db),
):
    """获取指定部门（含兼任部门）在职员工的品种列表（去重），用于签到表二级筛选。"""
    from app.modules.hr.models import Employee
    # 签到表人员口径：仅体现部门+兼任部门，不引入「未分类」人员
    rows = (await session.execute(
        select(Employee.variety, Employee.concurrent_variety)
        .where(
            Employee.is_deleted == False,
            Employee.status == '在职',
            or_(
                Employee.department == department,
                Employee.concurrent_departments.ilike(f"%{department}%"),
            ),
        )
    )).all()
    varieties: list[str] = []
    for variety, concurrent in rows:
        for raw in (variety, concurrent):
            if not raw:
                continue
            for v in str(raw).split("、"):
                v = v.strip()
                if v and v != "——" and v not in varieties:
                    varieties.append(v)
    return success_response(data=sorted(varieties))


async def _query_leave_counts(session: AsyncSession, department: str) -> tuple[int, int]:
    """统计部门病假/产假人数（无则 0），供通知/签到表人数栏展示。"""
    sick = maternity = 0
    try:
        result = await session.execute(
            text(
                "SELECT status, count(*) FROM hr.employees "
                "WHERE department = :dept AND is_deleted = false "
                "AND status IN ('病假', '产假') GROUP BY status"
            ),
            {"dept": department},
        )
        for row in result.fetchall():
            if row[0] == "病假":
                sick = row[1]
            elif row[0] == "产假":
                maternity = row[1]
    except Exception:
        pass
    return sick, maternity


@router.post("/training-sign-in-sheet", summary="生成培训签到表")
async def export_training_sign_in_sheet(
    payload: TrainingSignInSheetInput,
    session: AsyncSession = Depends(get_db),
):
    """根据填写的培训信息自动生成培训签到表 Word 文档。

    超过一页（20人）时自动分页，所有页面合并在一个 docx 中。
    """
    sick, maternity = await _query_leave_counts(session, payload.department)
    payload.sick_count = sick
    payload.maternity_count = maternity
    safe_date = str(payload.training_date).replace("-", "")
    try:
        buffer: BytesIO = generate_training_sign_in_sheet(payload)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_filename = f"training_sign_in_sheet_{safe_date}.docx"

    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=utf-8''{quote('培训签到表_' + safe_date + '.docx')}"
        },
    )


class AssessmentScoreInput(BaseModel):
    training_content: str = ""
    training_date: str = ""
    department: str = ""
    scores: list[dict] = []


@router.post("/training-assessment-scores/export", summary="导出实操考核成绩单")
async def export_assessment_scores(payload: AssessmentScoreInput):
    """根据填写的培训信息和员工成绩，生成实操考核成绩单 Word 文档。"""
    from app.modules.hr.assessment_score_generator import (
        generate_assessment_score_sheet,
    )

    buf = generate_assessment_score_sheet(
        training_content=payload.training_content,
        training_date=payload.training_date,
        department=payload.department,
        scores=payload.scores,
    )
    safe_date = payload.training_date.replace("-", "") if payload.training_date else "nodate"

    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f"attachment; filename*=utf-8''{quote(f'考核成绩单_{safe_date}.docx')}"
        },
    )


@router.post("/training-notification", summary="生成培训通知")
async def export_training_notification(
    payload: TrainingNotificationInput,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    session: AsyncSession = Depends(get_db),
):
    """根据填写的培训信息生成培训通知 Word 文档。培训台账需通过「添加到培训台账」按钮手动录入。"""
    sick, maternity = await _query_leave_counts(session, payload.department)
    payload.sick_count = sick
    payload.maternity_count = maternity
    try:
        buffer: BytesIO = generate_training_notification(payload)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(payload.training_date).replace("-", "")
    filename = f"training_notification_{safe_date}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/training-notification/generate-assessment", summary="AI生成考核题目（仅笔试）")
async def generate_assessment_questions(
    file: UploadFile,
    assessment_method: str = Form("笔试"),
    subject: str = Form(""),
    session: AsyncSession = Depends(get_db),
):
    """上传培训材料文件，AI 自动生成笔试试卷（结果同步入共享题库）。问答考核不再 AI 出题，改从题库选题。"""
    if not file.filename:
        raise HTTPException(400, "请上传文件")
    if assessment_method != "笔试":
        raise HTTPException(400, "问答考核已停用 AI 出题，请从题库选题")
    try:
        content = await file.read()
        # 尝试解码文件内容
        text = ""
        if file.filename.endswith(".txt"):
            text = content.decode("utf-8")
        elif file.filename.endswith(".docx"):
            from io import BytesIO

            from docx import Document
            doc = Document(BytesIO(content))
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        elif file.filename.endswith(".doc"):
            # 转换旧版 .doc 文件（兼容 macOS / Linux）
            import asyncio
            import os
            import shutil
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as tmp:
                tmp.write(content)
                tmp_path = tmp.name
            try:
                text = ""
                if shutil.which("textutil"):
                    out_path = tmp_path + ".txt"
                    proc = await asyncio.create_subprocess_exec(
                        "textutil", "-convert", "txt", tmp_path, "-output", out_path,
                        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
                    )
                    await asyncio.wait_for(proc.communicate(), timeout=30)
                    if proc.returncode == 0 and os.path.exists(out_path):
                        with open(out_path) as f:
                            text = f.read()
                    if os.path.exists(out_path):
                        os.unlink(out_path)
                elif shutil.which("antiword"):
                    proc = await asyncio.create_subprocess_exec(
                        "antiword", tmp_path,
                        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
                    )
                    stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=30)
                    if proc.returncode == 0:
                        text = stdout.decode("utf-8", errors="ignore")
                elif shutil.which("catdoc"):
                    proc = await asyncio.create_subprocess_exec(
                        "catdoc", tmp_path,
                        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
                    )
                    stdout, _ = await asyncio.wait_for(proc.communicate(), timeout=30)
                    if proc.returncode == 0:
                        text = stdout.decode("utf-8", errors="ignore")
                if not text:
                    text = "（提示：服务器未安装 .doc 转换工具，请上传 .docx 或 .txt 格式文件）"
            finally:
                os.unlink(tmp_path)
        else:
            text = content.decode("utf-8", errors="ignore")

        if not text.strip():
            raise HTTPException(400, "无法提取文件内容，请确认文件不为空")

        # 截断过长文本
        text = text[:8000]

        # 使用 AI 生成题目
        from app.modules.hr.ai_service import AiChatService
        from app.modules.hr.config import HR_AI_API_KEY, HR_AI_MODEL
        api_key = HR_AI_API_KEY
        model = HR_AI_MODEL or "deepseek-chat"

        if api_key:
            service = AiChatService(api_key=api_key, model=model)
            prompt = f"""你是一个药厂培训考核出题人。根据以下培训材料生成4道选择题，参考题库风格：

风格要求（非常重要）：
1. 题目简短精炼，10-20字以内，直接问操作要点或关键知识
2. 答案简洁明了，15字以内，一句话说清
3. 参考示例：问"调平依据什么工具？"答"使用水平尺"、问"拿取砝码需佩戴什么？"答"佩戴手套操作"
4. 每题10分，总计40分
5. 题目必须基于材料内容，不要编造

请以JSON格式返回：
{{
  "title": "考核标题",
  "total_score": 40,
  "questions": [
    {{"type": "问答", "question": "简短题目", "answer": "简短答案", "score": 10}}
  ]
}}

培训主题：{subject}
培训材料：
{text}"""

            messages = [{"role": "user", "content": prompt}]
            full_response = ""
            async for chunk in service.stream_chat(messages):
                if chunk.get("type") == "content":
                    full_response += chunk["text"]

            import json
            json_start = full_response.find("{")
            json_end = full_response.rfind("}") + 1
            if json_start >= 0 and json_end > json_start:
                result = json.loads(full_response[json_start:json_end])
            else:
                raise ValueError("AI 返回格式无法解析")
        else:
            raise HTTPException(400, "服务端未配置 HR_AI_API_KEY，无法生成笔试试卷")

        # 生成结果同步入共享题库（AI生成来源），考核矩阵/题库选题可复用；
        # 题库写入失败不连累生成结果
        from app.modules.hr.ai_exam_service import save_questions_to_bank

        questions = result.get("questions") or []
        try:
            bank_inserted = await save_questions_to_bank(
                session, questions, subject or "培训考核", subject=subject or ""
            )
        except Exception:
            import logging
            logging.getLogger(__name__).warning("题库写入失败，不影响生成结果", exc_info=True)
            await session.rollback()
            bank_inserted = 0
        return success_response(data={**result, "bank_inserted": bank_inserted}, message=f"生成成功，{bank_inserted} 题已入题库")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"生成失败: {str(e)}")


@router.post("/training-evaluations/export-admin", summary="导出培训效果评估表（台账管理端）")
async def export_training_evaluation_admin(
    department: str = Form(""),
    training_subject: str = Form(""),
    training_date: str = Form(""),
    training_method: str = Form(""),
    trainer_name: str = Form(""),
    assessment_method: str = Form(""),
    expected_count: int = Form(0),
    actual_count: int = Form(0),
    exam_count: int = Form(0),
    excellent_count: int = Form(0),
    qualified_count: int = Form(0),
    unqualified_count: int = Form(0),
    session: AsyncSession = Depends(get_db),
):
    """导出培训效果评估表，自动关联年度计划和台账成绩。"""
    from app.modules.hr.evaluation_document_generator import (
        TrainingEvaluationInput,
        generate_training_evaluation,
    )

    training_date_val = date.fromisoformat(training_date) if training_date else None
    plan_expected = expected_count
    plan_method = training_method or None
    plan_trainer = trainer_name or None
    plan_assessment = assessment_method or None
    plan_duration = None

    # 从年度计划补全空缺数据
    plan_target = None
    if department and training_subject:
        for subj_pat in (training_subject, f"%{training_subject}%", "%"):
            plan_row = (await session.execute(
                text("SELECT pi.trainee_count, pi.assessment_method, pi.training_method, pi.duration_hours, pi.content_and_textbook, pi.target_audience FROM hr.annual_training_plan_items pi JOIN hr.annual_training_plans p ON pi.plan_id = p.id WHERE p.department = :dept AND pi.content_and_textbook ILIKE :subj AND pi.is_deleted = false AND p.is_deleted = false LIMIT 1"),
                {"dept": department, "subj": subj_pat},
            )).fetchone()
            if plan_row:
                if not plan_expected and plan_row[0]:
                    plan_expected = int(plan_row[0])
                plan_assessment = plan_assessment or plan_row[1]
                plan_method = plan_method or plan_row[2]
                plan_duration = plan_duration or plan_row[3]
                plan_target = plan_row[5]  # target_audience
                break

    # 从台账统计成绩（混合文字+数字）
    if department and training_subject:
        if department and training_subject:
            scores_result = await session.execute(
                text("SELECT tl.assessment_result FROM hr.training_ledgers tl JOIN hr.employees e ON tl.employee_number = e.employee_number AND e.is_deleted = false WHERE tl.is_deleted = false AND e.department = :dept AND tl.training_subject ILIKE :subj AND tl.assessment_result IS NOT NULL AND tl.assessment_result != ''"),
                {"dept": department, "subj": f"%{training_subject}%"},
            )
            scores = [row[0] for row in scores_result.fetchall() if row[0]]
            if scores:
                exam_count = len(scores)
                excellent = qualified = unqualified = 0
                for s in scores:
                    try:
                        n = float(s)
                        if n >= 90: excellent += 1
                        elif n >= 80: qualified += 1
                        else: unqualified += 1
                    except ValueError:
                        if s in ("不合格", "unqualified", "fail"): unqualified += 1
                        elif s in ("合格", "qualified", "pass"): qualified += 1
                        elif s in ("优秀", "优", "excellent"): excellent += 1
                excellent_count = excellent
                qualified_count = qualified
                unqualified_count = unqualified

    # 从台账查培训对象（学员姓名列表），无则从年度计划取 target_audience
    trainee_names: list[str] = []
    if department and training_subject:
        names_result = await session.execute(
            text("SELECT e.name FROM hr.training_ledgers tl JOIN hr.employees e ON tl.employee_number = e.employee_number AND e.is_deleted = false WHERE tl.is_deleted = false AND e.department = :dept AND tl.training_subject ILIKE :subj"),
            {"dept": department, "subj": f"%{training_subject}%"},
        )
        trainee_names = list(dict.fromkeys(row[0] for row in names_result.fetchall() if row[0]))
    if not trainee_names and plan_target:
        trainee_names = [plan_target]

    # 应到人数：部门在职员工数，排除病假/产假
    sick_count = maternity_count = 0
    if department:
        emp_count = (await session.execute(
            text("SELECT count(*) FROM hr.employees WHERE department = :dept AND is_deleted = false AND status = '在职'"),
            {"dept": department},
        )).scalar()
        sick_count, maternity_count = await _query_leave_counts(session, department)
        if not emp_count:
            emp_count = (await session.execute(
                text("SELECT count(*) FROM hr.onboarding_records WHERE department = :dept AND is_deleted = false AND is_employed = '是'"),
                {"dept": department},
            )).scalar()
        plan_expected = emp_count or 0

    # 实到人数 = 有成绩的人数（参加考核才算实到）
    actual_count = exam_count

    # 计算参与率和合格率
    participation_rate = None
    pass_rate = None
    if plan_expected and plan_expected > 0:
        participation_rate = f"{round(actual_count / plan_expected * 100)}%"
    if exam_count and exam_count > 0:
        pass_count = (excellent_count or 0) + (qualified_count or 0)
        pass_rate = f"{round(pass_count / exam_count * 100)}%"

    payload = TrainingEvaluationInput(
        subject=training_subject or "培训",
        training_date=training_date_val,
        training_method=plan_method,
        trainer=plan_trainer,
        assessment_method=plan_assessment,
        duration_hours=plan_duration,
        expected_count=plan_expected or None,
        actual_count=actual_count or None,
        exam_count=exam_count or None,
        excellent_count=excellent_count or None,
        qualified_count=qualified_count or None,
        unqualified_count=unqualified_count or None,
        trainee_names=trainee_names,
        participation_rate=participation_rate,
        pass_rate=pass_rate,
        department=department or None,
        # 没有病假/产假时自动填 0；未按部门统计时才留空
        sick_leave=sick_count if department else None,
        maternity_leave=maternity_count if department else None,
    )
    buffer: BytesIO = generate_training_evaluation(payload)
    safe_date = training_date.replace("-", "") if training_date else "nodate"
    filename = f"training_evaluation_{safe_date}.docx"
    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{quote(filename)}"},
    )


@router.post("/training-evaluation", summary="生成培训效果评估表")
async def export_training_evaluation(
    payload: TrainingEvaluationInput,
):
    """根据填写的培训信息自动生成培训效果评估表 Word 文档。"""
    buffer: BytesIO = generate_training_evaluation(payload)

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(payload.training_date).replace("-", "") if payload.training_date else "nodate"
    safe_filename = f"training_evaluation_{safe_date}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f"attachment; filename=\"{safe_filename}\"; filename*=utf-8''{quote('培训效果评估表_' + safe_date + '.docx')}"
        },
    )


@router.post("/onboarding-evaluation", summary="生成员工上岗评估表")
async def export_onboarding_evaluation(
    payload: OnboardingEvaluationInput,
):
    """根据填写的评估信息自动生成员工上岗评估表 Excel 文档。"""
    buffer: BytesIO = generate_onboarding_evaluation(payload)

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(payload.evaluation_date).replace("-", "") if payload.evaluation_date else "nodate"
    filename = f"onboarding_evaluation_{safe_date}.xlsx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Department Routes ───

@router.get("/departments", summary="部门列表")
async def list_departments(
    keyword: str | None = Query(None, description="部门名称或编码关键词"),
    page_params: PageParams = Depends(),
    service: DepartmentService = Depends(get_department_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    # 数据范围限制：只返回授权部门
    if hr_scope.scoped_departments:
        if len(hr_scope.scoped_departments) == 1:
            keyword = next(iter(hr_scope.scoped_departments))
            departments, total = await service.list_departments(
                keyword=keyword, page=page_params.page, page_size=page_params.page_size,
            )
        else:
            # 多部门：全量拉取 + 过滤 + 内存分页
            all_deps, _ = await service.list_departments(
                keyword=None, page=1, page_size=1000,
            )
            filtered = [d for d in all_deps if d.name in hr_scope.scoped_departments]
            total = len(filtered)
            start = (page_params.page - 1) * page_params.page_size
            departments = filtered[start:start + page_params.page_size]
    else:
        departments, total = await service.list_departments(
            keyword=keyword, page=page_params.page, page_size=page_params.page_size,
        )
    data = [
        DepartmentResponse.model_validate(d).model_dump(mode="json")
        for d in departments
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/departments", summary="创建部门")
async def create_department(
    payload: DepartmentCreate,
    service: DepartmentService = Depends(get_department_service),
):
    department = await service.create_department(payload)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
        message="部门创建成功",
        status_code=201,
    )


@router.get("/departments/{department_id}", summary="部门详情")
async def get_department(
    department_id: UUID,
    service: DepartmentService = Depends(get_department_service),
):
    department = await service.get_department(department_id)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
    )


@router.put("/departments/{department_id}", summary="更新部门")
async def update_department(
    department_id: UUID,
    payload: DepartmentUpdate,
    service: DepartmentService = Depends(get_department_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    existing = await service.get_department(department_id)
    target_name = payload.name or existing.name
    hr_scope.ensure_dept_writable([existing.name, target_name])
    department = await service.update_department(department_id, payload)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
        message="部门更新成功",
    )


@router.delete("/departments/{department_id}", summary="删除部门")
async def delete_department(
    department_id: UUID,
    service: DepartmentService = Depends(get_department_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    existing = await service.get_department(department_id)
    hr_scope.ensure_dept_writable([existing.name])
    await service.delete_department(department_id)
    return success_response(message="部门删除成功")


# ─── Position Routes ───

@router.get("/positions", summary="职位列表（按部门筛选）")
async def list_positions(
    department: str | None = Query(None, description="部门名称筛选"),
    session: AsyncSession = Depends(get_db),
):
    """返回职位列表，可按部门筛选，包含关联的培训大类。"""
    from app.modules.hr.models import HrPosition, PositionTraining

    q = select(HrPosition).where(HrPosition.is_deleted == False)
    if department:
        q = q.where(HrPosition.department == department)
    q = q.order_by(HrPosition.sort_order)

    result = await session.execute(q)
    rows = result.scalars().all()

    # 批量查询所有岗位关联的培训大类
    dept_pos_keys = {(r.department, r.name) for r in rows}
    pt_map: dict[tuple[str, str], list[str]] = {}
    if dept_pos_keys:
        pt_q = select(PositionTraining).where(
            PositionTraining.is_deleted == False,
            PositionTraining.training_category != "",
        )
        pt_rows = (await session.execute(pt_q)).scalars().all()
        for pt in pt_rows:
            key = (pt.department, pt.position_name)
            if key in dept_pos_keys:
                pt_map.setdefault(key, []).append(pt.training_category)

    data = [
        {
            "id": str(r.id),
            "department": r.department,
            "name": r.name,
            "categories": list(dict.fromkeys(pt_map.get((r.department, r.name), []))),
        }
        for r in rows
    ]
    return success_response(data=data)


class PositionCreate(BaseModel):
    department: str
    name: str


@router.post("/positions", summary="新建职位")
async def create_position(
    payload: PositionCreate,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """手动新建一个职位，写入 hr.positions 表。"""
    hr_scope.ensure_dept_writable([payload.department])
    from app.modules.hr.models import HrPosition

    pos = HrPosition(department=payload.department, name=payload.name)
    session.add(pos)
    await session.flush()
    return success_response(
        data={"id": str(pos.id), "department": pos.department, "name": pos.name},
        message="职位创建成功",
        status_code=201,
    )


@router.delete("/positions/by-name/{position_name}", summary="按名称删除职位")
async def delete_position_by_name(
    position_name: str,
    department: str = Query(..., description="部门名称"),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """删除指定部门和名称的职位，同时清除关联的 SOP 目录条目。"""
    hr_scope.ensure_dept_writable([department])
    from app.modules.hr.models import HrPosition

    pos = (await session.execute(
        select(HrPosition).where(
            HrPosition.department == department,
            HrPosition.name == position_name,
            HrPosition.is_deleted == False,
        )
    )).scalar_one_or_none()
    if not pos:
        raise HTTPException(404, "职位不存在")
    # 软删除职位及关联培训内容（不做物理删除，保留历史可恢复）
    await session.execute(
        text("UPDATE hr.position_trainings SET is_deleted = true WHERE department = :d AND position_name = :p AND is_deleted = false"),
        {"d": department, "p": position_name},
    )
    await session.execute(
        text("UPDATE hr.sop_catalog SET is_deleted = true WHERE department = :d AND position_name = :p AND is_deleted = false"),
        {"d": department, "p": position_name},
    )
    pos.is_deleted = True
    await session.flush()
    return success_response(message="删除成功")


@router.get("/positions/departments", summary="职位表中所有部门")
async def list_position_departments(
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import HrPosition
    depts = {row[0] for row in (await session.execute(select(HrPosition.department).where(HrPosition.is_deleted == False).distinct().order_by(HrPosition.department))).all()}
    if hr_scope.scoped_departments:
        depts = depts & hr_scope.scoped_departments
    return success_response(data=sorted(depts))


# ─── Position Training Routes ───

@router.get("/position-trainings", summary="岗位培训内容列表")
async def list_position_trainings(
    department: str | None = Query(None, description="部门筛选"),
    position_name: str | None = Query(None, description="岗位名称筛选"),
    session: AsyncSession = Depends(get_db),
):
    """按部门和岗位查询关联的培训内容（SOP/文件）。"""
    from app.modules.hr.models import PositionTraining

    q = select(PositionTraining).where(PositionTraining.is_deleted == False)
    if department:
        q = q.where(PositionTraining.department == department)
    if position_name:
        q = q.where(or_(
            PositionTraining.position_name == position_name,
            PositionTraining.position_name.endswith(position_name),
        ))
    q = q.order_by(PositionTraining.department, PositionTraining.position_name, PositionTraining.sort_order)

    result = await session.execute(q)
    rows = result.scalars().all()
    data = [
        {
            "id": str(r.id),
            "position_name": r.position_name,
            "department": r.department,
            "variety": r.variety,
            "training_category": r.training_category,
            "trainer": r.trainer,
            "training_method": r.training_method,
            "sop_number": r.sop_number,
            "file_name": r.file_name,
        }
        for r in rows
    ]
    return success_response(data=data)


class PositionTrainingCreate(BaseModel):
    department: str
    position_name: str
    training_category: str
    sop_number: str | None = None
    file_name: str


@router.post("/position-trainings", summary="新建岗位培训内容")
async def create_position_training(
    payload: PositionTrainingCreate,
    session: AsyncSession = Depends(get_db),
):
    """手动新建一条岗位培训关联记录。"""
    from app.modules.hr.models import PositionTraining, SopCatalog

    pt = PositionTraining(
        department=payload.department,
        position_name=payload.position_name,
        training_category=payload.training_category,
        sop_number=payload.sop_number,
        file_name=payload.file_name,
    )
    session.add(pt)

    # 同步写入 SOP 目录
    sc = SopCatalog(
        department=payload.department,
        category=payload.training_category,
        sop_number=payload.sop_number,
        file_name=payload.file_name,
        position_name=payload.position_name,
    )
    session.add(sc)

    await session.flush()
    return success_response(data={"id": str(pt.id)}, message="创建成功", status_code=201)


@router.get("/position-trainings/departments", summary="岗位培训内容中的部门列表")
async def list_pt_departments(
    session: AsyncSession = Depends(get_db),
):
    from app.modules.hr.models import PositionTraining

    result = await session.execute(
        select(PositionTraining.department)
        .where(PositionTraining.is_deleted == False)
        .distinct()
        .order_by(PositionTraining.department)
    )
    return success_response(data=[row[0] for row in result.all()])


# ─── Team Routes ───

@router.get("/teams", summary="班组列表")
async def list_teams(
    department_id: UUID | None = Query(None, description="部门筛选"),
    keyword: str | None = Query(None, description="班组名称或编码关键词"),
    page_params: PageParams = Depends(),
    service: TeamService = Depends(get_team_service),
):
    teams, total = await service.list_teams(
        department_id=department_id,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        TeamResponse.model_validate(t).model_dump(mode="json")
        for t in teams
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/teams", summary="创建班组")
async def create_team(
    payload: TeamCreate,
    service: TeamService = Depends(get_team_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import HrDepartment
    dept = (await service.department_repo.session.execute(
        select(HrDepartment).where(HrDepartment.id == payload.department_id)
    )).scalar_one_or_none()
    if dept:
        hr_scope.ensure_dept_writable([dept.name])
    team = await service.create_team(payload)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
        message="班组创建成功",
        status_code=201,
    )


@router.get("/teams/{team_id}", summary="班组详情")
async def get_team(
    team_id: UUID,
    service: TeamService = Depends(get_team_service),
):
    team = await service.get_team(team_id)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
    )


@router.put("/teams/{team_id}", summary="更新班组")
async def update_team(
    team_id: UUID,
    payload: TeamUpdate,
    service: TeamService = Depends(get_team_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    existing = await service.get_team(team_id)
    if existing and existing.department_id:
        dept = await service.department_repo.get_by_id(existing.department_id)
        if dept:
            hr_scope.ensure_dept_writable([dept.name])
    team = await service.update_team(team_id, payload)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
        message="班组更新成功",
    )


@router.delete("/teams/{team_id}", summary="删除班组")
async def delete_team(
    team_id: UUID,
    service: TeamService = Depends(get_team_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    existing = await service.get_team(team_id)
    if existing and existing.department_id:
        dept = await service.department_repo.get_by_id(existing.department_id)
        if dept:
            hr_scope.ensure_dept_writable([dept.name])
    await service.delete_team(team_id)
    return success_response(message="班组删除成功")


# ─── OffboardingRecord Routes ───

@router.get("/offboarding-records", summary="离职记录列表")
async def list_offboarding_records(
    employee_id: UUID | None = Query(None, description="员工ID筛选"),
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    page_params: PageParams = Depends(),
    service: OffboardingRecordService = Depends(get_offboarding_service),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    if hr_scope.is_unrestricted:
        records, total = await service.list_records(
            employee_id=employee_id,
            keyword=keyword,
            page=page_params.page,
            page_size=page_params.page_size,
        )
    else:
        # 数据范围受限：离职记录无部门字段，join 员工表强制收敛到本部门/本人
        from app.modules.hr.models import Employee, OffboardingRecord
        stmt = (
            select(OffboardingRecord)
            .where(OffboardingRecord.is_deleted == False)
            .options(selectinload(OffboardingRecord.employee))
            .join(Employee, OffboardingRecord.employee_id == Employee.id)
        )
        if employee_id: stmt = stmt.where(OffboardingRecord.employee_id == employee_id)
        if keyword: stmt = stmt.where(Employee.name.ilike(f"%{keyword}%") | Employee.employee_number.ilike(f"%{keyword}%"))
        stmt = hr_scope.apply_list_scope(stmt, Employee)
        total = (await session.execute(select(func.count()).select_from(stmt.subquery()))).scalar() or 0
        records = list((await session.execute(
            stmt.order_by(OffboardingRecord.created_at.desc())
            .offset((page_params.page - 1) * page_params.page_size)
            .limit(page_params.page_size)
        )).scalars().all())
    data = [
        OffboardingRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/offboarding-records", summary="创建离职记录")
async def create_offboarding_record(
    payload: OffboardingRecordCreate,
    service: OffboardingRecordService = Depends(get_offboarding_service),
):
    record = await service.create_record(payload)
    # 手动构建响应，避免触发未加载的 relationship
    data = {
        "id": str(record.id),
        "employee_id": str(record.employee_id),
        "offboarding_date": (
            record.offboarding_date.isoformat()
            if record.offboarding_date else None
        ),
        "offboarding_type": record.offboarding_type,
        "reason": record.reason,
        "handover_status": record.handover_status,
        "notes": record.notes,
        "created_at": record.created_at.isoformat() if record.created_at else None,
        "updated_at": record.updated_at.isoformat() if record.updated_at else None,
    }
    return success_response(
        data=data,
        message="离职记录创建成功，员工状态已更新为离职",
        status_code=201,
    )


@router.get("/offboarding-records/{record_id}", summary="离职记录详情")
async def get_offboarding_record(
    record_id: UUID,
    service: OffboardingRecordService = Depends(get_offboarding_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    record = await service.get_record(record_id)
    await _ensure_record_in_scope(hr_scope, record)
    return success_response(
        data=OffboardingRecordResponse.model_validate(record).model_dump(mode="json"),
    )


@router.put("/offboarding-records/{record_id}", summary="更新离职记录")
async def update_offboarding_record(
    record_id: UUID,
    payload: OffboardingRecordUpdate,
    service: OffboardingRecordService = Depends(get_offboarding_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    record = await service.get_record(record_id)
    await _ensure_record_in_scope(hr_scope, record)
    record = await service.update_record(record_id, payload)
    return success_response(
        data=OffboardingRecordResponse.model_validate(record).model_dump(mode="json"),
        message="离职记录更新成功",
    )


@router.delete("/offboarding-records/{record_id}", summary="删除离职记录")
async def delete_offboarding_record(
    record_id: UUID,
    service: OffboardingRecordService = Depends(get_offboarding_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    record = await service.get_record(record_id)
    await _ensure_record_in_scope(hr_scope, record)
    await service.delete_record(record_id)
    return success_response(message="离职记录删除成功")


# ─── OnboardingRecord Routes ───

@router.get("/onboarding-records", summary="老厂入职台账列表")
async def list_onboarding_records(
    department: str | None = Query(None, description="部门筛选"),
    position: str | None = Query(None, description="岗位筛选"),
    is_employed: str | None = Query("是", description="是否在职筛选，默认仅显示在职"),
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    sort_by: str = Query("hire_date", description="排序字段"),
    sort_order: str = Query("desc", description="排序方向"),
    page_params: PageParams = Depends(),
    service: OnboardingRecordService = Depends(get_onboarding_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    if not hr_scope.is_unrestricted:
        if hr_scope.scoped_departments:
            if len(hr_scope.scoped_departments) == 1:
                department = next(iter(hr_scope.scoped_departments))
            # 多部门：department 保持 None，service 不过滤，由下方 post-filter 处理
        else:
            keyword = hr_scope.employee_number
    records, total = await service.list_records(
        department=department,
        position=position,
        is_employed=is_employed,
        keyword=keyword,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    if not hr_scope.is_unrestricted:
        if not hr_scope.scoped_departments:
            # self_only：keyword 为前缀匹配，再按工号精确收敛到本人
            records = [r for r in records if r.employee_number == hr_scope.employee_number]
            total = len(records)
        elif len(hr_scope.scoped_departments) > 1:
            # 多部门：按授权部门集合过滤
            records = [r for r in records if r.department in hr_scope.scoped_departments]
            total = len(records)
    data = [
        OnboardingRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.get("/onboarding-records/{record_id}", summary="入职记录详情")
async def get_onboarding_record(
    record_id: UUID,
    service: OnboardingRecordService = Depends(get_onboarding_service),
):
    record = await service.get_record(record_id)
    return success_response(
        data=OnboardingRecordResponse.model_validate(record).model_dump(mode="json"),
    )


@router.delete("/onboarding-records/{record_id}", summary="删除入职台账记录")
async def delete_onboarding_record(
    record_id: UUID,
    service: OnboardingRecordService = Depends(get_onboarding_service),
):
    await service.delete_record(record_id)
    return success_response(message="删除成功")


# ─── DepartureRecord Routes ───

@router.get("/departure-records", summary="老厂离职台账列表")
async def list_departure_records(
    department: str | None = Query(None, description="部门筛选"),
    offboarding_type: str | None = Query(None, description="离职类型筛选"),
    keyword: str | None = Query(None, description="姓名/部门/职位关键词"),
    sort_by: str = Query("offboarding_date", description="排序字段"),
    sort_order: str = Query("desc", description="排序方向"),
    page_params: PageParams = Depends(),
    service: DepartureRecordService = Depends(get_departure_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    if not hr_scope.is_unrestricted:
        if not hr_scope.scoped_departments:
            # 离职台账无工号字段，self_only 无法定位本人记录
            raise ForbiddenException("数据范围限制：仅可访问本人相关数据")
        if len(hr_scope.scoped_departments) == 1:
            department = next(iter(hr_scope.scoped_departments))
        else:
            # 多部门：SQL 层过滤（先过滤后分页，页内容不再稀疏/为空）
            scoped_departments = set(hr_scope.scoped_departments)
    records, total = await service.list_records(
        department=department,
        departments=scoped_departments if len(hr_scope.scoped_departments) > 1 else None,
        offboarding_type=offboarding_type,
        keyword=keyword,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        DepartureRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/departure-records", summary="创建离职台账记录")
async def create_departure_record(
    payload: DepartureRecordCreate,
    service: DepartureRecordService = Depends(get_departure_service),
    has_sensitive: bool = Depends(check_sensitive_permission),
):
    record = await service.create_record(payload)
    return success_response(
        data=mask_sensitive_fields(
            DepartureRecordResponse.model_validate(record).model_dump(mode="json"), has_sensitive
        ),
        message="离职台账记录创建成功",
        status_code=201,
    )


@router.get("/departure-records/{record_id}", summary="离职台账记录详情")
async def get_departure_record(
    record_id: UUID,
    service: DepartureRecordService = Depends(get_departure_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
    has_sensitive: bool = Depends(check_sensitive_permission),
):
    record = await service.get_record(record_id)
    if not hr_scope.is_unrestricted:
        hr_scope.ensure_dept_writable([record.department])
    return success_response(
        data=mask_sensitive_fields(
            DepartureRecordResponse.model_validate(record).model_dump(mode="json"), has_sensitive
        ),
    )


@router.put("/departure-records/{record_id}", summary="更新离职台账记录")
async def update_departure_record(
    record_id: UUID,
    payload: DepartureRecordUpdate,
    service: DepartureRecordService = Depends(get_departure_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
    has_sensitive: bool = Depends(check_sensitive_permission),
):
    record = await service.get_record(record_id)
    if not hr_scope.is_unrestricted:
        hr_scope.ensure_dept_writable([record.department])
    record = await service.update_record(record_id, payload)
    return success_response(
        data=mask_sensitive_fields(
            DepartureRecordResponse.model_validate(record).model_dump(mode="json"), has_sensitive
        ),
        message="离职台账记录更新成功",
    )


@router.delete("/departure-records/{record_id}", summary="删除离职台账记录")
async def delete_departure_record(
    record_id: UUID,
    service: DepartureRecordService = Depends(get_departure_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    record = await service.get_record(record_id)
    if not hr_scope.is_unrestricted:
        hr_scope.ensure_dept_writable([record.department])
    await service.delete_record(record_id)
    return success_response(message="离职台账记录删除成功")


@router.post("/departure-records/{record_id}/preview-certificate", summary="预览离职证明")
async def preview_departure_certificate(
    record_id: UUID,
    service: DepartureRecordService = Depends(get_departure_service),
):
    from fastapi.responses import HTMLResponse

    from app.modules.hr.termination_certificate_generator import (
        generate_termination_certificate_html,
    )
    record = await service.get_record(record_id)

    # 兜底：离职台账缺少的字段从员工档案补（同名多人时不做猜测，避免错打他人证件号）
    id_number = record.id_card or ""
    entry_date = record.livo_entry_date or record.factory_entry_date or None
    if (not id_number or not entry_date) and record.name:
        from app.modules.hr.models import Employee
        emp_rows = list((await service.repo.session.execute(
            select(Employee).where(Employee.name == record.name, Employee.is_deleted == False)
        )).scalars().all())
        emp_row = emp_rows[0] if len(emp_rows) == 1 else None
        if emp_row:
            if not id_number:
                id_number = emp_row.id_card or ""
            if not entry_date:
                entry_date = emp_row.livo_entry_date or emp_row.factory_entry_date or emp_row.hire_date

    html = generate_termination_certificate_html(
        name=record.name or "", id_number=id_number or "",
        department=record.department or "", position=record.position or "",
        entry_date=entry_date or "",
        leave_date=record.offboarding_date or "",
        leave_reason=record.offboarding_type or "个人原因",
    )
    return HTMLResponse(content=html)


@router.post("/departure-records/{record_id}/send-certificate", summary="发送离职证明邮件")
async def send_departure_certificate(
    record_id: UUID, employee_email: str = Form(...),
    service: DepartureRecordService = Depends(get_departure_service),
    session: AsyncSession = Depends(get_db),
):
    from app.modules.hr.mail_service import send_email
    from app.modules.hr.models import EmailLog
    from app.modules.hr.termination_certificate_generator import (
        generate_termination_certificate_pdf,
    )
    record = await service.get_record(record_id)
    name = record.name or "员工"

    # 兜底：离职台账缺少的字段从员工档案补（同名多人时不做猜测，避免错打他人证件号）
    id_number = record.id_card or ""
    entry_date = record.livo_entry_date or record.factory_entry_date or None
    if (not id_number or not entry_date) and record.name:
        from app.modules.hr.models import Employee
        emp_rows = list((await service.repo.session.execute(
            select(Employee).where(Employee.name == record.name, Employee.is_deleted == False)
        )).scalars().all())
        emp_row = emp_rows[0] if len(emp_rows) == 1 else None
        if emp_row:
            if not id_number:
                id_number = emp_row.id_card or ""
            if not entry_date:
                entry_date = emp_row.livo_entry_date or emp_row.factory_entry_date or emp_row.hire_date

    pdf_buf = generate_termination_certificate_pdf(
        name=name, id_number=id_number or "",
        department=record.department or "", position=record.position or "",
        entry_date=entry_date or "",
        leave_date=record.offboarding_date or "",
        leave_reason=record.offboarding_type or "个人原因",
    )
    filename = f"解除劳动关系证明_{name}.pdf"
    subj = "解除劳动关系证明"
    html = f"<html><body style=\"font-family:sans-serif;padding:20px;\"><h2>解除劳动关系证明</h2><p>{name}，您好！</p><p>附件是您的解除劳动关系证明，请查收。</p></body></html>"
    try:
        await send_email(to=employee_email, subject=subj, html_body=html, attachments=[(filename, pdf_buf.read())], session=session); st, err = "sent", None
    except Exception as e:
        st, err = "failed", str(e)
    session.add(EmailLog(email_type="departure_cert", employee_name=name, recipient=employee_email, subject=subj, status=st, error_message=err))
    await session.commit()
    if st == "failed": raise HTTPException(500, f"发送失败: {err}")
    return success_response(message="离职证明已发送")


# ─── TrainingLedger Routes ───

@router.get("/training-ledgers", summary="培训台账列表")
async def list_training_ledgers(
    employee_number: str | None = Query(None, description="工号筛选"),
    date_from: date | None = Query(None, description="培训日期起"),
    date_to: date | None = Query(None, description="培训日期止"),
    page_params: PageParams = Depends(),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    if hr_scope.is_unrestricted:
        records, total = await service.list_records(
            employee_number=employee_number,
            date_from=date_from,
            date_to=date_to,
            page=page_params.page,
            page_size=page_params.page_size,
            sort_by="training_date",
            sort_order="asc",
        )
    else:
        # 数据范围受限：join 员工表强制收敛到本部门/本人，查别部门工号返回空
        from app.modules.hr.models import Employee, OnboardingRecord, TrainingLedger
        departed_subq = select(OnboardingRecord.employee_number).where(
            OnboardingRecord.is_deleted == False, OnboardingRecord.is_employed == "否")
        stmt = (select(TrainingLedger)
            .join(Employee, TrainingLedger.employee_number == Employee.employee_number)
            .where(TrainingLedger.is_deleted == False, Employee.is_deleted == False,
                   TrainingLedger.employee_number.not_in(departed_subq)))
        if employee_number: stmt = stmt.where(TrainingLedger.employee_number == employee_number)
        if date_from: stmt = stmt.where(TrainingLedger.training_date >= date_from)
        if date_to: stmt = stmt.where(TrainingLedger.training_date <= date_to)
        stmt = hr_scope.apply_list_scope(stmt, Employee)
        total = (await session.execute(select(func.count()).select_from(stmt.subquery()))).scalar() or 0
        records = list((await session.execute(
            stmt.order_by(TrainingLedger.training_date.asc())
            .offset((page_params.page - 1) * page_params.page_size)
            .limit(page_params.page_size)
        )).scalars().all())
    data = [
        TrainingLedgerResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/training-ledgers", summary="创建培训台账记录")
async def create_training_ledger(
    payload: TrainingLedgerCreate,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    record = await service.create_record(payload)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
        message="培训台账记录创建成功",
        status_code=201,
    )


class BatchScoreUpdate(BaseModel):
    id: str
    assessment_result: str


class BatchScoreRequest(BaseModel):
    records: list[BatchScoreUpdate]


@router.post("/training-ledgers/batch-scores", summary="批量保存培训成绩")
async def batch_update_scores(
    body: BatchScoreRequest,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """批量更新培训台账中的考核成绩字段（受限用户仅可更新授权范围内记录）。"""
    updated = 0
    for rec in body.records:
        if not hr_scope.is_unrestricted:
            row = (await session.execute(
                text("SELECT department FROM hr.training_ledgers WHERE id = :id AND is_deleted = false"),
                {"id": rec.id},
            )).fetchone()
            if not row:
                continue
            if not hr_scope.scoped_departments:
                raise ForbiddenException("数据范围限制：仅可访问本人相关数据")
            if row[0] not in hr_scope.scoped_departments:
                raise ForbiddenException(f"数据范围限制：仅可操作授权部门记录（{row[0]}）")
        result = await session.execute(
            text("UPDATE hr.training_ledgers SET assessment_result = :r, updated_at = now() WHERE id = :id AND is_deleted = false"),
            {"r": rec.assessment_result, "id": rec.id},
        )
        updated += result.rowcount
    await session.commit()
    return success_response(data={"updated": updated}, message=f"已更新 {updated} 条")


# ─── TrainingLedgerPage Routes (must be before /{record_id}) ───

@router.get("/training-ledgers/pages", summary="已创建的培训台账页面列表")
async def list_training_ledger_pages(
    service: TrainingLedgerPageService = Depends(get_training_ledger_page_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    pages_with_dept = await service.list_pages_with_department()
    # 数据范围受限时只保留授权部门/本人的台账页面
    if not hr_scope.is_unrestricted:
        if hr_scope.scoped_departments:
            pages_with_dept = [(p, d) for p, d in pages_with_dept if d in hr_scope.scoped_departments]
        else:
            pages_with_dept = [(p, d) for p, d in pages_with_dept if p.employee_number == hr_scope.employee_number]
    data = [
        {
            "id": str(page.id),
            "employee_number": page.employee_number,
            "employee_name": page.employee_name,
            "department": dept or "未知部门",
            "created_at": page.created_at.isoformat() if page.created_at else None,
            "updated_at": page.updated_at.isoformat() if page.updated_at else None,
        }
        for page, dept in pages_with_dept
    ]
    return success_response(data=data)


@router.post("/training-ledgers/pages", summary="创建培训台账页面")
async def create_training_ledger_page(
    payload: TrainingLedgerPageCreate,
    service: TrainingLedgerPageService = Depends(get_training_ledger_page_service),
):
    page = await service.create_page(payload)
    return success_response(
        data=TrainingLedgerPageResponse(
            id=page.id,
            employee_number=page.employee_number,
            employee_name=page.employee_name,
            department=None,
            created_at=page.created_at,
            updated_at=page.updated_at,
        ).model_dump(mode="json"),
        message="培训台账页面创建成功",
        status_code=201,
    )


def _generate_training_ledger_excel(employee: dict, records: list[dict]) -> BytesIO:
    """Generate training ledger Excel based on employee training ledger format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "员工培训台账"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=16)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    ws.merge_cells("A1:G1")
    ws["A1"] = "丽珠集团福州福兴医药有限公司"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = "员工培训台账"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center_align
    ws.row_dimensions[2].height = 24

    ws["A3"] = "姓名"
    ws["A3"].font = bold_font
    ws["A3"].alignment = center_align
    ws["A3"].border = thin_border
    ws["B3"] = employee.get("name", "")
    ws["B3"].border = thin_border
    ws["C3"] = "性别"
    ws["C3"].font = bold_font
    ws["C3"].alignment = center_align
    ws["C3"].border = thin_border
    ws["D3"] = employee.get("gender", "")
    ws["D3"].border = thin_border
    ws["E3"] = "工作卡号"
    ws["E3"].font = bold_font
    ws["E3"].alignment = center_align
    ws["E3"].border = thin_border
    ws.merge_cells("F3:G3")
    ws["F3"] = employee.get("employee_number", "")
    ws["F3"].border = thin_border
    ws["G3"].border = thin_border

    ws["A4"] = "部门"
    ws["A4"].font = bold_font
    ws["A4"].alignment = center_align
    ws["A4"].border = thin_border
    ws["B4"] = employee.get("department", "")
    ws["B4"].border = thin_border
    ws["C4"] = "岗位/职务"
    ws["C4"].font = bold_font
    ws["C4"].alignment = center_align
    ws["C4"].border = thin_border
    ws["D4"] = employee.get("position", "")
    ws["D4"].border = thin_border
    ws["E4"] = "入厂时间"
    ws["E4"].font = bold_font
    ws["E4"].alignment = center_align
    ws["E4"].border = thin_border
    ws.merge_cells("F4:G4")
    ws["F4"] = employee.get("factory_entry_date") or employee.get("hire_date", "")
    ws["F4"].border = thin_border
    ws["G4"].border = thin_border

    ws["A5"] = "岗位变动"
    ws["A5"].font = bold_font
    ws["A5"].alignment = center_align
    ws["A5"].border = thin_border
    ws.merge_cells("B5:G5")
    ws["B5"] = employee.get("transfer_history", "无")
    ws["B5"].border = thin_border
    for c in range(3, 8):
        ws.cell(row=5, column=c).border = thin_border

    ws["A6"] = "记录"
    ws["A6"].font = bold_font
    ws["A6"].alignment = center_align
    ws["A6"].border = thin_border
    ws.merge_cells("B6:G6")
    ws["B6"] = ""
    ws["B6"].border = thin_border
    for c in range(3, 8):
        ws.cell(row=6, column=c).border = thin_border

    headers = ["年月日", "培训课程", "培训方式", "课时", "培训单位/培训师", "考核成绩", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[7].height = 24

    for idx, record in enumerate(records, 8):
        values = [
            record.get("training_date", ""),
            record.get("training_subject", ""),
            record.get("training_method", ""),
            record.get("duration_hours", ""),
            record.get("trainer", ""),
            record.get("assessment_result", ""),
            record.get("remarks", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align if col in (1, 3, 4, 6, 7) else left_align

    while len(records) < 12:
        row = 8 + len(records)
        for col in range(1, 8):
            ws.cell(row=row, column=col, value="").border = thin_border
        records.append({})

    footer_row = 8 + len(records)
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    ws.cell(row=footer_row, column=1, value="备注：笔试考核设置为满分100分，考试合格线为80分。")
    ws.cell(row=footer_row, column=1).alignment = left_align
    ws.cell(row=footer_row, column=1).border = thin_border
    for c in range(2, 8):
        ws.cell(row=footer_row, column=c).border = thin_border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/training-ledgers/export", summary="导出培训台账Excel")
async def export_training_ledger(
    employee_number: str = Query(..., description="员工工号"),
    ledger_service: TrainingLedgerService = Depends(get_training_ledger_service),
    employee_service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """根据员工数据生成并导出培训台账 Excel 文件。"""
    employee = await employee_service.get_employee_by_number(employee_number)
    if not employee:
        raise HTTPException(status_code=404, detail="未找到该员工")
    hr_scope.ensure_can_access_employee(employee)

    records, _ = await ledger_service.list_records(
        employee_number=employee_number,
        page=1,
        page_size=1000,
        sort_by="training_date",
        sort_order="asc",
    )

    employee_dict = EmployeeResponse.model_validate(employee).model_dump(mode="json")
    record_dicts = [
        TrainingLedgerResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]

    buffer = _generate_training_ledger_excel(employee_dict, record_dicts)
    buffer.seek(0)

    safe_name = employee.name or "unknown"
    filename = f"{safe_name}培训台账.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"
        },
    )


@router.get("/training-ledgers/admin", summary="管理员培训台账总览")
async def ledger_admin(department: str | None = Query(None), training_subject: str | None = Query(None), date_from: date | None = Query(None), date_to: date | None = Query(None), page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200), session: AsyncSession = Depends(get_db), hr_scope: HrAccessContext = Depends(get_hr_scope)):
    from app.modules.hr.models import Employee, TrainingLedger
    cols = (TrainingLedger.id, TrainingLedger.employee_number, Employee.name.label("employee_name"), TrainingLedger.training_subject, TrainingLedger.training_date, TrainingLedger.training_method, TrainingLedger.trainer, TrainingLedger.assessment_result, Employee.department)
    stmt = select(*cols).join(Employee, TrainingLedger.employee_number == Employee.employee_number).where(TrainingLedger.is_deleted == False, Employee.is_deleted == False)
    if hr_scope.is_unrestricted and department: stmt = stmt.where(Employee.department == department)
    stmt = hr_scope.apply_list_scope(stmt, Employee)
    if training_subject: stmt = stmt.where(TrainingLedger.training_subject.ilike(f"%{training_subject}%"))
    if date_from: stmt = stmt.where(TrainingLedger.training_date >= date_from)
    if date_to: stmt = stmt.where(TrainingLedger.training_date <= date_to)
    total = (await session.execute(select(func.count()).select_from(stmt.subquery()))).scalar() or 0
    rows = (await session.execute(stmt.order_by(TrainingLedger.training_date.desc()).offset((page-1)*page_size).limit(page_size))).all()
    return paginated_response(data=[{"id":str(r[0]),"employee_number":r[1],"employee_name":r[2],"training_subject":r[3],"training_date":str(r[4]) if r[4] else None,"training_method":r[5],"trainer":r[6],"assessment_result":r[7],"department":r[8]} for r in rows], page=page, page_size=page_size, total=total)


@router.get("/training-ledgers/admin/departments", summary="台账中的部门列表")
async def ledger_admin_departments(session: AsyncSession = Depends(get_db), hr_scope: HrAccessContext = Depends(get_hr_scope)):
    from app.modules.hr.models import Employee, TrainingLedger
    stmt = (select(Employee.department).select_from(TrainingLedger)
        .join(Employee, TrainingLedger.employee_number == Employee.employee_number)
        .where(TrainingLedger.is_deleted == False, Employee.is_deleted == False))
    # 数据范围受限角色的下拉框只给本部门/本人
    stmt = hr_scope.apply_list_scope(stmt, Employee)
    r = (await session.execute(stmt.distinct().order_by(Employee.department))).all()
    return success_response(data=[d[0] for d in r if d[0]])


@router.get("/training-ledgers/admin/subjects", summary="台账中的培训内容列表")
async def ledger_admin_subjects(department: str | None = Query(None), session: AsyncSession = Depends(get_db), hr_scope: HrAccessContext = Depends(get_hr_scope)):
    from app.modules.hr.models import Employee, TrainingLedger
    stmt = (select(TrainingLedger.training_subject).select_from(TrainingLedger)
        .join(Employee, TrainingLedger.employee_number == Employee.employee_number)
        .where(TrainingLedger.is_deleted == False, Employee.is_deleted == False))
    if hr_scope.is_unrestricted and department: stmt = stmt.where(Employee.department == department)
    stmt = hr_scope.apply_list_scope(stmt, Employee)
    r = (await session.execute(stmt.distinct().order_by(TrainingLedger.training_subject))).all()
    return success_response(data=[s[0] for s in r if s[0]])


@router.get("/training-ledgers/admin/stats", summary="培训台账统计")
async def ledger_admin_stats(department: str | None = Query(None), training_subject: str | None = Query(None), date_from: date | None = Query(None), date_to: date | None = Query(None), session: AsyncSession = Depends(get_db), hr_scope: HrAccessContext = Depends(get_hr_scope)):
    from app.modules.hr.models import Employee, TrainingLedger
    base = select(TrainingLedger.assessment_result).select_from(TrainingLedger).join(Employee, TrainingLedger.employee_number == Employee.employee_number).where(TrainingLedger.is_deleted == False, Employee.is_deleted == False)
    if hr_scope.is_unrestricted and department: base = base.where(Employee.department == department)
    base = hr_scope.apply_list_scope(base, Employee)
    if training_subject: base = base.where(TrainingLedger.training_subject == training_subject)
    if date_from: base = base.where(TrainingLedger.training_date >= date_from)
    if date_to: base = base.where(TrainingLedger.training_date <= date_to)
    rows = (await session.execute(base)).all()
    total = len(rows)
    qualified = sum(1 for r in rows if r[0] and _parse_score(r[0]) >= 80)
    return success_response(data={"total": total, "qualified": qualified, "pass_rate": f"{qualified/max(total,1)*100:.0f}%" if total else "0%"})

def _parse_score(v: str | None) -> int:
    if not v: return 0
    try: return int(float(v))
    except: return 0


class AddQuestionItems(BaseModel):
    items: list[dict]
    source: str = "手工录入"


@router.post("/question-bank", summary="添加题目")
async def add_question_bank_items(
    payload: AddQuestionItems,
    session: AsyncSession = Depends(get_db),
):
    """手动添加题目到题库。"""
    inserted = 0
    for q in payload.items:
        if not q.get("question"):
            continue
        await session.execute(
            text("INSERT INTO hr.question_bank (id, file_no, question, answer, score, source) VALUES (gen_random_uuid(), :fn, :q, :a, :s, :src)"),
            {"fn": q.get("file_no"), "q": q["question"], "a": q.get("answer"), "s": q.get("score", 10), "src": payload.source},
        )
        inserted += 1
    await session.commit()
    return success_response(data={"inserted": inserted}, message=f"已添加 {inserted} 题")


@router.get("/question-bank", summary="题库检索")
async def qbank_search(file_no: str | None = Query(None), keyword: str | None = Query(None), page: int = Query(1, ge=1), page_size: int = Query(200, ge=1, le=500), session: AsyncSession = Depends(get_db)):
    where = "WHERE is_deleted = false"
    params: dict = {"lim": page_size, "off": (page - 1) * page_size}
    if file_no: where += " AND file_no ILIKE :fn"; params["fn"] = f"%{file_no}%"
    if keyword: where += " AND (question ILIKE :kw OR subject ILIKE :kw)"; params["kw"] = f"%{keyword}%"
    r = await session.execute(text(f"SELECT id, file_no, question, answer, score, source, usage_count FROM hr.question_bank {where} ORDER BY usage_count DESC LIMIT :lim OFFSET :off"), params)
    total = (await session.execute(
        text(f"SELECT COUNT(*) FROM hr.question_bank {where}"),
        {k: v for k, v in params.items() if k != "lim" and k != "off"},
    )).scalar() or 0
    return success_response(
        data=[{"id":str(row[0]),"file_no":row[1],"question":row[2],"answer":row[3],"score":row[4],"source":row[5],"usage_count":row[6]} for row in r],
        meta={"total": total, "page": page, "page_size": page_size},
    )


@router.delete("/question-bank/{item_id}", summary="删除题目")
async def delete_question_bank_item(
    item_id: UUID,
    session: AsyncSession = Depends(get_db),
):
    from app.modules.hr.models import QuestionBank
    q = (await session.execute(
        select(QuestionBank).where(QuestionBank.id == item_id, QuestionBank.is_deleted == False)
    )).scalar_one_or_none()
    if not q:
        raise HTTPException(404, "题目不存在")
    q.is_deleted = True
    await session.flush()
    return success_response(message="已删除")


@router.post("/question-bank/import-docx", summary="从培训记录 docx 导入题库")
async def qbank_import_docx(
    file: UploadFile,
    session: AsyncSession = Depends(get_db),
):
    """上传培训记录 .docx，自动提取问答题目并导入题库。"""
    if not file.filename or not file.filename.endswith(".docx"):
        raise HTTPException(400, "仅支持 .docx 格式")
    try:
        content = await file.read()
    except Exception:
        raise HTTPException(400, "读取文件失败")

    from app.modules.hr.qa_docx_importer import parse_training_record

    try:
        subject, questions = parse_training_record(content, file.filename or "")
    except Exception as e:
        raise HTTPException(400, f"解析文档失败: {e}")

    if not questions:
        raise HTTPException(400, "未从文档中提取到问答题目，请确认文件包含问答记录表")

    imported = 0
    for q in questions:
        # 检查去重：同 file_no + question
        existing = (await session.execute(
            text("SELECT 1 FROM hr.question_bank WHERE file_no = :fn AND question = :q AND is_deleted = false"),
            {"fn": q["file_no"], "q": q["question"]},
        )).first()
        if existing:
            continue
        await session.execute(
            text("INSERT INTO hr.question_bank (id, file_no, question, answer, score, source) VALUES (gen_random_uuid(), :fn, :q, :a, :s, :src)"),
            {"fn": q["file_no"], "q": q["question"], "a": q["answer"], "s": q["score"], "src": "docx_import"},
        )
        imported += 1

    await session.flush()
    return success_response(
        data={"subject": subject, "total": len(questions), "imported": imported, "skipped": len(questions) - imported},
        message=f"导入完成：新增 {imported} 题，跳过 {len(questions) - imported} 题（重复）",
        status_code=201,
    )


# ─── QA Assessments ───

class QaAssessmentCreateBody(BaseModel):
    subject: str
    department: str | None = None
    training_date: str | None = None
    training_method: str | None = None
    assessment_method: str | None = "问答"
    trainer: str | None = None
    questions: list[dict] | None = None
    question_count: int = 0
    full_score: int = 100
    excellent_line: int = 90
    pass_line: int = 80
    trainee_names: list[str] = []


@router.post("/qa-assessments", summary="创建考核场次")
async def create_qa_assessment(
    payload: QaAssessmentCreateBody,
    session: AsyncSession = Depends(get_db),
):
    """创建问答考核场次，含选题快照和受训人员名单。"""
    import json

    total_score = sum((q.get("score", 10) or 10) for q in (payload.questions or []))
    if total_score == 0 and payload.full_score:
        total_score = payload.full_score

    def _date(v: str | None):
        if not v:
            return None
        try:
            return date.fromisoformat(v)
        except (ValueError, TypeError):
            return None

    result = await session.execute(
        text("""INSERT INTO hr.qa_assessments
            (subject, department, training_date, training_method, assessment_method, trainer,
             questions, question_count, full_score, excellent_line, pass_line, trainee_names)
            VALUES (:s, :d, :td, :tm, :am, :t, :q, :qc, :fs, :el, :pl, :tn)
            RETURNING id"""),
        {
            "s": payload.subject, "d": payload.department,
            "td": _date(payload.training_date), "tm": payload.training_method,
            "am": payload.assessment_method, "t": payload.trainer,
            "q": json.dumps(payload.questions, ensure_ascii=False) if payload.questions else None,
            "qc": payload.question_count,
            "fs": total_score, "el": payload.excellent_line, "pl": payload.pass_line,
            "tn": json.dumps(payload.trainee_names, ensure_ascii=False),
        },
    )
    assessment_id = result.scalar_one()
    await session.flush()
    return success_response(data={"id": str(assessment_id)}, message="考核场次创建成功", status_code=201)


@router.get("/qa-assessments", summary="考核场次列表")
async def list_qa_assessments(
    department: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """按部门查询考核场次列表。"""
    if not hr_scope.is_unrestricted:
        if not hr_scope.scoped_departments:
            # self_only 用户对考核场次无意义
            raise ForbiddenException("数据范围限制：仅可访问本人相关数据")
        # 数据范围受限角色强制只看授权部门，忽略前端传入的 department 参数
        department = next(iter(hr_scope.scoped_departments)) if len(hr_scope.scoped_departments) == 1 else None
    where = "WHERE is_deleted = false"
    params: dict = {"lim": page_size, "off": (page - 1) * page_size}
    if department:
        where += " AND department = :dept"
        params["dept"] = department
    elif not hr_scope.is_unrestricted and hr_scope.scoped_departments and len(hr_scope.scoped_departments) > 1:
        placeholders = ", ".join(f":dept_{i}" for i in range(len(hr_scope.scoped_departments)))
        where += f" AND department IN ({placeholders})"
        for i, d in enumerate(hr_scope.scoped_departments):
            params[f"dept_{i}"] = d
    total = (await session.execute(
        text(f"SELECT count(*) FROM hr.qa_assessments {where}"),
        {k: v for k, v in params.items() if k != "lim" and k != "off"},
    )).scalar() or 0
    rows = (await session.execute(
        text(f"SELECT id, subject, department, training_date, training_method, assessment_method, trainer, question_count, created_at FROM hr.qa_assessments {where} ORDER BY created_at DESC LIMIT :lim OFFSET :off"),
        params,
    )).fetchall()
    data = [
        {
            "id": str(r[0]), "subject": r[1], "department": r[2],
            "training_date": str(r[3]) if r[3] else None,
            "training_method": r[4], "assessment_method": r[5], "trainer": r[6],
            "question_count": r[7], "created_at": str(r[8]) if r[8] else None,
        }
        for r in rows
    ]
    return paginated_response(data=data, page=page, page_size=page_size, total=total)


class QaScoreSaveBody(BaseModel):
    assessed_date: str | None = None
    scores: list[dict] = []


@router.get("/qa-assessments/{assessment_id}", summary="考核场次详情")
async def get_qa_assessment(
    assessment_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """获取考核详情，含题目和成绩。"""
    import json as _json
    row = (await session.execute(
        text("SELECT id, subject, department, training_date, training_method, assessment_method, trainer, questions, question_count, full_score, excellent_line, pass_line, trainee_names, created_at FROM hr.qa_assessments WHERE id = :id AND is_deleted = false"),
        {"id": assessment_id},
    )).fetchone()
    if not row:
        raise HTTPException(404, "考核场次不存在")
    _ensure_qa_assessment_access(hr_scope, row[2])

    scores = (await session.execute(
        text("SELECT id, employee_name, employee_number, wrong_questions, total_score, grade, result_text, assessed_date FROM hr.qa_assessment_scores WHERE assessment_id = :aid AND is_deleted = false"),
        {"aid": assessment_id},
    )).fetchall()

    return success_response(data={
        "assessment": {
            "id": str(row[0]), "subject": row[1], "department": row[2],
            "training_date": str(row[3]) if row[3] else None,
            "training_method": row[4], "assessment_method": row[5], "trainer": row[6],
            "questions": (row[7] if isinstance(row[7], list) else _json.loads(row[7])) if row[7] else [],
            "question_count": row[8], "full_score": row[9],
            "excellent_line": row[10], "pass_line": row[11],
            "trainee_names": (row[12] if isinstance(row[12], list) else _json.loads(row[12])) if row[12] else [],
            "created_at": str(row[13]) if row[13] else None,
        },
        "scores": [
            {
                "id": str(s[0]), "employee_name": s[1], "employee_number": s[2],
                "wrong_questions": (s[3] if isinstance(s[3], list) else (_json.loads(s[3]) if s[3] else [])), "total_score": s[4],
                "grade": s[5], "result_text": s[6],
                "assessed_date": str(s[7]) if s[7] else None,
            }
            for s in scores
        ],
    })


@router.put("/qa-assessments/{assessment_id}/scores", summary="保存考核成绩")
async def save_qa_scores(
    assessment_id: UUID,
    payload: QaScoreSaveBody,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """保存或更新考核成绩。"""
    import json as _json
    from datetime import date as _date

    def _d(v: str | None):
        if not v:
            return None
        try:
            return _date.fromisoformat(v)
        except (ValueError, TypeError):
            return None

    # 获取考核场次题目信息用于准确计分
    assess_row = (await session.execute(
        text("SELECT questions, full_score, excellent_line, pass_line, department FROM hr.qa_assessments WHERE id = :id AND is_deleted = false"),
        {"id": assessment_id},
    )).fetchone()
    if not assess_row:
        raise HTTPException(404, "考核场次不存在或已删除")
    _ensure_qa_assessment_access(hr_scope, assess_row[4])
    questions_data: list[dict] = []
    full_score_val = 100
    if assess_row:
        qs = assess_row[0]
        questions_data = qs if isinstance(qs, list) else (_json.loads(qs) if isinstance(qs, str) else [])
        full_score_val = assess_row[1] or 100

    # 更新考核日期
    if payload.assessed_date:
        await session.execute(
            text("UPDATE hr.qa_assessments SET training_date = :d WHERE id = :id"),
            {"d": _d(payload.assessed_date), "id": assessment_id},
        )

    el = assess_row[2] or 90
    pl = assess_row[3] or 80
    for s in payload.scores:
        name = s.get("employee_name", "")
        if not name:
            continue
        wrong = s.get("wrong_questions") or []
        # 根据实际题目分值计算总分
        if questions_data:
            deduction = sum(
                (questions_data[i - 1].get("score", 10) if isinstance(questions_data[i - 1], dict) else 10)
                for i in wrong if isinstance(i, int) and 1 <= i <= len(questions_data)
            )
            score = max(0, full_score_val - deduction)
        else:
            per_q = full_score_val / max(len(questions_data) or 10, 1)
            score = max(0, full_score_val - len(wrong) * int(per_q))

        grade = "优秀" if score >= el else ("合格" if score >= pl else "不合格")
        wrong_sorted = sorted(wrong)
        result_text = f"第{'、'.join(str(i) for i in wrong_sorted)}题错误" if wrong else "全对"
        existing = (await session.execute(
            text("SELECT id FROM hr.qa_assessment_scores WHERE assessment_id = :aid AND employee_name = :nm AND is_deleted = false"),
            {"aid": assessment_id, "nm": name},
        )).first()
        if existing:
            await session.execute(
                text("UPDATE hr.qa_assessment_scores SET wrong_questions=:wq, total_score=:ts, assessed_date=:ad, grade=:g, result_text=:rt WHERE id=:id"),
                {"wq": _json.dumps(wrong), "ts": score, "ad": _d(payload.assessed_date), "g": grade, "rt": result_text, "id": existing[0]},
            )
        else:
            await session.execute(
                text("INSERT INTO hr.qa_assessment_scores (assessment_id, employee_name, employee_number, wrong_questions, total_score, grade, result_text, assessed_date) VALUES (:aid, :nm, :en, :wq, :ts, :g, :rt, :ad)"),
                {
                    "aid": assessment_id, "nm": name,
                    "en": s.get("employee_number", ""),
                    "wq": _json.dumps(wrong), "ts": score,
                    "g": grade, "rt": result_text,
                    "ad": _d(payload.assessed_date),
                },
            )

    await session.flush()

    # 同步写入培训台账
    assessment = (await session.execute(
        text("SELECT subject, training_date, training_method, trainer FROM hr.qa_assessments WHERE id = :id"),
        {"id": assessment_id},
    )).fetchone()
    if assessment:
        subj, train_date, method, trainer = assessment
        for s in payload.scores:
            name = s.get("employee_name", "")
            emp_no = s.get("employee_number", "")
            if not name:
                continue
            # 没工号时从员工表反查（同名多人时不做猜测，跳过该行）
            if not emp_no:
                emp_rows = (await session.execute(
                    text("SELECT employee_number FROM hr.employees WHERE name = :nm AND is_deleted = false"),
                    {"nm": name},
                )).fetchall()
                if len(emp_rows) == 1 and emp_rows[0][0]:
                    emp_no = emp_rows[0][0]
            if not emp_no:
                continue
            wrong = s.get("wrong_questions") or []
            if questions_data:
                deduction = sum(
                    (questions_data[i - 1].get("score", 10) if isinstance(questions_data[i - 1], dict) else 10)
                    for i in wrong if isinstance(i, int) and 1 <= i <= len(questions_data)
                )
                score = max(0, full_score_val - deduction)
            else:
                per_q = full_score_val / max(len(questions_data) or 10, 1)
                score = max(0, full_score_val - len(wrong) * int(per_q))
            ledger_date = _d(payload.assessed_date) or train_date or _date.today()
            exist = (await session.execute(
                text("SELECT id FROM hr.training_ledgers WHERE employee_number = :en AND training_date = :td AND training_subject = :ts AND is_deleted = false"),
                {"en": emp_no, "td": ledger_date, "ts": subj},
            )).first()
            if exist:
                await session.execute(
                    text("UPDATE hr.training_ledgers SET assessment_result = :ar, training_method = :tm, trainer = :t, updated_at = now() WHERE id = :id"),
                    {"ar": str(score), "tm": method, "t": trainer, "id": exist[0]},
                )
            else:
                await session.execute(
                    text("INSERT INTO hr.training_ledgers (id, employee_number, training_date, training_subject, training_method, trainer, assessment_result, source_type) VALUES (gen_random_uuid(), :en, :td, :ts, :tm, :t, :ar, 'qa_assessment')"),
                    {"en": emp_no, "td": ledger_date, "ts": subj, "tm": method, "t": trainer, "ar": str(score)},
                )

    await session.commit()
    return success_response(message="成绩保存成功，已同步培训台账")


@router.post("/qa-assessments/{assessment_id}/random-scores", summary="随机生成考核成绩")
async def random_qa_scores(
    assessment_id: UUID,
    excellent_ratio: float = Query(0.3, ge=0, le=1, description="优秀比例，默认30%"),
    pass_ratio: float | None = Query(None, ge=0, le=1, description="合格比例（缺省余量全为合格）"),
    excellent_line: int | None = Query(None, description="优秀线，缺省取场次配置"),
    pass_line: int | None = Query(None, description="合格线，缺省取场次配置"),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """按优秀/合格比例随机赋分。

    优秀率+合格率由人员自行填写，余量自动为不合格（分数落在合格线以下）。
    """
    import random as _random
    asmt = (await session.execute(
        text("SELECT id, full_score, excellent_line, pass_line, trainee_names, questions, department FROM hr.qa_assessments WHERE id = :id AND is_deleted = false"),
        {"id": assessment_id},
    )).fetchone()
    if not asmt:
        raise HTTPException(404, "考核场次不存在")
    _ensure_qa_assessment_access(hr_scope, asmt[6])
    full_score_val = asmt[1] or 100
    el = excellent_line if excellent_line is not None else (asmt[2] or 90)
    pl = pass_line if pass_line is not None else (asmt[3] or 80)
    questions_raw = asmt[5]
    questions = questions_raw if isinstance(questions_raw, list) else (json.loads(questions_raw) if questions_raw else [])
    trainee_raw = asmt[4]
    all_names = trainee_raw if isinstance(trainee_raw, list) else (json.loads(trainee_raw) if trainee_raw else [])

    # 去重 + 保证顺序
    unique_names = list(dict.fromkeys(str(n).strip() for n in all_names if str(n).strip()))
    total = len(unique_names)
    if total == 0:
        raise HTTPException(400, "没有受训人员")
    if pass_ratio is not None and excellent_ratio + pass_ratio > 1:
        raise HTTPException(400, "优秀率与合格率之和不能超过 1")
    # 区间边界校验：等级线不合法时扣分循环无法收敛，直接拒绝
    if el <= pl or pl < 1 or el < 1:
        raise HTTPException(400, f"优秀线/合格线配置不正确（优秀线 {el}，合格线 {pl}）")
    if el > full_score_val:
        raise HTTPException(400, f"优秀线 {el} 不能高于满分 {full_score_val}")
    # 四舍五入分配人数：默认 0.3/0.7 在 7 人场次应为 2 优 5 合格，
    # 截断会把余量静默打成不合格
    excellent_count = int(total * excellent_ratio + 0.5)
    pass_count = int(total * (pass_ratio or 0) + 0.5)
    indices = list(range(total))
    _random.shuffle(indices)
    excellent_set = set(indices[:excellent_count])
    pass_set: set[int] = set()
    if pass_ratio is not None:
        pass_set = set(indices[excellent_count:excellent_count + pass_count])

    scores = []
    q_scores = [q.get("score", 10) if isinstance(q, dict) else 10 for q in questions]
    total_q = len(q_scores)
    today = date.today()

    for i, name in enumerate(unique_names):
        is_excellent = i in excellent_set
        is_fail = pass_ratio is not None and i not in excellent_set and i not in pass_set
        # 目标区间：优秀 [el, 满分]，合格 [pl, el-1]，不合格 [0, pl-1]
        if is_fail:
            min_target, max_target = 0, pl - 1
        elif is_excellent:
            min_target, max_target = el, full_score_val
        else:
            min_target, max_target = pl, el - 1
        # 随机选择错题直到分数落到目标区间；
        # 题目分值较粗导致无组合落入区间时（如 4 题×25 分目标 [80,89]），
        # 取距区间最近的兜底组合，避免目标合格的学员直接得满分（优秀）
        wrong = []
        deduction = 0
        fallback_wrong: list[int] = []
        fallback_deduction = 0
        found_fallback = False
        available = list(range(total_q))
        _random.shuffle(available)
        for qi in available:
            potential = deduction + q_scores[qi]
            score_after = full_score_val - potential
            if score_after < min_target:
                # 低于区间下限：第一次越界即「最接近区间」的兜底组合
                if not found_fallback and score_after >= 0:
                    fallback_wrong = wrong + [qi + 1]
                    fallback_deduction = potential
                    found_fallback = True
                continue
            deduction = potential
            wrong.append(qi + 1)  # 1-indexed
            if score_after <= max_target:
                break
        score = max(0, full_score_val - deduction)
        if score > max_target and found_fallback:
            # 无组合落入区间：比较「区间上方的最近分数」与「下方的最近分数」取更近者
            dist_above = score - max_target
            dist_below = min_target - (full_score_val - fallback_deduction)
            if dist_below < dist_above:
                wrong = fallback_wrong
                deduction = fallback_deduction
                score = max(0, full_score_val - deduction)
        # 查找工号
        emp_no = ""
        emp = (await session.execute(
            text("SELECT employee_number FROM hr.employees WHERE name = :nm AND is_deleted = false LIMIT 1"),
            {"nm": name},
        )).fetchone()
        if emp:
            emp_no = emp[0]
        wrong.sort()
        result_text = f"第{'、'.join(str(i) for i in wrong)}题错误" if wrong else "全对"
        scores.append({
            "employee_name": name,
            "employee_number": emp_no,
            "wrong_questions": wrong,
            "total_score": score,
            "grade": "优秀" if score >= el else ("合格" if score >= pl else "不合格"),
            "result_text": result_text,
            "assessed_date": today,
        })

    # 写入数据库
    for s in scores:
        existing = (await session.execute(
            text("SELECT id FROM hr.qa_assessment_scores WHERE assessment_id = :aid AND employee_name = :nm"),
            {"aid": str(assessment_id), "nm": s["employee_name"]},
        )).fetchone()
        if existing:
            await session.execute(
                text("UPDATE hr.qa_assessment_scores SET wrong_questions = :wq, total_score = :ts, employee_number = :en, assessed_date = :ad, grade = :g, result_text = :rt WHERE id = :id"),
                {"wq": json.dumps(s["wrong_questions"]), "ts": s["total_score"], "en": s["employee_number"], "ad": s["assessed_date"], "g": s["grade"], "rt": s["result_text"], "id": existing[0]},
            )
        else:
            await session.execute(
                text("INSERT INTO hr.qa_assessment_scores (assessment_id, employee_name, employee_number, wrong_questions, total_score, grade, result_text, assessed_date) VALUES (:aid, :nm, :en, :wq, :ts, :g, :rt, :ad)"),
                {"aid": str(assessment_id), "nm": s["employee_name"], "en": s["employee_number"], "wq": json.dumps(s["wrong_questions"]), "ts": s["total_score"], "g": s["grade"], "rt": s["result_text"], "ad": s["assessed_date"]},
            )
    await session.commit()
    return success_response(data={"generated": len(scores), "scores": scores}, message=f"已随机生成 {len(scores)} 人成绩")


@router.post("/qa-assessments/{assessment_id}/sync-ledger", summary="同步成绩到培训台账")
async def sync_qa_to_ledger(
    assessment_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """将考核场次的成绩批量同步到培训台账。"""
    from datetime import date as _date

    # 获取考核场次
    assess = (await session.execute(
        text("SELECT subject, training_date, training_method, trainer, department FROM hr.qa_assessments WHERE id = :id AND is_deleted = false"),
        {"id": assessment_id},
    )).fetchone()
    if not assess:
        raise HTTPException(404, "考核场次不存在")
    _ensure_qa_assessment_access(hr_scope, assess[4])

    subj, train_date, method, trainer = assess[0], assess[1], assess[2], assess[3]
    td = train_date or _date.today()

    # 获取所有成绩
    scores = (await session.execute(
        text("SELECT employee_name, employee_number, total_score FROM hr.qa_assessment_scores WHERE assessment_id = :aid AND is_deleted = false"),
        {"aid": assessment_id},
    )).fetchall()

    if not scores:
        raise HTTPException(400, "该考核场次还没有成绩记录，请先保存成绩")

    synced = 0
    skipped_exist = 0
    no_emp = 0
    no_emp_names: list[str] = []

    for name, emp_no, score in scores:
        # 反查工号（同名多人时不做猜测）
        if not emp_no:
            emp_rows = (await session.execute(
                text("SELECT employee_number FROM hr.employees WHERE name = :nm AND is_deleted = false"),
                {"nm": name},
            )).fetchall()
            if len(emp_rows) == 1 and emp_rows[0][0]:
                emp_no = emp_rows[0][0]
        if not emp_no:
            no_emp += 1
            no_emp_names.append(name or "?")
            continue

        # 去重按 员工+科目+日期：同名科目不同场次的培训各建一条台账
        exist = (await session.execute(
            text("SELECT 1 FROM hr.training_ledgers WHERE employee_number = :en AND training_subject = :ts AND training_date = :td AND is_deleted = false"),
            {"en": emp_no, "ts": subj, "td": td},
        )).fetchone()
        if exist:
            skipped_exist += 1
            continue

        await session.execute(
            text("INSERT INTO hr.training_ledgers (id, employee_number, training_date, training_subject, training_method, trainer, assessment_result, source_type) VALUES (gen_random_uuid(), :en, :td, :ts, :tm, :t, :ar, 'qa_assessment')"),
            {"en": emp_no, "td": td, "ts": subj, "tm": method or "", "t": trainer or "", "ar": str(score or 0)},
        )
        synced += 1

    await session.commit()

    msg = f"已同步 {synced} 人到培训台账"
    if skipped_exist:
        msg += f"，{skipped_exist} 人已存在跳过"
    if no_emp:
        msg += f"，{no_emp} 人缺工号未同步"
        if no_emp_names:
            msg += f"（{'、'.join(no_emp_names[:5])}{'...' if len(no_emp_names) > 5 else ''}）"

    return success_response(
        data={"synced": synced, "skipped": skipped_exist, "no_emp": no_emp},
        message=msg,
    )


@router.delete("/qa-assessments/{assessment_id}", summary="删除考核场次")
async def delete_qa_assessment(
    assessment_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """软删除考核场次及关联成绩（受限用户仅可删除授权部门场次）。"""
    row = (await session.execute(
        text("SELECT department FROM hr.qa_assessments WHERE id = :id AND is_deleted = false"),
        {"id": assessment_id},
    )).fetchone()
    if not row:
        raise HTTPException(404, "考核场次不存在")
    _ensure_qa_assessment_access(hr_scope, row[0])
    await session.execute(text("UPDATE hr.qa_assessments SET is_deleted = true WHERE id = :id"), {"id": assessment_id})
    await session.execute(text("UPDATE hr.qa_assessment_scores SET is_deleted = true WHERE assessment_id = :id"), {"id": assessment_id})
    await session.flush()
    return success_response(message="考核场次已删除")


class QaRecordExportRequest(BaseModel):
    training_content: str = ""
    training_date: str = ""
    training_method: str = ""
    training_department: str = ""
    questions_json: str = "[]"
    trainee_names_json: str = "[]"
    scores_json: str = "[]"
    trainer_name: str = ""


@router.post("/training-notification/export-qa-record-with-scores", summary="导出含错题的完整实操记录表")
async def export_training_qa_record_with_scores(
    body: QaRecordExportRequest,
    session: AsyncSession = Depends(get_db),
):
    """导出含学员错题记录和评分的完整问答实操记录表。"""
    import json as _json

    from app.modules.hr.qa_record_generator import generate_qa_record

    questions = _json.loads(body.questions_json) if body.questions_json else []
    trainee_names = _json.loads(body.trainee_names_json) if body.trainee_names_json else []
    scores_raw = _json.loads(body.scores_json) if body.scores_json else []
    training_date_val = date.fromisoformat(body.training_date) if body.training_date else None

    # 构建带错题描述的 score entries
    score_entries = []
    for s in scores_raw:
        wrong_indices = s.get("wrong_questions", []) or []
        total_q = len(questions)
        if wrong_indices and total_q:
            # wrong_questions 约定为 1 基序号（与保存路径一致）
            wrong_nums = "、".join(str(i) for i in wrong_indices if isinstance(i, int) and 1 <= i <= total_q)
            result_text = f"第{wrong_nums}题错误，其他题目正确"
        else:
            result_text = "全对，所有题目回答正确"
        score_entries.append({
            "name": s.get("name", ""),
            "employee_number": "",
            "total_score": str(s.get("total_score", "")),
            "assessed_date": str(training_date_val) if training_date_val else "",
            "result_text": result_text,
        })

    # 尝试保存考核记录到数据库（失败不影响导出；同场次幂等，不重复插入）
    try:
        import uuid as _uuid
        # 同科目+部门+日期按最新场次复用（历史重复场次取最近一条，避免错题对不上题集）
        existing_row = (await session.execute(
            text(
                "SELECT id, excellent_line, pass_line FROM hr.qa_assessments "
                "WHERE subject = :s AND department = :d AND training_date = :date AND is_deleted = false "
                "ORDER BY created_at DESC LIMIT 1"
            ),
            {"s": body.training_content or "问答考核", "d": body.training_department, "date": training_date_val},
        )).fetchone()
        if existing_row:
            assessment_id = existing_row[0]
            el = existing_row[1] or 90
            pl = existing_row[2] or 80
            # 复用已有场次时同样把本次成绩写入（幂等：按姓名 upsert，不静默丢成绩）
            for s in scores_raw:
                name = s.get("name", "")
                if not name:
                    continue
                score = s.get("total_score", 0)
                wrong = _json.dumps(s.get("wrong_questions", []) or [])
                grade = "优秀" if score >= el else ("合格" if score >= pl else "不合格")
                prev = (await session.execute(
                    text("SELECT id FROM hr.qa_assessment_scores WHERE assessment_id = :aid AND employee_name = :nm AND is_deleted = false"),
                    {"aid": str(assessment_id), "nm": name},
                )).fetchone()
                if prev:
                    await session.execute(
                        text("UPDATE hr.qa_assessment_scores SET wrong_questions = :wrong, total_score = :score, assessed_date = :date, grade = :g WHERE id = :id"),
                        {"wrong": wrong, "score": score, "date": training_date_val, "g": grade, "id": prev[0]},
                    )
                else:
                    await session.execute(
                        text("INSERT INTO hr.qa_assessment_scores (id, assessment_id, employee_name, wrong_questions, total_score, assessed_date) VALUES (gen_random_uuid(), :aid, :name, :wrong, :score, :date)"),
                        {"aid": assessment_id, "name": name, "wrong": _json.dumps(s.get("wrong_questions", [])), "score": s.get("total_score", 0), "date": training_date_val},
                    )
        else:
            assessment_id = _uuid.uuid4()
            await session.execute(
                text("INSERT INTO hr.qa_assessments (id, subject, department, training_date, training_method, trainer, questions, trainee_names) VALUES (:id, :subject, :dept, :date, :method, :trainer, :questions, :trainees)"),
                {"id": assessment_id, "subject": body.training_content or "问答考核", "dept": body.training_department, "date": training_date_val, "method": body.training_method, "trainer": body.trainer_name, "questions": _json.dumps(questions), "trainees": _json.dumps(trainee_names)},
            )
            for s in scores_raw:
                name = s.get("name", "")
                if not name:
                    continue
                score = s.get("total_score", 0)
                grade = "优秀" if score >= 90 else ("合格" if score >= 80 else "不合格")
                await session.execute(
                    text("INSERT INTO hr.qa_assessment_scores (id, assessment_id, employee_name, wrong_questions, total_score, grade, assessed_date) VALUES (gen_random_uuid(), :aid, :name, :wrong, :score, :g, :date)"),
                    {"aid": assessment_id, "name": name, "wrong": _json.dumps(s.get("wrong_questions", [])), "score": score, "g": grade, "date": training_date_val},
                )
        await session.commit()
    except Exception:
        await session.rollback()

    buffer = generate_qa_record(
        training_content=body.training_content,
        training_date=training_date_val,
        training_method=body.training_method,
        training_department=body.training_department,
        questions=questions,
        trainee_names=trainee_names,
        scores=score_entries,
        trainer_name=body.trainer_name,
    )
    filename = f"问答实操记录表_{body.training_date}.docx"
    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{quote(filename)}"},
    )


@router.post("/training-notification/export-score-report", summary="导出成绩单")
async def export_score_report(body: QaRecordExportRequest, session: AsyncSession = Depends(get_db)):
    """导出考核成绩单，同时自动录入培训台账。"""
    import json as _json

    from app.modules.hr.score_report_generator import generate_score_report

    scores_raw = _json.loads(body.scores_json) if body.scores_json else []
    training_date_val = date.fromisoformat(body.training_date) if body.training_date else None
    scores = [{"name": s.get("name", ""), "department": body.training_department, "total_score": s.get("total_score", 0)} for s in scores_raw]

    # 自动将成绩写入培训台账
    for s in scores_raw:
        name = s.get("name", "")
        score = s.get("total_score", 0)
        if not name:
            continue
        # 查找该员工的工号（与其他成绩路径统一走 hr.employees）；
        # 同名多人时优先在职、按工号确定序，避免成绩挂到错误员工名下
        emp_row = (await session.execute(
            text(
                "SELECT employee_number FROM hr.employees WHERE name = :n AND is_deleted = false "
                "ORDER BY CASE WHEN status = '在职' THEN 0 ELSE 1 END, employee_number LIMIT 1"
            ),
            {"n": name},
        )).fetchone()
        emp_no = emp_row[0] if emp_row else None
        if emp_no:
            # 检查是否已有该场次培训记录，有则更新成绩（同名科目不同日期各建一条）
            existing = (await session.execute(
                text("SELECT id FROM hr.training_ledgers WHERE employee_number = :en AND training_subject = :subj AND training_date = :date AND is_deleted = false LIMIT 1"),
                {"en": emp_no, "subj": body.training_content, "date": training_date_val},
            )).fetchone()
            if existing:
                await session.execute(
                    text("UPDATE hr.training_ledgers SET assessment_result = :r, updated_at = now() WHERE id = :id"),
                    {"r": str(score), "id": existing[0]},
                )
            else:
                await session.execute(
                    text("INSERT INTO hr.training_ledgers (id, employee_number, training_subject, training_date, assessment_result, source_type) VALUES (gen_random_uuid(), :en, :subj, :date, :r, 'manual')"),
                    {"en": emp_no, "subj": body.training_content, "date": training_date_val, "r": str(score)},
                )
    await session.commit()

    buffer = generate_score_report(training_content=body.training_content, training_date=body.training_date, department=body.training_department, scores=scores)
    filename = f"成绩单_{body.training_date}.docx"
    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{quote(filename)}"},
    )


@router.post("/training-notification/export-qa-record", summary="导出问答实操记录表")
async def export_training_qa_record(
    training_content: str = Form("", description="培训内容"),
    training_purpose: str = Form("", description="培训目的"),
    training_date: str = Form("", description="培训日期"),
    training_method: str = Form("", description="培训方式"),
    training_department: str = Form("", description="受训部门"),
    questions_json: str = Form("[]", description="题目JSON"),
    trainee_names_json: str = Form("[]", description="学员姓名JSON"),
):
    """从培训通知页直接导出问答实操记录表（无评分数据，供打印后手写）。"""
    import json as _json

    from app.modules.hr.qa_record_generator import generate_qa_record

    questions = _json.loads(questions_json) if questions_json else []
    trainee_names = _json.loads(trainee_names_json) if trainee_names_json else []
    training_date_val = date.fromisoformat(training_date) if training_date else None

    buffer = generate_qa_record(
        training_content=training_content,
        training_purpose=training_purpose,
        training_date=training_date_val,
        training_method=training_method,
        training_department=training_department,
        questions=questions,
        trainee_names=trainee_names,
    )
    filename = f"问答实操记录表_{training_date}.docx"
    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{quote(filename)}"},
    )


@router.get("/qa-assessments/{assessment_id}/export-record", summary="导出问答记录表")
async def export_qa_record(
    assessment_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """导出问答实操记录表 docx。"""
    import json as _json

    from app.modules.hr.qa_record_generator import generate_qa_record

    row = (await session.execute(
        text("SELECT subject, department, training_date, training_method, trainer, questions, trainee_names FROM hr.qa_assessments WHERE id = :id AND is_deleted = false"),
        {"id": assessment_id},
    )).fetchone()
    if not row:
        raise HTTPException(404, "考核场次不存在")
    _ensure_qa_assessment_access(hr_scope, row[1])

    scores = (await session.execute(
        text("SELECT employee_name, employee_number, wrong_questions, total_score, assessed_date FROM hr.qa_assessment_scores WHERE assessment_id = :aid AND is_deleted = false"),
        {"aid": assessment_id},
    )).fetchall()

    questions = row[5] if isinstance(row[5], list) else (_json.loads(row[5]) if row[5] else [])
    trainee_names = row[6] if isinstance(row[6], list) else (_json.loads(row[6]) if row[6] else [])

    score_entries = []
    for s in scores:
        wrong = s[2] if isinstance(s[2], list) else (_json.loads(s[2]) if s[2] else [])
        wrong_set = set(wrong)
        total_q = len(questions)
        # wrong_questions 存储为 1 基序号，渲染需用 1 基比较（原 0 基比较整体错位一位）
        wrong_nums = "、".join(str(i) for i in range(1, total_q + 1) if i in wrong_set) if total_q else ""
        result_text = f"第{wrong_nums}题错误，其他题目正确" if wrong_nums else "全对，所有题目回答正确"
        score_entries.append({
            "name": s[0],
            "employee_number": s[1] or "",
            "total_score": str(s[3]) if s[3] is not None else "",
            "assessed_date": str(s[4]) if s[4] else "",
            "result_text": result_text,
        })

    try:
        buffer = generate_qa_record(
            training_content=row[0] or "",
            training_department=row[1] or "",
            training_date=row[2],
            training_method=row[3] or "",
            trainer_name=row[4] or "",
            questions=questions,
            trainee_names=trainee_names,
            scores=score_entries,
        )
    except Exception as e:
        raise HTTPException(400, f"生成记录表失败: {e}")

    def _iter(): buffer.seek(0); yield buffer.read()
    return StreamingResponse(_iter(), media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": "attachment; filename=qa_record.docx"})


@router.get("/qa-assessments/{assessment_id}/export-evaluation", summary="导出培训效果评估表")
async def export_qa_evaluation(
    assessment_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """导出培训效果评估表 docx。自动从成绩数据汇总统计。"""
    from app.modules.hr.evaluation_document_generator import (
        TrainingEvaluationInput,
        generate_training_evaluation,
    )

    row = (await session.execute(
        text("SELECT subject, department, training_date, training_method, trainer, assessment_method, excellent_line, pass_line FROM hr.qa_assessments WHERE id = :id AND is_deleted = false"),
        {"id": assessment_id},
    )).fetchone()
    if not row:
        raise HTTPException(404, "考核场次不存在")
    _ensure_qa_assessment_access(hr_scope, row[1])

    scores = (await session.execute(
        text("SELECT employee_name, total_score FROM hr.qa_assessment_scores WHERE assessment_id = :aid AND is_deleted = false"),
        {"aid": assessment_id},
    )).fetchall()

    # 从成绩汇总统计（等级线取场次配置，缺省 90/80）
    excellent_line = row[6] or 90
    pass_line = row[7] or 80
    total = len(scores)
    excellent = sum(1 for s in scores if (s[1] or 0) >= excellent_line)
    qualified = sum(1 for s in scores if pass_line <= (s[1] or 0) < excellent_line)
    unqualified = sum(1 for s in scores if (s[1] or 0) < pass_line)

    # 应到人数 = 部门总人数
    dept_count = (await session.execute(
        text("SELECT count(*) FROM hr.employees WHERE department = :d AND is_deleted = false"),
        {"d": row[1] or ""},
    )).scalar()
    expected = dept_count if dept_count else None

    pass_rate = f"{(excellent + qualified) / max(total, 1) * 100:.0f}%" if total > 0 else ""
    participation = f"{total / max(expected or total, 1) * 100:.0f}%" if total > 0 else ""

    payload = TrainingEvaluationInput(
        subject=row[0] or "",
        training_date=row[2],
        training_method=row[3] or "",
        trainer=row[4] or "",
        assessment_method=row[5] or "",
        trainee_names=[s[0] for s in scores],
        expected_count=expected,
        actual_count=total if total > 0 else None,
        exam_count=total if total > 0 else None,
        excellent_count=excellent if total > 0 else None,
        qualified_count=qualified if total > 0 else None,
        unqualified_count=unqualified if total > 0 else None,
        pass_rate=pass_rate or None,
        participation_rate=participation or None,
    )
    try:
        buffer = generate_training_evaluation(payload)
    except Exception as e:
        raise HTTPException(400, f"生成评估表失败: {e}")

    def _iter(): buffer.seek(0); yield buffer.read()
    return StreamingResponse(_iter(), media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": "attachment; filename=qa_evaluation.docx"})


@router.get("/qa-assessments/{assessment_id}/export-scores", summary="导出成绩单")
async def export_qa_scores(
    assessment_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """导出成绩单 docx（使用成绩单模板），并自动同步成绩到培训台账。"""
    from datetime import date as _date

    from app.modules.hr.score_report_generator import generate_score_report

    row = (await session.execute(
        text("SELECT subject, department, training_date, training_method, trainer FROM hr.qa_assessments WHERE id = :id AND is_deleted = false"),
        {"id": assessment_id},
    )).fetchone()
    if not row:
        raise HTTPException(404, "考核场次不存在")
    _ensure_qa_assessment_access(hr_scope, row[1])

    subj, dept, train_date, method, trainer = row

    scores = (await session.execute(
        text("SELECT employee_name, employee_number, total_score FROM hr.qa_assessment_scores WHERE assessment_id = :aid AND is_deleted = false ORDER BY employee_name"),
        {"aid": assessment_id},
    )).fetchall()

    dept = dept or ""
    score_list = [{"name": s[0], "total_score": s[2] or 0, "department": dept} for s in scores]

    # 同步到培训台账
    def _d(v):
        if not v: return None
        try: return _date.fromisoformat(str(v))
        except: return None

    td = _d(train_date) or _date.today()
    for s in scores:
        name, emp_no, score = s[0], s[1], s[2] or 0
        # 没工号时从员工表反查
        if not emp_no:
            emp = (await session.execute(
                text("SELECT employee_number FROM hr.employees WHERE name = :nm AND is_deleted = false LIMIT 1"),
                {"nm": name},
            )).fetchone()
            if emp and emp[0]:
                emp_no = emp[0]
        if not emp_no:
            continue
        exist = (await session.execute(
            text("SELECT 1 FROM hr.training_ledgers WHERE employee_number = :en AND training_date = :td AND training_subject = :ts AND is_deleted = false"),
            {"en": emp_no, "td": td, "ts": subj},
        )).first()
        if not exist:
            await session.execute(
                text("INSERT INTO hr.training_ledgers (id, employee_number, training_date, training_subject, training_method, trainer, assessment_result, source_type) VALUES (gen_random_uuid(), :en, :td, :ts, :tm, :t, :ar, 'qa_assessment')"),
                {"en": emp_no, "td": td, "ts": subj, "tm": method, "t": trainer, "ar": str(score)},
            )
    await session.commit()

    try:
        buffer = generate_score_report(
            training_content=row[0] or "",
            training_date=str(row[2] or ""),
            department=row[1] or "",
            scores=score_list,
        )
    except FileNotFoundError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(400, f"生成成绩单失败: {e}")

    def _iter(): buffer.seek(0); yield buffer.read()
    return StreamingResponse(_iter(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": "attachment; filename=score_report.docx"})


@router.get("/training-ledgers/{record_id}", summary="培训台账记录详情")
async def get_training_ledger(
    record_id: UUID,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    record = await service.get_record(record_id)
    await _ensure_ledger_access(record, hr_scope, session)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
    )


@router.put("/training-ledgers/{record_id}", summary="更新培训台账记录")
async def update_training_ledger(
    record_id: UUID,
    payload: TrainingLedgerUpdate,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    record = await service.get_record(record_id)
    await _ensure_ledger_access(record, hr_scope, session)
    record = await service.update_record(record_id, payload)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
        message="培训台账记录更新成功",
    )


@router.delete("/training-ledgers/{record_id}", summary="删除培训台账记录")
async def delete_training_ledger(
    record_id: UUID,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    record = await service.get_record(record_id)
    await _ensure_ledger_access(record, hr_scope, session)
    await service.delete_record(record_id)
    return success_response(message="培训台账记录删除成功")


# ─── AnnualTrainingPlan Routes ───

@router.post("/annual-training-plans/upload", summary="上传年度培训计划")
async def upload_annual_training_plan(
    file: UploadFile,
    service: EmployeeService = Depends(get_employee_service),
):
    """上传 Excel 年度培训计划，按年度+部门自动分类为计划项。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 格式")
    try:
        content = await file.read()
        result = await service.upload_annual_plan(content)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return success_response(
        data=result,
        message=f"新增 {result['created']} 条，更新 {result['updated']} 条"
        + (f"，{len(result['errors'])} 条失败" if result.get('errors') else ""),
    )


@router.get("/annual-training-plans", summary="年度培训计划列表")
async def list_annual_training_plans(
    year: int | None = Query(None, description="年度筛选"),
    department: str | None = Query(None, description="部门筛选"),
    page_params: PageParams = Depends(),
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    if not hr_scope.is_unrestricted:
        if not hr_scope.scoped_departments:
            # 年度计划按部门归属，self_only 无对应数据
            raise ForbiddenException("数据范围限制：仅可访问本人相关数据")
        if len(hr_scope.scoped_departments) == 1:
            department = next(iter(hr_scope.scoped_departments))
        # 多部门：department 保持 None，service 不过滤，由下方 post-filter 处理
    plans, total = await service.list_plans(
        year=year,
        department=department,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    if not hr_scope.is_unrestricted and len(hr_scope.scoped_departments) > 1:
        plans = [p for p in plans if p.department in hr_scope.scoped_departments]
        # 分页总数按范围统计全局数量（不能用过滤后的本页长度，否则翻页失效）
        from app.modules.hr.models import AnnualTrainingPlan
        count_stmt = select(func.count()).select_from(AnnualTrainingPlan).where(
            AnnualTrainingPlan.department.in_(hr_scope.scoped_departments),
            AnnualTrainingPlan.is_deleted == False,  # noqa: E712
        )
        if year is not None:
            count_stmt = count_stmt.where(AnnualTrainingPlan.year == year)
        total = (await session.execute(count_stmt)).scalar() or 0
    # 批量查每个计划的培训完成进度
    plan_ids = [p.id for p in plans]
    progress_map: dict = {}
    if plan_ids:
        rows = (await session.execute(
            text("SELECT plan_id, count(*) as total, sum(CASE WHEN tracking_status = '完成' THEN 1 ELSE 0 END) as done FROM hr.annual_training_plan_items WHERE plan_id = ANY(:pids) AND is_deleted = false GROUP BY plan_id"),
            {"pids": plan_ids},
        )).fetchall()
        progress_map = {str(r[0]): f"{int(r[2] or 0)}/{int(r[1])} 已完成" for r in rows}

    data = [
        {
            **AnnualTrainingPlanResponse.model_validate(p).model_dump(mode="json"),
            "training_progress": progress_map.get(str(p.id), ""),
        }
        for p in plans
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/annual-training-plans", summary="创建年度培训计划")
async def create_annual_training_plan(
    payload: AnnualTrainingPlanCreate,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    plan = await service.create_plan(payload)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
        message="年度培训计划创建成功",
        status_code=201,
    )


@router.get("/annual-training-plans/{plan_id}", summary="年度培训计划详情")
async def get_annual_training_plan(
    plan_id: UUID,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    plan = await service.get_plan(plan_id)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
    )


@router.put("/annual-training-plans/{plan_id}", summary="更新年度培训计划")
async def update_annual_training_plan(
    plan_id: UUID,
    payload: AnnualTrainingPlanUpdate,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    plan = await service.update_plan(plan_id, payload)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
        message="年度培训计划更新成功",
    )


@router.delete("/annual-training-plans/{plan_id}", summary="删除年度培训计划")
async def delete_annual_training_plan(
    plan_id: UUID,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    await service.delete_plan(plan_id)
    return success_response(message="年度培训计划删除成功")


@router.get("/annual-plan-items", summary="全部年度计划明细（扁平列表）")
async def list_all_annual_plan_items(
    year: int | None = Query(None, description="年度筛选"),
    department: str | None = Query(None, description="部门筛选"),
    keyword: str | None = Query(None, description="培训内容关键词"),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """返回所有年度计划明细的扁平列表，关联部门信息，用于表格展示。"""
    conditions = ["i.is_deleted = false", "p.is_deleted = false"]
    params: dict = {}
    if not hr_scope.is_unrestricted:
        if not hr_scope.scoped_departments:
            # 年度计划按部门归属，self_only 无对应数据
            raise ForbiddenException("数据范围限制：仅可访问本人相关数据")
        if len(hr_scope.scoped_departments) == 1:
            department = next(iter(hr_scope.scoped_departments))
        # 多部门：department 保持 None，参数在下方用 IN 子句处理
    if year is not None:
        conditions.append("p.year = :year")
        params["year"] = year
    if department:
        conditions.append("p.department ILIKE :dept")
        params["dept"] = f"%{department}%"
    elif not hr_scope.is_unrestricted and hr_scope.scoped_departments and len(hr_scope.scoped_departments) > 1:
        placeholders = ", ".join(f":dept_{i}" for i in range(len(hr_scope.scoped_departments)))
        conditions.append(f"p.department IN ({placeholders})")
        for i, d in enumerate(hr_scope.scoped_departments):
            params[f"dept_{i}"] = d
    if keyword:
        conditions.append("i.content_and_textbook ILIKE :kw")
        params["kw"] = f"%{keyword}%"

    where = " AND ".join(conditions)
    sql = text(f"""
        SELECT i.id, i.month, i.content_and_textbook, i.target_audience,
               i.position_and_count, i.training_method, i.duration_hours,
               i.confirmer, i.confirm_date, i.remarks, i.tracking_status,
               i.location, i.assessment_method, i.notes,
               p.department, p.year, p.id as plan_id
        FROM hr.annual_training_plan_items i
        JOIN hr.annual_training_plans p ON i.plan_id = p.id
        WHERE {where}
        ORDER BY p.department, i.month
    """)
    rows = (await session.execute(sql, params)).all()

    return success_response(data=[
        {
            "id": str(r[0]), "month": r[1], "content_and_textbook": r[2],
            "target_audience": r[3], "position_and_count": r[4],
            "training_method": r[5], "duration_hours": r[6],
            "confirmer": r[7], "confirm_date": str(r[8]) if r[8] else None,
            "remarks": r[9], "tracking_status": r[10],
            "location": r[11], "assessment_method": r[12], "notes": r[13],
            "department": r[14], "year": r[15], "plan_id": str(r[16]),
        }
        for r in rows
    ])


@router.get("/annual-training-plans/{plan_id}/items", summary="年度计划明细列表")
async def list_annual_training_plan_items(
    plan_id: UUID,
    service: AnnualTrainingPlanItemService = Depends(get_annual_training_plan_item_service),
):
    items = await service.list_items(plan_id)
    data = [
        AnnualTrainingPlanItemResponse.model_validate(i).model_dump(mode="json")
        for i in items
    ]
    return success_response(data=data)


class CreatePlanItemBody(BaseModel):
    month: str | None = None
    content_and_textbook: str | None = None
    target_audience: str | None = None
    position_and_count: str | None = None
    training_method: str | None = None
    assessment_method: str | None = None
    location: str | None = None
    duration_hours: float | None = None
    confirm_date: str | None = None
    notes: str | None = None
    remarks: str | None = None


@router.post("/annual-training-plans/{plan_id}/items", summary="新增年度计划明细")
async def create_annual_training_plan_item(
    plan_id: UUID,
    payload: CreatePlanItemBody,
    session: AsyncSession = Depends(get_db),
):
    from datetime import date as dt_date

    from app.modules.hr.models import AnnualTrainingPlanItem

    item = AnnualTrainingPlanItem(plan_id=plan_id, **payload.model_dump(exclude_none=True))
    if payload.confirm_date:
        try:
            item.confirm_date = dt_date.fromisoformat(payload.confirm_date)
        except ValueError:
            pass
    session.add(item)
    await session.flush()
    return success_response(
        data=AnnualTrainingPlanItemResponse.model_validate(item).model_dump(mode="json"),
        message="创建成功",
        status_code=201,
    )


@router.delete("/annual-training-plans/{plan_id}/items/{item_id}", summary="删除年度计划明细")
async def delete_annual_training_plan_item(
    plan_id: UUID,
    item_id: UUID,
    session: AsyncSession = Depends(get_db),
):
    """软删除一条年度计划明细。"""
    from app.modules.hr.models import AnnualTrainingPlanItem
    item = (await session.execute(
        select(AnnualTrainingPlanItem).where(
            AnnualTrainingPlanItem.id == item_id,
            AnnualTrainingPlanItem.plan_id == plan_id,
            AnnualTrainingPlanItem.is_deleted == False,
        )
    )).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "明细不存在")
    # 软删除计划明细；员工的培训台账是历史记录，不做任何级联删除
    item.is_deleted = True
    await session.flush()
    return success_response(message="删除成功")


@router.put("/annual-training-plans/{plan_id}/items/batch", summary="批量更新年度计划明细")
async def batch_update_annual_training_plan_items(
    plan_id: UUID,
    payload: AnnualTrainingPlanItemBatchUpdate,
    service: AnnualTrainingPlanItemService = Depends(get_annual_training_plan_item_service),
):
    items = await service.batch_update_items(plan_id, payload)
    data = [
        AnnualTrainingPlanItemResponse.model_validate(i).model_dump(mode="json")
        for i in items
    ]
    return success_response(
        data=data,
        message="年度计划明细更新成功",
    )


def _generate_annual_plan_excel(plan: dict, items: list[dict]) -> BytesIO:
    """Generate annual training plan Excel based on 7.7 template format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "年度培训计划"

    # Styles
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=16)

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 10

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{plan['year']} 年培训计划"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Department row
    ws.merge_cells("A2:I2")
    ws["A2"] = f"部门：{plan['department']}"
    ws["A2"].font = bold_font
    ws["A2"].alignment = left_align
    ws.row_dimensions[2].height = 22

    # Header row
    headers = ["序号", "培训季度及课时", "培训内容及使用教材", "培训对象",
               "授课单位及授课人", "考核方式", "培训跟踪", "确认人/日期", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[3].height = 28

    # Data rows
    for idx, item in enumerate(items, 1):
        row = 3 + idx
        quarter = item.get("month") or ""
        hours = item.get("duration_hours")
        quarter_hours = f"{quarter}\n{hours}课时" if hours else quarter

        values = [
            idx,
            quarter_hours,
            item.get("content_and_textbook") or "",
            item.get("target_audience") or "",
            item.get("position_and_count") or "",
            item.get("training_method") or "",
            item.get("tracking_status") or "",
            f"{item.get('confirmer') or ''}{' / ' + str(item.get('confirm_date')) if item.get('confirm_date') else ''}",
            item.get("remarks") or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align if col in (1, 2, 6, 7) else left_align
        ws.row_dimensions[row].height = 36

    # Pad to at least 12 rows
    while len(items) < 12:
        row = 3 + len(items) + 1
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = thin_border
        ws.row_dimensions[row].height = 36
        items.append({})

    # Footer row
    footer_row = 4 + len(items) + 1
    ws.merge_cells(f"A{footer_row}:E{footer_row}")
    ws.cell(row=footer_row, column=1, value="制表人/日期：")
    ws.cell(row=footer_row, column=1).alignment = left_align
    ws.cell(row=footer_row, column=1).border = thin_border
    for c in range(2, 6):
        ws.cell(row=footer_row, column=c).border = thin_border

    ws.merge_cells(f"F{footer_row}:I{footer_row}")
    ws.cell(row=footer_row, column=6, value="部门负责人/日期：")
    ws.cell(row=footer_row, column=6).alignment = left_align
    ws.cell(row=footer_row, column=6).border = thin_border
    for c in range(7, 10):
        ws.cell(row=footer_row, column=c).border = thin_border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.post("/annual-training-plans/complete-by-content", summary="按培训内容标记完成")
async def complete_plan_by_content(
    payload: dict = None,
    session: AsyncSession = Depends(get_db),
):
    """根据培训内容标记所有匹配的年度计划明细为已完成。"""
    from app.modules.hr.models import AnnualTrainingPlanItem
    content = payload.get("content", "") if payload else ""
    if not content:
        raise HTTPException(400, "缺少培训内容")
    items = (await session.execute(
        select(AnnualTrainingPlanItem).where(
            AnnualTrainingPlanItem.content_and_textbook == content,
            AnnualTrainingPlanItem.is_deleted == False,
        )
    )).scalars().all()
    for item in items:
        item.tracking_status = "完成"
    await session.flush()
    return success_response(data={"updated": len(items)}, message=f"已标记 {len(items)} 条为完成")


@router.get("/annual-training-plans/{plan_id}/export", summary="导出年度培训计划Excel")
async def export_annual_training_plan(
    plan_id: UUID,
    plan_service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
    item_service: AnnualTrainingPlanItemService = Depends(get_annual_training_plan_item_service),
):
    """根据年度计划数据生成并导出Excel文件（7.7年度培训计划格式）。"""
    plan = await plan_service.get_plan(plan_id)
    items = await item_service.list_items(plan_id)

    plan_dict = AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json")
    item_dicts = [
        AnnualTrainingPlanItemResponse.model_validate(i).model_dump(mode="json")
        for i in items
    ]

    buffer = _generate_annual_plan_excel(plan_dict, item_dicts)
    buffer.seek(0)

    safe_dept = plan.department.replace(" ", "_")
    filename = f"{plan.year}年度培训计划_{safe_dept}.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"
        },
    )

# ─── Trainer Routes ───


@router.post("/trainers/upload", summary="上传内训师台账")
async def upload_trainers(
    file: UploadFile,
    service: EmployeeService = Depends(get_employee_service),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """上传 Excel 内训师名单，按姓名+部门自动新增或更新。受限用户仅可导入本部门数据。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 格式")
    allowed = None if hr_scope.is_unrestricted else hr_scope.scoped_departments
    if not hr_scope.is_unrestricted and not allowed:
        raise ForbiddenException("数据范围限制：仅可操作本人相关数据")
    try:
        content = await file.read()
        result = await service.upload_trainers(content, allowed_departments=allowed)
    except ValueError as e:
        raise HTTPException(400, str(e))
    msg = f"新增 {result['created']}，更新 {result['updated']}"
    if result.get("skipped"):
        msg += f"，跳过 {result['skipped']}（非本部门）"
    return success_response(data=result, message=msg)


@router.get("/trainers", summary="内训师台账列表", response_model=TrainerListResponse)
async def list_trainers(
    department: str | None = Query(None),
    keyword: str | None = Query(None),
    is_level1: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.core.response import paginated_response
    from app.modules.hr.models import HrTrainer

    if not hr_scope.is_unrestricted:
        if not hr_scope.scoped_departments:
            # 内训师按部门归属，self_only 无对应数据
            raise ForbiddenException("数据范围限制：仅可访问本人相关数据")
        if len(hr_scope.scoped_departments) == 1:
            department = next(iter(hr_scope.scoped_departments))
        # 多部门：department 保持 None，在 query 层用 IN 处理
    query = select(HrTrainer).where(HrTrainer.is_deleted == False)
    count_q = select(func.count()).select_from(HrTrainer).where(HrTrainer.is_deleted == False)
    if department:
        query = query.where(HrTrainer.department == department)
        count_q = count_q.where(HrTrainer.department == department)
    elif not hr_scope.is_unrestricted and hr_scope.scoped_departments:
        query = query.where(HrTrainer.department.in_(hr_scope.scoped_departments))
        count_q = count_q.where(HrTrainer.department.in_(hr_scope.scoped_departments))
    if is_level1:
        query = query.where(HrTrainer.is_level1 == is_level1)
        count_q = count_q.where(HrTrainer.is_level1 == is_level1)
    if keyword:
        query = query.where(
            or_(HrTrainer.name.ilike(f"%{keyword}%"), HrTrainer.trainable_departments.ilike(f"%{keyword}%"))
        )
        count_q = count_q.where(
            or_(HrTrainer.name.ilike(f"%{keyword}%"), HrTrainer.trainable_departments.ilike(f"%{keyword}%"))
        )

    total = (await session.execute(count_q)).scalar() or 0
    rows = (await session.execute(query.order_by(HrTrainer.department, HrTrainer.name).offset((page - 1) * page_size).limit(page_size))).scalars().all()

    return paginated_response(
        data=[TrainerResponse.model_validate(r).model_dump(mode="json") for r in rows],
        page=page, page_size=page_size, total=total,
    )


class TrainerCreate(BaseModel):
    name: str
    department: str | None = None
    trainable_departments: str | None = None
    qualification_scope: str | None = None
    certification_date: date | None = None
    confirmation_date: date | None = None
    confirmation_reminder: date | None = None
    is_level1: str | None = None
    admin: str | None = None
    remarks: str | None = None
    period: str | None = None


@router.post("/trainers", summary="新增内训师")
async def create_trainer(
    payload: TrainerCreate,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import HrTrainer
    # 本部门数据范围：部门为空时取用户唯一授权部门，指定部门必须在授权范围内
    if not hr_scope.is_unrestricted:
        if not payload.department and len(hr_scope.scoped_departments) == 1:
            payload.department = next(iter(hr_scope.scoped_departments))
        if not payload.department:
            raise HTTPException(400, "部门不能为空（数据范围限制下无法自动归属）")
        hr_scope.ensure_dept_writable([payload.department])
    elif not payload.department:
        raise HTTPException(400, "部门不能为空")
    t = HrTrainer(**payload.model_dump())
    session.add(t)
    await session.flush()
    return success_response(data=TrainerResponse.model_validate(t).model_dump(mode="json"), message="创建成功", status_code=201)


@router.put("/trainers/{trainer_id}", summary="更新内训师")
async def update_trainer(
    trainer_id: UUID,
    payload: TrainerCreate,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import HrTrainer
    t = (await session.execute(select(HrTrainer).where(HrTrainer.id == trainer_id, HrTrainer.is_deleted == False))).scalar_one_or_none()
    if not t: raise HTTPException(404, "内训师不存在")
    hr_scope.ensure_dept_writable([t.department, payload.department])
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(t, k, v)
    await session.flush()
    return success_response(data=TrainerResponse.model_validate(t).model_dump(mode="json"), message="更新成功")


@router.delete("/trainers/{trainer_id}", summary="删除内训师")
async def delete_trainer(
    trainer_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import HrTrainer
    t = (await session.execute(select(HrTrainer).where(HrTrainer.id == trainer_id, HrTrainer.is_deleted == False))).scalar_one_or_none()
    if not t:
        raise HTTPException(404, "内训师不存在")
    hr_scope.ensure_dept_writable([t.department])
    t.is_deleted = True
    await session.flush()
    await session.commit()
    return success_response(message="删除成功")


@router.delete("/trainers", summary="清空内训师台账")
async def clear_trainers(
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    if not hr_scope.is_unrestricted:
        raise ForbiddenException("数据范围限制：仅全量数据权限可清空台账")
    await session.execute(text("DELETE FROM hr.trainers"))
    await session.commit()
    return success_response(message="清空成功")


# ─── DeptTrainingPersonnel Routes ───


@router.get("/dept-training-personnel", summary="部门培训人员表列表")
async def list_dept_training_personnel(
    department: str | None = Query(None, description="部门筛选"),
    keyword: str | None = Query(None, description="关键词搜索"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import DeptTrainingPersonnel

    if not hr_scope.is_unrestricted:
        if not hr_scope.scoped_departments:
            # 部门培训人员按部门归属，self_only 无对应数据
            raise ForbiddenException("数据范围限制：仅可访问本人相关数据")
        if len(hr_scope.scoped_departments) == 1:
            department = next(iter(hr_scope.scoped_departments))
        # 多部门：department 保持 None，在 query 层用 IN 处理
    query = select(DeptTrainingPersonnel).where(DeptTrainingPersonnel.is_deleted == False)
    count_q = select(func.count()).select_from(DeptTrainingPersonnel).where(DeptTrainingPersonnel.is_deleted == False)

    if department:
        query = query.where(
            or_(
                DeptTrainingPersonnel.department == department,
                DeptTrainingPersonnel.display_department == department,
            )
        )
        count_q = count_q.where(
            or_(
                DeptTrainingPersonnel.department == department,
                DeptTrainingPersonnel.display_department == department,
            )
        )
    elif not hr_scope.is_unrestricted and hr_scope.scoped_departments:
        query = query.where(
            or_(
                DeptTrainingPersonnel.department.in_(hr_scope.scoped_departments),
                DeptTrainingPersonnel.display_department.in_(hr_scope.scoped_departments),
            )
        )
        count_q = count_q.where(
            or_(
                DeptTrainingPersonnel.department.in_(hr_scope.scoped_departments),
                DeptTrainingPersonnel.display_department.in_(hr_scope.scoped_departments),
            )
        )
    if keyword:
        query = query.where(
            or_(
                DeptTrainingPersonnel.display_department.ilike(f"%{keyword}%"),
                DeptTrainingPersonnel.training_admin.ilike(f"%{keyword}%"),
                DeptTrainingPersonnel.department_head.ilike(f"%{keyword}%"),
                DeptTrainingPersonnel.level1_trainer.ilike(f"%{keyword}%"),
            )
        )
        count_q = count_q.where(
            or_(
                DeptTrainingPersonnel.display_department.ilike(f"%{keyword}%"),
                DeptTrainingPersonnel.training_admin.ilike(f"%{keyword}%"),
                DeptTrainingPersonnel.department_head.ilike(f"%{keyword}%"),
                DeptTrainingPersonnel.level1_trainer.ilike(f"%{keyword}%"),
            )
        )

    total = (await session.execute(count_q)).scalar() or 0
    rows = (
        await session.execute(
            query.order_by(DeptTrainingPersonnel.display_department, DeptTrainingPersonnel.variety)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    ).scalars().all()

    return paginated_response(
        data=[DeptTrainingPersonnelResponse.model_validate(r).model_dump(mode="json") for r in rows],
        page=page,
        page_size=page_size,
        total=total,
    )


class DeptTrainingPersonnelCreateBody(BaseModel):
    display_department: str
    variety: str | None = None
    department: str
    training_admin: str | None = None
    department_head: str | None = None
    level1_trainer: str | None = None


@router.post("/dept-training-personnel", summary="新增部门培训人员")
async def create_dept_training_personnel(
    payload: DeptTrainingPersonnelCreateBody,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import DeptTrainingPersonnel

    hr_scope.ensure_dept_writable([payload.department, payload.display_department])
    t = DeptTrainingPersonnel(**payload.model_dump())
    session.add(t)
    await session.flush()
    return success_response(
        data=DeptTrainingPersonnelResponse.model_validate(t).model_dump(mode="json"),
        message="创建成功",
        status_code=201,
    )


@router.put("/dept-training-personnel/{item_id}", summary="更新部门培训人员")
async def update_dept_training_personnel(
    item_id: UUID,
    payload: DeptTrainingPersonnelCreateBody,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import DeptTrainingPersonnel

    t = (
        await session.execute(
            select(DeptTrainingPersonnel).where(
                DeptTrainingPersonnel.id == item_id,
                DeptTrainingPersonnel.is_deleted == False,
            )
        )
    ).scalar_one_or_none()
    if not t:
        raise HTTPException(404, "记录不存在")
    hr_scope.ensure_dept_writable([t.department, t.display_department, payload.department, payload.display_department])
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(t, k, v)
    await session.flush()
    return success_response(
        data=DeptTrainingPersonnelResponse.model_validate(t).model_dump(mode="json"),
        message="更新成功",
    )


@router.delete("/dept-training-personnel/{item_id}", summary="删除部门培训人员")
async def delete_dept_training_personnel(
    item_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import DeptTrainingPersonnel

    t = (
        await session.execute(
            select(DeptTrainingPersonnel).where(
                DeptTrainingPersonnel.id == item_id,
                DeptTrainingPersonnel.is_deleted == False,
            )
        )
    ).scalar_one_or_none()
    if not t:
        raise HTTPException(404, "记录不存在")
    hr_scope.ensure_dept_writable([t.department, t.display_department])
    t.is_deleted = True
    await session.flush()
    return success_response(message="删除成功")


@router.post("/dept-training-personnel/upload", summary="上传部门培训人员表")
async def upload_dept_training_personnel(
    file: UploadFile,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """上传 Excel 部门培训人员表，按体现部门+品种去重 upsert。受限用户仅可导入本部门数据。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 格式")

    try:
        content = await file.read()
        from io import BytesIO

        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

        created = 0
        updated = 0
        skipped = 0

        for row in rows:
            display_dept = str(row[1]).strip() if row[1] else None
            if not display_dept:
                continue
            variety = str(row[2]).strip() if len(row) > 2 and row[2] else None
            department = str(row[3]).strip() if len(row) > 3 and row[3] else display_dept
            # 本部门数据范围：非本部门行跳过
            if not hr_scope.is_unrestricted:
                try:
                    hr_scope.ensure_dept_writable([display_dept, department])
                except Exception:
                    skipped += 1
                    continue
            training_admin = str(row[4]).strip() if len(row) > 4 and row[4] else None
            department_head = str(row[5]).strip() if len(row) > 5 and row[5] else None
            level1_trainer = str(row[6]).strip() if len(row) > 6 and row[6] else None

            from app.modules.hr.models import DeptTrainingPersonnel

            # 按体现部门+品种去重查找
            existing_q = select(DeptTrainingPersonnel).where(
                DeptTrainingPersonnel.is_deleted == False,
                DeptTrainingPersonnel.display_department == display_dept,
            )
            if variety:
                existing_q = existing_q.where(DeptTrainingPersonnel.variety == variety)
            else:
                existing_q = existing_q.where(DeptTrainingPersonnel.variety.is_(None))

            existing = (await session.execute(existing_q)).scalar_one_or_none()

            if existing:
                existing.department = department
                existing.training_admin = training_admin
                existing.department_head = department_head
                existing.level1_trainer = level1_trainer
                updated += 1
            else:
                t = DeptTrainingPersonnel(
                    display_department=display_dept,
                    variety=variety,
                    department=department,
                    training_admin=training_admin,
                    department_head=department_head,
                    level1_trainer=level1_trainer,
                )
                session.add(t)
                created += 1

        await session.flush()
    except Exception as e:
        raise HTTPException(400, f"导入失败: {str(e)}")

    msg = f"新增 {created}，更新 {updated}"
    if skipped:
        msg += f"，跳过 {skipped}（非本部门）"
    return success_response(
        data={"created": created, "updated": updated, "skipped": skipped},
        message=msg,
    )


# ─── SOP Catalog Routes ───

@router.delete("/sop-catalog/{item_id}", summary="删除SOP目录条目")
async def delete_sop_catalog_item(
    item_id: UUID,
    session: AsyncSession = Depends(get_db),
):
    from app.modules.hr.models import SopCatalog
    item = (await session.execute(
        select(SopCatalog).where(SopCatalog.id == item_id, SopCatalog.is_deleted == False)
    )).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "SOP条目不存在")
    # 软删除目录条目；仅软删除与该条目直接关联的岗位培训行（按 SOP 编号/文件名匹配），
    # 不动手工维护的同岗位其他培训内容
    from app.modules.hr.models import PositionTraining
    pt_conds = [PositionTraining.department == item.department]
    if item.position_name:
        pt_conds.append(PositionTraining.position_name == item.position_name)
    if item.sop_number:
        pt_conds.append(PositionTraining.sop_number == item.sop_number)
    elif item.file_name:
        pt_conds.append(PositionTraining.file_name == item.file_name)
    linked = (await session.execute(
        select(PositionTraining).where(
            *pt_conds, PositionTraining.is_deleted == False  # noqa: E712
        )
    )).scalars().all()
    for pt in linked:
        pt.is_deleted = True
    item.is_deleted = True
    await session.flush()
    return success_response(message="删除成功")


@router.post("/sop-catalog/upload", summary="上传SOP目录")
async def upload_sop_catalog(
    file: UploadFile,
    service: EmployeeService = Depends(get_employee_service),
):
    """上传 Excel SOP 目录，按 SOP编号 自动新增或更新。"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "仅支持 .xlsx / .xls 格式")
    try:
        content = await file.read()
        result = await service.upload_sop_catalog(content)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return success_response(data=result, message=f"新增 {result['created']}，更新 {result['updated']}")


@router.get("/sop-catalog", summary="SOP目录列表", response_model=SopCatalogListResponse)
async def list_sop_catalog(
    department: str | None = Query(None),
    category: str | None = Query(None),
    keyword: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    session: AsyncSession = Depends(get_db),
):
    from app.core.response import paginated_response
    from app.modules.hr.models import SopCatalog

    query = select(SopCatalog).where(SopCatalog.is_deleted == False)
    count_q = select(func.count()).select_from(SopCatalog).where(SopCatalog.is_deleted == False)
    if department:
        query = query.where(SopCatalog.department == department)
        count_q = count_q.where(SopCatalog.department == department)
    if category:
        query = query.where(SopCatalog.category == category)
        count_q = count_q.where(SopCatalog.category == category)
    if keyword:
        query = query.where(SopCatalog.file_name.ilike(f"%{keyword}%"))
        count_q = count_q.where(SopCatalog.file_name.ilike(f"%{keyword}%"))

    total = (await session.execute(count_q)).scalar() or 0
    rows = (await session.execute(query.order_by(SopCatalog.category, SopCatalog.file_name).offset((page - 1) * page_size).limit(page_size))).scalars().all()

    return paginated_response(
        data=[SopCatalogResponse.model_validate(r).model_dump(mode="json") for r in rows],
        page=page, page_size=page_size, total=total,
    )


@router.get("/sop-catalog/departments", summary="SOP目录部门列表")
async def list_sop_catalog_departments(
    session: AsyncSession = Depends(get_db),
):
    """返回 SOP 目录中所有不重复的部门名称。"""
    from app.modules.hr.models import SopCatalog

    result = await session.execute(
        select(SopCatalog.department)
        .where(SopCatalog.is_deleted == False, SopCatalog.department.isnot(None))
        .distinct()
        .order_by(SopCatalog.department)
    )
    departments = [row[0] for row in result.all()]
    return success_response(data=departments)


@router.get("/sop-catalog/categories", summary="SOP目录分类列表")
async def list_sop_catalog_categories(
    department: str | None = Query(None, description="按部门筛选"),
    session: AsyncSession = Depends(get_db),
):
    """返回 SOP 目录中所有不重复的分类名称，可按部门筛选。"""
    from app.modules.hr.models import SopCatalog

    stmt = (
        select(SopCatalog.category)
        .where(SopCatalog.is_deleted == False, SopCatalog.category.isnot(None))
        .distinct()
        .order_by(SopCatalog.category)
    )
    if department:
        stmt = stmt.where(SopCatalog.department == department)
    result = await session.execute(stmt)
    categories = [row[0] for row in result.all()]
    return success_response(data=categories)


@router.post("/training-evaluations/upsert", summary="同步评估补录的记录（培训内容+部门+应到人数）")
async def upsert_training_evaluation(
    training_content: str = Form(...),
    department: str = Form(...),
    expected_count: int = Form(0),
    training_method: str = Form(""),
    trainer_name: str = Form(""),
    assessment_method: str = Form(""),
    session: AsyncSession = Depends(get_db),
):
    """从培训通知面板同步：按培训内容+部门 upsert。"""
    existing = (await session.execute(
        text("SELECT id FROM hr.training_evaluations WHERE training_content = :c AND department = :d AND is_deleted = false"),
        {"c": training_content, "d": department}
    )).fetchone()
    if existing:
        await session.execute(
            text("UPDATE hr.training_evaluations SET expected_count = :n, training_method = :tm, trainer_name = :tn, assessment_method = :am, updated_at = now() WHERE id = :id"),
            {"n": expected_count, "id": existing[0], "tm": training_method, "tn": trainer_name, "am": assessment_method}
        )
        msg = "updated"
    else:
        await session.execute(
            text("INSERT INTO hr.training_evaluations (id, training_content, department, expected_count, training_method, trainer_name, assessment_method) VALUES (gen_random_uuid(), :c, :d, :n, :tm, :tn, :am)"),
            {"c": training_content, "d": department, "n": expected_count, "tm": training_method, "tn": trainer_name, "am": assessment_method}
        )
        msg = "created"
    await session.commit()
    return success_response(data={"status": msg}, message="同步成功")



# ─── Exam Papers Routes ───

class SaveExamPaperRequest(BaseModel):
    subject: str
    department: str | None = None
    training_date: str | None = None
    training_method: str | None = None
    questions: dict | None = None
    full_score: int = 100
    pass_line: int = 80
    choice_count: int = 0
    true_false_count: int = 0
    multi_choice_count: int = 0
    fill_blank_count: int = 0


@router.post("/exam-papers", summary="保存试卷")
async def save_exam_paper(
    payload: SaveExamPaperRequest,
    session: AsyncSession = Depends(get_db),
):
    """保存 AI 生成或手工组卷的笔试试卷，供后续复用下载。"""
    from app.modules.hr.models import ExamPaper
    paper = ExamPaper(
        subject=payload.subject,
        department=payload.department,
        training_date=date.fromisoformat(payload.training_date) if payload.training_date else None,
        training_method=payload.training_method,
        questions=payload.questions,
        full_score=payload.full_score,
        pass_line=payload.pass_line,
        choice_count=payload.choice_count,
        true_false_count=payload.true_false_count,
        multi_choice_count=payload.multi_choice_count,
        fill_blank_count=payload.fill_blank_count,
    )
    session.add(paper)
    await session.commit()
    return success_response(data={"id": str(paper.id)}, message="试卷已保存", status_code=201)


@router.get("/exam-papers", summary="试卷列表")
async def list_exam_papers(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """列出已保存的笔试试卷。"""
    from app.modules.hr.models import ExamPaper
    q = select(ExamPaper).where(ExamPaper.is_deleted == False).order_by(ExamPaper.created_at.desc())
    total_q = select(func.count()).select_from(ExamPaper).where(ExamPaper.is_deleted == False)
    if not hr_scope.is_unrestricted:
        if not hr_scope.scoped_departments:
            # 试卷按培训部门归属，self_only 无对应数据
            raise ForbiddenException("数据范围限制：仅可访问本人相关数据")
        # 数据范围受限角色强制只看授权部门试卷
        q = q.where(ExamPaper.department.in_(hr_scope.scoped_departments))
        total_q = total_q.where(ExamPaper.department.in_(hr_scope.scoped_departments))
    total = (await session.execute(total_q)).scalar() or 0
    papers = (await session.execute(q.offset((page - 1) * page_size).limit(page_size))).scalars().all()
    data = [{
        "id": str(p.id),
        "subject": p.subject,
        "department": p.department,
        "training_date": str(p.training_date) if p.training_date else None,
        "training_method": p.training_method,
        "full_score": p.full_score,
        "pass_line": p.pass_line,
        "choice_count": p.choice_count,
        "true_false_count": p.true_false_count,
        "multi_choice_count": p.multi_choice_count,
        "fill_blank_count": p.fill_blank_count,
        "source": p.source,
        "created_at": str(p.created_at) if p.created_at else None,
    } for p in papers]
    return paginated_response(data=data, page=page, page_size=page_size, total=total)


@router.get("/exam-papers/{paper_id}/download", summary="下载试卷")
async def download_exam_paper(
    paper_id: UUID,
    session: AsyncSession = Depends(get_db),
):
    """根据已保存试卷的题目快照重新生成 Word 文档下载。"""
    from app.modules.hr.exam_paper_generator import generate_exam_paper
    from app.modules.hr.models import ExamPaper

    q = select(ExamPaper).where(ExamPaper.id == paper_id, ExamPaper.is_deleted == False)
    r = await session.execute(q)
    paper = r.scalar_one_or_none()
    if not paper:
        raise HTTPException(404, "试卷不存在")

    buffer = generate_exam_paper(paper)
    filename = f"笔试试卷_{paper.subject}.docx"
    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{quote(filename)}"},
    )


# ─── TransferRecord Routes ───

@router.get("/transfers", summary="员工异动记录列表")
async def list_transfers(
    employee_id: UUID = Query(...),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """查询某员工的异动记录（分页）。"""
    from sqlalchemy import func

    from app.modules.hr.models import Employee, TransferRecord
    if not hr_scope.is_unrestricted:
        emp = (await session.execute(
            select(Employee).where(Employee.id == employee_id, Employee.is_deleted == False)
        )).scalar_one_or_none()
        if not emp:
            raise HTTPException(404, "员工不存在")
        hr_scope.ensure_can_access_employee(emp)
    base = select(TransferRecord).where(
        TransferRecord.employee_id == employee_id, TransferRecord.is_deleted == False
    )
    total = (await session.execute(select(func.count()).select_from(base.subquery()))).scalar() or 0
    r = await session.execute(
        base.order_by(TransferRecord.effective_date.desc()).offset((page - 1) * page_size).limit(page_size)
    )
    records = r.scalars().all()
    return paginated_response(
        data=[{
            "id": str(t.id), "transfer_type": t.transfer_type,
            "from_department": t.from_department, "to_department": t.to_department,
            "from_position": t.from_position, "to_position": t.to_position,
            "effective_date": str(t.effective_date) if t.effective_date else None,
            "reason": t.reason, "created_at": str(t.created_at),
        } for t in records],
        page=page, page_size=page_size, total=total,
    )


# ─── Employee Tags ───


@router.get("/employee-tags", summary="员工标签列表")
async def list_employee_tags(
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import EmployeeTag
    creator = hr_scope.user.name or hr_scope.user.employee_number or ""
    result = await session.execute(
        select(EmployeeTag.tag_name, func.count().label('cnt'))
        .where(EmployeeTag.is_deleted == False, EmployeeTag.created_by == creator)
        .group_by(EmployeeTag.tag_name)
        .order_by(EmployeeTag.tag_name)
    )
    tags = [{"tag_name": row[0], "count": row[1]} for row in result.fetchall()]
    return success_response(data=tags)


@router.post("/employee-tags", summary="操作员工标签")
async def save_employee_tag(
    payload: dict,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import EmployeeTag
    creator = hr_scope.user.name or hr_scope.user.employee_number or ""
    emp_no = payload.get("employee_number", "")
    tag_name = payload.get("tag_name", "")
    action = payload.get("action", "add")
    if not emp_no or not tag_name:
        raise HTTPException(400, "员工工号和标签名不能为空")
    if action == "remove":
        # 删除该工号+标签的全部行（不限创建人）：转训自动打上的标签
        # 在员工档案中必须能被移除，否则删了又出现
        await session.execute(
            text("UPDATE hr.employee_tags SET is_deleted = true WHERE employee_number = :en AND tag_name = :tn AND is_deleted = false"),
            {"en": emp_no, "tn": tag_name},
        )
    else:
        # ON CONFLICT DO NOTHING：与活跃行部分唯一索引配合，并发打标不再 500
        # （部分索引推断必须显式给出 index_where）
        from sqlalchemy.dialects.postgresql import insert as pg_insert
        await session.execute(
            pg_insert(EmployeeTag)
            .values(employee_number=emp_no, tag_name=tag_name, created_by=creator)
            .on_conflict_do_nothing(
                index_elements=["employee_number", "tag_name"],
                index_where=text("is_deleted = false"),
            ),
        )
    await session.commit()
    return success_response(message="操作成功")


@router.get("/employee-tags/by-employee", summary="查询员工标签")
async def get_employee_tags(
    employee_number: str = Query(...),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import Employee as EmpModel
    emp = (await session.execute(
        select(EmpModel).where(
            EmpModel.employee_number == employee_number,
            EmpModel.is_deleted == False,  # noqa: E712
        )
    )).scalars().first()
    if emp:
        hr_scope.ensure_can_access_employee(emp)
    result = await session.execute(
        text("SELECT tag_name, created_by FROM hr.employee_tags WHERE employee_number = :en AND is_deleted = false"),
        {"en": employee_number},
    )
    tags = [{"tag_name": row[0], "created_by": row[1]} for row in result.fetchall()]
    return success_response(data=tags)


# ─── Employee Classifications（分类清单，下拉选项模式） ───


@router.get("/employee-classifications", summary="员工分类清单")
async def list_employee_classifications(
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """返回当前用户的分类清单（含每个分类下的人数）。"""
    from app.modules.hr.models import EmployeeClassification, EmployeeTag
    creator = hr_scope.user.name or hr_scope.user.employee_number or ""
    rows = (await session.execute(
        select(EmployeeClassification)
        .where(
            EmployeeClassification.is_deleted == False,  # noqa: E712
            EmployeeClassification.created_by == creator,
        )
        .order_by(EmployeeClassification.created_at)
    )).scalars().all()
    counts = dict((await session.execute(
        select(EmployeeTag.tag_name, func.count())
        .where(EmployeeTag.is_deleted == False, EmployeeTag.created_by == creator)
        .group_by(EmployeeTag.tag_name)
    )).all())
    return success_response(data=[
        {"id": str(c.id), "name": c.name, "count": counts.get(c.name, 0)}
        for c in rows
    ])


@router.post("/employee-classifications", summary="新增分类")
async def create_employee_classification(
    payload: dict,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    from app.modules.hr.models import EmployeeClassification
    creator = hr_scope.user.name or hr_scope.user.employee_number or ""
    name = str(payload.get("name", "")).strip()
    if not name:
        raise HTTPException(400, "分类名称不能为空")
    exist = (await session.execute(
        select(EmployeeClassification).where(
            EmployeeClassification.is_deleted == False,  # noqa: E712
            EmployeeClassification.created_by == creator,
            EmployeeClassification.name == name,
        )
    )).scalar_one_or_none()
    if exist:
        raise HTTPException(400, "该分类已存在")
    c = EmployeeClassification(name=name, created_by=creator)
    session.add(c)
    await session.commit()
    return success_response(data={"id": str(c.id), "name": c.name}, message="已新增", status_code=201)


@router.get("/employee-classifications/{class_id}/members", summary="分类下的人员")
async def list_classification_members(
    class_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """查看某分类下收纳的具体人员。"""
    from app.modules.hr.models import EmployeeClassification
    creator = hr_scope.user.name or hr_scope.user.employee_number or ""
    c = (await session.execute(
        select(EmployeeClassification).where(
            EmployeeClassification.id == class_id,
            EmployeeClassification.is_deleted == False,  # noqa: E712
            EmployeeClassification.created_by == creator,
        )
    )).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "分类不存在")
    rows = (await session.execute(
        text("""
            SELECT e.name, e.employee_number, e.department, e.position
            FROM hr.employee_tags t
            JOIN hr.employees e ON e.employee_number = t.employee_number AND e.is_deleted = false
            WHERE t.tag_name = :tn AND t.created_by = :cb AND t.is_deleted = false
            ORDER BY e.department, e.employee_number
        """),
        {"tn": c.name, "cb": creator},
    )).fetchall()
    return success_response(data=[
        {"name": r[0], "employee_number": r[1], "department": r[2], "position": r[3]}
        for r in rows
    ])


@router.post("/employee-classifications/{class_id}/remove-members", summary="批量移除分类人员")
async def remove_classification_members(
    class_id: UUID,
    payload: dict,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """把选中的人员从该分类中移除（仅解除分类，不影响员工档案）。"""
    from app.modules.hr.models import EmployeeClassification
    creator = hr_scope.user.name or hr_scope.user.employee_number or ""
    c = (await session.execute(
        select(EmployeeClassification).where(
            EmployeeClassification.id == class_id,
            EmployeeClassification.is_deleted == False,  # noqa: E712
            EmployeeClassification.created_by == creator,
        )
    )).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "分类不存在")
    emp_nos = [str(x).strip() for x in payload.get("employee_numbers", []) if str(x).strip()]
    if not emp_nos:
        raise HTTPException(400, "请选择要移除的人员")
    result = await session.execute(
        text("""
            UPDATE hr.employee_tags SET is_deleted = true
            WHERE tag_name = :tn AND created_by = :cb AND is_deleted = false
              AND employee_number = ANY(:nos)
        """),
        {"tn": c.name, "cb": creator, "nos": emp_nos},
    )
    await session.commit()
    return success_response(data={"removed": result.rowcount}, message=f"已移除 {result.rowcount} 人")


@router.delete("/employee-classifications/{class_id}", summary="删除分类")
async def delete_employee_classification(
    class_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """删除分类清单项，同时解除该分类下所有员工的标签。"""
    from app.modules.hr.models import EmployeeClassification
    creator = hr_scope.user.name or hr_scope.user.employee_number or ""
    c = (await session.execute(
        select(EmployeeClassification).where(
            EmployeeClassification.id == class_id,
            EmployeeClassification.is_deleted == False,  # noqa: E712
            EmployeeClassification.created_by == creator,
        )
    )).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "分类不存在")
    c.is_deleted = True
    await session.execute(
        text("UPDATE hr.employee_tags SET is_deleted = true WHERE tag_name = :tn AND created_by = :cb AND is_deleted = false"),
        {"tn": c.name, "cb": creator},
    )
    await session.commit()
    return success_response(message="已删除")


@router.post("/transfers", summary="创建员工异动记录")
async def create_transfer(
    payload: TransferCreate,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """新增员工异动记录（受限用户仅可操作授权范围内员工）。"""
    from app.modules.hr.models import Employee as EmpModel, TransferRecord
    emp = (await session.execute(
        select(EmpModel).where(
            EmpModel.id == payload.employee_id,
            EmpModel.is_deleted == False,  # noqa: E712
        )
    )).scalar_one_or_none()
    if emp:
        hr_scope.ensure_can_access_employee(emp)
    t = TransferRecord(
        employee_id=payload.employee_id,
        transfer_type=payload.transfer_type,
        from_department=payload.from_department,
        to_department=payload.to_department,
        from_position=payload.from_position,
        to_position=payload.to_position,
        effective_date=payload.effective_date,
        reason=payload.reason,
    )
    session.add(t)
    await session.flush()
    return success_response(data={"id": str(t.id)}, message="异动记录创建成功", status_code=201)


# ─── 月度绩效考核 ───

@router.get("/performance-evaluations", summary="月度绩效考核列表")
async def list_performance_evaluations(
    month: str | None = Query(None, description="考核月份"),
    department: str | None = Query(None, description="部门"),
    status: str | None = Query(None, description="状态"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    service = PerformanceEvaluationService(session)
    params = PerformanceListParams(
        evaluation_month=month, department=department,
        status=status, page=page, page_size=page_size,
    )
    scoped = None if hr_scope.is_unrestricted else hr_scope.scoped_departments
    items, total = await service.list_evaluations(params, departments=scoped)
    return success_response(data={"items": items, "total": total, "page": page, "page_size": page_size})


@router.get("/performance-evaluations/my", summary="我的绩效考核")
async def my_performance_evaluations(
    month: str | None = Query(None, description="考核月份"),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """部门负责人：看自己的自评 / 分管领导：看所管部门的评估"""
    service = PerformanceEvaluationService(session)
    user_name = hr_scope.user.name or ""
    items, total = await service.list_evaluations(
        PerformanceListParams(evaluation_month=month, page=1, page_size=100),
        departments=hr_scope.scoped_departments if not hr_scope.is_unrestricted else None,
    )
    # 过滤：只显示当前用户是部门负责人或分管领导的记录
    filtered = [
        it for it in items
        if it["department_head"] == user_name or it.get("evaluator_leader") == user_name
    ]
    return success_response(data={"items": filtered, "total": len(filtered)})


@router.post("/performance-evaluations", summary="创建月度绩效考核")
async def create_performance_evaluation(
    payload: PerformanceEvaluationCreate,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    if payload.department:
        hr_scope.ensure_dept_writable([payload.department])
    service = PerformanceEvaluationService(session)
    try:
        result = await service.create_evaluation(payload)
        return success_response(data=result, message="考核记录创建成功", status_code=201)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/performance-evaluations/auto-create", summary="批量生成当月考核")
async def auto_create_performance_evaluations(
    month: str = Query(..., description="考核月份 YYYY-MM"),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """为授权范围内部门生成当月考核记录（已有则跳过）。"""
    if not hr_scope.is_unrestricted and not hr_scope.scoped_departments:
        raise ForbiddenException("数据范围限制：仅可访问本人相关数据")
    from app.modules.hr.models import HrDepartment as HrDept
    depts = (await session.execute(
        select(HrDept.name).where(HrDept.is_deleted == False)  # noqa: E712
    )).scalars().all()
    if not hr_scope.is_unrestricted:
        depts = [d for d in depts if d in hr_scope.scoped_departments]
    service = PerformanceEvaluationService(session)
    created = await service.auto_create_for_month(month, list(depts))
    return success_response(data={"created": len(created)}, message=f"已为 {len(created)} 个部门生成考核")


@router.get("/performance-evaluations/{evaluation_id}", summary="考核详情")
async def get_performance_evaluation(
    evaluation_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    service = PerformanceEvaluationService(session)
    result = await service.get_evaluation(evaluation_id)
    if result is None:
        raise HTTPException(status_code=404, detail="考核记录不存在")
    dept = result.get("department") if isinstance(result, dict) else getattr(result, "department", None)
    if dept:
        hr_scope.ensure_dept_writable([dept])
    # 领导评分入口标记：仅分管领导本人（或管理权限）可评，前端据此禁用输入与按钮
    if isinstance(result, dict):
        from app.platform.permission.deps import get_user_permissions

        perms = await get_user_permissions(str(hr_scope.user.id), session)
        is_manager = "hr:performance:manage" in perms
        leader_name = result.get("evaluator_leader")
        result["can_edit_leader"] = is_manager or (
            bool(leader_name) and leader_name == hr_scope.user.name
        )
    return success_response(data=result)


@router.put("/performance-evaluations/{evaluation_id}", summary="更新考核内容")
async def update_performance_evaluation(
    evaluation_id: UUID,
    payload: PerformanceEvaluationUpdate,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """更新考核基本信息及指标项（仅草稿状态可编辑；受限用户仅限授权部门）。"""
    service = PerformanceEvaluationService(session)
    current = await service.get_evaluation(evaluation_id)
    if current is None:
        raise HTTPException(status_code=404, detail="考核记录不存在")
    dept = current.get("department") if isinstance(current, dict) else getattr(current, "department", None)
    if dept:
        hr_scope.ensure_dept_writable([dept])
    try:
        result = await service.update_evaluation(evaluation_id, payload)
        return success_response(data=result, message="考核已更新")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/performance-evaluations/{evaluation_id}/submit-self", summary="提交自评")
async def submit_self_evaluation(
    evaluation_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    service = PerformanceEvaluationService(session)
    try:
        result = await service.submit_self(evaluation_id, hr_scope.user.name or "")
        return success_response(data=result, message="自评已提交")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except PermissionError as e:
        raise HTTPException(status_code=403, detail=str(e))


@router.post("/performance-evaluations/{evaluation_id}/submit-leader", summary="提交领导评分")
async def submit_leader_evaluation(
    evaluation_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    service = PerformanceEvaluationService(session)
    try:
        result = await service.submit_leader(evaluation_id, hr_scope.user.name or "")
        return success_response(data=result, message="领导评分已提交")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except PermissionError as e:
        raise HTTPException(status_code=403, detail=str(e))


# ─── 考核项目配置 ───

@router.get("/performance-categories", summary="考核项目列表")
async def list_performance_categories(session: AsyncSession = Depends(get_db)):
    repo = PerformanceEvaluationRepository(session)
    cats = await repo.list_categories()
    return success_response(data=[{
        "id": str(c.id), "name": c.name, "weight": c.weight,
        "evaluator": c.evaluator, "is_active": c.is_active,
        "sort_order": c.sort_order,
        "created_at": c.created_at.isoformat() if c.created_at else None,
    } for c in cats])


@router.post("/performance-categories", summary="创建考核项目")
async def create_performance_category(
    payload: PerformanceCategoryCreate, session: AsyncSession = Depends(get_db),
):
    from app.modules.hr.models import PerformanceCategory
    repo = PerformanceEvaluationRepository(session)
    cat = PerformanceCategory(name=payload.name, weight=payload.weight, evaluator=payload.evaluator, sort_order=payload.sort_order)
    cat = await repo.create_category(cat)
    return success_response(data={"id": str(cat.id)}, message="考核项目已创建", status_code=201)


@router.put("/performance-categories/{category_id}", summary="更新考核项目")
async def update_performance_category(
    category_id: UUID, payload: PerformanceCategoryUpdate, session: AsyncSession = Depends(get_db),
):
    repo = PerformanceEvaluationRepository(session)
    cat = await repo.get_category_by_id(category_id)
    if not cat: raise HTTPException(status_code=404, detail="考核项目不存在")
    for field in ("name", "weight", "evaluator", "is_active", "sort_order"):
        v = getattr(payload, field, None)
        if v is not None: setattr(cat, field, v)
    await repo.update_category(cat)
    return success_response(data={"id": str(cat.id)}, message="已更新")


@router.delete("/performance-categories/{category_id}", summary="删除考核项目")
async def delete_performance_category(category_id: UUID, session: AsyncSession = Depends(get_db)):
    repo = PerformanceEvaluationRepository(session)
    await repo.delete_category(category_id)
    return success_response(message="已删除")


@router.get("/performance-categories/{category_id}/dept-weights", summary="获取部门权重列表")
async def get_dept_weights(category_id: UUID, session: AsyncSession = Depends(get_db)):
    repo = PerformanceEvaluationRepository(session)
    dws = await repo.get_dept_weights(category_id)
    return success_response(data=[{"department": d.department, "weight": d.weight} for d in dws])


@router.post("/performance-categories/{category_id}/dept-weights", summary="保存部门权重")
async def save_dept_weights(
    category_id: UUID,
    payload: dict = Body(...),
    session: AsyncSession = Depends(get_db),
):
    repo = PerformanceEvaluationRepository(session)
    for item in payload.get("weights", []):
        await repo.upsert_dept_weight(category_id, item["department"], item["weight"])
    return success_response(message="部门权重已保存")


# ─── 考核项目评分 ───

@router.get("/performance-evaluations/{evaluation_id}/category-scores", summary="获取考核项目评分")
async def get_category_scores(
    evaluation_id: UUID,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """评分矩阵数据：每行带 can_edit（当前用户是否为该项目负责人/管理权限）。"""
    from app.platform.permission.deps import get_user_permissions

    repo = PerformanceEvaluationRepository(session)
    # 获取考核所属部门
    from app.modules.hr.models import MonthlyPerformanceEvaluation
    ev = (await session.execute(
        select(MonthlyPerformanceEvaluation).where(MonthlyPerformanceEvaluation.id == evaluation_id)
    )).scalar_one_or_none()
    dept = ev.department if ev else ""
    perms = await get_user_permissions(str(hr_scope.user.id), session)
    is_manager = "hr:performance:manage" in perms
    scores = await repo.get_category_scores(evaluation_id)
    cats = await repo.list_categories()
    score_map = {s.category_id: s for s in scores}
    result = []
    for cat in cats:
        if not cat.is_active:
            continue
        # 部门权重优先，否则用项目默认权重
        dw = await repo.get_dept_weight(cat.id, dept)
        weight = dw.weight if dw else cat.weight
        s = score_map.get(cat.id)
        result.append({
            "id": str(s.id) if s else "",
            "evaluation_id": str(evaluation_id),
            "category_id": str(cat.id),
            "category_name": cat.name,
            "evaluator": cat.evaluator,
            "can_edit": is_manager or (bool(cat.evaluator) and cat.evaluator == hr_scope.user.name),
            "weight": weight,
            "score": s.score if s else None,
            "scored_by": s.scored_by if s else None,
            "scored_at": s.scored_at.isoformat() if s and s.scored_at else None,
        })
    return success_response(data=result)


async def _recalc_evaluation_total(session: AsyncSession, evaluation_id: UUID) -> None:
    """重算月度考核部门加权总分：Σ(权重×分数)/Σ(权重)，仅统计已评分项目。

    部门×项目子集不同（A 部门有环保项目、B 部门没有），权重即该部门该
    项目的 PerformanceDeptWeight；未评分的项目不参与计算。
    """
    from app.modules.hr.models import (
        MonthlyPerformanceEvaluation,
        PerformanceCategoryScore,
    )

    rows = (await session.execute(
        select(PerformanceCategoryScore).where(
            PerformanceCategoryScore.evaluation_id == evaluation_id,
            PerformanceCategoryScore.is_deleted == False,  # noqa: E712
            PerformanceCategoryScore.score.isnot(None),
        )
    )).scalars().all()
    total_weight = 0.0
    weighted_sum = 0.0
    for r in rows:
        w = float(r.weight or 0)
        if w <= 0:
            continue
        total_weight += w
        weighted_sum += w * float(r.score or 0)
    final = round(weighted_sum / total_weight, 2) if total_weight > 0 else None
    ev = (await session.execute(
        select(MonthlyPerformanceEvaluation).where(
            MonthlyPerformanceEvaluation.id == evaluation_id
        )
    )).scalar_one_or_none()
    if ev:
        ev.final_score = final
        await session.flush()


@router.post("/performance-evaluations/{evaluation_id}/category-scores", summary="批量提交考核项目评分")
async def save_category_scores(
    evaluation_id: UUID, payload: CategoryScoreBatchInput,
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """仅项目负责人（或 hr:performance:manage 权限）可保存本项目评分。

    多人评分模型：每个考核项目有独立负责人，评分对象是部门；部门×项目
    子集由 PerformanceDeptWeight 决定（有的项目 A 部门有、B 部门没有）。
    权重一律以部门权重配置为准（忽略客户端传入值）。
    保存后自动重算该考核的部门加权总分 final_score。
    """
    from app.modules.hr.models import MonthlyPerformanceEvaluation, PerformanceCategory
    from app.platform.permission.deps import get_user_permissions

    repo = PerformanceEvaluationRepository(session)
    ev = (await session.execute(
        select(MonthlyPerformanceEvaluation).where(
            MonthlyPerformanceEvaluation.id == evaluation_id,
            MonthlyPerformanceEvaluation.is_deleted == False,  # noqa: E712
        )
    )).scalar_one_or_none()
    if not ev:
        raise HTTPException(404, "考核记录不存在")
    # 数据范围：受限用户仅可评授权部门的考核
    hr_scope.ensure_dept_writable([ev.department])

    perms = await get_user_permissions(str(hr_scope.user.id), session)
    is_manager = "hr:performance:manage" in perms
    category_ids = {s.category_id for s in payload.scores}
    categories = (await session.execute(
        select(PerformanceCategory).where(PerformanceCategory.id.in_(category_ids))
    )).scalars().all()
    cat_map = {c.id: c for c in categories}
    for s in payload.scores:
        cat = cat_map.get(s.category_id)
        if not cat:
            raise HTTPException(400, f"考核项目不存在: {s.category_id}")
        if not is_manager:
            if not cat.evaluator:
                raise HTTPException(403, f"「{cat.name}」未配置评分人，请联系管理员")
            if cat.evaluator != hr_scope.user.name:
                raise HTTPException(
                    403, f"「{cat.name}」由 {cat.evaluator} 负责评分，您无权提交"
                )
        # 权重以 PerformanceDeptWeight 配置为准，忽略客户端传入的 weight
        dw = await repo.get_dept_weight(s.category_id, ev.department)
        weight = dw.weight if dw else cat.weight
        await repo.upsert_category_score(
            evaluation_id, s.category_id, s.score, hr_scope.user.name, weight,
        )
    await _recalc_evaluation_total(session, evaluation_id)
    return success_response(message="评分已保存")


def _safe_sheet_name(name: str) -> str:
    """Excel 工作表名去非法字符并截断（≤31 字符）。"""
    import re as _re

    name = _re.sub(r"[\\/*?:\[\]]", "", name)[:31].strip()
    return name or "项目"


@router.get("/performance-reports", summary="绩效汇总报表（按项目一张表）")
async def export_performance_report(
    month: str = Query(..., description="考核月份 YYYY-MM"),
    session: AsyncSession = Depends(get_db),
    hr_scope: HrAccessContext = Depends(get_hr_scope),
):
    """汇总报表：一个工作簿，每个考核项目一个工作表（行=部门评分明细），
    另有「部门总分」总表。适配多人评分模型：部门×项目子集各不相同。
    数据范围受限用户只导出授权部门。"""
    from io import BytesIO

    from openpyxl import Workbook

    from app.modules.hr.models import (
        MonthlyPerformanceEvaluation,
        PerformanceCategory,
        PerformanceCategoryScore,
    )

    stmt = select(MonthlyPerformanceEvaluation).where(
        MonthlyPerformanceEvaluation.evaluation_month == month,
        MonthlyPerformanceEvaluation.is_deleted == False,  # noqa: E712
    )
    # 考核主表没有员工字段（actual_department/employee_number），
    # 不能用 apply_list_scope（会 AttributeError）；按部门手动收敛
    if not hr_scope.is_unrestricted:
        if hr_scope.data_scope == "self_only":
            raise ForbiddenException("数据范围限制：仅可访问本人相关数据")
        if not hr_scope.scoped_departments:
            raise ForbiddenException("数据范围限制：无法确定您的授权部门，请联系管理员")
        stmt = stmt.where(
            MonthlyPerformanceEvaluation.department.in_(hr_scope.scoped_departments)
        )
    evs = (await session.execute(stmt)).scalars().all()
    if not evs:
        raise HTTPException(404, f"{month} 无考核数据")
    ev_by_id = {e.id: e for e in evs}
    scores = (await session.execute(
        select(PerformanceCategoryScore).where(
            PerformanceCategoryScore.evaluation_id.in_(ev_by_id.keys()),
            PerformanceCategoryScore.is_deleted == False,  # noqa: E712
        )
    )).scalars().all()
    cats = (await session.execute(
        select(PerformanceCategory).where(
            PerformanceCategory.is_deleted == False  # noqa: E712
        )
    )).scalars().all()

    wb = Workbook()
    ws_total = wb.active
    ws_total.title = "部门总分"
    ws_total.append(["部门", "部门负责人", "加权总分"])
    for ev in evs:
        ws_total.append([ev.department, ev.department_head, ev.final_score])
    for cat in cats:
        ws = wb.create_sheet(title=_safe_sheet_name(cat.name))
        ws.append(["部门", "权重%", "分数", "评分人", "评分时间"])
        for s in scores:
            if s.category_id != cat.id:
                continue
            ev = ev_by_id.get(s.evaluation_id)
            ws.append([
                ev.department if ev else "",
                s.weight,
                s.score,
                s.scored_by,
                s.scored_at.strftime("%Y-%m-%d %H:%M") if s.scored_at else "",
            ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                "attachment; "
                f"filename*=UTF-8''{quote(f'绩效汇总_{month}.xlsx')}"
            )
        },
    )


