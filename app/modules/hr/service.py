"""HR business workflows live here."""

import logging
from datetime import date, datetime
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import DuplicateException, NotFoundException
from app.modules.hr.models import (
    AnnualTrainingPlan,
    AnnualTrainingPlanItem,
    DepartureRecord,
    Employee,
    HrDepartment,
    OffboardingRecord,
    OnboardingRecord,
    Team,
    TrainingLedger,
    TrainingLedgerPage,
)
from app.modules.hr.repository import (
    AnnualTrainingPlanItemRepository,
    AnnualTrainingPlanRepository,
    DepartmentRepository,
    DepartureRecordRepository,
    EmployeeRepository,
    OffboardingRecordRepository,
    OnboardingRecordRepository,
    TeamRepository,
    TrainingLedgerPageRepository,
    TrainingLedgerRepository,
)
from app.modules.hr.schemas import (
    AnnualTrainingPlanCreate,
    AnnualTrainingPlanItemBatchUpdate,
    AnnualTrainingPlanUpdate,
    DepartmentCreate,
    DepartmentUpdate,
    DepartureRecordCreate,
    DepartureRecordUpdate,
    EmployeeCreate,
    EmployeeUpdate,
    OffboardingRecordCreate,
    OffboardingRecordUpdate,
    TeamCreate,
    TeamUpdate,
    TrainingLedgerCreate,
    TrainingLedgerUpdate,
)

logger = logging.getLogger(__name__)

# ─── Services ───

class EmployeeService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = EmployeeRepository(session)

    async def get_employee(self, employee_id: UUID) -> Employee:
        employee = await self.repo.get_by_id(employee_id)
        if not employee:
            raise NotFoundException("员工", str(employee_id))
        return employee

    async def get_employee_by_number(self, employee_number: str) -> Employee:
        employee = await self.repo.get_by_employee_number(employee_number)
        if not employee:
            raise NotFoundException("员工", employee_number)
        return employee

    async def create_employee(self, data: EmployeeCreate) -> Employee:
        existing = await self.repo.get_by_employee_number(data.employee_number)
        if existing:
            raise DuplicateException("工号", data.employee_number)

        employee = Employee(**data.model_dump())
        employee.status = "在职"

        result = await self.repo.create(employee)
        return result

    # ── Excel 列名 → 模型字段名 映射 ──
    _UPLOAD_COLUMN_MAP: dict[str, str] = {
        "工号": "employee_number",
        "姓名": "name",
        "域账号": "domain_account",
        "部门": "department",
        "班组": "team",
        "职位": "position",
        "岗位类别": "job_category",
        "级别": "level",
        "兼任部门": "concurrent_departments",
        "资格": "qualifications",
        "资格类型": "qualification_type",
        "性别": "gender",
        "籍贯": "native_place",
        "政治面貌": "political_status",
        "婚姻状况": "marital_status",
        "户籍类型": "household_type",
        "用工性质": "status_category",
        "出生年份": "birth_year",
        "出生月份": "birth_month",
        "出生日": "birth_day",
        "年龄": "age",
        "参加工作时间": "work_start_date",
        "进厂时间": "factory_entry_date",
        "进丽珠时间": "livo_entry_date",
        "入职日期": "hire_date",
        "毕业时间": "graduation_date",
        "工龄": "work_years",
        "厂龄": "factory_tenure",
        "公司工龄": "company_tenure",
        "学历": "education",
        "分类": "classification",
        "毕业学校": "school",
        "专业": "major",
        "身份证号": "id_card",
        "身份证有效期": "id_card_expiry",
        "身份证地址": "id_card_address",
        "现居住地址": "current_address",
        "合同类型": "contract_type",
        "合同开始日期": "contract_start_date",
        "合同结束日期": "contract_end_date",
        "合同开始日期2": "contract_start_2",
        "合同结束日期2": "contract_end_2",
        "合同开始日期3": "contract_start_3",
        "合同结束日期3": "contract_end_3",
        "合同开始日期4": "contract_start_4",
        "合同结束日期4": "contract_end_4",
        "手机": "phone",
        "邮箱": "email",
        "紧急联系人": "emergency_contact_name",
        "紧急联系人电话": "emergency_contact_phone",
        "紧急联系人关系": "emergency_contact_relation",
        "银行卡号": "bank_account",
        "培训编号": "training_id",
        "调动历史": "transfer_history",
        "备注": "remarks",
    }

    _DATE_FIELDS: set[str] = {
        "work_start_date", "factory_entry_date", "livo_entry_date",
        "hire_date", "graduation_date", "id_card_expiry",
        "contract_start_date", "contract_end_date",
        "contract_start_2", "contract_end_2",
        "contract_start_3", "contract_end_3",
        "contract_start_4", "contract_end_4",
    }

    _INT_FIELDS: set[str] = {
        "birth_year", "birth_month", "birth_day", "age", "work_years",
    }

    # 常见日期格式（按优先级从高到低排序）
    _DATE_FORMATS: list[str] = [
        "%Y-%m-%d",       # 2024-01-15
        "%Y/%m/%d",       # 2024/01/15
        "%Y.%m.%d",       # 2024.01.15
        "%Y年%m月%d日",    # 2024年01月15日
        "%Y%m%d",         # 20240115
        "%Y-%m-%d %H:%M:%S",       # 2024-01-15 00:00:00
        "%Y/%m/%d %H:%M:%S",       # 2024/01/15 00:00:00
        "%d/%m/%Y",       # 15/01/2024
        "%m/%d/%Y",       # 01/15/2024
    ]

    @staticmethod
    def _parse_date_value(val: object) -> date | None:
        """将各种格式的日期值统一转换为 date 对象。返回 None 表示无法解析。"""
        from datetime import date as date_cls, datetime as datetime_cls

        if isinstance(val, datetime_cls):
            return val.date()
        if isinstance(val, date_cls):
            return val
        if isinstance(val, (int, float)):
            # Excel 日期序列号（以 1899-12-30 为第 0 天）
            try:
                from datetime import timedelta
                excel_epoch = date_cls(1899, 12, 30)
                return excel_epoch + timedelta(days=int(val))
            except (ValueError, OverflowError):
                return None
        if isinstance(val, str):
            val = val.strip()
            if not val:
                return None
            for fmt in EmployeeService._DATE_FORMATS:
                try:
                    return datetime_cls.strptime(val, fmt).date()
                except ValueError:
                    continue
        return None

    async def upload_employees(self, file_bytes: bytes) -> dict:
        """从 Excel 文件批量导入员工，按工号 upsert。返回 {created, updated, errors}。"""
        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("文件为空")

        header = [str(c).strip() if c else "" for c in rows[0]]
        # 建立列索引
        col_map: dict[int, str] = {}
        for idx, col_name in enumerate(header):
            field = self._UPLOAD_COLUMN_MAP.get(col_name)
            if field:
                col_map[idx] = field

        if "employee_number" not in col_map.values():
            raise ValueError("缺少「工号」列，无法导入")

        created = 0
        updated = 0
        errors: list[str] = []

        for row_idx, row in enumerate(rows[1:], start=2):
            if all(c is None for c in row):
                continue
            try:
                data: dict = {}
                for col_idx, field_name in col_map.items():
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is None or (isinstance(val, str) and val.strip() == ""):
                        continue
                    if isinstance(val, str):
                        val = val.strip()
                    if field_name in self._DATE_FIELDS:
                        parsed = self._parse_date_value(val)
                        if parsed is None:
                            continue  # 无法解析的日期跳过
                        val = parsed
                    elif field_name in self._INT_FIELDS:
                        try:
                            val = int(float(str(val)))
                        except (ValueError, TypeError):
                            continue
                    elif field_name == "qualifications":
                        val = [v.strip() for v in str(val).split(",") if v.strip()]
                    data[field_name] = val

                # 必填字段默认值
                data.setdefault("position", "")
                data.setdefault("department", "")

                if "employee_number" not in data:
                    errors.append(f"第{row_idx}行: 缺少工号")
                    continue

                existing = await self.repo.get_by_employee_number(data["employee_number"])
                if existing:
                    await self.repo.upsert_by_employee_number(data)
                    updated += 1
                else:
                    if "hire_date" not in data:
                        data["hire_date"] = date.today()
                    if "status" not in data:
                        data["status"] = "待审批"
                    await self.repo.upsert_by_employee_number(data)
                    created += 1
            except Exception as e:
                errors.append(f"第{row_idx}行: {e}")

        return {"created": created, "updated": updated, "errors": errors}

    async def approve_employee(self, employee_number: str) -> Employee:
        employee = await self.repo.get_by_employee_number(employee_number)
        if not employee:
            raise NotFoundException("员工", employee_number)
        if employee.status != "待审批":
            raise DuplicateException("审批", "该员工已审批完成")

        employee.status = "在职"
        result = await self.repo.update(employee)
        return result

    async def update_employee(self, employee_id: UUID, data: EmployeeUpdate) -> Employee:
        employee = await self.get_employee(employee_id)
        update_data = data.model_dump(exclude_unset=True)

        if "employee_number" in update_data:
            existing = await self.repo.get_by_employee_number(update_data["employee_number"])
            if existing and existing.id != employee_id:
                raise DuplicateException("工号", update_data["employee_number"])

        for field, value in update_data.items():
            setattr(employee, field, value)

        result = await self.repo.update(employee)
        return result

    async def delete_employee(self, employee_id: UUID) -> None:
        employee = await self.get_employee(employee_id)
        await self.repo.soft_delete(employee)

    async def list_employees(
        self,
        *,
        department: str | None = None,
        status: str | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
        sort_by: str = "sort_order",
        sort_order: str = "asc",
    ) -> tuple[list[Employee], int]:
        return await self.repo.list_employees(
            department=department,
            status=status,
            keyword=keyword,
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            sort_order=sort_order,
        )

    # ── 年度培训计划上传 ──

    _PLAN_COLUMN_MAP: dict[str, str] = {
        "年度": "year",
        "年份": "year",
        "部门": "department",
        "月份": "month",
        "培训人数": "trainee_count",
        "课时": "duration_hours",
        "培训内容及使用教材": "content_and_textbook",
        "培训内容": "content_and_textbook",
        "培训对象": "target_audience",
        "参加岗位/参加人数": "position_and_count",
        "培训方式": "training_method",
        "培训学时": "training_hours",
        "确认者": "confirmer",
    }

    async def upload_annual_plan(self, file_bytes: bytes) -> dict:
        """从 Excel 批量导入年度培训计划，按年度+部门自动分类。"""
        from io import BytesIO
        from openpyxl import load_workbook
        from app.modules.hr.models import AnnualTrainingPlan, AnnualTrainingPlanItem

        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("文件为空")
        header = [str(c).strip() if c else "" for c in rows[0]]
        col_map = {idx: self._PLAN_COLUMN_MAP[h] for idx, h in enumerate(header) if h in self._PLAN_COLUMN_MAP}

        plan_cache: dict[tuple, AnnualTrainingPlan] = {}
        created, updated, errors = 0, 0, []

        for row_idx, row in enumerate(rows[1:], start=2):
            if all(c is None for c in row):
                continue
            try:
                data = {}
                for ci, fn in col_map.items():
                    v = row[ci] if ci < len(row) else None
                    if v is None or (isinstance(v, str) and v.strip() == ""):
                        continue
                    if isinstance(v, str):
                        v = v.strip()
                    if fn == "year":
                        v = int(float(str(v)))
                    if fn in ("trainee_count",):
                        v = int(float(str(v)))
                    data[fn] = v

                dept = data.get("department", "")
                year = data.get("year", 2026)
                if not dept:
                    errors.append(f"第{row_idx}行: 缺少部门"); continue

                # 找或创建年度计划
                cache_key = (year, dept)
                plan = plan_cache.get(cache_key)
                if not plan:
                    q = select(AnnualTrainingPlan).where(
                        AnnualTrainingPlan.year == year,
                        AnnualTrainingPlan.department == dept,
                        AnnualTrainingPlan.is_deleted == False,
                    )
                    r = await self.repo.session.execute(q)
                    plan = r.scalar_one_or_none()
                    if not plan:
                        plan = AnnualTrainingPlan(year=year, department=dept, status="草稿")
                        self.repo.session.add(plan)
                        await self.repo.session.flush()
                    plan_cache[cache_key] = plan

                # 添加计划项
                item_data = {k: v for k, v in data.items() if k not in ("year", "department")}
                item = AnnualTrainingPlanItem(plan_id=plan.id, **item_data)
                self.repo.session.add(item)
                await self.repo.session.flush()
                created += 1
            except Exception as e:
                errors.append(f"第{row_idx}行: {e}")

        return {"created": created, "updated": updated, "errors": errors}

class DepartmentService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = DepartmentRepository(session)

    async def get_department(self, department_id: UUID) -> HrDepartment:
        department = await self.repo.get_by_id(department_id)
        if not department:
            raise NotFoundException("部门", str(department_id))
        return department

    async def create_department(self, data: DepartmentCreate) -> HrDepartment:
        existing = await self.repo.get_by_code(data.code)
        if existing:
            raise DuplicateException("部门编码", data.code)

        department = HrDepartment(**data.model_dump())
        result = await self.repo.create(department)
        return result

    async def update_department(self, department_id: UUID, data: DepartmentUpdate) -> HrDepartment:
        department = await self.get_department(department_id)
        update_data = data.model_dump(exclude_unset=True)

        if "code" in update_data:
            existing = await self.repo.get_by_code(update_data["code"])
            if existing and existing.id != department_id:
                raise DuplicateException("部门编码", update_data["code"])

        for field, value in update_data.items():
            setattr(department, field, value)

        result = await self.repo.update(department)
        return result

    async def delete_department(self, department_id: UUID) -> None:
        department = await self.get_department(department_id)
        code = department.code
        await self.repo.soft_delete(department)

    async def list_departments(
        self,
        *,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[HrDepartment], int]:
        departments, total = await self.repo.list_departments(
            keyword=keyword,
            page=page,
            page_size=page_size,
        )
        # Attach employee count to each department
        from sqlalchemy import select, func
        from app.modules.hr.models import Employee
        for dept in departments:
            count = await self.repo.session.scalar(
                select(func.count()).select_from(Employee).where(
                    Employee.department == dept.name,
                    Employee.is_deleted.is_(False),
                )
            )
            dept.employee_count = count or 0
        return departments, total

class TeamService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = TeamRepository(session)
        self.department_repo = DepartmentRepository(session)

    async def get_team(self, team_id: UUID) -> Team:
        team = await self.repo.get_by_id(team_id)
        if not team:
            raise NotFoundException("班组", str(team_id))
        return team

    async def create_team(self, data: TeamCreate) -> Team:
        department = await self.department_repo.get_by_id(data.department_id)
        if not department:
            raise NotFoundException("部门", str(data.department_id))

        team = Team(**data.model_dump())
        result = await self.repo.create(team)
        return result

    async def update_team(self, team_id: UUID, data: TeamUpdate) -> Team:
        team = await self.get_team(team_id)
        update_data = data.model_dump(exclude_unset=True)

        if "department_id" in update_data:
            department = await self.department_repo.get_by_id(update_data["department_id"])
            if not department:
                raise NotFoundException("部门", str(update_data["department_id"]))

        for field, value in update_data.items():
            setattr(team, field, value)

        result = await self.repo.update(team)
        return result

    async def delete_team(self, team_id: UUID) -> None:
        team = await self.get_team(team_id)
        await self.repo.soft_delete(team)

    async def list_teams(
        self,
        *,
        department_id: UUID | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[Team], int]:
        return await self.repo.list_teams(
            department_id=department_id,
            keyword=keyword,
            page=page,
            page_size=page_size,
        )

class OffboardingRecordService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = OffboardingRecordRepository(session)
        self.employee_repo = EmployeeRepository(session)

    async def get_record(self, record_id: UUID) -> OffboardingRecord:
        record = await self.repo.get_by_id(record_id)
        if not record:
            raise NotFoundException("离职记录", str(record_id))
        return record

    async def create_record(self, data: OffboardingRecordCreate) -> OffboardingRecord:
        employee = await self.employee_repo.get_by_id(data.employee_id)
        if not employee:
            raise NotFoundException("员工", str(data.employee_id))

        record = OffboardingRecord(**data.model_dump())
        record = await self.repo.create(record)

        # 自动将员工状态更新为离职
        employee.status = "离职"
        await self.employee_repo.update(employee)

        return record

    async def update_record(self, record_id: UUID, data: OffboardingRecordUpdate) -> OffboardingRecord:
        record = await self.get_record(record_id)
        update_data = data.model_dump(exclude_unset=True)

        for field, value in update_data.items():
            setattr(record, field, value)

        result = await self.repo.update(record)

        return result

    async def delete_record(self, record_id: UUID) -> None:
        record = await self.get_record(record_id)
        await self.repo.soft_delete(record)

    async def list_records(
        self,
        *,
        employee_id: UUID | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[OffboardingRecord], int]:
        return await self.repo.list_records(
            employee_id=employee_id,
            keyword=keyword,
            page=page,
            page_size=page_size,
        )

class OnboardingRecordService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = OnboardingRecordRepository(session)

    async def get_record(self, record_id: UUID) -> OnboardingRecord:
        record = await self.repo.get_by_id(record_id)
        if not record:
            raise NotFoundException("入职记录", str(record_id))
        return record

    async def list_records(
        self,
        *,
        department: str | None = None,
        position: str | None = None,
        is_employed: str | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
        sort_by: str = "hire_date",
        sort_order: str = "desc",
    ) -> tuple[list[OnboardingRecord], int]:
        return await self.repo.list_records(
            department=department,
            position=position,
            is_employed=is_employed,
            keyword=keyword,
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            sort_order=sort_order,
        )

class DepartureRecordService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = DepartureRecordRepository(session)

    async def get_record(self, record_id: UUID) -> DepartureRecord:
        record = await self.repo.get_by_id(record_id)
        if not record:
            raise NotFoundException("离职台账记录", str(record_id))
        return record

    async def list_records(
        self,
        *,
        department: str | None = None,
        offboarding_type: str | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
        sort_by: str = "offboarding_date",
        sort_order: str = "desc",
    ) -> tuple[list[DepartureRecord], int]:
        return await self.repo.list_records(
            department=department,
            offboarding_type=offboarding_type,
            keyword=keyword,
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            sort_order=sort_order,
        )

    async def create_record(self, data: DepartureRecordCreate) -> DepartureRecord:
        record = DepartureRecord(**data.model_dump())
        return await self.repo.create(record)

    async def update_record(self, record_id: UUID, data: DepartureRecordUpdate) -> DepartureRecord:
        record = await self.get_record(record_id)
        update_data = data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(record, field, value)
        return await self.repo.update(record)

    async def delete_record(self, record_id: UUID) -> None:
        record = await self.get_record(record_id)
        await self.repo.soft_delete(record)

class TrainingLedgerService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = TrainingLedgerRepository(session)

    async def get_record(self, record_id: UUID) -> TrainingLedger:
        record = await self.repo.get_by_id(record_id)
        if not record:
            raise NotFoundException("培训台账记录", str(record_id))
        return record

    async def create_record(self, data: TrainingLedgerCreate) -> TrainingLedger:
        # 去重：同员工+同日期+同主题视为重复
        existing = await self.repo.session.execute(
            select(TrainingLedger).where(
                TrainingLedger.employee_number == data.employee_number,
                TrainingLedger.training_date == data.training_date,
                TrainingLedger.training_subject == data.training_subject,
                TrainingLedger.is_deleted == False,
            )
        )
        dup = existing.scalar_one_or_none()
        if dup:
            return dup
        record = TrainingLedger(**data.model_dump())
        return await self.repo.create(record)

    async def update_record(
        self, record_id: UUID, data: TrainingLedgerUpdate
    ) -> TrainingLedger:
        record = await self.get_record(record_id)
        update_data = data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(record, field, value)
        return await self.repo.update(record)

    async def delete_record(self, record_id: UUID) -> None:
        record = await self.get_record(record_id)
        await self.repo.soft_delete(record)

    async def list_records(
        self,
        *,
        employee_number: str | None = None,
        date_from: date | None = None,
        date_to: date | None = None,
        page: int = 1,
        page_size: int = 20,
        sort_by: str = "training_date",
        sort_order: str = "asc",
    ) -> tuple[list[TrainingLedger], int]:
        return await self.repo.list_records(
            employee_number=employee_number,
            date_from=date_from,
            date_to=date_to,
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            sort_order=sort_order,
        )

    async def create_from_notification(
        self,
        *,
        employee_number: str,
        training_date: date,
        training_subject: str,
        training_method: str | None,
        trainer: str | None,
        source_id: str | None = None,
    ) -> TrainingLedger | None:
        """当培训通知包含特定员工时，自动创建培训台账记录。"""
        if source_id:
            existing = await self.repo.get_by_source("notification", source_id)
            if existing:
                return existing

        record = TrainingLedger(
            employee_number=employee_number,
            training_date=training_date,
            training_subject=training_subject,
            training_method=training_method,
            trainer=trainer,
            source_type="notification",
            source_id=source_id,
        )
        return await self.repo.create(record)

class TrainingLedgerPageService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = TrainingLedgerPageRepository(session)

    async def list_pages(self) -> list[TrainingLedgerPage]:
        return await self.repo.list_pages()

    async def list_pages_with_department(self) -> list[tuple[TrainingLedgerPage, str | None]]:
        return await self.repo.list_pages_with_department()

    async def create_page(self, data) -> TrainingLedgerPage:
        existing = await self.repo.get_by_employee_number(data.employee_number)
        if existing:
            raise DuplicateException("培训台账页面", data.employee_number)
        page = TrainingLedgerPage(**data.model_dump())
        return await self.repo.create(page)

class AnnualTrainingPlanService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = AnnualTrainingPlanRepository(session)
        self.item_repo = AnnualTrainingPlanItemRepository(session)

    async def get_plan(self, plan_id: UUID) -> AnnualTrainingPlan:
        plan = await self.repo.get_by_id(plan_id)
        if not plan:
            raise NotFoundException("年度培训计划", str(plan_id))
        return plan

    async def create_plan(self, data: AnnualTrainingPlanCreate) -> AnnualTrainingPlan:
        existing = await self.repo.get_by_year_and_department(data.year, data.department)
        if existing:
            raise DuplicateException("年度培训计划", f"{data.year}年-{data.department}")
        plan = AnnualTrainingPlan(**data.model_dump())
        return await self.repo.create(plan)

    async def update_plan(self, plan_id: UUID, data: AnnualTrainingPlanUpdate) -> AnnualTrainingPlan:
        plan = await self.get_plan(plan_id)
        update_data = data.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(plan, field, value)
        return await self.repo.update(plan)

    async def delete_plan(self, plan_id: UUID) -> None:
        plan = await self.get_plan(plan_id)
        await self.repo.soft_delete(plan)

    async def list_plans(
        self,
        *,
        year: int | None = None,
        department: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[AnnualTrainingPlan], int]:
        return await self.repo.list_plans(
            year=year,
            department=department,
            page=page,
            page_size=page_size,
        )

class AnnualTrainingPlanItemService:
    def __init__(self, session: AsyncSession) -> None:
        self.repo = AnnualTrainingPlanItemRepository(session)
        self.plan_repo = AnnualTrainingPlanRepository(session)

    async def list_items(self, plan_id: UUID) -> list[AnnualTrainingPlanItem]:
        return await self.repo.list_items(plan_id)

    async def batch_update_items(
        self, plan_id: UUID, data: AnnualTrainingPlanItemBatchUpdate
    ) -> list[AnnualTrainingPlanItem]:
        plan = await self.plan_repo.get_by_id(plan_id)
        if not plan:
            raise NotFoundException("年度培训计划", str(plan_id))

        # 删除旧明细
        await self.repo.delete_by_plan_id(plan_id)

        # 创建新明细
        results: list[AnnualTrainingPlanItem] = []
        for idx, item_data in enumerate(data.items):
            item = AnnualTrainingPlanItem(
                plan_id=plan_id,
                sort_order=idx,
                **item_data.model_dump(exclude={"sort_order"}),
            )
            created = await self.repo.create(item)
            results.append(created)
        return results
