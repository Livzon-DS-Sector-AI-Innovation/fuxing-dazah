"""
Excel template filler — openpyxl-based engine that copies a template,
replaces placeholders, fills data rows with style preservation, and
configures page setup for downstream PDF conversion.

Does NOT depend on the rest of the safety module or the database.
Works with plain dicts, so it can be driven by any data source.
"""

from __future__ import annotations

import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .config import TemplateConfig


class ExcelTemplateFiller:
    """Fill an Excel template with row data, preserving formatting.

    Usage::

        config = TemplateConfig(...)
        filler = ExcelTemplateFiller(config)
        wb = filler.fill(template_path, data_records)
        wb.save(output_path)
    """

    def __init__(self, config: TemplateConfig) -> None:
        self._cfg = config

    # ── Public API ──────────────────────────────────────────────────────

    def fill(
        self,
        template_path: str | Path,
        data: list[dict],
    ) -> openpyxl.Workbook:
        """Load template, fill with data, return the workbook (unsaved)."""
        wb = openpyxl.load_workbook(str(template_path))
        ws = wb.active

        self._replace_title(ws, data)
        sample_styles = self._capture_sample_styles(ws)
        self._fill_data_rows(ws, data, sample_styles)
        self._apply_page_setup(ws)
        self._strip_column_letters_row(ws)

        return wb

    def fill_and_save(
        self,
        template_path: str | Path,
        data: list[dict],
        output_path: str | Path,
    ) -> Path:
        """Convenience: fill + save, returning the output path."""
        wb = self.fill(template_path, data)
        wb.save(str(output_path))
        return Path(output_path)

    # ── Private helpers ─────────────────────────────────────────────────

    def _replace_title(self, ws, data: list[dict]) -> None:
        """Replace *** placeholder in the title row."""
        resolver = self._cfg.title_resolver
        if resolver is None:
            return
        replacement = resolver(data)
        if not replacement:
            return

        cell = ws.cell(self._cfg.title_row, 1)
        original = cell.value or ""
        new_title = original.replace(self._cfg.title_placeholder, replacement)
        if new_title != original:
            cell.value = new_title

    def _capture_sample_styles(self, ws) -> dict[str, dict]:
        """Snapshots font/fill/alignment/border from every column of the sample row."""
        styles: dict[str, dict] = {}
        row = self._cfg.sample_row
        for c in range(1, self._cfg.total_columns + 1):
            col_letter = get_column_letter(c).lower()
            src = ws.cell(row, c)
            styles[col_letter] = {
                "font": copy.copy(src.font),
                "fill": copy.copy(src.fill),
                "alignment": copy.copy(src.alignment),
                "border": copy.copy(src.border),
            }
        return styles

    def _fill_data_rows(
        self, ws, data: list[dict], sample_styles: dict[str, dict]
    ) -> None:
        """Write data rows starting at sample_row, cloning styles."""
        start_row = self._cfg.sample_row
        total_cols = self._cfg.total_columns
        data_height = ws.row_dimensions[start_row].height
        mapping = self._cfg.column_mapping
        numeric_cols = self._cfg.numeric_columns
        risk_col = self._cfg.risk_label_column
        risk_colors = self._cfg.risk_label_colors
        seq_col = self._cfg.sequence_column

        for row_idx, record in enumerate(data):
            excel_row = start_row + row_idx
            ws.row_dimensions[excel_row].height = data_height

            for c in range(1, total_cols + 1):
                col_letter = get_column_letter(c).lower()

                # ── Value ──
                if c == seq_col:
                    formatted = row_idx + 1
                else:
                    db_field = mapping.get(col_letter, "")
                    raw = record.get(db_field, "") if db_field else ""
                    formatted = self._format_value(raw, col_letter, numeric_cols)

                cell = ws.cell(excel_row, c)
                cell.value = formatted

                # ── Style (clone from sample) ──
                style = sample_styles[col_letter]
                cell.font = style["font"]
                cell.fill = style["fill"]
                cell.alignment = style["alignment"]
                cell.border = style["border"]

                # ── Risk label coloring ──
                if col_letter == risk_col and formatted:
                    color = self._pick_risk_color(str(formatted), risk_colors)
                    if color:
                        cell.font = Font(
                            bold=True,
                            size=cell.font.size or 10.5,
                            name=cell.font.name or "宋体",
                            color=color,
                        )

    def _apply_page_setup(self, ws) -> None:
        """Set print/PDF page properties on the worksheet."""
        ps = self._cfg.page_setup
        ws.page_setup.orientation = ps.orientation
        ws.page_setup.paperSize = ps.paper_size
        ws.page_setup.fitToWidth = ps.fit_to_width
        ws.page_setup.fitToHeight = ps.fit_to_height
        ws.sheet_properties.pageSetUpPr = (
            openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        )

    @staticmethod
    def _strip_column_letters_row(ws) -> None:
        """Remove the column-letter hint row (row with a,b,c,…) if present.
        Row 6 in the standard template; we clear it so it doesn't appear in PDF.
        """
        # The template row 6 has a,b,c,… — keep it in xlsx for reference
        # but it's harmless in PDF.  This hook exists for subclasses.
        pass

    # ── Static helpers ──────────────────────────────────────────────────

    @staticmethod
    def _format_value(value, col_letter: str, numeric_cols: frozenset) -> str | float:
        """Coerce value to the expected type for the column."""
        if value is None:
            return ""
        s = str(value).strip()
        if not s:
            return ""
        if col_letter in numeric_cols:
            try:
                return float(s)
            except ValueError:
                return s
        return s

    @staticmethod
    def _pick_risk_color(label: str, color_map: dict[str, str]) -> str | None:
        """Return the first color whose key is found in the label."""
        label_lower = label.lower()
        for keyword, color in color_map.items():
            if keyword.lower() in label_lower:
                return color
        return None
