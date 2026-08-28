"""检定到期提醒 API 端点。"""

from __future__ import annotations

from fastapi import Depends, Query
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.response import success_response
from app.modules.meter import service
from app.modules.meter.api._router import router

# ═══════════════════════════════════════════
# 检定到期提醒
# ═══════════════════════════════════════════


@router.get("/calibration/alerts", summary="检定到期提醒")
async def calibration_alerts(
    days_before: int = Query(default=30, ge=0, le=365, description="0=截止今天(含超期), >0=未来N天内到期"),
    department: str | None = Query(default=None, description="部门筛选"),
    source: str | None = Query(default=None, pattern="^(instrument|gas_detector)$", description="数据源筛选"),
    db: AsyncSession = Depends(get_db),
) -> JSONResponse:
    alerts = await service.get_calibration_alerts(db, days_before=days_before, department=department, source=source)
    return success_response(alerts)



@router.get("/calibration/alerts/export-excel", summary="导出检定到期提醒为 Excel")
async def export_calibration_alerts_excel(
    days_before: int = Query(default=30, ge=0, le=365, description="0=截止今天(含超期), >0=未来N天内到期"),
    department: str | None = Query(default=None, description="部门筛选"),
    source: str | None = Query(default=None, pattern="^(instrument|gas_detector)$", description="数据源筛选"),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    import io as io_mod

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    alerts = await service.get_calibration_alerts(db, days_before=days_before, department=department, source=source)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "检定到期提醒"

    headers = ["来源", "编号", "名称", "位置", "部门", "下次检定日期", "距到期天数"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, a in enumerate(alerts, 2):
        source_label = "计量器具" if a["source"] == "instrument" else "探测器"
        days_text = ""
        if a["days_until_due"] is not None:
            d = a["days_until_due"]
            if d < 0:
                days_text = f"已过期 {abs(d)} 天"
            elif d == 0:
                days_text = "今天到期"
            else:
                days_text = f"{d} 天"
        values = [
            source_label,
            a.get("serial_number", ""),
            a.get("instrument_name", ""),
            a.get("location", ""),
            a.get("department", ""),
            a["next_calibration_date"].isoformat() if a.get("next_calibration_date") else "",
            days_text,
        ]
        for col_idx, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row_idx in range(2, len(alerts) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)  # type: ignore[union-attr]

    buf = io_mod.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=calibration_alerts_export.xlsx"},
    )
